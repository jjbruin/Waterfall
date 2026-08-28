"""NAV engine — hypothetical liquidation at the valuation date (Phase 3).

The NAV calculation is the deal's Cap waterfall run once, at the cycle's
as-of date, on liquidation proceeds anchored to the actual balance sheet:

    concluded value
    - debt              (ISBS Interim BS debt accounts at the BS snapshot)
    + current assets    (curated ISBS BS lines — suggested, AM-editable)
    - current liabilities
    = net proceeds to distribute
    -> Cap_WF walk      (accrued pref injected from the Excel-validated pref
                         walk, capital from seeded accounting states, IRR
                         lookbacks from the deal's own IRR steps)
    = PSC NAV / OP NAV

Decisions encoded (Aug 28, 2026):
  - BS curation: the app SUGGESTS current asset/liability lines (assets
    1000-1199; liabilities 2000-2149 excluding debt accounts); the asset
    manager includes/excludes per line. Selections persist per record, carry
    forward from the prior cycle, and YoY treatment changes are flagged.
  - Cost-basis deals: value derived from accounting capital balances plus
    accrued pref at the valuation date (plus debt, for the property value).
  - Portfolios: children are valued per property; NAV runs ONCE at the
    parent on rolled-up value and a consolidated balance sheet.
  - Dual-tranche PE falls out of waterfall step order (IRR gates included).
  - Pref amounts: Excel-exact walk (build_pref_balance_detail — Act/Act,
    12/31 compounding) injected into the seeded states so NAV_Calc ties the
    Accrued_Pref / OP_Pref tabs to the penny.
  - Publish: the app is the system of record — approved NAVs write the
    `valuations` row and the Val_IS_{year} forecast directly.
"""

from __future__ import annotations

import json
import logging
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

import pandas as pd
from sqlalchemy import text

from config import DEBT_BS_ACCTS

logger = logging.getLogger(__name__)

CAP_WF = "Cap_WF"


# ============================================================
# Balance sheet curation
# ============================================================

def _default_bs_include(account_type: str, account: str) -> bool:
    """Suggested inclusion for the NAV's current asset / liability adjustment."""
    try:
        acct_num = int(str(account).strip())
    except (TypeError, ValueError):
        return False
    at = str(account_type or "").strip().lower()
    if at == "assets":
        return 1000 <= acct_num < 1200
    if at == "liabilities":
        return 2000 <= acct_num < 2150 and str(acct_num) not in DEBT_BS_ACCTS
    return False


def _load_selections(engine, record_id: int) -> Dict[str, bool]:
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT account, included FROM valuation_bs_selections WHERE record_id = :r
        """), {"r": record_id}).fetchall()
    return {str(r[0]): bool(r[1]) for r in rows}


def _prior_cycle_selections(engine, record_id: int) -> Dict[str, bool]:
    """Carry-forward: the same deal's stored selections from the latest prior cycle."""
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT r.vcode, c.year FROM valuation_records r
            JOIN valuation_cycles c ON c.id = r.cycle_id WHERE r.id = :r
        """), {"r": record_id}).fetchone()
        if not row:
            return {}
        prior = conn.execute(text("""
            SELECT s.account, s.included
            FROM valuation_bs_selections s
            JOIN valuation_records pr ON pr.id = s.record_id
            JOIN valuation_cycles pc ON pc.id = pr.cycle_id
            WHERE pr.vcode = :v AND pc.year < :y
            ORDER BY pc.year DESC
        """), {"v": row[0], "y": int(row[1])}).fetchall()
    out: Dict[str, bool] = {}
    for acct, inc in prior:  # newest cycle first; keep first occurrence
        out.setdefault(str(acct), bool(inc))
    return out


def save_bs_selections(engine, record_id: int, selections: Dict[str, Any],
                       username: str) -> Dict[str, Any]:
    from flask_app.services.valuation_service import (
        ensure_valuation_tables, _require_not_approved, _now)
    ensure_valuation_tables(engine)
    _require_not_approved(engine, record_id)
    with engine.begin() as conn:
        for account, included in (selections or {}).items():
            result = conn.execute(text("""
                UPDATE valuation_bs_selections
                SET included = :inc, updated_by = :u, updated_at = :now
                WHERE record_id = :r AND account = :a
            """), {"inc": 1 if included else 0, "u": username, "now": _now(),
                   "r": record_id, "a": str(account)})
            if result.rowcount == 0:
                conn.execute(text("""
                    INSERT INTO valuation_bs_selections (record_id, account, included, updated_by)
                    VALUES (:r, :a, :inc, :u)
                """), {"r": record_id, "a": str(account),
                       "inc": 1 if included else 0, "u": username})
    return {"status": "saved", "count": len(selections or {})}


# ============================================================
# Inputs — record, family, balance sheet, value resolution
# ============================================================

def _load_record(engine, record_id: int) -> Dict[str, Any]:
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT r.*, c.year AS cycle_year, c.as_of_date
            FROM valuation_records r JOIN valuation_cycles c ON c.id = r.cycle_id
            WHERE r.id = :i
        """), {"i": record_id}).fetchone()
        if not row:
            raise ValueError(f"Valuation record {record_id} not found")
        rec = dict(row._mapping)
        children = conn.execute(text("""
            SELECT vcode, concluded_value, method, status
            FROM valuation_records
            WHERE cycle_id = :c AND parent_vcode = :v
        """), {"c": rec["cycle_id"], "v": rec["vcode"]}).fetchall()
    rec["children"] = [dict(ch._mapping) for ch in children]
    return rec


def _bs_snapshot(data: dict, vcodes: List[str], as_of: pd.Timestamp):
    """Consolidated Interim BS lines across vcodes at each vcode's latest
    reported date on or before as_of. Returns (lines, snapshot_dates)."""
    from flask_app.services.financials_service import _prepare_isbs

    frames = []
    snapshot_dates: Dict[str, Optional[str]] = {}
    for vc in vcodes:
        isbs = _prepare_isbs(data.get("isbs_raw"), vc)
        if isbs.empty:
            snapshot_dates[vc] = None
            continue
        bs = isbs[isbs["vSource"] == "Interim BS"]
        if bs.empty:
            snapshot_dates[vc] = None
            continue
        periods = sorted(bs["dtEntry_parsed"].dropna().unique())
        snap = next((pd.Timestamp(p) for p in reversed(periods) if pd.Timestamp(p) <= as_of), None)
        if snap is None:
            snap = pd.Timestamp(periods[-1])
        snapshot_dates[vc] = snap.strftime("%Y-%m-%d")
        frames.append(bs[bs["dtEntry_parsed"] == snap])

    if not frames:
        return [], snapshot_dates
    allbs = pd.concat(frames, ignore_index=True)
    grouped = allbs.groupby(["vAccountType", "vAccount", "vDescription"], dropna=False)["mAmount"].sum().reset_index()
    type_order = {"Assets": 0, "Liabilities": 1, "Equity": 2}
    grouped["_ord"] = grouped["vAccountType"].map(lambda t: type_order.get(str(t), 3))
    grouped = grouped.sort_values(["_ord", "vAccount"])

    lines = []
    for _, r in grouped.iterrows():
        amt = float(r["mAmount"])
        if abs(amt) < 0.005:
            continue
        lines.append({
            "account_type": str(r["vAccountType"]) if pd.notna(r["vAccountType"]) else "",
            "account": str(r["vAccount"]) if pd.notna(r["vAccount"]) else "",
            "description": str(r["vDescription"]) if pd.notna(r["vDescription"]) else "",
            "amount": amt,
        })
    return lines, snapshot_dates


def get_nav_inputs(engine, record_id: int, data: dict) -> Dict[str, Any]:
    """The NAV tab's editable inputs: curated BS lines (with suggestion /
    carry-forward / change flags), debt, and the value resolution."""
    from flask_app.services.valuation_service import ensure_valuation_tables
    ensure_valuation_tables(engine)

    rec = _load_record(engine, record_id)
    as_of = pd.Timestamp(rec["as_of_date"])
    vcodes = [rec["vcode"]] + [c["vcode"] for c in rec["children"]]

    lines, snapshot_dates = _bs_snapshot(data, vcodes, as_of)
    stored = _load_selections(engine, record_id)
    prior = _prior_cycle_selections(engine, record_id)

    debt = 0.0
    for ln in lines:
        if ln["account"] in DEBT_BS_ACCTS:
            debt += abs(ln["amount"])

    ca = cl = 0.0
    out_lines = []
    for ln in lines:
        acct = ln["account"]
        is_debt = acct in DEBT_BS_ACCTS
        default_inc = _default_bs_include(ln["account_type"], acct)
        included = stored.get(acct, prior.get(acct, default_inc))
        if is_debt:
            included = False
        selectable = (ln["account_type"] in ("Assets", "Liabilities")) and not is_debt
        changed_vs_prior = (acct in prior and prior[acct] != included) or (
            acct not in prior and bool(prior) and included != default_inc)
        if included and ln["account_type"] == "Assets":
            ca += ln["amount"]
        elif included and ln["account_type"] == "Liabilities":
            cl += abs(ln["amount"])
        out_lines.append({
            **ln,
            "included": bool(included),
            "default_included": default_inc,
            "is_debt": is_debt,
            "selectable": selectable,
            "changed_vs_prior": bool(changed_vs_prior),
        })

    value, value_source = _resolve_value(rec)

    return {
        "record_id": record_id,
        "vcode": rec["vcode"],
        "cycle_year": int(rec["cycle_year"]),
        "as_of_date": str(rec["as_of_date"]),
        "vcodes": vcodes,
        "snapshot_dates": snapshot_dates,
        "classification": rec.get("classification_override") or rec.get("classification"),
        "lines": out_lines,
        "debt": debt,
        "current_assets": ca,
        "current_liabilities": cl,
        "concluded_value": value,
        "value_source": value_source,
        "children": rec["children"],
        "has_prior_selections": bool(prior),
    }


def _resolve_value(rec: Dict[str, Any]):
    """Entered value > children rollup > (cost derivation happens in compute)."""
    if rec.get("concluded_value") is not None:
        return float(rec["concluded_value"]), "entered"
    child_vals = [c["concluded_value"] for c in rec["children"] if c.get("concluded_value") is not None]
    if child_vals and len(child_vals) == len(rec["children"]) and rec["children"]:
        return float(sum(child_vals)), "children_rollup"
    cls = rec.get("classification_override") or rec.get("classification")
    if cls == "cost":
        return None, "cost_derived"  # derived during compute (needs capital + accrued pref)
    return None, "missing"


# ============================================================
# The NAV computation
# ============================================================

def _cap_wf_steps(data: dict, vcode: str) -> pd.DataFrame:
    wf = data.get("wf")
    if wf is None or wf.empty:
        return pd.DataFrame()
    steps = wf[(wf["vcode"].astype(str) == str(vcode)) & (wf["vmisc"] == CAP_WF)]
    return steps.sort_values("iOrder") if not steps.empty else steps


def _pref_walks(engine, data: dict, vcode: str, as_of: date,
                steps: pd.DataFrame) -> Dict[str, dict]:
    """Excel-exact pref walk per Cap_WF Pref-step investor.

    PSC-side investors use the report's own rate priority (deal_terms
    pe_coupon first); OP-side investors get their Pref step's rate passed as
    an explicit override so the PE coupon can never bleed onto the OP side.
    """
    from flask_app.services.reports_service import build_pref_balance_detail

    walks: Dict[str, dict] = {}
    if steps.empty:
        return walks
    pref_steps = steps[steps["vState"] == "Pref"]
    rate_col = "nPercent_dec" if "nPercent_dec" in pref_steps.columns else "nPercent"
    for _, s in pref_steps.iterrows():
        pc = str(s.get("PropCode", "")).strip()
        if not pc or pc in walks:
            continue
        rate = float(s[rate_col]) if pd.notna(s.get(rate_col)) else 0.0
        override = rate if pc.upper().startswith("OP") else None
        try:
            walks[pc] = build_pref_balance_detail(
                vcode, pc, as_of, data.get("acct"), data.get("inv"),
                wf_steps=data.get("wf"), pref_rate_override=override)
        except Exception:
            logger.warning(f"Pref walk failed for {vcode}/{pc}", exc_info=True)
            walks[pc] = {"header": {"investor_id": pc, "accrued_pref": 0.0,
                                    "investment_balance": 0.0, "pref_rate": rate},
                         "rows": []}
    return walks


def _step_refs(engine, vcode: str) -> Dict[int, str]:
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT iorder, agreement_ref FROM valuation_step_refs
            WHERE vcode = :v AND wf_type = :t
        """), {"v": vcode, "t": CAP_WF}).fetchall()
    return {int(r[0]): r[1] for r in rows if r[1]}


def compute_nav(engine, record_id: int, data: dict, username: str,
                save: bool = True) -> Dict[str, Any]:
    """Run the liquidation walk and (optionally) persist the result."""
    from waterfall import seed_states_from_accounting, run_waterfall
    from flask_app.services.valuation_service import ensure_valuation_tables, _now
    ensure_valuation_tables(engine)

    rec = _load_record(engine, record_id)
    vcode = rec["vcode"]
    as_of_ts = pd.Timestamp(rec["as_of_date"])
    as_of = as_of_ts.date()

    steps = _cap_wf_steps(data, vcode)
    if steps.empty:
        raise ValueError(f"{vcode} has no Cap_WF waterfall — the NAV walk needs one (see Waterfall Setup)")

    inputs = get_nav_inputs(engine, record_id, data)
    notes: List[str] = []

    # Pref walks (Excel-exact) + seeded capital states
    walks = _pref_walks(engine, data, vcode, as_of, steps)
    states = seed_states_from_accounting(
        data.get("acct"), data.get("inv"), data.get("wf"), vcode, cutoff_date=as_of)

    # Cost-basis derivation when no value was entered
    value = inputs["concluded_value"]
    value_source = inputs["value_source"]
    cost_components = None
    if value is None and value_source == "cost_derived":
        capital = sum(st.capital_outstanding for st in states.values())
        accrued = sum(w["header"].get("accrued_pref", 0) or 0 for w in walks.values())
        value = inputs["debt"] + capital + accrued
        cost_components = {"debt": inputs["debt"], "capital_balances": capital,
                           "accrued_pref": accrued}
        notes.append(
            f"Cost-basis value derived: debt {inputs['debt']:,.0f} + capital balances "
            f"{capital:,.0f} + accrued pref {accrued:,.0f} = {value:,.0f}")
    if value is None:
        raise ValueError(
            "No concluded value — enter one on the Assumptions tab"
            + (" (or complete the child property values)" if rec["children"] else ""))

    net_proceeds = value - inputs["debt"] + inputs["current_assets"] - inputs["current_liabilities"]
    if net_proceeds < 0:
        notes.append(f"Net proceeds are negative ({net_proceeds:,.0f}) — the walk distributes nothing")

    # Inject the Excel-exact pref balances into the seeded states, then stop
    # the engine from re-accruing under its own convention.
    for pc, walk in walks.items():
        st = states.get(pc)
        if st is None:
            from models import InvestorState
            st = InvestorState(propcode=pc)
            states[pc] = st
        accrued = float(walk["header"].get("accrued_pref", 0) or 0)
        pool = st.get_pool("initial")
        for tier in pool.pref_tiers:
            tier.pref_unpaid_compounded = 0.0
            tier.pref_accrued_current_year = 0.0
            if hasattr(tier, "pref_accrued_prior_year"):
                tier.pref_accrued_prior_year = 0.0
        st.pref_unpaid_compounded = accrued
        pool.last_accrual_date = as_of

    period_cash = pd.DataFrame([{"event_date": as_of, "cash_available": max(0.0, net_proceeds)}])
    allocs, end_states = run_waterfall(data.get("wf"), vcode, CAP_WF, period_cash,
                                       initial_states=states)

    refs = _step_refs(engine, vcode)
    walk_lines = []
    remaining = max(0.0, net_proceeds)
    if allocs is not None and not getattr(allocs, "empty", True):
        for _, a in allocs.iterrows():
            allocated = float(a.get("Allocated", 0) or 0)
            remaining = float(a.get("RemainingAfter", remaining))
            walk_lines.append({
                "iorder": int(a.get("iOrder", 0)),
                "agreement_ref": refs.get(int(a.get("iOrder", 0)), ""),
                "recipient": str(a.get("PropCode", "")),
                "step": str(a.get("vState", "")),
                "label": str(a.get("vtranstype", "")),
                "rate": float(a["nPercent"]) if pd.notna(a.get("nPercent")) else None,
                "allocated": allocated,
                "remaining_after": remaining,
            })

    psc_nav = sum(l["allocated"] for l in walk_lines if not l["recipient"].upper().startswith("OP"))
    op_nav = sum(l["allocated"] for l in walk_lines if l["recipient"].upper().startswith("OP"))

    pref_summaries = {
        pc: {
            "investor_id": pc,
            "pref_rate": w["header"].get("pref_rate"),
            "investment_balance": w["header"].get("investment_balance"),
            "accrued_pref": w["header"].get("accrued_pref"),
        } for pc, w in walks.items()
    }
    irr_steps = steps[steps["vState"] == "IRR"]
    rate_col = "nPercent_dec" if "nPercent_dec" in irr_steps.columns else "nPercent"
    lookback_rates = {
        str(s["PropCode"]).strip(): float(s[rate_col])
        for _, s in irr_steps.iterrows() if pd.notna(s.get(rate_col))
    }
    if lookback_rates:
        notes.append(
            "IRR lookbacks discount actual payment dates (acquisition fees excluded, "
            "per app-wide policy). The legacy Excel packages used month-end dates, so "
            "small lookback differences vs prior manual packages are the date "
            "convention, not an error.")

    result = {
        "record_id": record_id,
        "vcode": vcode,
        "as_of_date": str(rec["as_of_date"]),
        "value": value,
        "value_source": value_source if value_source != "cost_derived" else "cost_derived",
        "cost_components": cost_components,
        "debt": inputs["debt"],
        "current_assets": inputs["current_assets"],
        "current_liabilities": inputs["current_liabilities"],
        "net_proceeds": net_proceeds,
        "walk": walk_lines,
        "psc_nav": psc_nav,
        "op_nav": op_nav,
        "pref": pref_summaries,
        "lookback_rates": lookback_rates,
        "notes": notes,
        "computed_by": username,
        "computed_at": _now(),
    }

    if save:
        from flask_app.serializers import safe_json
        inputs_json = json.dumps(safe_json({
            "value": value, "value_source": value_source,
            "cost_components": cost_components,
            "debt": inputs["debt"], "current_assets": inputs["current_assets"],
            "current_liabilities": inputs["current_liabilities"],
            "snapshot_dates": inputs["snapshot_dates"],
            "bs_lines": inputs["lines"], "pref": pref_summaries,
            "lookback_rates": lookback_rates, "notes": notes,
        }), default=str)
        walk_json = json.dumps(safe_json(walk_lines), default=str)
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM valuation_nav_results WHERE record_id = :r"),
                         {"r": record_id})
            conn.execute(text("""
                INSERT INTO valuation_nav_results
                    (record_id, inputs_json, walk_json, net_proceeds, psc_nav, op_nav,
                     computed_by, computed_at)
                VALUES (:r, :ij, :wj, :np, :pn, :on, :u, :now)
            """), {"r": record_id, "ij": inputs_json, "wj": walk_json,
                   "np": net_proceeds, "pn": psc_nav, "on": op_nav,
                   "u": username, "now": _now()})

    return result


def get_nav(engine, record_id: int) -> Optional[Dict[str, Any]]:
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT inputs_json, walk_json, net_proceeds, psc_nav, op_nav,
                   computed_by, computed_at
            FROM valuation_nav_results WHERE record_id = :r
        """), {"r": record_id}).fetchone()
    if not row:
        return None
    inputs = json.loads(row[0]) if row[0] else {}
    return {
        **inputs,
        "walk": json.loads(row[1]) if row[1] else [],
        "net_proceeds": row[2],
        "psc_nav": row[3],
        "op_nav": row[4],
        "computed_by": row[5],
        "computed_at": str(row[6]) if row[6] is not None else None,
    }


def nav_results_for_cycle(engine, cycle_id: int) -> Dict[str, dict]:
    """{vcode: {psc_nav, op_nav, net_proceeds}} for committee views."""
    from flask_app.services.valuation_service import ensure_valuation_tables
    ensure_valuation_tables(engine)
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT r.vcode, n.psc_nav, n.op_nav, n.net_proceeds
            FROM valuation_nav_results n
            JOIN valuation_records r ON r.id = n.record_id
            WHERE r.cycle_id = :c
        """), {"c": cycle_id}).fetchall()
    return {str(r[0]): {"psc_nav": r[1], "op_nav": r[2], "net_proceeds": r[3]} for r in rows}


def set_step_ref(engine, vcode: str, iorder: int, agreement_ref: str,
                 username: str) -> Dict[str, Any]:
    """LLC agreement citation for a Cap_WF step (e.g. '8.2(a)'). Keyed by
    (vcode, wf_type, iOrder) so it survives waterfall re-saves."""
    from flask_app.services.valuation_service import ensure_valuation_tables
    ensure_valuation_tables(engine)
    with engine.begin() as conn:
        result = conn.execute(text("""
            UPDATE valuation_step_refs SET agreement_ref = :ref
            WHERE vcode = :v AND wf_type = :t AND iorder = :o
        """), {"ref": agreement_ref or None, "v": vcode, "t": CAP_WF, "o": int(iorder)})
        if result.rowcount == 0:
            conn.execute(text("""
                INSERT INTO valuation_step_refs (vcode, wf_type, iorder, agreement_ref)
                VALUES (:v, :t, :o, :ref)
            """), {"v": vcode, "t": CAP_WF, "o": int(iorder), "ref": agreement_ref or None})
    return {"status": "saved"}


# ============================================================
# Publish — the app as system of record
# ============================================================

def publish_record(engine, record_id: int, data: dict, username: str) -> Dict[str, Any]:
    """Write the approved NAV into the valuations table and stage the linked
    Argus import into forecasts as Val_IS_{year}. Requires committee approval
    and a computed NAV."""
    from flask_app.services.valuation_service import _now

    rec = _load_record(engine, record_id)
    if rec.get("status") != "approved":
        raise ValueError("Only committee-approved valuations can be published")
    nav = get_nav(engine, record_id)
    if not nav:
        raise ValueError("Compute the NAV before publishing")

    vcode = rec["vcode"]
    year = int(rec["cycle_year"])
    as_of = pd.Timestamp(rec["as_of_date"])
    as_of_str = as_of.strftime("%Y-%m-%d")

    inv = data.get("inv")
    prop_name = ""
    if inv is not None and not inv.empty:
        m = inv[inv["vcode"].astype(str) == vcode]
        if not m.empty:
            prop_name = str(m.iloc[0].get("Investment_Name", ""))

    published = {"valuations_row": True, "forecast_rows": 0}
    with engine.begin() as conn:
        # Existing rows store dtValuation in mixed formats ('12/31/2025 0:00'
        # from MRI, ISO from prior publishes) — match by parsed date, delete
        # by the exact stored string.
        existing = conn.execute(text("""
            SELECT dtValuation FROM valuations WHERE UPPER(vCode) = :v
        """), {"v": vcode.upper()}).fetchall()
        for (dt_str,) in existing:
            parsed = pd.to_datetime(dt_str, errors="coerce")
            if pd.notna(parsed) and parsed.date() == as_of.date():
                conn.execute(text("""
                    DELETE FROM valuations WHERE UPPER(vCode) = :v AND dtValuation = :d
                """), {"v": vcode.upper(), "d": dt_str})
        conn.execute(text("""
            INSERT INTO valuations
                (vCode, vPropertyName, dtValuation, vMethod, mAnnualNOI, fCapRate,
                 nTermCapRate, nDiscountRateForEquityInterest, mIncomeCapConcludedValue,
                 mDebtValue, mEquityValue, mMezzanineValue, nCostSaleRate)
            VALUES (:v, :name, :d, :method, :noi, :cap, :tcap, :disc, :val, :debt,
                    :equity, :mezz, :cos)
        """), {
            "v": vcode, "name": prop_name, "d": as_of_str,
            "method": rec.get("method"), "noi": rec.get("direct_cap_noi"),
            "cap": rec.get("cap_rate"), "tcap": rec.get("term_cap_rate"),
            "disc": rec.get("discount_rate"), "val": nav.get("value"),
            "debt": nav.get("debt"), "equity": nav.get("net_proceeds"),
            "mezz": nav.get("psc_nav"), "cos": rec.get("cost_of_sale_pct"),
        })

        # Val_IS_{year} forecast rows from the linked Argus import
        if rec.get("argus_import_id"):
            vsource = f"Val_IS_{year}"
            rows = conn.execute(text("""
                SELECT period_date, coa_account, amount_norm FROM argus_cashflows
                WHERE import_id = :iid AND vcode = :v AND coa_account IS NOT NULL
            """), {"iid": int(rec["argus_import_id"]), "v": vcode}).fetchall()
            if rows:
                conn.execute(text("""
                    DELETE FROM forecasts WHERE Vcode = :v AND vSource = :s
                """), {"v": vcode, "s": vsource})
                try:
                    from flask import current_app
                    pro_yr_base = int(current_app.config.get("PRO_YR_BASE_DEFAULT", year))
                except Exception:
                    pro_yr_base = year
                for period_date, coa, amount_norm in rows:
                    dt = pd.Timestamp(period_date)
                    conn.execute(text("""
                        INSERT INTO forecasts (Vcode, Date, vAccount, vSource, mAmount, Pro_Yr)
                        VALUES (:v, :d, :a, :s, :m, :p)
                    """), {
                        "v": vcode,
                        "d": f"{dt.month}/{dt.day}/{dt.year}",
                        "a": int(coa),
                        "s": vsource,
                        # forecasts carries the MRI sign convention
                        # (revenue negative); amount_norm is the forecast
                        # convention (revenue positive) — negate.
                        "m": -float(amount_norm or 0),
                        "p": dt.year - pro_yr_base,
                    })
                published["forecast_rows"] = len(rows)

        conn.execute(text("""
            UPDATE valuation_records SET published_at = :now, published_by = :u
            WHERE id = :i
        """), {"now": _now(), "u": username, "i": record_id})

    # Refresh caches so downstream consumers see the new rows immediately
    try:
        from flask_app.services import data_service, compute_service
        data_service.refresh_table("valuations")
        if published["forecast_rows"]:
            data_service.refresh_table("forecasts")
        compute_service.clear_cache(vcode)
    except Exception:
        logger.warning("Cache refresh after publish failed", exc_info=True)

    return {"status": "published", **published}


# ============================================================
# Auditor package Excel
# ============================================================

def generate_nav_package(engine, record_id: int, data: dict) -> bytes:
    """One workbook per deal, mirroring the manual auditor package:
    NAV_Calc / Bal_Sht / Accrued_Pref (+OP_Pref) / IRR_Lookbacks /
    LLC_Waterfall / Loader (when an Argus import is linked)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask_app.services.valuation_service import ensure_valuation_tables
    ensure_valuation_tables(engine)

    rec = _load_record(engine, record_id)
    vcode = rec["vcode"]
    as_of = pd.Timestamp(rec["as_of_date"]).date()
    nav = get_nav(engine, record_id)
    if not nav:
        raise ValueError("Compute the NAV before generating the auditor package")

    inv = data.get("inv")
    prop_name = vcode
    if inv is not None and not inv.empty:
        m = inv[inv["vcode"].astype(str) == vcode]
        if not m.empty:
            prop_name = str(m.iloc[0].get("Investment_Name", vcode))

    steps = _cap_wf_steps(data, vcode)
    walks = _pref_walks(engine, data, vcode, as_of, steps)

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    curr = "#,##0.00"
    curr0 = "#,##0"
    pct = "0.00%"

    wb = Workbook()

    def _header_row(ws, r, cols):
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(row=r, column=ci, value=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        return r + 1

    # ---------- NAV_Calc ----------
    ws = wb.active
    ws.title = "NAV_Calc"
    ws.cell(row=1, column=1, value=prop_name).font = title_font
    ws.cell(row=2, column=1, value=vcode)
    ws.cell(row=1, column=4, value=f"As of {as_of.strftime('%m/%d/%Y')}").font = bold
    ws.cell(row=3, column=1, value="PSC - NAV Calculation").font = bold

    r = 5
    def _line(label, amount, fmt=curr, indent=2, is_bold=False):
        nonlocal r
        c1 = ws.cell(row=r, column=indent, value=label)
        c2 = ws.cell(row=r, column=4, value=float(amount) if amount is not None else None)
        c2.number_format = fmt
        if is_bold:
            c1.font = bold
            c2.font = bold
        r += 1

    _line("Concluded Value" + (" (cost basis, derived)" if nav.get("value_source") == "cost_derived" else ""),
          nav.get("value"))
    _line("Less: Loan Balance", -(nav.get("debt") or 0))
    _line("Plus: Current Assets", nav.get("current_assets"))
    _line("Less: Current Liabilities", -(nav.get("current_liabilities") or 0))
    _line("Net Proceeds to Distribute", nav.get("net_proceeds"), is_bold=True)
    r += 1
    ws.cell(row=r, column=2, value="Waterfall:").font = bold
    r += 1
    r = _header_row(ws, r, ["Ref", "Recipient", "Step", "", "Amount", "Remaining"])
    for line in nav.get("walk", []):
        ws.cell(row=r, column=1, value=line.get("agreement_ref") or "")
        ws.cell(row=r, column=2, value=line.get("recipient"))
        label = line.get("label") or line.get("step")
        rate = line.get("rate")
        if rate and line.get("step") in ("Pref", "IRR"):
            label = f"{label} ({rate:.2%})" if rate < 1 else label
        elif line.get("step") in ("Share", "Tag"):
            label = f"{label}"
        ws.cell(row=r, column=3, value=f"{line.get('step')} - {label}")
        ws.cell(row=r, column=5, value=float(line.get("allocated", 0))).number_format = curr
        ws.cell(row=r, column=6, value=float(line.get("remaining_after", 0))).number_format = curr
        r += 1
    r += 1
    _line("PSC Liquidated NAV", nav.get("psc_nav"), is_bold=True)
    _line("OP Liquidated NAV", nav.get("op_nav"), is_bold=True)
    for note in nav.get("notes", []) or []:
        ws.cell(row=r, column=2, value=note)
        r += 1
    for col, width in (("A", 8), ("B", 30), ("C", 44), ("D", 16), ("E", 16), ("F", 16)):
        ws.column_dimensions[col].width = width

    # ---------- Bal_Sht ----------
    ws2 = wb.create_sheet("Bal_Sht")
    ws2.cell(row=1, column=1, value=f"{prop_name} — Balance Sheet (Interim BS)").font = title_font
    snap = nav.get("snapshot_dates") or {}
    ws2.cell(row=2, column=1, value="Snapshot: " + ", ".join(f"{k} {v}" for k, v in snap.items() if v))
    rr = _header_row(ws2, 4, ["Type", "Acct", "Line Item", "Amount", "In NAV Adj?", "Changed vs Prior Cycle"])
    for ln in nav.get("bs_lines", []) or []:
        ws2.cell(row=rr, column=1, value=ln.get("account_type"))
        ws2.cell(row=rr, column=2, value=ln.get("account"))
        ws2.cell(row=rr, column=3, value=ln.get("description"))
        ws2.cell(row=rr, column=4, value=float(ln.get("amount", 0))).number_format = curr
        flag = "Debt" if ln.get("is_debt") else ("Yes" if ln.get("included") else "")
        ws2.cell(row=rr, column=5, value=flag)
        ws2.cell(row=rr, column=6, value="CHANGED" if ln.get("changed_vs_prior") else "")
        rr += 1
    rr += 1
    for label, val in (("Debt", nav.get("debt")), ("Current Assets", nav.get("current_assets")),
                       ("Current Liabilities", nav.get("current_liabilities"))):
        ws2.cell(row=rr, column=3, value=label).font = bold
        ws2.cell(row=rr, column=4, value=float(val or 0)).number_format = curr
        ws2.cell(row=rr, column=4).font = bold
        rr += 1
    for col, width in (("A", 12), ("B", 8), ("C", 36), ("D", 16), ("E", 12), ("F", 20)):
        ws2.column_dimensions[col].width = width

    # ---------- Pref sheets ----------
    pref_cols = ["InvestmentID", "InvestorID", "EffectiveDate", "Amt", "Typename",
                 "Investment_Balance", "Compounded Pref", "Inv + Comp", "DaysSinceLast",
                 "Current Due", "Accrued Pref", "Total Due", "Pref Paid", "Remaining Accrual"]
    for pc, walk in walks.items():
        sheet_name = "OP_Pref" if pc.upper().startswith("OP") else "Accrued_Pref"
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name}_{pc}"[:31]
        wsp = wb.create_sheet(sheet_name)
        h = walk.get("header", {})
        wsp.cell(row=1, column=1, value=f"{prop_name} — {pc}").font = title_font
        wsp.cell(row=2, column=1, value="Pref Rate:").font = bold
        wsp.cell(row=2, column=2, value=h.get("pref_rate", 0)).number_format = pct
        wsp.cell(row=2, column=4, value="Investment Balance:").font = bold
        wsp.cell(row=2, column=5, value=h.get("investment_balance", 0)).number_format = curr0
        wsp.cell(row=3, column=4, value="Accrued:").font = bold
        wsp.cell(row=3, column=5, value=h.get("accrued_pref", 0)).number_format = curr
        wsp.cell(row=4, column=4, value="Total:").font = bold
        wsp.cell(row=4, column=5, value=h.get("total", 0)).number_format = curr
        pr = _header_row(wsp, 6, pref_cols)
        for row in walk.get("rows", []):
            for ci, col in enumerate(pref_cols, 1):
                v = row.get(col)
                cell = wsp.cell(row=pr, column=ci, value=v)
                if col not in ("InvestmentID", "InvestorID", "EffectiveDate", "Typename", "DaysSinceLast"):
                    cell.number_format = curr
            pr += 1
        wsp.column_dimensions["C"].width = 12
        wsp.column_dimensions["E"].width = 30

    # ---------- IRR_Lookbacks ----------
    ws3 = wb.create_sheet("IRR_Lookbacks")
    ws3.cell(row=1, column=1, value=f"{prop_name} — IRR Lookbacks").font = title_font
    col_base = 1
    lookbacks = nav.get("lookback_rates", {}) or {}
    from waterfall import seed_states_from_accounting
    states = seed_states_from_accounting(
        data.get("acct"), data.get("inv"), data.get("wf"), vcode, cutoff_date=as_of)
    for pc, target in lookbacks.items():
        st = states.get(pc)
        ws3.cell(row=2, column=col_base, value=f"{pc} — {target:.0%} IRR Lookback").font = bold
        rr = _header_row(ws3, 3, ["Date", "Amount", "Label"])
        first_row = rr
        cfs = list(st.cashflows) if st else []
        labels = list(st.cashflow_labels) if st and len(st.cashflow_labels) == len(cfs) else [""] * len(cfs)
        # Terminal rows: this walk's pref + capital at the NAV date, then the lookback
        pref_cap = sum(l["allocated"] for l in nav.get("walk", [])
                       if l["recipient"] == pc and l["step"] in ("Pref", "Initial"))
        lookback_amt = sum(l["allocated"] for l in nav.get("walk", [])
                           if l["recipient"] == pc and l["step"] == "IRR")
        rows_data = [(d, a, lbl) for (d, a), lbl in zip(cfs, labels)]
        rows_data.append((as_of, pref_cap, "Pref + Initial Capital (NAV walk)"))
        rows_data.append((as_of, lookback_amt, f"{target:.0%} IRR Lookback (NAV walk)"))
        for d, a, lbl in rows_data:
            ws3.cell(row=rr, column=col_base, value=pd.Timestamp(d).strftime("%m/%d/%Y"))
            ws3.cell(row=rr, column=col_base + 1, value=float(a)).number_format = curr
            ws3.cell(row=rr, column=col_base + 2, value=lbl)
            rr += 1
        # Live XIRR check (requires Excel 365/2021+; _xlfn prefix per app convention)
        amt_col = ws3.cell(row=first_row, column=col_base + 1).column_letter
        date_col = ws3.cell(row=first_row, column=col_base).column_letter
        ws3.cell(row=rr + 1, column=col_base, value="XIRR check:").font = bold
        f = (f"=_xlfn.IFERROR(_xlfn.XIRR({amt_col}{first_row}:{amt_col}{rr - 1},"
             f"DATEVALUE({date_col}{first_row}:{date_col}{rr - 1})),\"N/A\")")
        ws3.cell(row=rr + 1, column=col_base + 1, value=f).number_format = pct
        col_base += 4

    # ---------- LLC_Waterfall ----------
    ws4 = wb.create_sheet("LLC_Waterfall")
    ws4.cell(row=1, column=1, value=f"{prop_name} — Cap Waterfall (as modeled)").font = title_font
    refs = _step_refs(engine, vcode)
    rr = _header_row(ws4, 3, ["Ref", "iOrder", "Recipient", "Step", "FXRate", "Rate", "Description"])
    for _, s in steps.iterrows():
        io = int(s["iOrder"]) if pd.notna(s.get("iOrder")) else 0
        ws4.cell(row=rr, column=1, value=refs.get(io, ""))
        ws4.cell(row=rr, column=2, value=io)
        ws4.cell(row=rr, column=3, value=str(s.get("PropCode", "")))
        ws4.cell(row=rr, column=4, value=str(s.get("vState", "")))
        fx = s.get("FXRate")
        ws4.cell(row=rr, column=5, value=float(fx) if pd.notna(fx) else None)
        rate_col = "nPercent_dec" if "nPercent_dec" in steps.columns else "nPercent"
        rt = s.get(rate_col)
        c = ws4.cell(row=rr, column=6, value=float(rt) if pd.notna(rt) else None)
        c.number_format = pct
        ws4.cell(row=rr, column=7, value=str(s.get("vtranstype", "")))
        rr += 1
    rr += 1
    with engine.connect() as conn:
        excerpts = conn.execute(text("""
            SELECT filename FROM valuation_documents
            WHERE record_id = :r AND doc_type = 'llc_excerpt'
        """), {"r": record_id}).fetchall()
    ws4.cell(row=rr, column=1,
             value=("LLC agreement excerpts on file: " + "; ".join(e[0] for e in excerpts))
             if excerpts else
             "No LLC agreement excerpt uploaded — attach the relevant section as evidence.").font = bold
    ws4.column_dimensions["G"].width = 34

    # ---------- Loader (Argus forecast rows) ----------
    if rec.get("argus_import_id"):
        from flask_app.services import argus_service
        try:
            from flask import current_app
            pro_yr_base = int(current_app.config.get("PRO_YR_BASE_DEFAULT", rec["cycle_year"]))
        except Exception:
            pro_yr_base = int(rec["cycle_year"])
        fc = argus_service.get_forecast_df_by_id(engine, vcode, int(rec["argus_import_id"]), pro_yr_base)
        if fc is not None and not fc.empty:
            ws5 = wb.create_sheet("Loader")
            vsource = f"Val_IS_{rec['cycle_year']}"
            rr = _header_row(ws5, 1, ["Vcode", "dtEntry", "vSource", "vAccount", "mAmount", "Year"])
            for _, row in fc.sort_values(["event_date", "vAccount"]).iterrows():
                dt = pd.Timestamp(row["event_date"])
                ws5.cell(row=rr, column=1, value=vcode)
                ws5.cell(row=rr, column=2, value=dt.strftime("%b-%Y"))
                ws5.cell(row=rr, column=3, value=vsource)
                ws5.cell(row=rr, column=4, value=int(row["vAccount"]))
                ws5.cell(row=rr, column=5, value=-float(row["mAmount_norm"])).number_format = curr
                ws5.cell(row=rr, column=6, value=int(dt.year))
                rr += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_cycle_packages_zip(engine, cycle_id: int, data: dict) -> bytes:
    """One auditor workbook per record with a computed NAV, zipped."""
    import zipfile
    from flask_app.services.valuation_service import get_cycle_dashboard

    dash = get_cycle_dashboard(engine, cycle_id, data)
    buf = BytesIO()
    added = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for rec in dash["records"]:
            if rec["status"] == "excluded":
                continue
            if get_nav(engine, rec["id"]) is None:
                continue
            try:
                content = generate_nav_package(engine, rec["id"], data)
            except Exception:
                logger.warning(f"Package failed for {rec['vcode']}", exc_info=True)
                continue
            safe = "".join(ch for ch in rec["deal_name"] if ch.isalnum() or ch in " -_").strip()
            zf.writestr(f"NAV_{rec['vcode']}_{safe}.xlsx", content)
            added += 1
        if added == 0:
            zf.writestr("README.txt", "No records with a computed NAV in this cycle yet.")
    return buf.getvalue()


# ============================================================
# Phase 4 — Tie-out / reconciliation checks
# ============================================================

def run_record_checks(engine, record_id: int, data: dict) -> Dict[str, Any]:
    """Automatic tie-outs for one valuation record.

    Severity: fail (blocks a defensible package) > warn (needs a look) >
    info (worth knowing) > ok. Each check reports even when it passes so
    the reviewer can see what was tested.
    """
    from config import REVENUE_ACCTS, EXPENSE_ACCTS
    from flask_app.services.valuation_service import ensure_valuation_tables
    ensure_valuation_tables(engine)

    rec = _load_record(engine, record_id)
    vcode = rec["vcode"]
    year = int(rec["cycle_year"])
    as_of = pd.Timestamp(rec["as_of_date"])
    classification = rec.get("classification_override") or rec.get("classification")
    checks: List[Dict[str, str]] = []

    def _add(key, severity, message):
        checks.append({"key": key, "severity": severity, "message": message})

    # --- Waterfall structure ---
    steps = _cap_wf_steps(data, vcode)
    if steps.empty:
        _add("cap_wf", "fail", "No Cap_WF waterfall — the NAV walk cannot run (see Waterfall Setup)")
    else:
        _add("cap_wf", "ok", f"Cap_WF present ({len(steps)} steps)")
        pref_steps = steps[steps["vState"] == "Pref"]
        if pref_steps.empty:
            _add("pref_step", "warn", "Cap_WF has no Pref step — no accrued pref will be paid in the walk")
        # IRR lookback completeness: deal_terms says a lookback exists
        try:
            with engine.connect() as conn:
                dt_row = conn.execute(text(
                    "SELECT irr_lookback FROM deal_terms WHERE UPPER(vcode) = :v"
                ), {"v": vcode.upper()}).fetchone()
            expected_lookback = float(dt_row[0]) if dt_row and dt_row[0] is not None else None
        except Exception:
            expected_lookback = None
        irr_steps = steps[steps["vState"] == "IRR"]
        if expected_lookback and expected_lookback > 0 and irr_steps.empty:
            _add("irr_step", "warn",
                 f"deal_terms carries a {expected_lookback:.0%} IRR lookback but the Cap_WF has no IRR step — "
                 "the NAV walk will skip the lookback")
        elif not irr_steps.empty:
            rate_col = "nPercent_dec" if "nPercent_dec" in irr_steps.columns else "nPercent"
            rates = sorted({float(r) for r in irr_steps[rate_col].dropna()})
            _add("irr_step", "ok", "IRR lookback step(s) at " + ", ".join(f"{r:.0%}" for r in rates))

    # --- Assumption completeness ---
    if classification != "cost":
        missing = [label for label, field in (
            ("method", "method"), ("concluded value", "concluded_value"),
            ("going-in cap rate", "cap_rate"),
        ) if rec.get(field) in (None, "")]
        if rec["children"] and rec.get("concluded_value") in (None, "") and "concluded value" in missing:
            child_vals = [c for c in rec["children"] if c.get("concluded_value") is not None]
            if len(child_vals) == len(rec["children"]):
                missing.remove("concluded value")  # rollup covers it
        if missing:
            _add("assumptions", "warn", "Missing: " + ", ".join(missing))
        else:
            _add("assumptions", "ok", "Assumptions complete")
    else:
        _add("assumptions", "ok", "Cost basis — value derived from capital balances + accrued pref")

    # --- Children values (portfolio parents) ---
    if rec["children"]:
        missing_children = [c["vcode"] for c in rec["children"] if c.get("concluded_value") is None]
        if missing_children:
            _add("children", "warn",
                 f"{len(missing_children)} child propert{'y' if len(missing_children) == 1 else 'ies'} "
                 f"missing a concluded value: {', '.join(missing_children)}")
        else:
            _add("children", "ok", f"All {len(rec['children'])} child property values entered")

    # --- Evidence ---
    with engine.connect() as conn:
        doc_types = {r[0] for r in conn.execute(text(
            "SELECT DISTINCT doc_type FROM valuation_documents WHERE record_id = :r"
        ), {"r": record_id}).fetchall()}
        open_q = conn.execute(text(
            "SELECT COUNT(*) FROM valuation_questions WHERE record_id = :r AND status = 'open'"
        ), {"r": record_id}).fetchone()[0]
    if classification == "third_party":
        if "appraisal" in doc_types:
            _add("appraisal_doc", "ok", "Appraisal document on file")
        else:
            _add("appraisal_doc", "warn", "Third-party classification but no appraisal PDF uploaded")
        if rec.get("argus_import_id"):
            _add("argus", "ok", "Argus projection linked")
        else:
            _add("argus", "warn", "No Argus projection linked — the Budget Review valuation column and "
                                  "the Val_IS forecast publish will be empty")
    if "llc_excerpt" not in doc_types:
        _add("llc_excerpt", "info", "No LLC agreement excerpt uploaded — the auditor package cites the "
                                    "modeled waterfall without agreement evidence")
    if open_q:
        _add("questions", "warn", f"{open_q} reviewer question(s) still open")
    else:
        _add("questions", "ok", "No open reviewer questions")

    # --- AI cross-check ---
    try:
        from flask_app.services.valuation_ai_service import get_ai_summary
        ai = get_ai_summary(engine, record_id)
    except Exception:
        ai = None
    if ai:
        mismatches = [c["field"] for c in ai.get("checks", []) if c.get("match") is False]
        if mismatches:
            _add("ai_crosscheck", "warn",
                 "Entered assumptions differ from the appraisal per the AI summary: " + ", ".join(mismatches))
        else:
            _add("ai_crosscheck", "ok", "Entered assumptions match the AI-extracted appraisal values")
    elif classification == "third_party" and "appraisal" in doc_types:
        _add("ai_crosscheck", "info", "AI appraisal summary not generated yet — no automated cross-check")

    # --- Direct-cap sanity: value vs NOI / cap ---
    val, noi, cap = rec.get("concluded_value"), rec.get("direct_cap_noi"), rec.get("cap_rate")
    if val and noi and cap:
        implied = float(noi) / float(cap)
        diff_pct = (float(val) - implied) / implied
        if abs(diff_pct) > 0.10:
            _add("value_vs_cap", "warn",
                 f"Concluded value {float(val):,.0f} is {diff_pct:+.1%} vs NOI/cap implied {implied:,.0f} — "
                 "fine for a DCF conclusion, but worth confirming")
        else:
            _add("value_vs_cap", "ok",
                 f"Value within {diff_pct:+.1%} of the NOI/cap implied {implied:,.0f}")

    # --- Argus year-1 NOI vs entered NOI ---
    if rec.get("argus_import_id") and noi:
        try:
            from flask_app.services import argus_service
            try:
                from flask import current_app
                pro_yr_base = int(current_app.config.get("PRO_YR_BASE_DEFAULT", year))
            except Exception:
                pro_yr_base = year
            fc = argus_service.get_forecast_df_by_id(engine, vcode, int(rec["argus_import_id"]), pro_yr_base)
            if fc is not None and not fc.empty:
                fc = fc.copy()
                fc["event_date"] = pd.to_datetime(fc["event_date"])
                start = fc["event_date"].min()
                yr1 = fc[fc["event_date"] < start + pd.DateOffset(months=12)]
                noi_accts = REVENUE_ACCTS | EXPENSE_ACCTS
                argus_noi = float(yr1[yr1["vAccount"].astype(int).isin(noi_accts)]["mAmount_norm"].sum())
                diff = (argus_noi - float(noi)) / float(noi) if noi else 0
                if abs(diff) > 0.05:
                    _add("argus_noi", "warn",
                         f"Argus year-1 NOI {argus_noi:,.0f} is {diff:+.1%} vs the entered direct-cap NOI "
                         f"{float(noi):,.0f}")
                else:
                    _add("argus_noi", "ok",
                         f"Argus year-1 NOI {argus_noi:,.0f} within {diff:+.1%} of the entered NOI")
        except Exception:
            logger.warning(f"argus_noi check failed for {vcode}", exc_info=True)

    # --- Balance sheet staleness + NAV state ---
    nav = get_nav(engine, record_id)
    if nav:
        for vc, snap in (nav.get("snapshot_dates") or {}).items():
            if not snap:
                _add("bs_staleness", "warn", f"{vc}: no Interim BS reported — NAV uses no balance sheet")
                continue
            age_days = (as_of - pd.Timestamp(snap)).days
            if age_days > 92:
                _add("bs_staleness", "warn",
                     f"{vc}: balance sheet snapshot {snap} is {age_days} days before the valuation date")
            else:
                _add("bs_staleness", "ok", f"{vc}: balance sheet as of {snap}")
        changed = [l["account"] for l in (nav.get("bs_lines") or []) if l.get("changed_vs_prior")]
        if changed:
            _add("bs_changed", "info",
                 "Balance sheet treatment changed vs the prior cycle on account(s): " + ", ".join(changed))
        # Stale NAV vs later assumption edits
        try:
            if rec.get("updated_at") and nav.get("computed_at") and \
                    str(nav["computed_at"]) < str(rec["updated_at"]):
                _add("nav_stale", "warn", "The record changed after the NAV was computed — recompute")
            else:
                _add("nav_stale", "ok", "NAV is current with the record")
        except Exception:
            pass
        # Swing vs prior-year published Pref NAV
        mri_val = data.get("mri_val")
        if mri_val is not None and not mri_val.empty and nav.get("psc_nav"):
            df = mri_val.copy()
            vcol = "vCode" if "vCode" in df.columns else "vcode"
            df["_dt"] = pd.to_datetime(df["dtValuation"], errors="coerce")
            prior = df[(df[vcol].astype(str).str.strip().str.upper() == vcode.upper())
                       & (df["_dt"].dt.year == year - 1)]
            if not prior.empty:
                prior_mezz = pd.to_numeric(prior.iloc[0].get("mMezzanineValue"), errors="coerce")
                if pd.notna(prior_mezz) and prior_mezz > 0:
                    swing = (float(nav["psc_nav"]) - float(prior_mezz)) / float(prior_mezz)
                    sev = "info" if abs(swing) > 0.25 else "ok"
                    _add("nav_vs_prior", sev,
                         f"PSC NAV {float(nav['psc_nav']):,.0f} vs prior-year {float(prior_mezz):,.0f} "
                         f"({swing:+.1%})")
    elif rec.get("status") in ("signed_off", "approved"):
        _add("nav_computed", "warn", "No NAV computed yet for a record in committee review")
    else:
        _add("nav_computed", "info", "No NAV computed yet")

    counts = {"fail": 0, "warn": 0, "info": 0, "ok": 0}
    for c in checks:
        counts[c["severity"]] = counts.get(c["severity"], 0) + 1
    return {"record_id": record_id, "vcode": vcode, "checks": checks, "counts": counts}
