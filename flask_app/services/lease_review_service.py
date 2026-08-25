"""
lease_review_service.py
Lease review service for due diligence — extraction, validation, and analysis.

Supports multi-property portfolio reviews. Data stored in PostgreSQL/SQLite
via the standard database layer.
"""

import copy
import hashlib
import json
import re
import logging
import os
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MATERIAL_LEASE_SF_THRESHOLD = 10_000  # SF threshold for "material" lease
MATERIAL_LEASE_RENT_THRESHOLD = 100_000  # Annual rent threshold

# Document type classification by filename pattern
DOC_TYPE_PATTERNS = [
    (r'(?i)lease(?!.*(?:abstract|amend|memo))', 'Original Lease'),
    (r'(?i)(first|second|third|fourth|fifth|\d+)\s*(amendment|amend)', 'Amendment'),
    (r'(?i)amend', 'Amendment'),
    (r'(?i)commencement', 'Commencement Letter'),
    (r'(?i)option\s*letter', 'Option Letter'),
    (r'(?i)snda', 'SNDA'),
    (r'(?i)memo\s*of\s*lease', 'Memorandum of Lease'),
    (r'(?i)waiver', 'Waiver Letter'),
    (r'(?i)estoppel', 'Estoppel'),
    (r'(?i)consent', 'Consent Letter'),
    (r'(?i)move.?in', 'Move-In Notice'),
    (r'(?i)delivery', 'Delivery Notice'),
    (r'(?i)rent\s*commencement', 'Rent Commencement'),
    (r'(?i)change.*(?:notice|contact|office|address)', 'Notice Change'),
    (r'(?i)coi|certificate.*insurance', 'COI'),
    (r'(?i)open.*business', 'Opening Notice'),
    (r'(?i)short\s*form', 'Short Form Lease'),
    (r'(?i)co.?tenancy.*(?:cure|notice)', 'Co-Tenancy Notice'),
    (r'(?i)cam\s*dispute', 'CAM Dispute'),
    (r'(?i)architect|certif', 'Certification'),
    (r'(?i)cctv|install', 'Installation Notice'),
]


def classify_document(filename: str) -> str:
    """Classify a lease document by its filename."""
    for pattern, doc_type in DOC_TYPE_PATTERNS:
        if re.search(pattern, filename):
            return doc_type
    return 'Other'


def parse_doc_date(filename: str) -> Optional[str]:
    """Extract date from filename like '2024.03.28_Bealls-Lease.pdf'."""
    m = re.match(r'(\d{4})[.\-](\d{2})[.\-](\d{2})', filename)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


# ---------------------------------------------------------------------------
# Database DDL — called from ensure_pg_tables / create_additional_tables
# ---------------------------------------------------------------------------

LEASE_DDL_PG = [
    """
    CREATE TABLE IF NOT EXISTS lease_reviews (
        id              SERIAL PRIMARY KEY,
        property_name   TEXT NOT NULL,
        property_address TEXT,
        total_gla       DOUBLE PRECISION,
        total_annual_rent DOUBLE PRECISION,
        total_tenants   INTEGER,
        rent_roll_date  TEXT,
        prospect_property_id INTEGER,
        status          TEXT DEFAULT 'in_progress',
        source_folder   TEXT,
        created_by      TEXT,
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_tenants (
        id              SERIAL PRIMARY KEY,
        review_id       INTEGER NOT NULL REFERENCES lease_reviews(id),
        tenant_name     TEXT NOT NULL,
        suite           TEXT,
        square_feet     DOUBLE PRECISION,
        lease_type      TEXT,
        lease_start     TEXT,
        lease_end       TEXT,
        term_months     INTEGER,
        monthly_rent    DOUBLE PRECISION,
        monthly_rent_per_sf DOUBLE PRECISION,
        annual_rent     DOUBLE PRECISION,
        annual_rent_per_sf DOUBLE PRECISION,
        annual_recoveries_per_sf DOUBLE PRECISION,
        annual_misc_per_sf DOUBLE PRECISION,
        annual_sales_override DOUBLE PRECISION,
        security_deposit DOUBLE PRECISION,
        is_vacant       BOOLEAN DEFAULT FALSE,
        is_material     BOOLEAN DEFAULT FALSE,
        has_cotenancy   BOOLEAN DEFAULT FALSE,
        has_exclusive_use BOOLEAN DEFAULT FALSE,
        extraction_status TEXT DEFAULT 'pending',
        extraction_json TEXT,
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_documents (
        id              SERIAL PRIMARY KEY,
        tenant_id       INTEGER NOT NULL REFERENCES lease_tenants(id),
        review_id       INTEGER NOT NULL REFERENCES lease_reviews(id),
        filename        TEXT NOT NULL,
        file_path       TEXT,
        doc_type        TEXT,
        doc_date        TEXT,
        page_count      INTEGER,
        extracted_text   TEXT,
        extraction_status TEXT DEFAULT 'pending',
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_rent_steps (
        id              SERIAL PRIMARY KEY,
        tenant_id       INTEGER NOT NULL REFERENCES lease_tenants(id),
        effective_date  TEXT,
        monthly_rent    DOUBLE PRECISION,
        annual_rent     DOUBLE PRECISION,
        rent_per_sf     DOUBLE PRECISION,
        source_doc      TEXT,
        source_page     TEXT,
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_cotenancy (
        id              SERIAL PRIMARY KEY,
        tenant_id       INTEGER NOT NULL REFERENCES lease_tenants(id),
        review_id       INTEGER NOT NULL REFERENCES lease_reviews(id),
        clause_text     TEXT,
        trigger_description TEXT,
        trigger_threshold TEXT,
        cure_period_days INTEGER,
        alt_rent_formula TEXT,
        termination_right BOOLEAN DEFAULT FALSE,
        termination_notice_days INTEGER,
        sunset_provision TEXT,
        is_curable      BOOLEAN DEFAULT TRUE,
        waiver_mechanism TEXT,
        source_doc      TEXT,
        source_page     TEXT,
        notes           TEXT,
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_cotenancy_refs (
        id              SERIAL PRIMARY KEY,
        cotenancy_id    INTEGER NOT NULL REFERENCES lease_cotenancy(id),
        tenant_id       INTEGER NOT NULL,
        referenced_tenant_name TEXT NOT NULL,
        referenced_tenant_id INTEGER,
        reference_type  TEXT DEFAULT 'named',
        notes           TEXT
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_exclusive_use (
        id              SERIAL PRIMARY KEY,
        tenant_id       INTEGER NOT NULL REFERENCES lease_tenants(id),
        restriction_text TEXT,
        restricted_use  TEXT,
        radius_feet     DOUBLE PRECISION,
        source_doc      TEXT,
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_options (
        id              SERIAL PRIMARY KEY,
        tenant_id       INTEGER NOT NULL REFERENCES lease_tenants(id),
        option_type     TEXT NOT NULL,
        option_number   INTEGER,
        total_options   INTEGER,
        term_years      DOUBLE PRECISION,
        notice_days     INTEGER,
        notice_deadline TEXT,
        rent_terms      TEXT,
        auto_renewal    BOOLEAN DEFAULT FALSE,
        exercised       BOOLEAN DEFAULT FALSE,
        option_start    TEXT,
        option_end      TEXT,
        source_doc      TEXT,
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_validation (
        id              SERIAL PRIMARY KEY,
        tenant_id       INTEGER NOT NULL REFERENCES lease_tenants(id),
        field_name      TEXT NOT NULL,
        source_type     TEXT DEFAULT 'rent_roll',
        seller_value    TEXT,
        lease_value     TEXT,
        status          TEXT DEFAULT 'pending',
        source_doc      TEXT,
        notes           TEXT,
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_abstract_sections (
        id              SERIAL PRIMARY KEY,
        tenant_id       INTEGER NOT NULL REFERENCES lease_tenants(id),
        section_key     TEXT NOT NULL,
        section_title   TEXT NOT NULL,
        content         TEXT,
        lease_ref       TEXT,
        sort_order      INTEGER DEFAULT 0,
        updated_by      TEXT,
        updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(tenant_id, section_key)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_tenant_sales (
        id              SERIAL PRIMARY KEY,
        tenant_id       INTEGER NOT NULL REFERENCES lease_tenants(id),
        review_id       INTEGER NOT NULL REFERENCES lease_reviews(id),
        year            INTEGER NOT NULL,
        sales_amount    DOUBLE PRECISION NOT NULL DEFAULT 0,
        month_start     INTEGER NOT NULL DEFAULT 1,
        month_end       INTEGER NOT NULL DEFAULT 12,
        months_covered  INTEGER NOT NULL DEFAULT 12,
        comment         TEXT,
        source          TEXT DEFAULT 'ai_extract',
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(tenant_id, year)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_tenant_aliases (
        id              SERIAL PRIMARY KEY,
        alias_name      TEXT NOT NULL UNIQUE,
        canonical_name  TEXT NOT NULL,
        created_by      TEXT,
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_space_events (
        id              SERIAL PRIMARY KEY,
        review_id       INTEGER NOT NULL REFERENCES lease_reviews(id),
        event_type      TEXT NOT NULL,
        effective_date  TEXT NOT NULL,
        source_tenant_ids TEXT NOT NULL,
        description     TEXT,
        status          TEXT DEFAULT 'planned',
        created_by      TEXT,
        created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_space_event_results (
        id              SERIAL PRIMARY KEY,
        event_id        INTEGER NOT NULL REFERENCES lease_space_events(id),
        result_tenant_id INTEGER,
        tenant_name     TEXT,
        suite           TEXT,
        square_feet     DOUBLE PRECISION,
        monthly_rent    DOUBLE PRECISION,
        annual_rent     DOUBLE PRECISION,
        rent_per_sf     DOUBLE PRECISION,
        lease_start     TEXT,
        lease_end       TEXT,
        is_vacant       BOOLEAN DEFAULT FALSE,
        notes           TEXT
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS lease_market_assumptions (
        id                    SERIAL PRIMARY KEY,
        review_id             INTEGER NOT NULL REFERENCES lease_reviews(id),
        lease_type            TEXT NOT NULL,
        market_rent_psf       DOUBLE PRECISION,
        annual_rent_growth    DOUBLE PRECISION,
        renewal_probability   DOUBLE PRECISION,
        renewal_downtime_months INTEGER DEFAULT 0,
        renewal_ti_psf        DOUBLE PRECISION,
        renewal_lc_pct        DOUBLE PRECISION,
        renewal_rent_spread   DOUBLE PRECISION,
        renewal_term_years    INTEGER DEFAULT 5,
        new_downtime_months   INTEGER DEFAULT 6,
        new_ti_psf            DOUBLE PRECISION,
        new_lc_pct            DOUBLE PRECISION,
        new_rent_spread       DOUBLE PRECISION,
        new_term_years        INTEGER DEFAULT 10,
        free_rent_months      INTEGER DEFAULT 0,
        annual_expense_growth DOUBLE PRECISION,
        created_by            TEXT,
        created_at            TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at            TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(review_id, lease_type)
    )
    """,
]

# SQLite variants (no SERIAL, use INTEGER PRIMARY KEY AUTOINCREMENT)
LEASE_DDL_SQLITE = [
    ddl.replace('SERIAL PRIMARY KEY', 'INTEGER PRIMARY KEY AUTOINCREMENT')
       .replace('DOUBLE PRECISION', 'REAL')
       .replace('BOOLEAN', 'INTEGER')
       .replace('REFERENCES lease_reviews(id)', '')
       .replace('REFERENCES lease_tenants(id)', '')
       .replace('REFERENCES lease_cotenancy(id)', '')
    for ddl in LEASE_DDL_PG
]


def _migrate_add_column(engine, table: str, column: str, col_type: str):
    """Add a column to a table if it doesn't exist. Works on both PG and SQLite."""
    from sqlalchemy import text
    with engine.begin() as conn:
        if engine.dialect.name == 'postgresql':
            row = conn.execute(text("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = :tbl AND column_name = :col
            """), {'tbl': table, 'col': column}).fetchone()
            if row is None:
                conn.execute(text(
                    f"ALTER TABLE {table} ADD COLUMN {column} {col_type}"))
        else:
            cols = [r[1] for r in conn.execute(
                text(f"PRAGMA table_info({table})")).fetchall()]
            if column not in cols:
                conn.execute(text(
                    f"ALTER TABLE {table} ADD COLUMN {column} {col_type}"))


def _migrate_nullable(engine, table: str, column: str):
    """Drop NOT NULL constraint on a column. PG only — SQLite ignores constraints."""
    from sqlalchemy import text
    if engine.dialect.name != 'postgresql':
        return  # SQLite doesn't enforce NOT NULL on ALTER; no action needed
    with engine.begin() as conn:
        try:
            conn.execute(text(
                f"ALTER TABLE {table} ALTER COLUMN {column} DROP NOT NULL"))
        except Exception:
            pass  # Already nullable or column doesn't exist


def ensure_lease_tables(engine):
    """Create lease review tables and migrate missing columns."""
    from sqlalchemy import text, inspect
    with engine.connect() as conn:
        # Detect dialect for correct DDL
        dialect = engine.dialect.name
        if dialect == 'sqlite':
            for ddl in LEASE_DDL_SQLITE:
                conn.execute(text(ddl))
        else:
            for ddl in LEASE_DDL_PG:
                conn.execute(text(ddl))

        conn.commit()

    # Migrate: add prospect_property_id and rent_roll_date if missing
    with engine.begin() as conn:
        if engine.dialect.name == 'postgresql':
            for col in ['prospect_property_id', 'rent_roll_date']:
                row = conn.execute(text("""
                    SELECT column_name FROM information_schema.columns
                    WHERE table_name = 'lease_reviews' AND column_name = :col
                """), {'col': col}).fetchone()
                if row is None:
                    ctype = 'INTEGER' if col == 'prospect_property_id' else 'TEXT'
                    conn.execute(text(
                        f"ALTER TABLE lease_reviews ADD COLUMN {col} {ctype}"))
        else:
            cols = [r[1] for r in conn.execute(text("PRAGMA table_info(lease_reviews)")).fetchall()]
            if 'prospect_property_id' not in cols:
                conn.execute(text(
                    "ALTER TABLE lease_reviews ADD COLUMN prospect_property_id INTEGER"))
            if 'rent_roll_date' not in cols:
                conn.execute(text(
                    "ALTER TABLE lease_reviews ADD COLUMN rent_roll_date TEXT"))

    # Phase 1D: workflow_step + step_data on lease_reviews
    _migrate_add_column(engine, 'lease_reviews', 'workflow_step', "TEXT DEFAULT 'setup'")
    _migrate_add_column(engine, 'lease_reviews', 'step_data', 'TEXT')

    # Phase 1A: rent_roll_source on lease_tenants
    _migrate_add_column(engine, 'lease_tenants', 'rent_roll_source', 'TEXT')

    # Phase 1E: per-tenant approval columns on lease_tenants
    _migrate_add_column(engine, 'lease_tenants', 'approval_status', "TEXT DEFAULT 'pending'")
    _migrate_add_column(engine, 'lease_tenants', 'approved_by', 'TEXT')
    _migrate_add_column(engine, 'lease_tenants', 'approved_at', 'TIMESTAMP')
    _migrate_add_column(engine, 'lease_tenants', 'analyst_notes', 'TEXT')

    # Rent roll extended columns on lease_tenants
    _migrate_add_column(engine, 'lease_tenants', 'monthly_rent_per_sf', 'DOUBLE PRECISION')
    _migrate_add_column(engine, 'lease_tenants', 'annual_rent_per_sf', 'DOUBLE PRECISION')
    _migrate_add_column(engine, 'lease_tenants', 'annual_recoveries_per_sf', 'DOUBLE PRECISION')
    _migrate_add_column(engine, 'lease_tenants', 'annual_misc_per_sf', 'DOUBLE PRECISION')
    _migrate_add_column(engine, 'lease_tenants', 'annual_sales_override', 'DOUBLE PRECISION')

    # Phase 2A: option_start / option_end on lease_options
    _migrate_add_column(engine, 'lease_options', 'option_start', 'TEXT')
    _migrate_add_column(engine, 'lease_options', 'option_end', 'TEXT')

    # Exclusive use: review_id so rows survive a tenant rebuild, clause_role to
    # separate a tenant's own exclusive from a restriction it is subject to,
    # and carve_outs for the existing-tenant exceptions that decide whether a
    # violation is actionable.
    _migrate_add_column(engine, 'lease_exclusive_use', 'review_id', 'INTEGER')
    _migrate_add_column(engine, 'lease_exclusive_use', 'clause_role', 'TEXT')
    _migrate_add_column(engine, 'lease_exclusive_use', 'carve_outs', 'TEXT')

    # Phase 1B: file_hash + uploaded_by on lease_documents; file_data for PDF storage
    _migrate_add_column(engine, 'lease_documents', 'file_hash', 'TEXT')
    _migrate_add_column(engine, 'lease_documents', 'uploaded_by', 'TEXT')
    _migrate_add_column(engine, 'lease_documents', 'file_data',
                        'BYTEA' if engine.dialect.name == 'postgresql' else 'BLOB')

    # Phase 1B+: Allow NULL tenant_id for unmatched documents
    _migrate_nullable(engine, 'lease_documents', 'tenant_id')

    # Per-document extraction JSON for consolidation
    _migrate_add_column(engine, 'lease_documents', 'extraction_json', 'TEXT')

    # Phase: Space planning — new columns on lease_tenants
    _migrate_add_column(engine, 'lease_tenants', 'tenant_status', "TEXT DEFAULT 'active'")
    _migrate_add_column(engine, 'lease_tenants', 'replaced_by_event_id', 'INTEGER')
    _migrate_add_column(engine, 'lease_tenants', 'successor_tenant_id', 'INTEGER')

    logger.info("Lease review tables ensured")


def create_lease_tables_sqlite(conn):
    """Create lease review tables on SQLite."""
    cursor = conn.cursor()
    for ddl in LEASE_DDL_SQLITE:
        cursor.execute(ddl)
    conn.commit()
    logger.info("Lease review tables created on SQLite")


# ---------------------------------------------------------------------------
# Folder scanning — discover properties, tenants, and lease PDFs
# ---------------------------------------------------------------------------

def scan_property_folder(
    base_path: str,
    property_name: str,
    property_address: str = '',
) -> Dict[str, Any]:
    """Scan a property's DD folder to discover tenants and lease documents.

    Expected folder structure:
        base_path/
            Tenant Leases/
                <Tenant Name>/
                    YYYY.MM.DD_Tenant-DocType.pdf
                    ...
                ! Tenant Lease Abstracts/
                    ...
            Rent Roll/
                *.xlsx
    """
    base = Path(base_path)
    leases_dir = base / 'Tenant Leases'
    abstracts_dir = leases_dir / '! Tenant Lease Abstracts'

    result = {
        'property_name': property_name,
        'property_address': property_address,
        'tenants': [],
        'abstracts': [],
        'cotenancy_file': None,
        'rent_roll_files': [],
    }

    # Find rent roll files
    rent_roll_dir = base / 'Rent Roll'
    if rent_roll_dir.exists():
        result['rent_roll_files'] = [
            str(f) for f in rent_roll_dir.glob('*.xlsx')
            if not f.name.startswith('~$')
        ]

    # Find cotenancy file
    if abstracts_dir.exists():
        for f in abstracts_dir.glob('*Co-Tenancy*'):
            result['cotenancy_file'] = str(f)
            break

    # Scan tenant folders
    if leases_dir.exists():
        for tenant_dir in sorted(leases_dir.iterdir()):
            if not tenant_dir.is_dir():
                continue
            if tenant_dir.name.startswith('!') or tenant_dir.name == 'REA Docs':
                continue

            pdfs = []
            for pdf in tenant_dir.rglob('*.pdf'):
                # Skip COI and Prior Owner Docs subfolders
                rel = pdf.relative_to(tenant_dir)
                parts = rel.parts
                if any(p.upper() in ('COI', 'PRIOR OWNER DOCS') for p in parts[:-1]):
                    continue
                pdfs.append({
                    'filename': pdf.name,
                    'path': str(pdf),
                    'doc_type': classify_document(pdf.name),
                    'doc_date': parse_doc_date(pdf.name),
                })

            result['tenants'].append({
                'folder_name': tenant_dir.name,
                'documents': sorted(pdfs, key=lambda d: d.get('doc_date') or ''),
                'doc_count': len(pdfs),
            })

    # Scan abstracts
    if abstracts_dir.exists():
        for f in abstracts_dir.glob('*.xlsx'):
            if not f.name.startswith('~$') and 'Co-Tenancy' not in f.name:
                result['abstracts'].append(str(f))

    return result


# ---------------------------------------------------------------------------
# Rent roll parsing
# ---------------------------------------------------------------------------

def parse_rent_roll(file_path: str) -> pd.DataFrame:
    """Parse a Windsor-style rent roll Excel file into a DataFrame.

    Returns columns: suite, tenant_name, lease_type, square_feet,
    lease_start, lease_end, term_months, monthly_rent, rent_per_sf_month,
    annual_rent, rent_per_sf_year, annual_recoveries_per_sf,
    annual_misc_per_sf, security_deposit, is_vacant.

    Derives missing gross/per-SF values from what's provided.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for r in range(1, ws.max_row + 1):
        col1 = ws.cell(r, 1).value
        col2 = ws.cell(r, 2).value
        col3 = ws.cell(r, 3).value

        # Skip header/summary rows — data rows start with property code
        if col1 is None or str(col1).startswith('Total') or str(col1).startswith('Future'):
            continue
        if str(col1).strip() in ('Rent Roll', '', 'Occupied', 'Vacant', 'Total'):
            continue
        if col2 is None:
            continue
        # Skip header rows
        if str(col2).strip() in ('Unit(s)', ''):
            continue
        # Check for property code pattern (e.g. f4wind)
        if not re.match(r'^[a-zA-Z]\d', str(col1).strip()):
            continue

        tenant_name = str(col3) if col3 else ''
        is_vacant = 'VACANT' in tenant_name.upper()

        def to_float(v):
            try:
                return float(v) if v is not None else 0.0
            except (ValueError, TypeError):
                return 0.0

        def to_date_str(v):
            if v is None or pd.isna(v):
                return None
            if isinstance(v, datetime):
                return v.strftime('%Y-%m-%d')
            if isinstance(v, date):
                return v.isoformat()
            s = str(v).strip()
            if s in ('None', 'TBD', '', 'NaT'):
                return None
            return s

        sf = to_float(ws.cell(r, 5).value)
        monthly_rent = to_float(ws.cell(r, 9).value)
        monthly_rent_per_sf = to_float(ws.cell(r, 10).value)
        annual_rent = to_float(ws.cell(r, 11).value)
        annual_rent_per_sf = to_float(ws.cell(r, 12).value)
        annual_rec_per_sf = to_float(ws.cell(r, 13).value)
        annual_misc_per_sf = to_float(ws.cell(r, 14).value)
        security_deposit = to_float(ws.cell(r, 15).value)

        # Derive missing gross/per-SF values from what's provided
        if sf and sf > 0:
            if annual_rent and not annual_rent_per_sf:
                annual_rent_per_sf = annual_rent / sf
            elif annual_rent_per_sf and not annual_rent:
                annual_rent = annual_rent_per_sf * sf
            if monthly_rent and not monthly_rent_per_sf:
                monthly_rent_per_sf = monthly_rent / sf
            elif monthly_rent_per_sf and not monthly_rent:
                monthly_rent = monthly_rent_per_sf * sf
        # Cross-derive monthly/annual when one is missing
        if annual_rent and not monthly_rent:
            monthly_rent = annual_rent / 12
        elif monthly_rent and not annual_rent:
            annual_rent = monthly_rent * 12
        if sf and sf > 0:
            if annual_rent and not annual_rent_per_sf:
                annual_rent_per_sf = annual_rent / sf
            if monthly_rent and not monthly_rent_per_sf:
                monthly_rent_per_sf = monthly_rent / sf

        rows.append({
            'property_code': str(col1).strip(),
            'suite': str(col2).strip() if col2 else '',
            'tenant_name': tenant_name.strip(),
            'lease_type': str(ws.cell(r, 4).value or '').strip(),
            'square_feet': sf,
            'lease_start': to_date_str(ws.cell(r, 6).value),
            'lease_end': to_date_str(ws.cell(r, 7).value),
            'term_months': int(to_float(ws.cell(r, 8).value)),
            'monthly_rent': monthly_rent,
            'rent_per_sf_month': monthly_rent_per_sf,
            'annual_rent': annual_rent,
            'rent_per_sf_year': annual_rent_per_sf,
            'annual_recoveries_per_sf': annual_rec_per_sf,
            'annual_misc_per_sf': annual_misc_per_sf,
            'security_deposit': security_deposit,
            'is_vacant': is_vacant,
        })

    return pd.DataFrame(rows)


def _parse_pdf_rent_roll(file_obj) -> pd.DataFrame:
    """Extract rent roll table from a PDF using pdfplumber.

    Scans all pages, extracts tables, and returns the best candidate —
    the largest table whose headers contain a tenant/name keyword.
    Multi-page tables with matching columns are concatenated.
    """
    import io
    import pdfplumber

    if isinstance(file_obj, bytes):
        file_obj = io.BytesIO(file_obj)

    best_df = None
    best_size = 0

    with pdfplumber.open(file_obj) as pdf:
        # Collect all tables across all pages
        all_tables = []
        for page in pdf.pages:
            tables = page.extract_tables()
            for tbl in tables:
                if tbl and len(tbl) >= 2:
                    all_tables.append(tbl)

    if not all_tables:
        raise ValueError("No tables found in PDF. The rent roll may need "
                         "to be converted to Excel or CSV first.")

    def _has_tenant_header(headers):
        lower_h = [str(h).lower() for h in headers if h]
        return any(kw in h for h in lower_h
                   for kw in ['tenant', 'name', 'lessee', 'occupant'])

    # Try to find and concatenate tables with matching headers
    groups = {}  # header_key -> list of row lists
    for tbl in all_tables:
        header = tuple(str(c).strip() if c else '' for c in tbl[0])
        key = header
        if key not in groups:
            groups[key] = {'header': list(header), 'rows': []}
        groups[key]['rows'].extend(tbl[1:])

    for key, grp in groups.items():
        headers = grp['header']
        if not _has_tenant_header(headers):
            continue
        rows = grp['rows']
        size = len(rows)
        if size > best_size:
            best_size = size
            best_df = pd.DataFrame(rows, columns=headers)

    # Fallback: use the largest table regardless of header match
    if best_df is None:
        for tbl in sorted(all_tables, key=len, reverse=True):
            headers = [str(c).strip() if c else f'col_{i}'
                       for i, c in enumerate(tbl[0])]
            best_df = pd.DataFrame(tbl[1:], columns=headers)
            break

    if best_df is None or best_df.empty:
        raise ValueError("Could not extract rent roll data from PDF tables.")

    # Clean up: strip whitespace, drop fully empty rows
    for col in best_df.columns:
        if best_df[col].dtype == object:
            best_df[col] = best_df[col].apply(
                lambda x: str(x).strip() if x and str(x).strip() not in ('None', '') else None)
    best_df = best_df.dropna(how='all').reset_index(drop=True)

    logger.info(f"PDF rent roll: extracted {len(best_df)} rows, "
                f"{len(best_df.columns)} columns")
    return best_df


def parse_rent_roll_flexible(file_obj, filename: str = '') -> pd.DataFrame:
    """Parse an uploaded rent roll from Excel, CSV, or PDF using flexible column matching.

    Handles Argus-format rent rolls, generic Excel exports, CSVs, and PDF tables.
    Returns a DataFrame with standardized columns matching parse_rent_roll output.

    Column matching is fuzzy — looks for keywords in header row:
      tenant/name, suite/unit, sf/area/sqft, start, end/expir,
      base rent/annual rent, monthly rent, lease type/status, deposit
    """
    import io

    # Read into DataFrame
    lower_fn = (filename or '').lower()
    if lower_fn.endswith('.pdf'):
        df_raw = _parse_pdf_rent_roll(file_obj)
    elif lower_fn.endswith('.csv'):
        if isinstance(file_obj, (str, bytes)):
            df_raw = pd.read_csv(io.BytesIO(file_obj) if isinstance(file_obj, bytes)
                                 else io.StringIO(file_obj))
        else:
            df_raw = pd.read_csv(file_obj)
    else:
        # Excel — try each sheet until we find one with tenant data
        import openpyxl
        if isinstance(file_obj, bytes):
            file_obj = io.BytesIO(file_obj)
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
        except Exception:
            # Fallback: try pandas directly
            file_obj.seek(0)
            df_raw = pd.read_excel(file_obj, engine='openpyxl')
            wb = None

        if wb is not None:
            df_raw = None
            # Prioritize sheets with rent roll / tenant in the name
            sheet_order = sorted(wb.sheetnames,
                key=lambda s: (0 if any(kw in s.lower()
                    for kw in ['rent roll', 'tenant', 'rent_roll']) else 1))
            for sheet_name in sheet_order:
                ws = wb[sheet_name]
                data = list(ws.values)
                if not data:
                    continue
                # Find header row — look for individual cells containing column names
                # Argus exports use two-row headers; we want the row with names
                header_idx = None
                for i, row in enumerate(data[:30]):
                    if row is None:
                        continue
                    # Check individual cells (not concatenated row string)
                    # to avoid matching "Tenant Rent Roll" as both tenant+rent
                    # Normalise newlines/carriage returns to spaces (Presentation
                    # format cells like "DBA\nTenant Name")
                    cell_vals = [str(c).replace('\n', ' ').replace('\r', ' ')
                                 .lower().strip()
                                 for c in row if c is not None]
                    logger.info("parse_rent_roll_flexible: sheet=%s row=%d cells=%s",
                                sheet_name, i, cell_vals[:8])
                    # Best: a cell that is exactly or contains "tenant name"
                    if any('tenant name' in cv for cv in cell_vals):
                        header_idx = i
                        break
                    # Require a tenant-like keyword in one cell AND a data
                    # keyword in a DIFFERENT cell.
                    # Recognised tenant keywords: "tenant", "dba",
                    # or exactly "lease" (but NOT "lease from/to/type")
                    def _is_tenant_cell(cv: str) -> bool:
                        if 'tenant' in cv or 'dba' in cv:
                            return len(cv) < 30
                        if cv == 'lease':
                            return True
                        return False

                    has_tenant = any(_is_tenant_cell(cv) for cv in cell_vals)
                    if has_tenant:
                        other_cells = [cv for cv in cell_vals
                                       if not _is_tenant_cell(cv)]
                        if any(kw in cv for cv in other_cells
                               for kw in ['suite', 'unit', 'rent',
                                          'sf', 'sqft', 'area',
                                          'square', 'footage', 'charge']):
                            header_idx = i
                            break
                if header_idx is not None:
                    # Check if prior row is a sub-header (Argus two-row header)
                    # Argus format: row N-1 has category prefixes (Potential,
                    # Scheduled, etc.), row N has the column names (Base Rent,
                    # Start Date, etc.). Concatenate them to get full names
                    # like "Potential Base Rent", "Scheduled Base Rent".
                    raw_headers = list(data[header_idx])
                    # --- Merge PRIOR row (Argus two-row: prefixes above) ---
                    if header_idx > 0:
                        prior = data[header_idx - 1]
                        if prior:
                            prior_str = ' '.join(
                                str(c).lower() for c in prior if c)
                            if any(kw in prior_str for kw in [
                                    'lease', 'rent', 'potential', 'expense',
                                    'scheduled', 'absorption']):
                                for j, c in enumerate(prior):
                                    if c:
                                        p = str(c).strip()
                                        if raw_headers[j]:
                                            # Concatenate: "Potential" + "Base Rent"
                                            raw_headers[j] = (
                                                p + ' ' + str(raw_headers[j]).strip())
                                        else:
                                            raw_headers[j] = p
                    # --- Merge SUBSEQUENT rows (MRI multi-row: sub-headers below) ---
                    # MRI rent rolls have the main header in row N but rent/rec
                    # sub-headers in rows N+1, N+2 that fill empty columns.
                    data_start = header_idx + 1
                    sub_kws = ['rent', 'rec', 'misc', 'deposit', 'guarantee',
                               'per area', 'per sf', 'monthly', 'annual']
                    for offset in range(1, 3):
                        sub_idx = header_idx + offset
                        if sub_idx >= len(data):
                            break
                        sub_row = data[sub_idx]
                        if sub_row is None:
                            break
                        sub_str = ' '.join(
                            str(c).lower() for c in sub_row if c)
                        if not any(kw in sub_str for kw in sub_kws):
                            break
                        # This row is a sub-header — merge into raw_headers
                        for j, c in enumerate(sub_row):
                            if c is None:
                                continue
                            s = str(c).strip()
                            if j < len(raw_headers) and raw_headers[j]:
                                raw_headers[j] = (
                                    str(raw_headers[j]).strip() + ' ' + s)
                            elif j < len(raw_headers):
                                raw_headers[j] = s
                            else:
                                # Extend headers for extra columns
                                while len(raw_headers) < j:
                                    raw_headers.append(None)
                                raw_headers.append(s)
                        data_start = sub_idx + 1
                    logger.info("parse_rent_roll_flexible: merged headers=%s",
                                [h for h in raw_headers if h])
                    headers = [str(c).strip() if c else f'col_{j}'
                               for j, c in enumerate(raw_headers)]
                    # Deduplicate column names — append _2, _3 etc.
                    seen: Dict[str, int] = {}
                    for j, h in enumerate(headers):
                        if h in seen:
                            seen[h] += 1
                            headers[j] = f'{h}_{seen[h]}'
                        else:
                            seen[h] = 1
                    rows = data[data_start:]
                    df_raw = pd.DataFrame(rows, columns=headers)
                    # If this sheet has enough rows, use it
                    if len(df_raw.dropna(how='all')) >= 2:
                        break
                    df_raw = None
            wb.close()
            if df_raw is None:
                raise ValueError("No rent roll data found in any sheet")

    if df_raw is None or df_raw.empty:
        raise ValueError("No data found in uploaded file")

    # Drop fully empty rows
    df_raw = df_raw.dropna(how='all').reset_index(drop=True)

    # Fuzzy column matching — normalise newlines in headers to spaces
    col_map = {}
    cols_lower = {c: c.replace('\n', ' ').lower().strip() for c in df_raw.columns}

    def _find_col(*keywords, exclude=None):
        for col, cl in cols_lower.items():
            if col in col_map.values():
                continue
            if exclude and any(e in cl for e in exclude):
                continue
            if any(kw in cl for kw in keywords):
                return col
        return None

    col_map['tenant_name'] = _find_col('tenant', 'name', 'lessee', 'dba', exclude=['group'])
    # Argus rent roll uses bare "lease" as the tenant name column
    if not col_map['tenant_name']:
        for col, cl in cols_lower.items():
            if cl == 'lease':
                col_map['tenant_name'] = col
                break
    col_map['suite'] = _find_col('suite', 'unit', 'space')
    col_map['square_feet'] = _find_col('area', 'sqft', 'sq ft', 'square', 'sf', 'gla',
                                       'footage')
    col_map['lease_type'] = _find_col('lease type', 'type', 'lease status', 'status')
    col_map['lease_start'] = _find_col('start date', 'lease start', 'commence',
                                       'begin', 'lease from')
    if not col_map['lease_start']:
        col_map['lease_start'] = _find_col('start', exclude=['date'])
    col_map['lease_end'] = _find_col('end date', 'lease end', 'expir', 'termin',
                                     'maturity', 'lease to')
    col_map['annual_rent'] = _find_col('scheduled base', 'potential base',
                                       'annual rent', 'base rent',
                                       exclude=['monthly', 'per sf', '/sf',
                                                'turnover', 'free',
                                                'miscellaneous', 'percentage',
                                                'absorption', 'rate',
                                                'per area'])
    col_map['monthly_rent'] = _find_col('monthly amount', 'monthly rent',
                                        'month rent',
                                        exclude=['per sf', '/sf', 'rate',
                                                 'per area'])
    # MRI rent roll: first "rent" column (no qualifier) is monthly rent
    if not col_map['monthly_rent']:
        for col, cl in cols_lower.items():
            if col in col_map.values():
                continue
            if cl == 'rent':
                col_map['monthly_rent'] = col
                break
    # MRI rent roll: second "rent" column (rent_2) is annual rent
    if not col_map['annual_rent']:
        for col, cl in cols_lower.items():
            if col in col_map.values():
                continue
            if cl.startswith('rent') and 'per' not in cl and 'rec' not in cl:
                col_map['annual_rent'] = col
                break
    col_map['annual_rent_per_sf'] = _find_col('annual rent per', 'annual base per',
                                               'annual $/sf', 'annual rate/sf',
                                               'annual rate', 'rent per sf',
                                               'rent/sf',
                                               exclude=['monthly', 'recover',
                                                        'misc', 'expense',
                                                        'future'])
    # MRI: "rent per area" columns — first is monthly, second is annual
    col_map['monthly_rent_per_sf'] = _find_col('monthly rent per', 'monthly $/sf',
                                                'monthly base per',
                                                exclude=['annual'])
    if not col_map['monthly_rent_per_sf']:
        col_map['monthly_rent_per_sf'] = _find_col('rent per area',
                                                     exclude=['recover', 'misc'])
    if not col_map['annual_rent_per_sf']:
        col_map['annual_rent_per_sf'] = _find_col('rent per area', 'per area',
                                                    exclude=['recover', 'misc'])
    col_map['annual_recoveries_per_sf'] = _find_col('recover', 'cam',
                                                     'reimburse', 'rec.',
                                                     exclude=['misc', 'annual rent',
                                                              'base rent'])
    col_map['annual_misc_per_sf'] = _find_col('misc', 'other per',
                                               exclude=['recover', 'annual rent',
                                                        'base rent'])
    col_map['security_deposit'] = _find_col('deposit', 'security')

    logger.info("parse_rent_roll_flexible: col_map=%s",
                {k: v for k, v in col_map.items() if v})

    if not col_map.get('tenant_name'):
        raise ValueError(
            f"Cannot find tenant name column. Headers: {list(df_raw.columns)}")

    def _safe_float(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return 0.0
        try:
            return float(str(val).replace(',', '').replace('$', '').strip())
        except (ValueError, TypeError):
            return 0.0

    def _safe_date(val):
        if val is None or pd.isna(val):
            return None
        if isinstance(val, (datetime, date)):
            return val.strftime('%Y-%m-%d')
        s = str(val).strip()
        if not s or s.lower() in ('none', 'nan', 'nat', 'tbd', ''):
            return None
        try:
            return pd.to_datetime(s).strftime('%Y-%m-%d')
        except Exception:
            return s

    result_rows = []
    for _, row in df_raw.iterrows():
        tname = row.get(col_map['tenant_name']) if col_map.get('tenant_name') else None
        if tname is None or (isinstance(tname, float) and pd.isna(tname)):
            continue
        tname = str(tname).strip()
        tname_lower = tname.lower()
        if not tname or tname_lower in ('total', 'totals', 'subtotal', 'grand total',
                                         'future', '', 'nan'):
            continue
        # Skip summary/total rows (e.g. "Total Area", "Current Leases")
        if any(kw in tname_lower for kw in ['total area', 'total unit',
               'current leases', 'future leases', 'month-to-month']):
            continue
        # Skip rows where tenant name is purely numeric (summary data)
        try:
            float(tname.replace(',', ''))
            continue  # purely numeric — not a tenant name
        except ValueError:
            pass

        is_vacant = 'VACANT' in tname.upper()
        sf = _safe_float(row.get(col_map.get('square_feet', ''), 0))
        ann_rent = _safe_float(row.get(col_map.get('annual_rent', ''), 0))
        mon_rent = _safe_float(row.get(col_map.get('monthly_rent', ''), 0))
        ann_rent_psf = _safe_float(row.get(col_map.get('annual_rent_per_sf', ''), 0))
        mon_rent_psf = _safe_float(row.get(col_map.get('monthly_rent_per_sf', ''), 0))
        ann_rec_psf = _safe_float(row.get(col_map.get('annual_recoveries_per_sf', ''), 0))
        ann_misc_psf = _safe_float(row.get(col_map.get('annual_misc_per_sf', ''), 0))

        # Derive missing gross/per-SF values
        if sf > 0:
            if ann_rent and not ann_rent_psf:
                ann_rent_psf = ann_rent / sf
            elif ann_rent_psf and not ann_rent:
                ann_rent = ann_rent_psf * sf
            if mon_rent and not mon_rent_psf:
                mon_rent_psf = mon_rent / sf
            elif mon_rent_psf and not mon_rent:
                mon_rent = mon_rent_psf * sf
        if ann_rent > 0 and mon_rent == 0:
            mon_rent = ann_rent / 12
        elif mon_rent > 0 and ann_rent == 0:
            ann_rent = mon_rent * 12
        if sf > 0:
            if ann_rent and not ann_rent_psf:
                ann_rent_psf = ann_rent / sf
            if mon_rent and not mon_rent_psf:
                mon_rent_psf = mon_rent / sf

        result_rows.append({
            'tenant_name': tname,
            'suite': str(row.get(col_map.get('suite', ''), '')).strip()
                    if col_map.get('suite') else '',
            'lease_type': str(row.get(col_map.get('lease_type', ''), '')).strip()
                         if col_map.get('lease_type') else 'Retail',
            'square_feet': sf,
            'lease_start': _safe_date(row.get(col_map.get('lease_start', '')))
                          if col_map.get('lease_start') else None,
            'lease_end': _safe_date(row.get(col_map.get('lease_end', '')))
                        if col_map.get('lease_end') else None,
            'term_months': 0,
            'monthly_rent': mon_rent,
            'rent_per_sf_month': mon_rent_psf,
            'annual_rent': ann_rent,
            'rent_per_sf_year': ann_rent_psf,
            'annual_recoveries_per_sf': ann_rec_psf,
            'annual_misc_per_sf': ann_misc_psf,
            'security_deposit': _safe_float(
                row.get(col_map.get('security_deposit', ''), 0))
                if col_map.get('security_deposit') else 0,
            'is_vacant': is_vacant,
        })

    if not result_rows:
        raise ValueError("No tenant rows found in uploaded file")

    return pd.DataFrame(result_rows)


def import_rent_roll_to_review(engine, review_id: int, rr_df: pd.DataFrame) -> int:
    """Import parsed rent roll data into an existing lease review.

    Clears existing tenants for the review and replaces with new data.
    Returns count of tenants imported.
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        # Verify review exists
        rev = conn.execute(text(
            "SELECT id FROM lease_reviews WHERE id = :rid"
        ), {'rid': review_id}).fetchone()
        if not rev:
            raise ValueError(f"Review {review_id} not found")

        # Exclusive use restrictions are keyed to tenant rows that are about
        # to be deleted, but they can be analyst-entered or seeded from the
        # exclusives spreadsheet -- neither of which a re-import reproduces.
        # Carry them across by tenant identity instead of dropping them.
        carried = conn.execute(text("""
            SELECT UPPER(TRIM(COALESCE(t.suite, ''))),
                   UPPER(TRIM(COALESCE(t.tenant_name, ''))),
                   e.restriction_text, e.restricted_use, e.radius_feet,
                   e.clause_role, e.carve_outs, e.source_doc
            FROM lease_exclusive_use e
            JOIN lease_tenants t ON t.id = e.tenant_id
            WHERE t.review_id = :rid
        """), {'rid': review_id}).fetchall()

        # Clear existing tenants and related data
        conn.execute(text(
            "DELETE FROM lease_validation WHERE tenant_id IN "
            "(SELECT id FROM lease_tenants WHERE review_id = :rid)"), {'rid': review_id})
        conn.execute(text(
            "DELETE FROM lease_cotenancy_refs WHERE cotenancy_id IN "
            "(SELECT id FROM lease_cotenancy WHERE review_id = :rid)"), {'rid': review_id})
        conn.execute(text(
            "DELETE FROM lease_cotenancy WHERE review_id = :rid"), {'rid': review_id})
        conn.execute(text(
            "DELETE FROM lease_documents WHERE review_id = :rid"), {'rid': review_id})
        conn.execute(text(
            "DELETE FROM lease_rent_steps WHERE tenant_id IN "
            "(SELECT id FROM lease_tenants WHERE review_id = :rid)"), {'rid': review_id})
        conn.execute(text(
            "DELETE FROM lease_options WHERE tenant_id IN "
            "(SELECT id FROM lease_tenants WHERE review_id = :rid)"), {'rid': review_id})
        conn.execute(text(
            "DELETE FROM lease_exclusive_use WHERE tenant_id IN "
            "(SELECT id FROM lease_tenants WHERE review_id = :rid)"), {'rid': review_id})
        conn.execute(text(
            "DELETE FROM lease_tenants WHERE review_id = :rid"), {'rid': review_id})

        # Insert tenants
        count = 0
        for _, row in rr_df.iterrows():
            tname = row.get('tenant_name', '')
            sf = float(row.get('square_feet', 0) or 0)
            ann_rent = float(row.get('annual_rent', 0) or 0)
            is_material = (sf >= MATERIAL_LEASE_SF_THRESHOLD or
                           ann_rent >= MATERIAL_LEASE_RENT_THRESHOLD)

            conn.execute(text("""
                INSERT INTO lease_tenants
                    (review_id, tenant_name, suite, square_feet, lease_type,
                     lease_start, lease_end, term_months, monthly_rent,
                     monthly_rent_per_sf, annual_rent, annual_rent_per_sf,
                     rent_per_sf, annual_recoveries_per_sf, annual_misc_per_sf,
                     security_deposit,
                     is_vacant, is_material, has_cotenancy, has_exclusive_use)
                VALUES (:rid, :tn, :su, :sf, :lt,
                        :ls, :le, :tm, :mr,
                        :mrpsf, :ar, :arpsf,
                        :rpsf, :arecpsf, :amiscpsf,
                        :sd,
                        :iv, :im, FALSE, FALSE)
            """), {
                'rid': review_id,
                'tn': tname,
                'su': row.get('suite', ''),
                'sf': sf,
                'lt': row.get('lease_type', 'Retail'),
                'ls': row.get('lease_start'),
                'le': row.get('lease_end'),
                'tm': int(row.get('term_months', 0) or 0),
                'mr': float(row.get('monthly_rent', 0) or 0),
                'mrpsf': float(row.get('rent_per_sf_month', 0) or 0),
                'ar': ann_rent,
                'arpsf': float(row.get('rent_per_sf_year', 0) or 0),
                'rpsf': float(row.get('rent_per_sf_year', 0) or 0),
                'arecpsf': float(row.get('annual_recoveries_per_sf', 0) or 0),
                'amiscpsf': float(row.get('annual_misc_per_sf', 0) or 0),
                'sd': float(row.get('security_deposit', 0) or 0),
                'iv': bool(row.get('is_vacant', False)),
                'im': is_material,
            })
            count += 1

        # Re-link carried exclusives to the rebuilt tenant rows.  Anything
        # whose tenant is no longer on the rent roll is genuinely gone.
        restored = 0
        for row in carried:
            match = conn.execute(text("""
                SELECT id FROM lease_tenants
                WHERE review_id = :rid
                  AND UPPER(TRIM(COALESCE(suite, ''))) = :su
                  AND UPPER(TRIM(COALESCE(tenant_name, ''))) = :tn
                LIMIT 1
            """), {'rid': review_id, 'su': row[0], 'tn': row[1]}).fetchone()
            if not match:
                continue
            conn.execute(text("""
                INSERT INTO lease_exclusive_use
                    (tenant_id, review_id, restriction_text, restricted_use,
                     radius_feet, clause_role, carve_outs, source_doc)
                VALUES (:tid, :rid, :rt, :ru, :rf, :cr, :co, :sd)
            """), {
                'tid': match[0], 'rid': review_id, 'rt': row[2], 'ru': row[3],
                'rf': row[4], 'cr': row[5], 'co': row[6], 'sd': row[7],
            })
            conn.execute(text(
                "UPDATE lease_tenants SET has_exclusive_use = TRUE WHERE id = :tid"),
                {'tid': match[0]})
            restored += 1
        dropped = len(carried) - restored

        # Update review totals
        total_gla = float(rr_df['square_feet'].sum()) if 'square_feet' in rr_df else 0
        total_rent = float(rr_df['annual_rent'].sum()) if 'annual_rent' in rr_df else 0
        conn.execute(text("""
            UPDATE lease_reviews
            SET total_gla = :gla, total_annual_rent = :rent,
                total_tenants = :cnt, updated_at = CURRENT_TIMESTAMP
            WHERE id = :rid
        """), {'gla': total_gla, 'rent': total_rent, 'cnt': count, 'rid': review_id})

        conn.commit()

    logger.info(f"Imported {count} tenants into review {review_id}")
    if carried:
        logger.info(
            f"Exclusive use: carried {restored} of {len(carried)} restrictions "
            f"across the re-import; {dropped} had no matching tenant"
        )
    return count


# ---------------------------------------------------------------------------
# Phase 1A: Non-destructive rent roll merge
# ---------------------------------------------------------------------------

def _fuzzy_match_tenant(suite_a: str, name_a: str, suite_b: str, name_b: str) -> bool:
    """Check if two tenant identifiers refer to the same tenant.

    Match by (suite exact) OR (name fuzzy containment when both non-empty).
    """
    sa = str(suite_a or '').strip().lower()
    sb = str(suite_b or '').strip().lower()
    na = str(name_a or '').strip().lower()
    nb = str(name_b or '').strip().lower()

    # Exact suite match (when suites are non-empty)
    if sa and sb and sa == sb:
        return True

    # Name containment match
    if na and nb and (na in nb or nb in na):
        return True

    return False


def merge_rent_roll_to_review(
    engine,
    review_id: int,
    rr_df: pd.DataFrame,
    source_label: str = 'seller_rent_roll',
) -> Dict[str, Any]:
    """Merge rent roll data into an existing review without destroying extraction data.

    Fuzzy-matches uploaded rows to existing tenants by (suite, tenant_name).
    - Matched tenants: Update tenant-level fields only; extraction data preserved.
    - New tenants: Insert new lease_tenants rows.
    - Missing tenants: Flag existing tenants not found in the upload (not deleted).

    Returns merge report: {matched, added, not_in_upload, details}.
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        rev = conn.execute(text(
            "SELECT id FROM lease_reviews WHERE id = :rid"
        ), {'rid': review_id}).fetchone()
        if not rev:
            raise ValueError(f"Review {review_id} not found")

        # Load existing tenants
        existing = conn.execute(text("""
            SELECT id, tenant_name, suite FROM lease_tenants
            WHERE review_id = :rid
        """), {'rid': review_id}).fetchall()

        existing_list = [{'id': r[0], 'name': r[1], 'suite': r[2]} for r in existing]

        matched_ids = set()
        matched_count = 0
        added_count = 0
        details = []

        for _, row in rr_df.iterrows():
            rr_name = str(row.get('tenant_name', '')).strip()
            rr_suite = str(row.get('suite', '')).strip()
            if not rr_name:
                continue

            # Try to match to an existing tenant
            match = None
            for ex in existing_list:
                if ex['id'] in matched_ids:
                    continue
                if _fuzzy_match_tenant(rr_suite, rr_name, ex['suite'], ex['name']):
                    match = ex
                    break

            sf = float(row.get('square_feet', 0) or 0)
            ann_rent = float(row.get('annual_rent', 0) or 0)
            mon_rent = float(row.get('monthly_rent', 0) or 0)
            rpsf = float(row.get('rent_per_sf_year', 0) or 0)
            mrpsf = float(row.get('rent_per_sf_month', 0) or 0)
            arecpsf = float(row.get('annual_recoveries_per_sf', 0) or 0)
            amiscpsf = float(row.get('annual_misc_per_sf', 0) or 0)
            is_material = (sf >= MATERIAL_LEASE_SF_THRESHOLD or
                           ann_rent >= MATERIAL_LEASE_RENT_THRESHOLD)

            if match:
                # Update tenant-level fields; DO NOT touch extraction data
                matched_ids.add(match['id'])
                conn.execute(text("""
                    UPDATE lease_tenants
                    SET tenant_name = :tn, suite = :su, square_feet = :sf,
                        lease_type = :lt, lease_start = :ls, lease_end = :le,
                        monthly_rent = :mr, monthly_rent_per_sf = :mrpsf,
                        annual_rent = :ar, annual_rent_per_sf = :arpsf,
                        rent_per_sf = :rpsf,
                        annual_recoveries_per_sf = :arecpsf,
                        annual_misc_per_sf = :amiscpsf,
                        security_deposit = :sd, is_vacant = :iv, is_material = :im,
                        rent_roll_source = :src, updated_at = CURRENT_TIMESTAMP
                    WHERE id = :tid
                """), {
                    'tn': rr_name, 'su': rr_suite, 'sf': sf,
                    'lt': row.get('lease_type', 'Retail'),
                    'ls': row.get('lease_start'), 'le': row.get('lease_end'),
                    'mr': mon_rent, 'mrpsf': mrpsf,
                    'ar': ann_rent, 'arpsf': rpsf, 'rpsf': rpsf,
                    'arecpsf': arecpsf, 'amiscpsf': amiscpsf,
                    'sd': float(row.get('security_deposit', 0) or 0),
                    'iv': bool(row.get('is_vacant', False)),
                    'im': is_material, 'src': source_label,
                    'tid': match['id'],
                })
                matched_count += 1
                details.append({'action': 'updated', 'tenant': rr_name, 'suite': rr_suite})
            else:
                # Insert new tenant
                conn.execute(text("""
                    INSERT INTO lease_tenants
                        (review_id, tenant_name, suite, square_feet, lease_type,
                         lease_start, lease_end, term_months, monthly_rent,
                         monthly_rent_per_sf, annual_rent, annual_rent_per_sf,
                         rent_per_sf, annual_recoveries_per_sf, annual_misc_per_sf,
                         security_deposit,
                         is_vacant, is_material, has_cotenancy, has_exclusive_use,
                         rent_roll_source)
                    VALUES (:rid, :tn, :su, :sf, :lt,
                            :ls, :le, :tm, :mr,
                            :mrpsf, :ar, :arpsf,
                            :rpsf, :arecpsf, :amiscpsf,
                            :sd,
                            :iv, :im, FALSE, FALSE, :src)
                """), {
                    'rid': review_id, 'tn': rr_name, 'su': rr_suite, 'sf': sf,
                    'lt': row.get('lease_type', 'Retail'),
                    'ls': row.get('lease_start'), 'le': row.get('lease_end'),
                    'tm': int(row.get('term_months', 0) or 0),
                    'mr': mon_rent, 'mrpsf': mrpsf,
                    'ar': ann_rent, 'arpsf': rpsf, 'rpsf': rpsf,
                    'arecpsf': arecpsf, 'amiscpsf': amiscpsf,
                    'sd': float(row.get('security_deposit', 0) or 0),
                    'iv': bool(row.get('is_vacant', False)),
                    'im': is_material, 'src': source_label,
                })
                added_count += 1
                details.append({'action': 'added', 'tenant': rr_name, 'suite': rr_suite})

        # Flag tenants not in the upload
        not_in_upload = []
        for ex in existing_list:
            if ex['id'] not in matched_ids:
                not_in_upload.append({'tenant': ex['name'], 'suite': ex['suite']})

        # Update review totals from current state
        totals = conn.execute(text("""
            SELECT COALESCE(SUM(square_feet), 0),
                   COALESCE(SUM(annual_rent), 0),
                   COUNT(*)
            FROM lease_tenants WHERE review_id = :rid
        """), {'rid': review_id}).fetchone()

        conn.execute(text("""
            UPDATE lease_reviews
            SET total_gla = :gla, total_annual_rent = :rent,
                total_tenants = :cnt, updated_at = CURRENT_TIMESTAMP
            WHERE id = :rid
        """), {'gla': totals[0], 'rent': totals[1], 'cnt': totals[2], 'rid': review_id})

        conn.commit()

    report = {
        'matched': matched_count,
        'added': added_count,
        'not_in_upload': len(not_in_upload),
        'not_in_upload_tenants': not_in_upload,
        'details': details,
    }
    logger.info(f"Merge rent roll review {review_id}: {report['matched']} matched, "
                f"{report['added']} added, {report['not_in_upload']} not in upload")
    return report


# ---------------------------------------------------------------------------
# Phase 1B: Incremental document upload
# ---------------------------------------------------------------------------

def upload_documents_to_review(
    engine,
    review_id: int,
    files: List[Tuple[str, bytes]],
    uploaded_by: str = 'system',
    folder_hints: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Upload multiple PDF documents to a review with dedup and auto-matching.

    Args:
        files: List of (filename, file_bytes) tuples.
        folder_hints: Optional list of subfolder names (one per file, same order)
                      used as additional signal for tenant matching.

    Returns dict with counts: {added, skipped_duplicate, unmatched, details}.
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        rev = conn.execute(text(
            "SELECT id FROM lease_reviews WHERE id = :rid"
        ), {'rid': review_id}).fetchone()
        if not rev:
            raise ValueError(f"Review {review_id} not found")

        # Load existing document hashes for dedup
        existing_hashes = set()
        rows = conn.execute(text("""
            SELECT file_hash FROM lease_documents
            WHERE review_id = :rid AND file_hash IS NOT NULL
        """), {'rid': review_id}).fetchall()
        for r in rows:
            existing_hashes.add(r[0])

        # Load tenants for auto-matching
        tenants = conn.execute(text("""
            SELECT id, tenant_name, suite FROM lease_tenants
            WHERE review_id = :rid AND is_vacant = false
        """), {'rid': review_id}).fetchall()

        added = 0
        skipped = 0
        unmatched = 0
        details = []

        for i, (filename, file_bytes) in enumerate(files):
            file_hash = hashlib.sha256(file_bytes).hexdigest()

            # Dedup check
            if file_hash in existing_hashes:
                skipped += 1
                details.append({'filename': filename, 'action': 'skipped_duplicate'})
                continue

            # Classify document type
            doc_type = classify_document(filename)
            doc_date = parse_doc_date(filename)

            # Fuzzy-match to tenant by filename + optional folder hint
            folder_hint = (folder_hints[i]
                           if folder_hints and i < len(folder_hints)
                           else None)
            tenant_id = _match_file_to_tenant(filename, tenants,
                                              folder_hint=folder_hint)

            if tenant_id is None:
                unmatched += 1
                # Store with NULL tenant_id — user can assign manually
                conn.execute(text("""
                    INSERT INTO lease_documents
                        (tenant_id, review_id, filename, doc_type, doc_date,
                         extraction_status, file_hash, uploaded_by, file_data)
                    VALUES (NULL, :rid, :fn, :dt, :dd,
                            'pending', :fh, :ub, :fd)
                """), {
                    'rid': review_id,
                    'fn': filename, 'dt': doc_type, 'dd': doc_date,
                    'fh': file_hash, 'ub': uploaded_by, 'fd': file_bytes,
                })
                existing_hashes.add(file_hash)
                added += 1
                details.append({
                    'filename': filename, 'action': 'unmatched',
                    'doc_type': doc_type,
                })
                continue

            conn.execute(text("""
                INSERT INTO lease_documents
                    (tenant_id, review_id, filename, doc_type, doc_date,
                     extraction_status, file_hash, uploaded_by, file_data)
                VALUES (:tid, :rid, :fn, :dt, :dd,
                        'pending', :fh, :ub, :fd)
            """), {
                'tid': tenant_id, 'rid': review_id,
                'fn': filename, 'dt': doc_type, 'dd': doc_date,
                'fh': file_hash, 'ub': uploaded_by, 'fd': file_bytes,
            })
            existing_hashes.add(file_hash)
            added += 1
            details.append({
                'filename': filename, 'action': 'added',
                'doc_type': doc_type, 'tenant_id': tenant_id,
            })

        conn.commit()

    report = {
        'added': added,
        'skipped_duplicate': skipped,
        'unmatched': unmatched,
        'details': details,
    }
    logger.info(f"Uploaded docs to review {review_id}: {added} added, "
                f"{skipped} deduped, {unmatched} unmatched")
    return report


def _match_file_to_tenant(
    filename: str,
    tenants: list,
    folder_hint: Optional[str] = None,
) -> Optional[int]:
    """Match a PDF filename to a tenant by name containment.

    If *folder_hint* is provided (the subfolder name the file was stored in),
    it is used as an additional matching signal.  A folder-hint match is tried
    first — if the subfolder name matches a tenant, that wins.

    Returns tenant_id or None.
    """
    fname_lower = filename.lower().replace('.pdf', '').replace('_', ' ').replace('-', ' ')
    # Remove common prefixes/suffixes like date stamps
    fname_clean = re.sub(r'^\d{4}[.\-]\d{2}[.\-]\d{2}[_\s]*', '', fname_lower).strip()

    # --- Try folder hint first (subfolder name = tenant name) ---
    if folder_hint and folder_hint.strip():
        hint = folder_hint.strip().lower().replace('_', ' ').replace('-', ' ')
        hint_match = None
        hint_len = 0
        for t in tenants:
            tname = str(t[1]).strip().lower()
            if not tname:
                continue
            if tname in hint or hint in tname:
                if len(tname) > hint_len:
                    hint_match = t[0]
                    hint_len = len(tname)
        if hint_match is not None:
            return hint_match

    # --- Fall back to filename matching ---
    best_match = None
    best_len = 0

    for t in tenants:
        tname = str(t[1]).strip().lower()
        if not tname:
            continue
        # Check if tenant name appears in filename or vice versa
        if tname in fname_clean or fname_clean in tname:
            if len(tname) > best_len:
                best_match = t[0]  # tenant_id
                best_len = len(tname)
        # Try first word match for multi-word names
        first_word = tname.split()[0] if tname.split() else ''
        if first_word and len(first_word) > 3 and first_word in fname_clean:
            if len(tname) > best_len:
                best_match = t[0]
                best_len = len(tname)

    return best_match


def get_unmatched_documents(engine, review_id: int) -> List[Dict[str, Any]]:
    """Return documents with no tenant assignment (tenant_id IS NULL)."""
    from sqlalchemy import text
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT id, filename, doc_type, doc_date, extraction_status, uploaded_by
            FROM lease_documents
            WHERE review_id = :rid AND tenant_id IS NULL
            ORDER BY filename
        """), {'rid': review_id}).fetchall()
        return [dict(r._mapping) for r in rows]


def assign_document_to_tenant(
    engine,
    review_id: int,
    doc_id: int,
    tenant_id: int,
) -> None:
    """Assign an unmatched document to a specific tenant."""
    from sqlalchemy import text

    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE lease_documents
            SET tenant_id = :tid
            WHERE id = :did AND review_id = :rid
        """), {'tid': tenant_id, 'did': doc_id, 'rid': review_id})


# ---------------------------------------------------------------------------
# Phase 1E: Per-tenant approval
# ---------------------------------------------------------------------------

def approve_tenant(
    engine,
    review_id: int,
    tenant_id: int,
    status: str,
    approved_by: str,
    notes: Optional[str] = None,
) -> Dict[str, Any]:
    """Set approval status for a tenant.

    Args:
        status: 'approved', 'flagged', or 'pending'
    """
    from sqlalchemy import text

    if status not in ('approved', 'flagged', 'pending'):
        raise ValueError(f"Invalid approval status: {status}")

    with engine.connect() as conn:
        # Verify tenant belongs to this review
        t = conn.execute(text("""
            SELECT id FROM lease_tenants
            WHERE id = :tid AND review_id = :rid
        """), {'tid': tenant_id, 'rid': review_id}).fetchone()
        if not t:
            raise ValueError(f"Tenant {tenant_id} not found in review {review_id}")

        update_params = {
            'status': status,
            'by': approved_by if status == 'approved' else None,
            'at': datetime.utcnow().isoformat() if status == 'approved' else None,
            'tid': tenant_id,
        }
        if notes is not None:
            conn.execute(text("""
                UPDATE lease_tenants
                SET approval_status = :status, approved_by = :by,
                    approved_at = :at, analyst_notes = :notes,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :tid
            """), {**update_params, 'notes': notes})
        else:
            conn.execute(text("""
                UPDATE lease_tenants
                SET approval_status = :status, approved_by = :by,
                    approved_at = :at, updated_at = CURRENT_TIMESTAMP
                WHERE id = :tid
            """), update_params)

        conn.commit()

    return {'tenant_id': tenant_id, 'status': status}


# ---------------------------------------------------------------------------
# Phase 1D: Workflow progress
# ---------------------------------------------------------------------------

def get_workflow_progress(engine, review_id: int) -> Dict[str, Any]:
    """Get workflow step progress metrics for a review."""
    from sqlalchemy import text

    with engine.connect() as conn:
        rev = conn.execute(text("""
            SELECT workflow_step, step_data, total_tenants
            FROM lease_reviews WHERE id = :rid
        """), {'rid': review_id}).fetchone()
        if not rev:
            raise ValueError(f"Review {review_id} not found")

        workflow_step = rev[0] or 'setup'
        step_data = json.loads(rev[1]) if rev[1] else {}
        total_tenants = rev[2] or 0

        # Count metrics
        tenant_counts = conn.execute(text("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN is_vacant = false THEN 1 ELSE 0 END) as occupied,
                SUM(CASE WHEN approval_status = 'approved' THEN 1 ELSE 0 END) as approved,
                SUM(CASE WHEN approval_status = 'flagged' THEN 1 ELSE 0 END) as flagged
            FROM lease_tenants WHERE review_id = :rid
        """), {'rid': review_id}).fetchone()

        doc_counts = conn.execute(text("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN extraction_status = 'extracted' THEN 1 ELSE 0 END) as extracted,
                SUM(CASE WHEN extraction_status = 'pending' THEN 1 ELSE 0 END) as pending
            FROM lease_documents WHERE review_id = :rid
        """), {'rid': review_id}).fetchone()

        val_count = conn.execute(text("""
            SELECT COUNT(*) FROM lease_validation
            WHERE tenant_id IN (SELECT id FROM lease_tenants WHERE review_id = :rid)
        """), {'rid': review_id}).fetchone()

    return {
        'current_step': workflow_step,
        'step_data': step_data,
        'tenants_imported': tenant_counts[0] or 0,
        'tenants_occupied': tenant_counts[1] or 0,
        'tenants_approved': tenant_counts[2] or 0,
        'tenants_flagged': tenant_counts[3] or 0,
        'docs_uploaded': doc_counts[0] or 0,
        'docs_extracted': doc_counts[1] or 0,
        'docs_pending': doc_counts[2] or 0,
        'validations_run': val_count[0] or 0,
    }


def update_workflow_step(engine, review_id: int, step: str) -> None:
    """Update the current workflow step for a review."""
    from sqlalchemy import text

    valid_steps = ['setup', 'rent_roll', 'documents', 'extraction',
                   'validation', 'review', 'complete']
    if step not in valid_steps:
        raise ValueError(f"Invalid step: {step}. Must be one of {valid_steps}")

    with engine.begin() as conn:
        # Load current step_data and add timestamp
        rev = conn.execute(text(
            "SELECT step_data FROM lease_reviews WHERE id = :rid"
        ), {'rid': review_id}).fetchone()
        step_data = json.loads(rev[0]) if rev and rev[0] else {}
        step_data[step] = datetime.utcnow().isoformat()

        conn.execute(text("""
            UPDATE lease_reviews
            SET workflow_step = :step, step_data = :sd,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = :rid
        """), {'step': step, 'sd': json.dumps(step_data), 'rid': review_id})


def create_review_manual(engine, property_name: str, property_address: str = '',
                         total_gla: float = 0, created_by: str = 'system',
                         prospect_property_id: int = None) -> int:
    """Create a lease review without folder scanning.

    Returns the review_id.
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        params = {
            'pn': property_name, 'pa': property_address,
            'gla': total_gla, 'ppid': prospect_property_id,
            'cb': created_by,
        }
        result = conn.execute(text("""
            INSERT INTO lease_reviews
                (property_name, property_address, total_gla,
                 prospect_property_id, status, created_by)
            VALUES (:pn, :pa, :gla, :ppid, 'in_progress', :cb)
            RETURNING id
        """), params)
        review_id = result.fetchone()[0]
        conn.commit()

    logger.info(f"Created manual review '{property_name}': review_id={review_id}")
    return review_id


# ---------------------------------------------------------------------------
# Cotenancy spreadsheet parsing
# ---------------------------------------------------------------------------

def parse_cotenancy_spreadsheet(file_path: str) -> Tuple[List[Dict], List[Dict]]:
    """Parse the Co-Tenancy and Exclusives spreadsheet.

    Returns (cotenancy_records, exclusive_use_records).
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)

    cotenancy = []
    if 'Co-Tenancy' in wb.sheetnames:
        ws = wb['Co-Tenancy']
        for r in range(4, ws.max_row + 1):
            tenant = ws.cell(r, 1).value
            if tenant is None:
                continue
            tenant = str(tenant).strip()
            clause = str(ws.cell(r, 4).value or '')
            has_clause = clause and 'No Lease provision' not in clause and clause != 'None'
            cotenancy.append({
                'tenant_name': tenant,
                'suite': str(ws.cell(r, 2).value or '').strip(),
                'square_feet': float(ws.cell(r, 3).value or 0),
                'clause_text': clause if has_clause else None,
                'has_cotenancy': has_clause,
            })

    exclusive = []
    if 'Exclusive Use' in wb.sheetnames:
        ws = wb['Exclusive Use']
        for r in range(4, ws.max_row + 1):
            tenant = ws.cell(r, 1).value
            if tenant is None:
                continue
            tenant = str(tenant).strip()
            restriction = str(ws.cell(r, 4).value or '')
            has_restriction = (restriction and
                               'No lease provision' not in restriction.lower() and
                               restriction != 'None')
            if has_restriction:
                exclusive.append({
                    'tenant_name': tenant,
                    'suite': str(ws.cell(r, 2).value or '').strip(),
                    'square_feet': float(ws.cell(r, 3).value or 0),
                    'restriction_text': restriction,
                })

    return cotenancy, exclusive


# ---------------------------------------------------------------------------
# PDF text extraction
# ---------------------------------------------------------------------------

def extract_pdf_text(source) -> Tuple[str, int]:
    """Extract text from a PDF using PyMuPDF. Returns (text, page_count).

    Args:
        source: Either a file path (str) or raw PDF bytes.
    """
    try:
        import pymupdf
    except ImportError:
        import fitz as pymupdf

    text_parts = []
    if isinstance(source, (bytes, bytearray, memoryview)):
        doc = pymupdf.open(stream=bytes(source), filetype="pdf")
    else:
        doc = pymupdf.open(source)
    page_count = len(doc)
    for page in doc:
        text_parts.append(page.get_text())
    doc.close()
    return '\n'.join(text_parts), page_count


# ---------------------------------------------------------------------------
# Claude API extraction
# ---------------------------------------------------------------------------

# A value is a number only if it STARTS with one, after an optional currency
# symbol or sign. "60 days" is 60; "$0.50 per sf" is 0.50; but "Greater of CPI
# increase or $0.50 per sf" is not 0.50 -- taking a number from the middle of a
# sentence silently records a figure the lease does not say, which is worse
# than recording nothing.
_NUM_LEAD = re.compile(r"^\(?\s*[-+]?\s*\$?\s*(\d[\d,]*(?:\.\d+)?)\s*\)?")


def _to_number(value):
    """Coerce an extracted value to a float, or None if it is not a number.

    Extraction returns prose wherever a lease is descriptive rather than
    numeric -- "Greater of CPI increase or $0.50 per sf" is a real answer to
    "rent per sf". Passing that into a DOUBLE PRECISION column raises a
    DataError on PostgreSQL, which aborts the surrounding transaction and
    takes the rest of the extraction run down with it. Unparseable values
    become NULL; the descriptive text belongs in a text column, not here.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    s = value.strip()
    if not s:
        return None
    m = _NUM_LEAD.match(s)
    if not m:
        return None
    try:
        n = float(m.group(1).replace(",", ""))
    except (TypeError, ValueError):
        return None
    lead = s[:m.end()]
    if "(" in lead or lead.lstrip().startswith("-"):
        n = -n
    return n


def _to_int(value):
    """Same as _to_number, rounded, for INTEGER columns."""
    n = _to_number(value)
    return None if n is None else int(round(n))


# Phrases a model (or a seeded spreadsheet) uses to say "there is no
# exclusive".  These are the absence of a restriction, not a restriction, and
# must not be stored as one.
_EXCLUSIVE_NEGATIVES = {
    'no lease provision', 'none', 'n/a', 'na', 'no', 'no exclusive',
    'no exclusive use', 'no exclusive use provision', 'not applicable',
    'no restriction', 'no restrictions', 'silent', 'not addressed',
}

EXTRACTION_PROMPT = """You are a commercial real estate lease analyst. Extract the following structured information from this lease document.

TENANT: {tenant_name}
SUITE: {suite}
DOCUMENT TYPE: {doc_type}

Return a JSON object with these fields (use null for fields not found):

{{
  "tenant_legal_name": "...",
  "suite": "...",
  "square_feet": number,
  "permitted_use": "...",
  "lease_commencement": "YYYY-MM-DD",
  "rent_commencement": "YYYY-MM-DD",
  "lease_expiration": "YYYY-MM-DD",
  "holdover_rate": "...",
  "rent_steps": [
    {{"effective_date": "YYYY-MM-DD", "monthly_rent": number, "annual_rent": number, "rent_per_sf": number}}
  ],
  "escalation_structure": "fixed $ / fixed % / CPI / other",
  "security_deposit": number,
  "cam_structure": "pro rata / fixed / gross",
  "cam_cap_pct": number or null,
  "admin_fee_pct": number or null,
  "tax_pass_through": true/false,
  "insurance_pass_through": true/false,
  "cotenancy": {{
    "has_clause": true/false,
    "named_cotenants": ["Tenant A", "Tenant B"],
    "trigger_threshold": "description of what triggers",
    "cure_period_days": number,
    "alt_rent_formula": "e.g. 50% of base rent or 2% of gross sales",
    "termination_right": true/false,
    "termination_notice_days": number,
    "sunset_or_waiver": "description of how right expires",
    "is_curable": true/false,
    "uncurable_scenario": "description if applicable"
  }},
  "exclusive_use": [
    {{
      "clause_role": "holder / subject",
      "restricted_use": "the use that is protected or prohibited, e.g. sale of pet supplies",
      "restriction_text": "verbatim text of the operative sentence",
      "carve_outs": "existing tenants or uses excepted from the restriction, or null",
      "radius_feet": number or null,
      "source_section": "e.g. Section 1.3 / Exhibit D / Addendum 2"
    }}
  ],
  "renewal_options": [
    {{
      "option_number": 1,
      "total_options": number,
      "term_years": number,
      "option_start": "YYYY-MM-DD or null",
      "option_end": "YYYY-MM-DD or null",
      "notice_days": number,
      "notice_deadline": "YYYY-MM-DD or null",
      "auto_renewal": true/false,
      "exercised": true/false,
      "rent_terms": "fair market / fixed increase / CPI"
    }}
  ],
  "termination_options": [
    {{
      "option_number": 1,
      "total_options": number,
      "earliest_termination_date": "YYYY-MM-DD or null",
      "notice_days": number,
      "notice_deadline": "YYYY-MM-DD or null",
      "exercised": true/false,
      "termination_fee": "description of fee or penalty if any",
      "conditions": "description of conditions required to exercise"
    }}
  ],
  "assignment": {{
    "consent_required": true/false,
    "tenant_released": true/false
  }},
  "ti_allowance": number or null,
  "sales_provisions": {{
    "sales_reporting_required": true/false,
    "reporting_frequency": "monthly / quarterly / annually / null",
    "reporting_deadline": "e.g. 30 days after period end / null",
    "audit_right": true/false,
    "percentage_rent": {{
      "has_clause": true/false,
      "breakpoint": number or null,
      "rate_pct": number or null
    }},
    "sales_performance_clauses": [
      {{
        "clause_type": "kick-out / reduced rent / recapture / radius restriction / other",
        "trigger": "description of sales threshold or condition that activates the clause",
        "threshold_amount": number or null,
        "threshold_period": "trailing 12 months / lease year / other",
        "consequence": "description of what happens — e.g. reduced rent to X, right to terminate, landlord recapture right",
        "beneficiary": "tenant / landlord",
        "notice_days": number or null,
        "cure_period_days": number or null
      }}
    ]
  }},
  "go_dark_provision": "...",
  "key_dates": {{
    "next_notice_deadline": "YYYY-MM-DD or null",
    "description": "..."
  }}
}}

IMPORTANT:
- Extract rent steps from the rent schedule if present
- For amendments, only return CHANGED fields; unchanged fields should be null
- Dates must be YYYY-MM-DD format
- Dollar amounts should be numbers (no $ signs)
- If this is an amendment, note which fields were modified
- For renewal_options: option_start/option_end are the beginning and ending dates of each renewal period. If not explicitly stated, derive from the prior term's expiration + term_years. Mark exercised=true if an amendment or exercise notice confirms the option was exercised.
- For termination_options: extract early termination rights, kick-out clauses, and similar provisions. earliest_termination_date is when the tenant can first terminate. Mark exercised=true if a termination notice was exercised.
- For exclusive_use: this is easy to miss, so search the whole document, not
  just a heading called "Exclusive". These provisions appear under Permitted
  Use, Exclusive Use, Use Restrictions, Prohibited Uses, Restrictive
  Covenants, Radius Restriction, Continuous Operation, or in an exhibit,
  addendum, rider or site-plan attachment. Return one object per distinct
  restriction -- a lease often has several, and an empty list means you found
  none. Set clause_role to "holder" when THIS tenant holds the exclusive (the
  landlord may not lease to a competing use) and "subject" when this tenant is
  bound by someone else's exclusive or by a prohibited-use list. Capture
  carve_outs whenever named existing tenants or uses are excepted, since they
  determine whether a violation is actionable. Quote restriction_text
  verbatim rather than paraphrasing.
- For sales_provisions: Look for any requirement to report gross sales, certified sales statements, or sales audits. Extract percentage rent (overage rent) breakpoints and rates. For sales_performance_clauses, capture any provision where tenant sales performance triggers a consequence — including tenant kick-out rights (right to terminate if sales fall below a threshold), landlord recapture rights, reduced/alternative rent tied to sales levels, and radius restrictions limiting competing stores. Each clause should specify who benefits (tenant or landlord) and the exact consequence.

DOCUMENT TEXT:
{text}"""


def extract_lease_terms_via_api(
    text: str,
    tenant_name: str,
    suite: str,
    doc_type: str,
    api_key: Optional[str] = None,
) -> Dict[str, Any]:
    """Call Claude API to extract structured lease terms from PDF text.

    Uses Haiku for cost efficiency on high-volume extraction.
    """
    import anthropic

    key = api_key or os.environ.get('ANTHROPIC_API_KEY')
    if not key:
        raise ValueError("ANTHROPIC_API_KEY not set")

    client = anthropic.Anthropic(api_key=key)

    # Truncate very long documents to stay within context
    max_chars = 180_000
    if len(text) > max_chars:
        text = text[:max_chars] + "\n\n[TRUNCATED — document exceeds extraction limit]"

    prompt = EXTRACTION_PROMPT.format(
        tenant_name=tenant_name,
        suite=suite,
        doc_type=doc_type,
        text=text,
    )

    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )

    response_text = message.content[0].text

    # Parse JSON from response
    try:
        # Try to find JSON block
        json_match = re.search(r'\{[\s\S]*\}', response_text)
        if json_match:
            return json.loads(json_match.group())
    except json.JSONDecodeError:
        logger.warning(f"Failed to parse JSON for {tenant_name}: {response_text[:200]}")

    return {'_raw_response': response_text, '_parse_error': True}


# ---------------------------------------------------------------------------
# Ingestion pipeline — load property into database
# ---------------------------------------------------------------------------

def ingest_property(
    engine,
    base_path: str,
    property_name: str,
    property_address: str = '',
    rent_roll_path: Optional[str] = None,
    created_by: str = 'system',
) -> int:
    """Ingest a property's lease data into the database.

    Steps:
    1. Scan folder structure
    2. Parse rent roll
    3. Parse cotenancy spreadsheet
    4. Create database records
    5. Catalog all PDF documents

    Returns the review_id.
    """
    from sqlalchemy import text

    scan = scan_property_folder(base_path, property_name, property_address)

    # Parse rent roll
    rr_df = pd.DataFrame()
    rr_path = rent_roll_path
    if not rr_path and scan['rent_roll_files']:
        # Use most recent rent roll
        rr_path = sorted(scan['rent_roll_files'])[-1]
    if rr_path:
        rr_df = parse_rent_roll(rr_path)

    # Parse cotenancy
    cotenancy_data = []
    exclusive_data = []
    if scan['cotenancy_file']:
        cotenancy_data, exclusive_data = parse_cotenancy_spreadsheet(
            scan['cotenancy_file']
        )

    # Build cotenancy lookup by tenant name (fuzzy)
    cot_lookup = {}
    for rec in cotenancy_data:
        cot_lookup[rec['tenant_name'].lower().strip()] = rec

    exc_lookup = {}
    for rec in exclusive_data:
        exc_lookup[rec['tenant_name'].lower().strip()] = rec

    with engine.connect() as conn:
        # Create review
        total_gla = float(rr_df['square_feet'].sum()) if len(rr_df) else 0
        total_rent = float(rr_df['annual_rent'].sum()) if len(rr_df) else 0
        total_tenants = len(rr_df) if len(rr_df) else len(scan['tenants'])

        result = conn.execute(text("""
            INSERT INTO lease_reviews
                (property_name, property_address, total_gla,
                 total_annual_rent, total_tenants, source_folder, created_by)
            VALUES (:pn, :pa, :gla, :rent, :cnt, :sf, :cb)
            RETURNING id
        """), {
            'pn': property_name, 'pa': property_address,
            'gla': total_gla, 'rent': total_rent,
            'cnt': total_tenants, 'sf': base_path, 'cb': created_by,
        })
        review_id = result.fetchone()[0]

        # Insert tenants from rent roll
        tenant_id_map = {}  # suite -> tenant_id
        if len(rr_df):
            for _, row in rr_df.iterrows():
                tname = row['tenant_name']
                suite = row['suite']
                sf = row['square_feet']
                ann_rent = row['annual_rent']

                # Check if material
                is_material = (sf >= MATERIAL_LEASE_SF_THRESHOLD or
                               ann_rent >= MATERIAL_LEASE_RENT_THRESHOLD)

                # Check cotenancy (fuzzy match)
                has_cot = _fuzzy_match_cotenancy(tname, cot_lookup)
                has_exc = _fuzzy_match_cotenancy(tname, exc_lookup)

                r = conn.execute(text("""
                    INSERT INTO lease_tenants
                        (review_id, tenant_name, suite, square_feet, lease_type,
                         lease_start, lease_end, term_months, monthly_rent,
                         annual_rent, rent_per_sf, security_deposit,
                         is_vacant, is_material, has_cotenancy, has_exclusive_use)
                    VALUES (:rid, :tn, :su, :sf, :lt,
                            :ls, :le, :tm, :mr,
                            :ar, :rpsf, :sd,
                            :iv, :im, :hc, :heu)
                    RETURNING id
                """), {
                    'rid': review_id, 'tn': tname, 'su': suite,
                    'sf': sf, 'lt': row['lease_type'],
                    'ls': row['lease_start'], 'le': row['lease_end'],
                    'tm': row['term_months'], 'mr': row['monthly_rent'],
                    'ar': ann_rent, 'rpsf': row['rent_per_sf_year'],
                    'sd': row['security_deposit'],
                    'iv': row['is_vacant'], 'im': is_material,
                    'hc': has_cot is not None, 'heu': has_exc is not None,
                })
                tenant_id = r.fetchone()[0]
                tenant_id_map[suite] = tenant_id

                # Insert cotenancy record if applicable
                if has_cot and has_cot.get('clause_text'):
                    conn.execute(text("""
                        INSERT INTO lease_cotenancy
                            (tenant_id, review_id, clause_text)
                        VALUES (:tid, :rid, :ct)
                    """), {
                        'tid': tenant_id, 'rid': review_id,
                        'ct': has_cot['clause_text'],
                    })

                # Insert exclusive use if applicable
                if has_exc:
                    conn.execute(text("""
                        INSERT INTO lease_exclusive_use
                            (tenant_id, restriction_text)
                        VALUES (:tid, :rt)
                    """), {
                        'tid': tenant_id,
                        'rt': has_exc['restriction_text'],
                    })

        # Match tenant folders to tenant records and catalog documents
        for tenant_folder in scan['tenants']:
            folder_name = tenant_folder['folder_name']
            # Match folder to a tenant_id via suite or name
            tenant_id = _match_folder_to_tenant(
                folder_name, tenant_id_map, rr_df
            )
            if tenant_id is None:
                logger.warning(f"No tenant match for folder: {folder_name}")
                continue

            for doc in tenant_folder['documents']:
                conn.execute(text("""
                    INSERT INTO lease_documents
                        (tenant_id, review_id, filename, file_path,
                         doc_type, doc_date)
                    VALUES (:tid, :rid, :fn, :fp, :dt, :dd)
                """), {
                    'tid': tenant_id, 'rid': review_id,
                    'fn': doc['filename'], 'fp': doc['path'],
                    'dt': doc['doc_type'], 'dd': doc['doc_date'],
                })

        conn.commit()

    logger.info(
        f"Ingested property '{property_name}': review_id={review_id}, "
        f"{len(tenant_id_map)} tenants, {total_gla:.0f} GLA"
    )
    return review_id


def _fuzzy_match_cotenancy(
    tenant_name: str,
    lookup: Dict[str, Dict],
) -> Optional[Dict]:
    """Fuzzy match tenant name to cotenancy lookup."""
    key = tenant_name.lower().strip()
    # Direct match
    if key in lookup:
        return lookup[key]
    # Partial match — check if lookup key is contained in tenant name or vice versa
    for lk, rec in lookup.items():
        # Strip numbers and common suffixes for comparison
        clean_key = re.sub(r'[#\d]+$', '', lk).strip()
        clean_name = re.sub(r'[#\d]+$', '', key).strip()
        if clean_key and clean_name:
            if clean_key in clean_name or clean_name in clean_key:
                return rec
    return None


def _match_folder_to_tenant(
    folder_name: str,
    tenant_id_map: Dict[str, int],
    rr_df: pd.DataFrame,
) -> Optional[int]:
    """Match a folder name like 'Ross Dress For Less 884' to a tenant_id."""
    if len(rr_df) == 0:
        return None

    # Known aliases: folder name -> rent roll name patterns
    FOLDER_ALIASES = {
        'afl': 'autism',
        'cornerstone church': 'cornerstone',
        'creamy rolls': 'creamy rolls',
        'hibachi grill & supreme buffet': 'hibachi',
        'hibachi grill': 'hibachi',
        'high five indoor playgournd': 'high five',
        'high five indoor playground': 'high five',
        'kohls': "kohl's",
        'velva nail spa': 'velva nail',
        'le nails': 'velva nail',
    }

    # Clean folder name — remove trailing numbers and parenthetical notes
    clean = re.sub(r'\s+\d+$', '', folder_name).strip()
    clean = re.sub(r'\s*\(former.*?\)', '', clean, flags=re.IGNORECASE).strip()
    clean_lower = clean.lower()

    # Try alias lookup
    for alias_key, alias_val in FOLDER_ALIASES.items():
        if alias_key in clean_lower:
            for _, row in rr_df.iterrows():
                rr_name = str(row['tenant_name']).lower().strip()
                if alias_val in rr_name:
                    return tenant_id_map.get(row['suite'])

    for _, row in rr_df.iterrows():
        rr_name = str(row['tenant_name']).lower().strip()
        suite = row['suite']
        # Direct name match
        if clean_lower in rr_name or rr_name in clean_lower:
            return tenant_id_map.get(suite)
        # Check folder name contains tenant or vice versa
        rr_clean = re.sub(r'[#\d]+$', '', rr_name).strip()
        if rr_clean and (rr_clean in clean_lower or clean_lower in rr_clean):
            return tenant_id_map.get(suite)

    return None


# ---------------------------------------------------------------------------
# Reset extraction data for re-extraction
# ---------------------------------------------------------------------------

def reset_extraction_data(engine, review_id: int) -> Dict[str, int]:
    """Clear all AI-extracted data for a review and reset documents to pending.

    Preserves: tenant roster (from rent roll), uploaded documents (PDFs),
    field resolutions, abstract sections.

    Clears: rent_steps, cotenancy + refs, exclusive_use, options,
    validation, tenant extraction_json/status, document extraction_status.
    """
    from sqlalchemy import text

    counts: Dict[str, int] = {}

    with engine.begin() as conn:
        # Get tenant IDs for this review
        tid_rows = conn.execute(text(
            "SELECT id FROM lease_tenants WHERE review_id = :rid"
        ), {'rid': review_id}).fetchall()
        tenant_ids = [r[0] for r in tid_rows]

        if not tenant_ids:
            return {'tenants': 0}

        # Build placeholder list for IN clause
        placeholders = ','.join(f':t{i}' for i in range(len(tenant_ids)))
        tid_params = {f't{i}': tid for i, tid in enumerate(tenant_ids)}
        base_params = {'rid': review_id, **tid_params}

        # Delete cotenancy refs (must go before cotenancy due to FK)
        r = conn.execute(text(f"""
            DELETE FROM lease_cotenancy_refs
            WHERE cotenancy_id IN (
                SELECT id FROM lease_cotenancy
                WHERE tenant_id IN ({placeholders})
            )
        """), tid_params)
        counts['cotenancy_refs'] = r.rowcount

        # Delete cotenancy
        r = conn.execute(text(f"""
            DELETE FROM lease_cotenancy
            WHERE tenant_id IN ({placeholders})
        """), tid_params)
        counts['cotenancy'] = r.rowcount

        # Delete rent steps
        r = conn.execute(text(f"""
            DELETE FROM lease_rent_steps
            WHERE tenant_id IN ({placeholders})
        """), tid_params)
        counts['rent_steps'] = r.rowcount

        # Delete options
        r = conn.execute(text(f"""
            DELETE FROM lease_options
            WHERE tenant_id IN ({placeholders})
        """), tid_params)
        counts['options'] = r.rowcount

        # Delete exclusive use
        r = conn.execute(text(f"""
            DELETE FROM lease_exclusive_use
            WHERE tenant_id IN ({placeholders})
        """), tid_params)
        counts['exclusive_use'] = r.rowcount

        # Delete validation
        r = conn.execute(text(f"""
            DELETE FROM lease_validation
            WHERE tenant_id IN ({placeholders})
        """), tid_params)
        counts['validation'] = r.rowcount

        # Reset tenant extraction status and JSON
        conn.execute(text(f"""
            UPDATE lease_tenants
            SET extraction_status = 'pending',
                extraction_json = NULL,
                has_cotenancy = FALSE,
                has_exclusive_use = FALSE,
                updated_at = CURRENT_TIMESTAMP
            WHERE id IN ({placeholders})
        """), tid_params)

        # Reset document extraction status (keep extracted_text for re-use)
        r = conn.execute(text("""
            UPDATE lease_documents
            SET extraction_status = CASE
                WHEN extracted_text IS NOT NULL THEN 'text_extracted'
                ELSE 'pending'
            END
            WHERE review_id = :rid
        """), {'rid': review_id})
        counts['documents_reset'] = r.rowcount
        counts['tenants'] = len(tenant_ids)

    logger.info(f"Reset extraction data for review {review_id}: {counts}")
    return counts


# ---------------------------------------------------------------------------
# Extract lease terms via Claude API (batch)
# ---------------------------------------------------------------------------

def extract_all_documents(engine, review_id: int, api_key: Optional[str] = None,
                          progress_callback=None):
    """Extract text from all PDFs and run Claude extraction for key documents.

    Prioritizes Original Lease and Amendment documents.
    progress_callback: optional callable(extracted_count, total, current_file)
    """
    from sqlalchemy import text as sql_text

    with engine.connect() as conn:
        # Get all documents for this review (pending or text_extracted)
        # Note: file_data excluded from bulk query to avoid OOM — fetched per-doc
        docs = conn.execute(sql_text("""
            SELECT d.id, d.tenant_id, d.filename, d.file_path,
                   d.doc_type, d.doc_date,
                   t.tenant_name, t.suite,
                   d.extraction_status, d.extracted_text
            FROM lease_documents d
            JOIN lease_tenants t ON t.id = d.tenant_id
            WHERE d.review_id = :rid
            AND d.extraction_status IN ('pending', 'text_extracted')
            ORDER BY d.tenant_id, d.doc_date
        """), {'rid': review_id}).fetchall()

        logger.info(f"Extracting {len(docs)} documents for review {review_id}")

        total_docs = len(docs)
        for doc_idx, doc in enumerate(docs):
            doc_id = doc[0]
            tenant_id = doc[1]
            file_path = doc[3]
            doc_type = doc[4]
            tenant_name = doc[6]
            suite = doc[7]
            current_status = doc[8]
            existing_text = doc[9]

            try:
                # Step 1: Extract PDF text (skip if already extracted)
                if current_status == 'text_extracted' and existing_text:
                    pdf_text = existing_text
                else:
                    # Use file_path if available, otherwise fetch file_data from DB
                    source = file_path
                    if not source or not os.path.exists(source):
                        row = conn.execute(sql_text(
                            "SELECT file_data FROM lease_documents WHERE id = :did"
                        ), {'did': doc_id}).fetchone()
                        if row and row[0]:
                            source = row[0]
                        else:
                            raise FileNotFoundError(
                                f"No file_path or file_data for doc {doc_id}"
                            )
                    pdf_text, page_count = extract_pdf_text(source)

                    conn.execute(sql_text("""
                        UPDATE lease_documents
                        SET extracted_text = :txt, page_count = :pc,
                            extraction_status = 'text_extracted'
                        WHERE id = :did
                    """), {'txt': pdf_text, 'pc': page_count, 'did': doc_id})

                # Step 2: Run Claude extraction for leases and amendments
                if doc_type in ('Original Lease', 'Amendment'):
                    terms = extract_lease_terms_via_api(
                        pdf_text, tenant_name, suite, doc_type, api_key
                    )

                    if not terms.get('_parse_error'):
                        # Store per-document extraction JSON + mark extracted
                        conn.execute(sql_text("""
                            UPDATE lease_documents
                            SET extraction_status = 'extracted',
                                extraction_json = :ej
                            WHERE id = :did
                        """), {'did': doc_id, 'ej': json.dumps(terms)})

                        # Mark tenant as having extraction data
                        conn.execute(sql_text("""
                            UPDATE lease_tenants
                            SET extraction_status = 'extracted',
                                updated_at = CURRENT_TIMESTAMP
                            WHERE id = :tid
                        """), {'tid': tenant_id})

                        # Store rent steps (with dedup)
                        if terms.get('rent_steps'):
                            for step in terms['rent_steps']:
                                # Dedup: skip if (tenant_id, effective_date) already exists
                                dup = conn.execute(sql_text("""
                                    SELECT id FROM lease_rent_steps
                                    WHERE tenant_id = :tid AND effective_date = :ed
                                    LIMIT 1
                                """), {
                                    'tid': tenant_id,
                                    'ed': step.get('effective_date'),
                                }).fetchone()
                                if dup:
                                    continue
                                conn.execute(sql_text("""
                                    INSERT INTO lease_rent_steps
                                        (tenant_id, effective_date,
                                         monthly_rent, annual_rent,
                                         rent_per_sf, source_doc)
                                    VALUES (:tid, :ed, :mr, :ar, :rpsf, :sd)
                                """), {
                                    'tid': tenant_id,
                                    'ed': step.get('effective_date'),
                                    'mr': _to_number(step.get('monthly_rent')),
                                    'ar': _to_number(step.get('annual_rent')),
                                    'rpsf': _to_number(step.get('rent_per_sf')),
                                    'sd': doc[2],  # filename
                                })

                        # Store cotenancy from extraction (with dedup by source_doc)
                        cot = terms.get('cotenancy', {})
                        if cot and cot.get('has_clause') and not conn.execute(sql_text("""
                            SELECT id FROM lease_cotenancy
                            WHERE tenant_id = :tid AND source_doc = :sd LIMIT 1
                        """), {'tid': tenant_id, 'sd': doc[2]}).fetchone():
                            cot_result = conn.execute(sql_text("""
                                INSERT INTO lease_cotenancy
                                    (tenant_id, review_id,
                                     trigger_description, trigger_threshold,
                                     cure_period_days, alt_rent_formula,
                                     termination_right, termination_notice_days,
                                     sunset_provision, is_curable,
                                     waiver_mechanism, source_doc)
                                VALUES (:tid, :rid,
                                        :td, :tt, :cpd, :arf,
                                        :tr, :tnd, :sp, :ic,
                                        :wm, :sd)
                                RETURNING id
                            """), {
                                'tid': tenant_id, 'rid': review_id,
                                'td': cot.get('trigger_threshold'),
                                'tt': cot.get('trigger_threshold'),
                                'cpd': _to_int(cot.get('cure_period_days')),
                                'arf': cot.get('alt_rent_formula'),
                                'tr': cot.get('termination_right', False),
                                'tnd': _to_int(cot.get('termination_notice_days')),
                                'sp': cot.get('sunset_or_waiver'),
                                'ic': cot.get('is_curable', True),
                                'wm': cot.get('sunset_or_waiver'),
                                'sd': doc[2],
                            })
                            cot_id = cot_result.fetchone()[0]

                            # Insert named co-tenant references
                            for ref_name in (cot.get('named_cotenants') or []):
                                conn.execute(sql_text("""
                                    INSERT INTO lease_cotenancy_refs
                                        (cotenancy_id, tenant_id,
                                         referenced_tenant_name)
                                    VALUES (:cid, :tid, :rtn)
                                """), {
                                    'cid': cot_id, 'tid': tenant_id,
                                    'rtn': ref_name,
                                })

                        # Store exclusive use restrictions.  A lease can carry
                        # several, so each is a row; dedup on the restriction
                        # itself so re-running extraction does not duplicate.
                        for exc in (terms.get('exclusive_use') or []):
                            if not isinstance(exc, dict):
                                continue
                            r_use = (exc.get('restricted_use') or '').strip()
                            r_text = (exc.get('restriction_text') or '').strip()
                            # Skip explicit negatives -- "no exclusive" is the
                            # absence of a row, not a restriction.
                            if not r_use and not r_text:
                                continue
                            if r_text.lower() in _EXCLUSIVE_NEGATIVES or                                r_use.lower() in _EXCLUSIVE_NEGATIVES:
                                continue
                            if conn.execute(sql_text("""
                                SELECT id FROM lease_exclusive_use
                                WHERE tenant_id = :tid AND source_doc = :sd
                                  AND COALESCE(restricted_use, '') = :ru
                                LIMIT 1
                            """), {'tid': tenant_id, 'sd': doc[2],
                                   'ru': r_use}).fetchone():
                                continue
                            radius = exc.get('radius_feet')
                            try:
                                radius = float(radius) if radius is not None else None
                            except (TypeError, ValueError):
                                radius = None
                            role = (exc.get('clause_role') or '').strip().lower()
                            if role not in ('holder', 'subject'):
                                role = None
                            conn.execute(sql_text("""
                                INSERT INTO lease_exclusive_use
                                    (tenant_id, review_id, restriction_text,
                                     restricted_use, radius_feet, clause_role,
                                     carve_outs, source_doc)
                                VALUES (:tid, :rid, :rt, :ru, :rf, :cr, :co, :sd)
                            """), {
                                'tid': tenant_id, 'rid': review_id,
                                'rt': r_text or None, 'ru': r_use or None,
                                'rf': radius, 'cr': role,
                                'co': (exc.get('carve_outs') or None),
                                'sd': doc[2],
                            })
                            conn.execute(sql_text("""
                                UPDATE lease_tenants SET has_exclusive_use = TRUE
                                WHERE id = :tid
                            """), {'tid': tenant_id})

                        # Store renewal options (with dedup by source_doc + option_number)
                        for opt in (terms.get('renewal_options') or []):
                            opt_dup = conn.execute(sql_text("""
                                SELECT id FROM lease_options
                                WHERE tenant_id = :tid AND source_doc = :sd
                                  AND option_type = 'renewal'
                                  AND option_number = :on
                                LIMIT 1
                            """), {
                                'tid': tenant_id, 'sd': doc[2],
                                'on': opt.get('option_number'),
                            }).fetchone()
                            if opt_dup:
                                # Update exercised status if exercise notice confirms it
                                if opt.get('exercised'):
                                    conn.execute(sql_text("""
                                        UPDATE lease_options SET exercised = TRUE
                                        WHERE id = :id
                                    """), {'id': opt_dup[0]})
                                continue
                            conn.execute(sql_text("""
                                INSERT INTO lease_options
                                    (tenant_id, option_type, option_number,
                                     total_options, term_years, notice_days,
                                     notice_deadline, rent_terms,
                                     auto_renewal, exercised,
                                     option_start, option_end, source_doc)
                                VALUES (:tid, 'renewal', :on, :to, :ty,
                                        :nd, :ndl, :rt, :ar, :ex,
                                        :os, :oe, :sd)
                            """), {
                                'tid': tenant_id,
                                'on': opt.get('option_number'),
                                'to': opt.get('total_options'),
                                'ty': _to_number(opt.get('term_years')),
                                'nd': _to_int(opt.get('notice_days')),
                                'ndl': opt.get('notice_deadline'),
                                'rt': opt.get('rent_terms'),
                                'ar': opt.get('auto_renewal', False),
                                'ex': opt.get('exercised', False),
                                'os': opt.get('option_start'),
                                'oe': opt.get('option_end'),
                                'sd': doc[2],
                            })

                        # Store termination options (with dedup by source_doc + option_number)
                        for opt in (terms.get('termination_options') or []):
                            opt_dup = conn.execute(sql_text("""
                                SELECT id FROM lease_options
                                WHERE tenant_id = :tid AND source_doc = :sd
                                  AND option_type = 'termination'
                                  AND option_number = :on
                                LIMIT 1
                            """), {
                                'tid': tenant_id, 'sd': doc[2],
                                'on': opt.get('option_number'),
                            }).fetchone()
                            if opt_dup:
                                if opt.get('exercised'):
                                    conn.execute(sql_text("""
                                        UPDATE lease_options SET exercised = TRUE
                                        WHERE id = :id
                                    """), {'id': opt_dup[0]})
                                continue
                            conn.execute(sql_text("""
                                INSERT INTO lease_options
                                    (tenant_id, option_type, option_number,
                                     total_options, term_years, notice_days,
                                     notice_deadline, rent_terms,
                                     exercised, option_start, option_end,
                                     source_doc)
                                VALUES (:tid, 'termination', :on, :to, NULL,
                                        :nd, :ndl, :rt,
                                        :ex, :os, NULL,
                                        :sd)
                            """), {
                                'tid': tenant_id,
                                'on': opt.get('option_number'),
                                'to': opt.get('total_options'),
                                'nd': _to_int(opt.get('notice_days')),
                                'ndl': opt.get('notice_deadline'),
                                'rt': opt.get('conditions') or opt.get('termination_fee') or '',
                                'ex': opt.get('exercised', False),
                                'os': opt.get('earliest_termination_date'),
                                'sd': doc[2],
                            })

                conn.commit()
                logger.info(f"Extracted: {doc[2]}")

            except Exception as e:
                logger.error(f"Error extracting {doc[2]}: {e}")
                # On PostgreSQL a failed statement aborts the transaction, so
                # every later statement -- including this one -- fails with
                # InFailedSqlTransaction and masks the real cause. Roll back
                # first so the status write lands and the loop can continue
                # with the remaining documents.
                try:
                    conn.rollback()
                except Exception:
                    pass
                try:
                    conn.execute(sql_text("""
                        UPDATE lease_documents
                        SET extraction_status = 'error'
                        WHERE id = :did
                    """), {'did': doc_id})
                    conn.commit()
                except Exception as inner:
                    logger.error(
                        "Could not mark doc %s as errored: %s", doc_id, inner
                    )
                    try:
                        conn.rollback()
                    except Exception:
                        pass

            # Report progress after each doc (success or error)
            if progress_callback:
                progress_callback(doc_idx + 1, total_docs, doc[2])

    # After all documents extracted, consolidate per-tenant
    consolidate_review_extractions(engine, review_id)


# ---------------------------------------------------------------------------
# Consolidation — merge base lease + amendments into current terms
# ---------------------------------------------------------------------------

def _merge_extraction_terms(base: Dict, amendment: Dict) -> Dict:
    """Layer an amendment's extracted terms onto the base/running state.

    Rules:
    - Scalar fields: amendment value replaces base if not None
    - rent_steps: union by effective_date (amendment overwrites same date)
    - renewal_options / termination_options: merge by option_number;
      amendment can update exercised status or add new options
    - cotenancy / sales_provisions / assignment / percentage_rent:
      amendment replaces entirely if present
    - exclusive_use: list of restrictions; amendment replaces the set
    - key_dates: amendment replaces if present
    """
    merged = copy.deepcopy(base)

    # Scalar fields — amendment non-null wins
    scalar_keys = [
        'tenant_legal_name', 'suite', 'square_feet', 'permitted_use',
        'lease_commencement', 'rent_commencement', 'lease_expiration',
        'holdover_rate', 'escalation_structure', 'security_deposit',
        'cam_structure', 'cam_cap_pct', 'admin_fee_pct',
        'tax_pass_through', 'insurance_pass_through',
        'ti_allowance', 'go_dark_provision',
    ]
    for key in scalar_keys:
        val = amendment.get(key)
        if val is not None:
            merged[key] = val

    # Rent steps — merge by effective_date
    if amendment.get('rent_steps'):
        existing_by_date = {}
        for step in (merged.get('rent_steps') or []):
            ed = step.get('effective_date')
            if ed:
                existing_by_date[ed] = step
        for step in amendment['rent_steps']:
            ed = step.get('effective_date')
            if ed:
                existing_by_date[ed] = step  # amendment overwrites same date
            else:
                existing_by_date[id(step)] = step  # undated step, just append
        merged['rent_steps'] = sorted(
            existing_by_date.values(),
            key=lambda s: s.get('effective_date') or '',
        )

    # Options — merge by (option_type_key, option_number)
    for opt_key in ('renewal_options', 'termination_options'):
        amend_opts = amendment.get(opt_key)
        if amend_opts:
            existing_by_num = {}
            for opt in (merged.get(opt_key) or []):
                existing_by_num[opt.get('option_number')] = opt
            for opt in amend_opts:
                num = opt.get('option_number')
                if num in existing_by_num:
                    # Update existing: amendment non-null fields win
                    for k, v in opt.items():
                        if v is not None:
                            existing_by_num[num][k] = v
                else:
                    existing_by_num[num] = opt
            merged[opt_key] = sorted(
                existing_by_num.values(),
                key=lambda o: o.get('option_number') or 0,
            )

    # Object fields — amendment replaces entirely if present and non-empty
    object_keys = [
        'cotenancy', 'sales_provisions',
        'percentage_rent', 'assignment', 'key_dates',
    ]

    # exclusive_use is a list of restrictions; an amendment that mentions any
    # replaces the set, since it restates the restriction as amended.
    if amendment.get('exclusive_use'):
        merged['exclusive_use'] = amendment['exclusive_use']
    for key in object_keys:
        val = amendment.get(key)
        if val and isinstance(val, dict) and any(v is not None for v in val.values()):
            merged[key] = val

    return merged


def consolidate_tenant_extractions(
    engine, tenant_id: int
) -> Optional[Dict]:
    """Consolidate per-document extractions into current effective terms.

    Loads all extracted documents for a tenant, ordered: Original Lease first,
    then amendments by date. Layers them to produce merged current terms.
    Stores result as tenant's extraction_json.

    Returns the consolidated terms dict, or None if no extractions.
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        docs = conn.execute(text("""
            SELECT id, doc_type, doc_date, extraction_json
            FROM lease_documents
            WHERE tenant_id = :tid
              AND extraction_status = 'extracted'
              AND extraction_json IS NOT NULL
            ORDER BY
                CASE WHEN doc_type = 'Original Lease' THEN 0 ELSE 1 END,
                doc_date ASC NULLS LAST
        """), {'tid': tenant_id}).fetchall()

        if not docs:
            return None

        # Start with first document (should be Original Lease)
        consolidated = {}
        for doc in docs:
            try:
                terms = json.loads(doc[3])
                if isinstance(terms, dict):
                    if not consolidated:
                        consolidated = copy.deepcopy(terms)
                    else:
                        consolidated = _merge_extraction_terms(consolidated, terms)
            except (json.JSONDecodeError, TypeError):
                continue

        if not consolidated:
            return None

        # Store consolidated terms on tenant
        conn.execute(text("""
            UPDATE lease_tenants
            SET extraction_json = :ej,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = :tid
        """), {'tid': tenant_id, 'ej': json.dumps(consolidated)})
        conn.commit()

        logger.info(f"Consolidated {len(docs)} extractions for tenant {tenant_id}")
        return consolidated


def consolidate_review_extractions(
    engine, review_id: int
) -> int:
    """Consolidate extractions for all tenants in a review.

    Returns count of tenants consolidated.
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        tenants = conn.execute(text("""
            SELECT DISTINCT t.id
            FROM lease_tenants t
            JOIN lease_documents d ON d.tenant_id = t.id
            WHERE t.review_id = :rid
              AND d.extraction_status = 'extracted'
              AND d.extraction_json IS NOT NULL
        """), {'rid': review_id}).fetchall()

    count = 0
    for (tid,) in tenants:
        result = consolidate_tenant_extractions(engine, tid)
        if result:
            count += 1

    logger.info(
        f"Consolidated extractions for {count}/{len(tenants)} tenants "
        f"in review {review_id}"
    )
    return count


# ---------------------------------------------------------------------------
# Argus rent roll validation parsing
# ---------------------------------------------------------------------------

def parse_argus_validation(file_path: str) -> pd.DataFrame:
    """Parse the Argus Rent Roll Validation spreadsheet.

    This is the seller/broker's representation of Argus inputs.
    Returns a DataFrame with: suite, tenant_name, source, commencement,
    expiration, sq_ft, rent_per_sf.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['RR Validation']

    rows = []
    current_suite = None
    current_tenant = None

    for r in range(4, ws.max_row + 1):
        col_b = ws.cell(r, 2).value  # Unit column
        col_c = ws.cell(r, 3).value  # Tenant name or empty
        col_d = ws.cell(r, 4).value  # Source (Argus / Rent Roll / Lease)

        if col_b and str(col_b).strip():
            suite = str(col_b).strip()
            # Check if this is a unit code (e.g. A200)
            if re.match(r'^[A-Z]\d', suite):
                current_suite = suite

        if col_c and str(col_c).strip() and str(col_c) != 'None':
            current_tenant = str(col_c).strip()

        if col_d and str(col_d).strip() in ('Argus', 'Rent Roll', 'Lease'):
            source = str(col_d).strip()

            def to_str(v):
                if v is None:
                    return None
                if isinstance(v, datetime):
                    return v.strftime('%Y-%m-%d')
                return str(v).strip()

            rows.append({
                'suite': current_suite,
                'tenant_name': current_tenant,
                'source': source,
                'commencement': to_str(ws.cell(r, 5).value),
                'expiration': to_str(ws.cell(r, 6).value),
                'sq_ft': ws.cell(r, 7).value,
                'rent_per_sf': to_str(ws.cell(r, 9).value),
            })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Validation — compare seller sources to lease-extracted ground truth
# ---------------------------------------------------------------------------

def validate_rent_roll(
    engine,
    review_id: int,
    argus_file: Optional[str] = None,
) -> List[Dict]:
    """Compare seller-provided documents to extracted lease terms (ground truth).

    The actual lease PDFs are ground truth. The rent roll, Argus model,
    and cotenancy schedule are seller-provided representations that must
    be validated against the lease documents.

    Validation sources compared against lease:
    - Rent Roll (MRI/Yardi export): tenant, suite, SF, dates, rent
    - Argus (broker's DCF model): commencement, expiration, SF, rent/SF
    - Cotenancy Schedule (seller abstract): clause summaries

    Stores results in lease_validation with source_type field.
    """
    from sqlalchemy import text

    results = []

    # Get the rent roll date for finding the correct rent step
    with engine.connect() as conn:
        review = conn.execute(text(
            "SELECT rent_roll_date FROM lease_reviews WHERE id = :rid"
        ), {'rid': review_id}).fetchone()
        rr_date = review[0] if review and review[0] else None

    # Parse Argus file if provided
    argus_df = pd.DataFrame()
    if argus_file and os.path.exists(argus_file):
        argus_df = parse_argus_validation(argus_file)

    with engine.connect() as conn:
        # Clear prior validation results for this review
        conn.execute(text("""
            DELETE FROM lease_validation
            WHERE tenant_id IN (
                SELECT id FROM lease_tenants WHERE review_id = :rid
            )
        """), {'rid': review_id})

        tenants = conn.execute(text("""
            SELECT id, tenant_name, suite, square_feet, lease_start,
                   lease_end, monthly_rent, annual_rent, rent_per_sf,
                   security_deposit, extraction_json, extraction_status
            FROM lease_tenants
            WHERE review_id = :rid AND is_vacant = false
            ORDER BY suite
        """), {'rid': review_id}).fetchall()

        for t in tenants:
            tenant_id = t[0]
            tenant_name = t[1]
            suite = t[2]
            rr_sf = t[3]
            rr_start = t[4]
            rr_end = t[5]
            rr_monthly = t[6]
            rr_annual = t[7]
            rr_rpsf = t[8]
            extraction_status = t[11]

            # --- Find the correct rent step from lease (ground truth) ---
            # Strategy: find the step effective on the rent roll date.
            # Many extracted steps have non-date effective_dates (e.g.
            # "Lease Year 7", "Rent Commencement Date"), so we need
            # a fallback when date-based matching fails.
            current_step = None
            step_match_method = None

            # 1. Try date-based match: step effective on or before RR date
            if rr_date:
                current_step = conn.execute(text("""
                    SELECT effective_date, monthly_rent, annual_rent, rent_per_sf
                    FROM lease_rent_steps
                    WHERE tenant_id = :tid
                    AND effective_date <= :rrd
                    AND effective_date LIKE '____-%'
                    ORDER BY effective_date DESC
                    LIMIT 1
                """), {'tid': tenant_id, 'rrd': rr_date}).fetchone()
                if current_step:
                    step_match_method = 'date'

            # 2. Fallback: find the step whose annual rent is closest
            #    to the rent roll amount (best match by value)
            if not current_step and rr_annual:
                all_steps = conn.execute(text("""
                    SELECT effective_date, monthly_rent, annual_rent, rent_per_sf
                    FROM lease_rent_steps
                    WHERE tenant_id = :tid AND annual_rent IS NOT NULL
                """), {'tid': tenant_id}).fetchall()
                if all_steps:
                    best = min(all_steps, key=lambda s: abs(
                        (float(s[2]) if s[2] else 0) - float(rr_annual)
                    ))
                    current_step = best
                    step_match_method = 'closest_rent'

            # 3. Last fallback: any step at all
            if not current_step:
                current_step = conn.execute(text("""
                    SELECT effective_date, monthly_rent, annual_rent, rent_per_sf
                    FROM lease_rent_steps
                    WHERE tenant_id = :tid
                    ORDER BY effective_date DESC
                    LIMIT 1
                """), {'tid': tenant_id}).fetchone()
                if current_step:
                    step_match_method = 'fallback'

            # Get extraction JSON for date/SF comparisons
            ext = {}
            if t[10]:
                try:
                    ext = json.loads(t[10]) if isinstance(t[10], str) else t[10]
                except (json.JSONDecodeError, TypeError):
                    pass

            # --- 1. Rent Roll vs Lease Validation ---
            if current_step:
                lease_monthly = current_step[1]
                lease_annual = current_step[2]
                lease_rpsf = current_step[3]
                step_date = current_step[0]

                rr_validations = [
                    ('monthly_rent', rr_monthly, lease_monthly,
                     _compare_amounts(rr_monthly, lease_monthly)),
                    ('annual_rent', rr_annual, lease_annual,
                     _compare_amounts(rr_annual, lease_annual)),
                    ('rent_per_sf', rr_rpsf, lease_rpsf,
                     _compare_amounts(rr_rpsf, lease_rpsf, tolerance=0.05)),
                ]

                # Validate SF
                lease_sf = ext.get('square_feet')
                if lease_sf:
                    rr_validations.append(
                        ('square_feet', rr_sf, lease_sf,
                         _compare_amounts(rr_sf, lease_sf, tolerance=10))
                    )

                # Validate lease expiration
                lease_exp = ext.get('lease_expiration')
                if lease_exp and rr_end:
                    rr_validations.append(
                        ('lease_expiration', rr_end, lease_exp,
                         _compare_dates(rr_end, lease_exp))
                    )

                for field, seller_val, lease_val, status in rr_validations:
                    if field.startswith(('monthly', 'annual', 'rent_per')):
                        note = f"Step {step_date} (matched by {step_match_method})"
                    else:
                        note = None
                    conn.execute(text("""
                        INSERT INTO lease_validation
                            (tenant_id, field_name, source_type,
                             seller_value, lease_value, status,
                             source_doc, notes)
                        VALUES (:tid, :fn, 'rent_roll',
                                :sv, :lv, :st, :sd, :notes)
                    """), {
                        'tid': tenant_id, 'fn': field,
                        'sv': str(seller_val) if seller_val is not None else None,
                        'lv': str(lease_val) if lease_val is not None else None,
                        'st': status, 'sd': 'rent_roll',
                        'notes': note,
                    })
                    results.append({
                        'tenant': tenant_name, 'suite': suite,
                        'field': field, 'source_type': 'rent_roll',
                        'seller_value': str(seller_val) if seller_val is not None else None,
                        'lease_value': str(lease_val) if lease_val is not None else None,
                        'status': status,
                        'notes': note,
                    })
            elif extraction_status != 'extracted':
                # No extraction yet — mark as pending
                results.append({
                    'tenant': tenant_name, 'suite': suite,
                    'field': 'all', 'source_type': 'rent_roll',
                    'seller_value': None, 'lease_value': None,
                    'status': 'pending',
                    'notes': 'Lease not yet extracted',
                })

            # --- 2. Argus vs Lease Validation ---
            if len(argus_df) and ext:
                # Find Argus rows for this suite
                argus_rows = argus_df[
                    (argus_df['suite'] == suite) &
                    (argus_df['source'] == 'Argus')
                ]
                if len(argus_rows):
                    ar = argus_rows.iloc[0]
                    # Argus expiration vs lease expiration
                    argus_exp = ar.get('expiration')
                    lease_exp = ext.get('lease_expiration')
                    if argus_exp and lease_exp:
                        status = _compare_dates(argus_exp, lease_exp)
                        conn.execute(text("""
                            INSERT INTO lease_validation
                                (tenant_id, field_name, source_type,
                                 seller_value, lease_value, status, source_doc)
                            VALUES (:tid, 'lease_expiration', 'argus',
                                    :sv, :lv, :st, 'argus')
                        """), {
                            'tid': tenant_id,
                            'sv': str(argus_exp), 'lv': str(lease_exp),
                            'st': status,
                        })
                        results.append({
                            'tenant': tenant_name, 'suite': suite,
                            'field': 'lease_expiration',
                            'source_type': 'argus',
                            'seller_value': str(argus_exp),
                            'lease_value': str(lease_exp),
                            'status': status,
                        })

                    # Argus SF vs lease SF
                    argus_sf = ar.get('sq_ft')
                    lease_sf = ext.get('square_feet')
                    if argus_sf and lease_sf:
                        status = _compare_amounts(argus_sf, lease_sf, tolerance=10)
                        conn.execute(text("""
                            INSERT INTO lease_validation
                                (tenant_id, field_name, source_type,
                                 seller_value, lease_value, status, source_doc)
                            VALUES (:tid, 'square_feet', 'argus',
                                    :sv, :lv, :st, 'argus')
                        """), {
                            'tid': tenant_id,
                            'sv': str(argus_sf), 'lv': str(lease_sf),
                            'st': status,
                        })
                        results.append({
                            'tenant': tenant_name, 'suite': suite,
                            'field': 'square_feet',
                            'source_type': 'argus',
                            'seller_value': str(argus_sf),
                            'lease_value': str(lease_sf),
                            'status': status,
                        })

                    # Argus rent/SF vs lease rent/SF
                    argus_rpsf = ar.get('rent_per_sf')
                    if argus_rpsf and current_step:
                        lease_rpsf = current_step[3]
                        try:
                            argus_rpsf_f = float(argus_rpsf)
                        except (ValueError, TypeError):
                            argus_rpsf_f = None
                        if argus_rpsf_f is not None and lease_rpsf:
                            status = _compare_amounts(
                                argus_rpsf_f, lease_rpsf, tolerance=0.05
                            )
                            conn.execute(text("""
                                INSERT INTO lease_validation
                                    (tenant_id, field_name, source_type,
                                     seller_value, lease_value, status,
                                     source_doc)
                                VALUES (:tid, 'rent_per_sf', 'argus',
                                        :sv, :lv, :st, 'argus')
                            """), {
                                'tid': tenant_id,
                                'sv': str(argus_rpsf_f),
                                'lv': str(lease_rpsf),
                                'st': status,
                            })
                            results.append({
                                'tenant': tenant_name, 'suite': suite,
                                'field': 'rent_per_sf',
                                'source_type': 'argus',
                                'seller_value': str(argus_rpsf_f),
                                'lease_value': str(lease_rpsf),
                                'status': status,
                            })

            # --- 3. Cotenancy Schedule vs Lease Validation ---
            cot_row = conn.execute(text("""
                SELECT clause_text, trigger_description, alt_rent_formula,
                       termination_right, is_curable
                FROM lease_cotenancy
                WHERE tenant_id = :tid
            """), {'tid': tenant_id}).fetchone()

            cot_ext = ext.get('cotenancy', {}) if ext else {}
            if cot_row and cot_ext and cot_ext.get('has_clause'):
                # Compare trigger descriptions
                if cot_row[1] and cot_ext.get('trigger_threshold'):
                    results.append({
                        'tenant': tenant_name, 'suite': suite,
                        'field': 'cotenancy_trigger',
                        'source_type': 'cotenancy_schedule',
                        'seller_value': cot_row[1],
                        'lease_value': cot_ext.get('trigger_threshold'),
                        'status': 'review',
                        'notes': 'Manual review recommended — compare trigger language',
                    })

                # Compare alt rent
                if cot_row[2] and cot_ext.get('alt_rent_formula'):
                    results.append({
                        'tenant': tenant_name, 'suite': suite,
                        'field': 'cotenancy_alt_rent',
                        'source_type': 'cotenancy_schedule',
                        'seller_value': cot_row[2],
                        'lease_value': cot_ext.get('alt_rent_formula'),
                        'status': 'review',
                        'notes': 'Manual review recommended — compare alt rent formulas',
                    })

        conn.commit()
    return results


def _compare_dates(val1: str, val2: str) -> str:
    """Compare two date strings. Returns 'match', 'minor', or 'mismatch'."""
    try:
        d1 = pd.to_datetime(val1)
        d2 = pd.to_datetime(val2)
    except (ValueError, TypeError):
        return 'pending'
    diff = abs((d1 - d2).days)
    if diff == 0:
        return 'match'
    if diff <= 31:
        return 'minor'
    return 'mismatch'


def _compare_amounts(
    rr_val: float,
    lease_val: float,
    tolerance: float = 1.0,
) -> str:
    """Compare two amounts. Returns 'match', 'minor', or 'mismatch'."""
    try:
        rr = float(rr_val) if rr_val else 0
        lv = float(lease_val) if lease_val else 0
    except (ValueError, TypeError):
        return 'pending'
    if rr == 0 and lv == 0:
        return 'match'
    if abs(rr - lv) <= tolerance:
        return 'match'
    if rr != 0 and abs(rr - lv) / abs(rr) < 0.02:
        return 'minor'
    return 'mismatch'


# ---------------------------------------------------------------------------
# Analysis — expiration histogram, cotenancy matrix, scenarios
# ---------------------------------------------------------------------------

def get_expiration_histogram(
    engine,
    review_id: int,
    years: int = 10,
) -> Dict[str, Any]:
    """Build lease expiration histogram by year.

    Returns:
        - yearly_data: list of dicts per year (year, expiring_sf, expiring_rent,
          pct_of_total_rent, avg_rent_per_sf, tenant_count)
        - material_leases: list of material leases per year with cotenancy flags
        - totals: total_gla, total_annual_rent
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        # Get review totals
        review = conn.execute(text("""
            SELECT total_gla, total_annual_rent
            FROM lease_reviews WHERE id = :rid
        """), {'rid': review_id}).fetchone()
        total_gla = review[0] or 0
        total_rent = review[1] or 0

        # Get all non-vacant tenants with lease end dates
        tenants = conn.execute(text("""
            SELECT id, tenant_name, suite, square_feet, lease_end,
                   annual_rent, rent_per_sf, is_material, has_cotenancy
            FROM lease_tenants
            WHERE review_id = :rid AND is_vacant = false
            AND lease_end IS NOT NULL
            ORDER BY lease_end
        """), {'rid': review_id}).fetchall()

    current_year = datetime.now().year
    end_year = current_year + years

    yearly = {}
    for yr in range(current_year, end_year + 1):
        yearly[yr] = {
            'year': yr,
            'tenants': [],
            'expiring_sf': 0,
            'expiring_rent': 0,
            'pct_of_total_rent': 0,
            'avg_rent_per_sf': 0,
            'tenant_count': 0,
        }

    material_by_year = {}

    for t in tenants:
        try:
            lease_end = pd.to_datetime(t[4])
            exp_year = lease_end.year
        except Exception:
            continue

        if exp_year < current_year or exp_year > end_year:
            continue

        sf = t[3] or 0
        rent = t[5] or 0
        rpsf = t[6] or 0

        yearly[exp_year]['tenants'].append(t[1])
        yearly[exp_year]['expiring_sf'] += sf
        yearly[exp_year]['expiring_rent'] += rent
        yearly[exp_year]['tenant_count'] += 1

        # Track material leases
        if t[7]:  # is_material
            if exp_year not in material_by_year:
                material_by_year[exp_year] = []
            material_by_year[exp_year].append({
                'tenant_name': t[1],
                'suite': t[2],
                'square_feet': sf,
                'annual_rent': rent,
                'rent_per_sf': rpsf,
                'lease_end': t[4],
                'has_cotenancy': bool(t[8]),
                'cotenancy_implication': (
                    'Departure may trigger co-tenancy clauses in other leases'
                    if t[8] else None
                ),
            })

    # Calculate percentages and averages
    yearly_data = []
    for yr in range(current_year, end_year + 1):
        d = yearly[yr]
        if total_rent > 0:
            d['pct_of_total_rent'] = round(d['expiring_rent'] / total_rent * 100, 1)
        if d['expiring_sf'] > 0:
            d['avg_rent_per_sf'] = round(d['expiring_rent'] / d['expiring_sf'], 2)
        del d['tenants']
        yearly_data.append(d)

    return {
        'yearly_data': yearly_data,
        'material_leases': material_by_year,
        'totals': {
            'total_gla': total_gla,
            'total_annual_rent': total_rent,
        },
    }


def get_cotenancy_matrix(engine, review_id: int) -> Dict[str, Any]:
    """Build the co-tenancy cross-reference matrix.

    Returns which tenants are named as co-tenants in which leases,
    plus the "who depends on whom" reverse lookup.
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        # Get all cotenancy records with refs
        cotenancy = conn.execute(text("""
            SELECT c.id, c.tenant_id, t.tenant_name, t.suite,
                   t.annual_rent, c.clause_text,
                   c.trigger_description, c.alt_rent_formula,
                   c.termination_right, c.cure_period_days,
                   c.sunset_provision, c.is_curable
            FROM lease_cotenancy c
            JOIN lease_tenants t ON t.id = c.tenant_id
            WHERE c.review_id = :rid
        """), {'rid': review_id}).fetchall()

        refs = conn.execute(text("""
            SELECT cr.cotenancy_id, cr.referenced_tenant_name
            FROM lease_cotenancy_refs cr
            JOIN lease_cotenancy c ON c.id = cr.cotenancy_id
            WHERE c.review_id = :rid
        """), {'rid': review_id}).fetchall()

        # Load global alias mappings: alias_name -> canonical_name
        alias_rows = conn.execute(text(
            "SELECT alias_name, canonical_name FROM lease_tenant_aliases"
        )).fetchall()

    alias_map = {r[0]: r[1] for r in alias_rows}

    # Build forward map: tenant -> list of named cotenants
    forward = {}
    cot_details = {}
    for c in cotenancy:
        tenant_name = c[2]
        forward[tenant_name] = []
        cot_details[tenant_name] = {
            'suite': c[3],
            'annual_rent': c[4],
            'clause_summary': c[5][:200] if c[5] else '',
            'trigger': c[6],
            'alt_rent': c[7],
            'termination_right': c[8],
            'cure_days': c[9],
            'sunset': c[10],
            'is_curable': c[11],
        }

    # Map refs — apply alias normalization
    cot_id_to_tenant = {c[0]: c[2] for c in cotenancy}
    for ref in refs:
        cot_id = ref[0]
        ref_name = alias_map.get(ref[1], ref[1])  # resolve alias
        tenant_name = cot_id_to_tenant.get(cot_id)
        if tenant_name:
            forward[tenant_name].append(ref_name)

    # Build reverse map: referenced tenant -> list of tenants that depend on them
    reverse = {}
    for tenant, named_cotenants in forward.items():
        for cotenant in named_cotenants:
            if cotenant not in reverse:
                reverse[cotenant] = []
            detail = cot_details.get(tenant, {})
            reverse[cotenant].append({
                'dependent_tenant': tenant,
                'annual_rent': detail.get('annual_rent', 0),
                'alt_rent': detail.get('alt_rent', ''),
                'termination_right': detail.get('termination_right', False),
            })

    # Calculate rent at risk per referenced tenant
    rent_at_risk = {}
    for cotenant, dependents in reverse.items():
        total_rent = sum(d.get('annual_rent', 0) or 0 for d in dependents)
        term_count = sum(1 for d in dependents if d.get('termination_right'))
        rent_at_risk[cotenant] = {
            'total_dependent_rent': total_rent,
            'dependent_count': len(dependents),
            'termination_eligible_count': term_count,
            'dependents': dependents,
        }

    return {
        'forward': forward,
        'reverse': reverse,
        'rent_at_risk': rent_at_risk,
        'details': cot_details,
    }


# ---------------------------------------------------------------------------
# Tenant Alias Management
# ---------------------------------------------------------------------------

def get_tenant_aliases(engine) -> List[Dict]:
    """Get all global tenant alias mappings."""
    from sqlalchemy import text
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT id, alias_name, canonical_name, created_by, created_at
            FROM lease_tenant_aliases
            ORDER BY canonical_name, alias_name
        """)).fetchall()
    return [{'id': r[0], 'alias_name': r[1], 'canonical_name': r[2],
             'created_by': r[3], 'created_at': str(r[4]) if r[4] else None}
            for r in rows]


def get_alias_map(engine) -> Dict[str, str]:
    """Load alias_name -> canonical_name lookup dict (cached helper)."""
    from sqlalchemy import text
    with engine.connect() as conn:
        rows = conn.execute(text(
            "SELECT alias_name, canonical_name FROM lease_tenant_aliases"
        )).fetchall()
    return {r[0]: r[1] for r in rows}


def save_tenant_alias(engine, alias_name: str,
                      canonical_name: str, created_by: str = None) -> Dict:
    """Create or update a global alias mapping."""
    from sqlalchemy import text
    with engine.connect() as conn:
        try:
            conn.execute(text("""
                INSERT INTO lease_tenant_aliases
                    (alias_name, canonical_name, created_by)
                VALUES (:alias, :canon, :by)
                ON CONFLICT (alias_name)
                DO UPDATE SET canonical_name = :canon, created_by = :by
            """), {'alias': alias_name, 'canon': canonical_name, 'by': created_by})
        except Exception:
            conn.execute(text("""
                INSERT OR REPLACE INTO lease_tenant_aliases
                    (alias_name, canonical_name, created_by)
                VALUES (:alias, :canon, :by)
            """), {'alias': alias_name, 'canon': canonical_name, 'by': created_by})
        conn.commit()
    return {'alias_name': alias_name, 'canonical_name': canonical_name}


def delete_tenant_alias(engine, alias_id: int):
    """Delete a global alias mapping."""
    from sqlalchemy import text
    with engine.connect() as conn:
        conn.execute(text(
            "DELETE FROM lease_tenant_aliases WHERE id = :id"
        ), {'id': alias_id})
        conn.commit()


def suggest_alias_matches(engine, review_id: int) -> List[Dict]:
    """Auto-detect likely alias groups from co-tenancy refs for a review.

    Uses normalization (strip parentheticals, punctuation) to find
    ref names that likely refer to the same entity.
    """
    from sqlalchemy import text
    import re

    with engine.connect() as conn:
        refs = conn.execute(text("""
            SELECT DISTINCT cr.referenced_tenant_name
            FROM lease_cotenancy_refs cr
            JOIN lease_cotenancy c ON c.id = cr.cotenancy_id
            WHERE c.review_id = :rid
            ORDER BY cr.referenced_tenant_name
        """), {'rid': review_id}).fetchall()

        tenants = conn.execute(text("""
            SELECT id, tenant_name FROM lease_tenants
            WHERE review_id = :rid AND tenant_status = 'active'
        """), {'rid': review_id}).fetchall()

        existing = conn.execute(text(
            "SELECT alias_name FROM lease_tenant_aliases"
        )).fetchall()

    ref_names = [r[0] for r in refs]
    tenant_names = {t[1]: t[0] for t in tenants}
    already_aliased = {r[0] for r in existing}

    def _normalize(name):
        """Strip parentheticals, punctuation, and normalize whitespace."""
        name = re.sub(r'\s*\(.*?\)\s*', '', name)
        name = re.sub(r'[\'\".,]', '', name)
        name = re.sub(r'\s+', ' ', name).strip()
        return name.lower()

    # Group by normalized name
    groups: Dict[str, list] = {}
    for name in ref_names:
        if name in already_aliased:
            continue
        norm = _normalize(name)
        if norm not in groups:
            groups[norm] = []
        groups[norm].append(name)

    suggestions = []
    for norm, variants in groups.items():
        if len(variants) < 2:
            continue
        # Find best matching roster tenant
        best_match = None
        for tname in tenant_names:
            if _normalize(tname) == norm:
                best_match = tname
                break
        suggestions.append({
            'variants': variants,
            'suggested_canonical': best_match or variants[0],
            'roster_match': best_match,
        })

    return suggestions


def get_scenario_analysis(engine, review_id: int) -> List[Dict]:
    """Model cascading co-tenancy scenarios for major tenant departures.

    For each tenant that is referenced as a co-tenant by others,
    model what happens if they close.
    """
    matrix = get_cotenancy_matrix(engine, review_id)

    scenarios = []
    for cotenant, risk in matrix['rent_at_risk'].items():
        if risk['dependent_count'] == 0:
            continue

        scenario = {
            'departing_tenant': cotenant,
            'dependent_count': risk['dependent_count'],
            'total_dependent_rent': risk['total_dependent_rent'],
            'termination_eligible': risk['termination_eligible_count'],
            'impacts': [],
        }

        for dep in risk['dependents']:
            detail = matrix['details'].get(dep['dependent_tenant'], {})
            scenario['impacts'].append({
                'tenant': dep['dependent_tenant'],
                'annual_rent': dep.get('annual_rent', 0),
                'alt_rent_formula': dep.get('alt_rent', ''),
                'can_terminate': dep.get('termination_right', False),
                'cure_days': detail.get('cure_days'),
                'sunset': detail.get('sunset'),
                'is_curable': detail.get('is_curable', True),
            })

        scenarios.append(scenario)

    # Sort by total rent at risk descending
    scenarios.sort(key=lambda s: s['total_dependent_rent'], reverse=True)
    return scenarios


# ---------------------------------------------------------------------------
# Excel report generation
# ---------------------------------------------------------------------------

def generate_lease_review_excel(engine, review_id: int) -> bytes:
    """Generate the comprehensive lease review Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                             fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                              fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                           fill_type="solid")
    currency_fmt = '#,##0'
    pct_fmt = '0.0%'
    thin_border = Border(
        bottom=Side(style='thin'),
    )

    from sqlalchemy import text

    with engine.connect() as conn:
        review = conn.execute(text("""
            SELECT property_name, property_address, total_gla,
                   total_annual_rent, total_tenants
            FROM lease_reviews WHERE id = :rid
        """), {'rid': review_id}).fetchone()

        tenants = conn.execute(text("""
            SELECT tenant_name, suite, square_feet, lease_type,
                   lease_start, lease_end, term_months, monthly_rent,
                   annual_rent, rent_per_sf, security_deposit,
                   is_vacant, is_material, has_cotenancy
            FROM lease_tenants
            WHERE review_id = :rid
            ORDER BY suite
        """), {'rid': review_id}).fetchall()

    # --- Sheet 1: Executive Summary ---
    ws = wb.active
    ws.title = "Executive Summary"
    ws['A1'] = f"Lease Review: {review[0]}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = review[1] or ''
    ws['A4'] = "Total GLA"
    ws['B4'] = review[2]
    ws['B4'].number_format = '#,##0'
    ws['A5'] = "Total Annual Rent"
    ws['B5'] = review[3]
    ws['B5'].number_format = '#,##0'
    ws['A6'] = "Total Tenants"
    ws['B6'] = review[4]
    ws['A7'] = "Tenants with Co-Tenancy"
    ws['B7'] = sum(1 for t in tenants if t[13])
    ws['A8'] = "Vacant Suites"
    ws['B8'] = sum(1 for t in tenants if t[11])
    for r in range(4, 9):
        ws.cell(r, 1).font = Font(bold=True)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18

    # --- Sheet 2: Lease Expiration Histogram ---
    histogram = get_expiration_histogram(engine, review_id)
    ws2 = wb.create_sheet("Lease Expirations")

    # Header
    ws2['A1'] = f"Lease Expiration Schedule — {review[0]}"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A2'] = f"Total Annual Rent: ${review[3]:,.0f}   |   Total GLA: {review[2]:,.0f} SF"

    headers = ['Year', 'Expiring SF', 'Expiring Rent', '% of Total Rent',
               'Avg Rent/SF', '# Tenants']
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(4, c, h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, yr_data in enumerate(histogram['yearly_data']):
        row = 5 + i
        ws2.cell(row, 1, yr_data['year'])
        ws2.cell(row, 2, yr_data['expiring_sf']).number_format = '#,##0'
        ws2.cell(row, 3, yr_data['expiring_rent']).number_format = '#,##0'
        ws2.cell(row, 4, yr_data['pct_of_total_rent'] / 100).number_format = '0.0%'
        ws2.cell(row, 5, yr_data['avg_rent_per_sf']).number_format = '#,##0.00'
        ws2.cell(row, 6, yr_data['tenant_count'])

    # Add bar chart
    chart = BarChart()
    chart.title = "Annual Lease Revenue Expiring"
    chart.y_axis.title = "Annual Rent ($)"
    chart.x_axis.title = "Year"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data_rows = len(histogram['yearly_data'])
    if data_rows > 0:
        data_ref = Reference(ws2, min_col=3, min_row=4,
                             max_row=4 + data_rows)
        cats = Reference(ws2, min_col=1, min_row=5,
                         max_row=4 + data_rows)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = "4472C4"
        ws2.add_chart(chart, f"A{5 + data_rows + 2}")

    # Material leases below chart
    mat_start = 5 + data_rows + 18
    ws2.cell(mat_start, 1, "Material Leases Maturing by Year").font = Font(
        bold=True, size=12)
    mat_headers = ['Year', 'Tenant', 'Suite', 'Square Feet', 'Annual Rent',
                   '$/SF', 'Lease End', 'Co-Tenancy Risk']
    for c, h in enumerate(mat_headers, 1):
        cell = ws2.cell(mat_start + 1, c, h)
        cell.font = header_font_white
        cell.fill = header_fill
    r = mat_start + 2
    for yr in sorted(histogram['material_leases'].keys()):
        for lease in histogram['material_leases'][yr]:
            ws2.cell(r, 1, yr)
            ws2.cell(r, 2, lease['tenant_name'])
            ws2.cell(r, 3, lease['suite'])
            ws2.cell(r, 4, lease['square_feet']).number_format = '#,##0'
            ws2.cell(r, 5, lease['annual_rent']).number_format = '#,##0'
            ws2.cell(r, 6, lease['rent_per_sf']).number_format = '#,##0.00'
            ws2.cell(r, 7, lease['lease_end'])
            cot_text = lease.get('cotenancy_implication', '')
            ws2.cell(r, 8, cot_text or 'None')
            if cot_text:
                ws2.cell(r, 8).fill = yellow_fill
            r += 1

    # Auto-width
    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # --- Sheet 3: Rent Roll Validation ---
    ws3 = wb.create_sheet("Rent Roll Validation")
    ws3['A1'] = "Seller Document Validation vs Lease Terms (Ground Truth)"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3['A2'] = "Lease PDFs are the authoritative source. Rent Roll and Argus are seller representations validated against actual lease terms."

    # Pull validation results from DB
    with engine.connect() as conn:
        val_rows = conn.execute(text("""
            SELECT t.tenant_name, t.suite, v.field_name, v.source_type,
                   v.seller_value, v.lease_value, v.status, v.notes
            FROM lease_validation v
            JOIN lease_tenants t ON t.id = v.tenant_id
            WHERE t.review_id = :rid
            ORDER BY t.tenant_name, v.source_type, v.field_name
        """), {'rid': review_id}).fetchall()

    val_headers = ['Tenant', 'Suite', 'Field', 'Source', 'Seller Value',
                   'Lease Value', 'Status', 'Notes']
    for c, h in enumerate(val_headers, 1):
        cell = ws3.cell(4, c, h)
        cell.font = header_font_white
        cell.fill = header_fill

    for i, v in enumerate(val_rows):
        row = 5 + i
        ws3.cell(row, 1, v[0])  # tenant_name
        ws3.cell(row, 2, v[1])  # suite
        ws3.cell(row, 3, v[2])  # field_name
        ws3.cell(row, 4, v[3])  # source_type
        # Format numeric values
        for col_idx, val_idx in [(5, 4), (6, 5)]:
            val = v[val_idx]
            try:
                num = float(val) if val else None
                if num is not None:
                    ws3.cell(row, col_idx, num).number_format = '#,##0.00'
                else:
                    ws3.cell(row, col_idx, val or '')
            except (ValueError, TypeError):
                ws3.cell(row, col_idx, val or '')
        ws3.cell(row, 7, v[6])  # status
        ws3.cell(row, 8, v[7] or '')  # notes

        # Color code status
        status = v[6]
        if status == 'match':
            ws3.cell(row, 7).fill = green_fill
        elif status == 'mismatch':
            ws3.cell(row, 7).fill = red_fill
        elif status == 'minor':
            ws3.cell(row, 7).fill = yellow_fill
        elif status == 'review':
            ws3.cell(row, 7).fill = yellow_fill

    col_widths = [30, 12, 18, 18, 18, 18, 12, 45]
    for c, w in enumerate(col_widths, 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    # --- Sheet 4: Co-Tenancy Detail ---
    ws4 = wb.create_sheet("Co-Tenancy Detail")
    ws4['A1'] = "Co-Tenancy Clause Analysis"
    ws4['A1'].font = Font(bold=True, size=14)
    ws4['A2'] = "Lease PDFs are ground truth. Seller cotenancy schedule validated against actual lease terms."

    cot_headers = ['Tenant', 'Suite', 'SF', 'Annual Rent',
                   'Trigger', 'Cure Period', 'Alt Rent Formula',
                   'Termination Right', 'Sunset/Waiver', 'Curable?',
                   'Named Co-Tenants']
    for c, h in enumerate(cot_headers, 1):
        cell = ws4.cell(4, c, h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Get cotenancy details from DB
    with engine.connect() as conn:
        cot_rows = conn.execute(text("""
            SELECT t.tenant_name, t.suite, t.square_feet, t.annual_rent,
                   c.trigger_description, c.cure_period_days,
                   c.alt_rent_formula, c.termination_right,
                   c.sunset_provision, c.is_curable, c.waiver_mechanism,
                   c.id
            FROM lease_cotenancy c
            JOIN lease_tenants t ON t.id = c.tenant_id
            WHERE c.review_id = :rid
            ORDER BY t.annual_rent DESC
        """), {'rid': review_id}).fetchall()

        # Get refs for each cotenancy
        all_refs = conn.execute(text("""
            SELECT cr.cotenancy_id, cr.referenced_tenant_name
            FROM lease_cotenancy_refs cr
            JOIN lease_cotenancy c ON c.id = cr.cotenancy_id
            WHERE c.review_id = :rid
        """), {'rid': review_id}).fetchall()

    ref_map = {}
    for r in all_refs:
        ref_map.setdefault(r[0], []).append(r[1])

    for i, c in enumerate(cot_rows):
        row = 5 + i
        ws4.cell(row, 1, c[0])   # tenant_name
        ws4.cell(row, 2, c[1])   # suite
        ws4.cell(row, 3, c[2]).number_format = '#,##0'
        ws4.cell(row, 4, c[3]).number_format = '#,##0'
        ws4.cell(row, 5, c[4] or 'See lease')  # trigger
        ws4.cell(row, 6, f"{c[5]} days" if c[5] else 'N/A')  # cure
        ws4.cell(row, 7, c[6] or 'N/A')  # alt rent
        ws4.cell(row, 8, 'Yes' if c[7] else 'No')  # termination
        ws4.cell(row, 9, c[8] or 'N/A')  # sunset
        curable = c[9]
        ws4.cell(row, 10, 'Yes' if curable else 'NO - UNCURABLE')
        if not curable:
            ws4.cell(row, 10).fill = red_fill
            ws4.cell(row, 10).font = Font(bold=True, color="9C0006")
        # Named cotenants
        refs = ref_map.get(c[11], [])
        ws4.cell(row, 11, ', '.join(refs) if refs else 'N/A')

        # Wrap text for readability
        for col in (5, 7, 9, 11):
            ws4.cell(row, col).alignment = Alignment(wrap_text=True,
                                                      vertical='top')

    col_widths = [25, 10, 10, 14, 35, 12, 30, 14, 35, 16, 35]
    for c, w in enumerate(col_widths, 1):
        ws4.column_dimensions[get_column_letter(c)].width = w

    # Add risk summary below
    risk_start = 5 + len(cot_rows) + 3
    ws4.cell(risk_start, 1, "Departing Tenant Risk Summary").font = Font(
        bold=True, size=12)
    risk_headers = ['Departing Tenant', 'Affected Count', 'Rent at Risk',
                    'Can Terminate']
    for c, h in enumerate(risk_headers, 1):
        cell = ws4.cell(risk_start + 1, c, h)
        cell.font = header_font_white
        cell.fill = header_fill

    matrix = get_cotenancy_matrix(engine, review_id)
    r = risk_start + 2
    for cotenant, risk in sorted(matrix['rent_at_risk'].items(),
                                  key=lambda x: x[1]['total_dependent_rent'],
                                  reverse=True):
        ws4.cell(r, 1, cotenant)
        ws4.cell(r, 2, risk['dependent_count'])
        ws4.cell(r, 3, risk['total_dependent_rent']).number_format = '#,##0'
        ws4.cell(r, 4, risk['termination_eligible_count'])
        r += 1

    # --- Sheet 5: Exclusive Use ---
    ws5 = wb.create_sheet("Exclusive Use")
    ws5['A1'] = "Exclusive Use Restrictions"
    ws5['A1'].font = Font(bold=True, size=14)

    with engine.connect() as conn:
        exc_rows = conn.execute(text("""
            SELECT t.tenant_name, t.suite, e.restriction_text
            FROM lease_exclusive_use e
            JOIN lease_tenants t ON t.id = e.tenant_id
            WHERE t.review_id = :rid
            ORDER BY t.tenant_name
        """), {'rid': review_id}).fetchall()

    exc_headers = ['Tenant', 'Suite', 'Exclusive Use Restriction']
    for c, h in enumerate(exc_headers, 1):
        cell = ws5.cell(3, c, h)
        cell.font = header_font_white
        cell.fill = header_fill
    for i, e in enumerate(exc_rows):
        row = 4 + i
        ws5.cell(row, 1, e[0])
        ws5.cell(row, 2, e[1])
        ws5.cell(row, 3, e[2]).alignment = Alignment(wrap_text=True)
    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 60

    # --- Sheet 6: Option Schedule ---
    ws6 = wb.create_sheet("Option Schedule")
    ws6['A1'] = "Renewal / Termination Options"
    ws6['A1'].font = Font(bold=True, size=14)

    with engine.connect() as conn:
        opt_rows = conn.execute(text("""
            SELECT t.tenant_name, t.suite, o.option_type,
                   o.option_number, o.total_options,
                   o.option_start, o.option_end, o.term_years,
                   o.notice_days, o.notice_deadline, o.rent_terms,
                   o.auto_renewal, o.exercised, o.source_doc
            FROM lease_options o
            JOIN lease_tenants t ON t.id = o.tenant_id
            WHERE t.review_id = :rid
            ORDER BY t.tenant_name, o.option_type, o.option_number
        """), {'rid': review_id}).fetchall()

    opt_headers = ['Tenant', 'Suite', 'Type', 'Option #', 'Total Options',
                   'Start', 'End', 'Term (Yrs)',
                   'Notice (Days)', 'Notice Deadline',
                   'Rent Terms / Conditions', 'Auto-Renew', 'Exercised', 'Source']
    for c, h in enumerate(opt_headers, 1):
        cell = ws6.cell(3, c, h)
        cell.font = header_font_white
        cell.fill = header_fill
    for i, o in enumerate(opt_rows):
        row = 4 + i
        for c, val in enumerate(o, 1):
            cell = ws6.cell(row, c, val)
            if c == 11:  # rent_terms
                cell.alignment = Alignment(wrap_text=True)

    opt_widths = [30, 12, 12, 10, 10, 14, 14, 10, 12, 16, 40, 10, 10, 30]
    for c, w in enumerate(opt_widths, 1):
        ws6.column_dimensions[get_column_letter(c)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Lease Risk Analysis — field resolutions + resolved data
# ---------------------------------------------------------------------------

LEASE_RESOLUTION_DDL_PG = """
CREATE TABLE IF NOT EXISTS lease_field_resolutions (
    id              SERIAL PRIMARY KEY,
    tenant_id       INTEGER NOT NULL REFERENCES lease_tenants(id),
    field_name      TEXT NOT NULL,
    resolved_value  TEXT,
    resolved_source TEXT DEFAULT 'analyst_override',
    resolved_by     TEXT,
    resolved_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE (tenant_id, field_name)
)
"""

LEASE_RESOLUTION_DDL_SQLITE = (
    LEASE_RESOLUTION_DDL_PG
    .replace('SERIAL PRIMARY KEY', 'INTEGER PRIMARY KEY AUTOINCREMENT')
    .replace('REFERENCES lease_tenants(id)', '')
)


def ensure_resolution_table(engine):
    """Create the lease_field_resolutions table if it doesn't exist."""
    from sqlalchemy import text
    ddl = (LEASE_RESOLUTION_DDL_SQLITE
           if engine.dialect.name == 'sqlite'
           else LEASE_RESOLUTION_DDL_PG)
    with engine.begin() as conn:
        conn.execute(text(ddl))


# Resolvable fields and which columns they map to on lease_tenants
RESOLVABLE_FIELDS = {
    'square_feet', 'annual_rent', 'monthly_rent', 'rent_per_sf',
    'lease_start', 'lease_end', 'security_deposit',
}


def resolve_field(
    engine,
    tenant_id: int,
    field_name: str,
    resolved_value: str,
    resolved_source: str = 'analyst_override',
    resolved_by: str = 'system',
) -> Dict[str, Any]:
    """Save an analyst's resolution for a specific field on a tenant.

    Uses UPSERT (INSERT ON CONFLICT UPDATE) to handle re-resolution.
    """
    from sqlalchemy import text

    if field_name not in RESOLVABLE_FIELDS:
        raise ValueError(f"Field '{field_name}' is not resolvable. "
                         f"Valid: {sorted(RESOLVABLE_FIELDS)}")

    ensure_resolution_table(engine)

    with engine.begin() as conn:
        if engine.dialect.name == 'postgresql':
            conn.execute(text("""
                INSERT INTO lease_field_resolutions
                    (tenant_id, field_name, resolved_value, resolved_source,
                     resolved_by, resolved_at)
                VALUES (:tid, :fn, :rv, :rs, :rb, CURRENT_TIMESTAMP)
                ON CONFLICT (tenant_id, field_name)
                DO UPDATE SET resolved_value = :rv, resolved_source = :rs,
                              resolved_by = :rb, resolved_at = CURRENT_TIMESTAMP
            """), {
                'tid': tenant_id, 'fn': field_name,
                'rv': resolved_value, 'rs': resolved_source, 'rb': resolved_by,
            })
        else:
            conn.execute(text("""
                INSERT OR REPLACE INTO lease_field_resolutions
                    (tenant_id, field_name, resolved_value, resolved_source,
                     resolved_by, resolved_at)
                VALUES (:tid, :fn, :rv, :rs, :rb, CURRENT_TIMESTAMP)
            """), {
                'tid': tenant_id, 'fn': field_name,
                'rv': resolved_value, 'rs': resolved_source, 'rb': resolved_by,
            })

    return {'tenant_id': tenant_id, 'field_name': field_name,
            'resolved_value': resolved_value, 'resolved_source': resolved_source}


def clear_resolution(engine, tenant_id: int, field_name: str) -> None:
    """Remove a field resolution, reverting to default data."""
    from sqlalchemy import text
    ensure_resolution_table(engine)
    with engine.begin() as conn:
        conn.execute(text("""
            DELETE FROM lease_field_resolutions
            WHERE tenant_id = :tid AND field_name = :fn
        """), {'tid': tenant_id, 'fn': field_name})


def get_resolved_tenants(
    engine, review_id: int, include_replaced: bool = False,
) -> List[Dict[str, Any]]:
    """Get all tenants for a review with analyst resolutions applied.

    For each field, returns:
    - The analyst's resolved value if one exists
    - Otherwise the default from lease_tenants

    Each tenant dict includes a 'resolutions' sub-dict showing
    which fields have analyst overrides.

    By default filters to active tenants only. Set include_replaced=True
    to include replaced/deleted tenants.
    """
    from sqlalchemy import text
    ensure_resolution_table(engine)

    status_filter = ""
    if not include_replaced:
        status_filter = "AND (tenant_status IS NULL OR tenant_status = 'active')"

    with engine.connect() as conn:
        tenants = conn.execute(text(f"""
            SELECT id, tenant_name, suite, square_feet, lease_type,
                   lease_start, lease_end, term_months, monthly_rent,
                   annual_rent, rent_per_sf, security_deposit,
                   is_vacant, is_material, has_cotenancy, has_exclusive_use,
                   extraction_status, approval_status, tenant_status,
                   successor_tenant_id, replaced_by_event_id
            FROM lease_tenants
            WHERE review_id = :rid {status_filter}
            ORDER BY suite
        """), {'rid': review_id}).fetchall()

        # Load all resolutions for this review's tenants
        tenant_ids = [t[0] for t in tenants]
        resolutions = {}
        if tenant_ids:
            # Build parameterized IN clause
            placeholders = ', '.join(f':t{i}' for i in range(len(tenant_ids)))
            params = {f't{i}': tid for i, tid in enumerate(tenant_ids)}
            params['rid'] = review_id
            res_rows = conn.execute(text(f"""
                SELECT tenant_id, field_name, resolved_value, resolved_source,
                       resolved_by
                FROM lease_field_resolutions
                WHERE tenant_id IN ({placeholders})
            """), params).fetchall()
            for r in res_rows:
                resolutions.setdefault(r[0], {})[r[1]] = {
                    'value': r[2], 'source': r[3], 'by': r[4],
                }

    result = []
    for t in tenants:
        tid = t[0]
        tenant_res = resolutions.get(tid, {})

        def resolved(field_name, default_val, idx=None):
            """Return resolved value if exists, else default."""
            if field_name in tenant_res:
                raw = tenant_res[field_name]['value']
                # Try numeric conversion for numeric fields
                if field_name in ('square_feet', 'annual_rent', 'monthly_rent',
                                  'rent_per_sf', 'security_deposit'):
                    try:
                        return float(raw) if raw else default_val
                    except (ValueError, TypeError):
                        return default_val
                return raw
            return default_val

        row = {
            'id': tid,
            'tenant_name': t[1],
            'suite': t[2],
            'square_feet': resolved('square_feet', t[3]),
            'lease_type': t[4],
            'lease_start': resolved('lease_start', t[5]),
            'lease_end': resolved('lease_end', t[6]),
            'term_months': t[7],
            'monthly_rent': resolved('monthly_rent', t[8]),
            'annual_rent': resolved('annual_rent', t[9]),
            'rent_per_sf': resolved('rent_per_sf', t[10]),
            'security_deposit': resolved('security_deposit', t[11]),
            'is_vacant': bool(t[12]),
            'is_material': bool(t[13]),
            'has_cotenancy': bool(t[14]),
            'has_exclusive_use': bool(t[15]),
            'extraction_status': t[16],
            'approval_status': t[17] or 'pending',
            'tenant_status': t[18] or 'active',
            'successor_tenant_id': t[19],
            'replaced_by_event_id': t[20],
            'resolutions': {
                fn: {'value': info['value'], 'source': info['source']}
                for fn, info in tenant_res.items()
            },
        }
        result.append(row)

    return result


def get_resolved_expiration_histogram(
    engine, review_id: int, years: int = 10,
) -> Dict[str, Any]:
    """Expiration histogram using analyst-resolved tenant data."""
    resolved = get_resolved_tenants(engine, review_id)
    occupied = [t for t in resolved if not t['is_vacant'] and t['lease_end']]

    total_gla = sum(t['square_feet'] or 0 for t in resolved if not t['is_vacant'])
    total_rent = sum(t['annual_rent'] or 0 for t in resolved if not t['is_vacant'])

    current_year = datetime.now().year
    end_year = current_year + years

    yearly = {}
    for yr in range(current_year, end_year + 1):
        yearly[yr] = {
            'year': yr, 'expiring_sf': 0, 'expiring_rent': 0,
            'pct_of_total_rent': 0, 'avg_rent_per_sf': 0, 'tenant_count': 0,
        }

    material_by_year = {}

    for t in occupied:
        try:
            lease_end = pd.to_datetime(t['lease_end'])
            exp_year = lease_end.year
        except Exception:
            continue

        if exp_year < current_year or exp_year > end_year:
            continue

        sf = t['square_feet'] or 0
        rent = t['annual_rent'] or 0

        yearly[exp_year]['expiring_sf'] += sf
        yearly[exp_year]['expiring_rent'] += rent
        yearly[exp_year]['tenant_count'] += 1

        if t['is_material']:
            material_by_year.setdefault(exp_year, []).append({
                'tenant_name': t['tenant_name'],
                'suite': t['suite'],
                'square_feet': sf,
                'annual_rent': rent,
                'rent_per_sf': t['rent_per_sf'] or 0,
                'lease_end': t['lease_end'],
                'has_cotenancy': t['has_cotenancy'],
                'cotenancy_implication': (
                    'Departure may trigger co-tenancy clauses in other leases'
                    if t['has_cotenancy'] else None
                ),
            })

    yearly_data = []
    for yr in range(current_year, end_year + 1):
        d = yearly[yr]
        if total_rent > 0:
            d['pct_of_total_rent'] = round(d['expiring_rent'] / total_rent * 100, 1)
        if d['expiring_sf'] > 0:
            d['avg_rent_per_sf'] = round(d['expiring_rent'] / d['expiring_sf'], 2)
        yearly_data.append(d)

    return {
        'yearly_data': yearly_data,
        'material_leases': material_by_year,
        'totals': {'total_gla': total_gla, 'total_annual_rent': total_rent},
    }


def get_resolved_cotenancy_matrix(engine, review_id: int) -> Dict[str, Any]:
    """Cotenancy matrix using analyst-resolved rent data for impact sizing."""
    from sqlalchemy import text

    # Get resolved tenants for rent overrides
    resolved = get_resolved_tenants(engine, review_id)
    rent_by_tid = {t['id']: t['annual_rent'] or 0 for t in resolved}

    with engine.connect() as conn:
        cotenancy = conn.execute(text("""
            SELECT c.id, c.tenant_id, t.tenant_name, t.suite,
                   c.clause_text, c.trigger_description, c.alt_rent_formula,
                   c.termination_right, c.cure_period_days,
                   c.sunset_provision, c.is_curable
            FROM lease_cotenancy c
            JOIN lease_tenants t ON t.id = c.tenant_id
            WHERE c.review_id = :rid
        """), {'rid': review_id}).fetchall()

        refs = conn.execute(text("""
            SELECT cr.cotenancy_id, cr.referenced_tenant_name
            FROM lease_cotenancy_refs cr
            JOIN lease_cotenancy c ON c.id = cr.cotenancy_id
            WHERE c.review_id = :rid
        """), {'rid': review_id}).fetchall()

    forward = {}
    cot_details = {}
    for c in cotenancy:
        tenant_name = c[2]
        resolved_rent = rent_by_tid.get(c[1], 0)
        forward[tenant_name] = []
        cot_details[tenant_name] = {
            'suite': c[3],
            'annual_rent': resolved_rent,
            'clause_summary': c[4][:200] if c[4] else '',
            'trigger': c[5],
            'alt_rent': c[6],
            'termination_right': c[7],
            'cure_days': c[8],
            'sunset': c[9],
            'is_curable': c[10],
        }

    cot_id_to_tenant = {c[0]: c[2] for c in cotenancy}
    for ref in refs:
        tenant_name = cot_id_to_tenant.get(ref[0])
        if tenant_name:
            forward[tenant_name].append(ref[1])

    reverse = {}
    for tenant, named_cotenants in forward.items():
        for cotenant in named_cotenants:
            reverse.setdefault(cotenant, [])
            detail = cot_details.get(tenant, {})
            reverse[cotenant].append({
                'dependent_tenant': tenant,
                'annual_rent': detail.get('annual_rent', 0),
                'alt_rent': detail.get('alt_rent', ''),
                'termination_right': detail.get('termination_right', False),
            })

    rent_at_risk = {}
    for cotenant, dependents in reverse.items():
        total_rent = sum(d.get('annual_rent', 0) or 0 for d in dependents)
        term_count = sum(1 for d in dependents if d.get('termination_right'))
        rent_at_risk[cotenant] = {
            'total_dependent_rent': total_rent,
            'dependent_count': len(dependents),
            'termination_eligible_count': term_count,
            'dependents': dependents,
        }

    return {
        'forward': forward, 'reverse': reverse,
        'rent_at_risk': rent_at_risk, 'details': cot_details,
    }


def get_resolved_scenario_analysis(engine, review_id: int) -> List[Dict]:
    """Scenario analysis using resolved rent data."""
    matrix = get_resolved_cotenancy_matrix(engine, review_id)

    scenarios = []
    for cotenant, risk in matrix['rent_at_risk'].items():
        if risk['dependent_count'] == 0:
            continue

        scenario = {
            'departing_tenant': cotenant,
            'dependent_count': risk['dependent_count'],
            'total_dependent_rent': risk['total_dependent_rent'],
            'termination_eligible': risk['termination_eligible_count'],
            'impacts': [],
        }

        for dep in risk['dependents']:
            detail = matrix['details'].get(dep['dependent_tenant'], {})
            scenario['impacts'].append({
                'tenant': dep['dependent_tenant'],
                'annual_rent': dep.get('annual_rent', 0),
                'alt_rent_formula': dep.get('alt_rent', ''),
                'can_terminate': dep.get('termination_right', False),
                'cure_days': detail.get('cure_days'),
                'sunset': detail.get('sunset'),
                'is_curable': detail.get('is_curable', True),
            })

        scenarios.append(scenario)

    scenarios.sort(key=lambda s: s['total_dependent_rent'], reverse=True)
    return scenarios


def get_risk_analysis_data(engine, review_id: int) -> Dict[str, Any]:
    """Get complete risk analysis data bundle for a review.

    Returns resolved tenants, validation results, expirations,
    cotenancy matrix, and scenario analysis — all using analyst-resolved data.
    """
    from sqlalchemy import text

    resolved = get_resolved_tenants(engine, review_id)
    expirations = get_resolved_expiration_histogram(engine, review_id)
    cotenancy = get_resolved_cotenancy_matrix(engine, review_id)
    scenarios = get_resolved_scenario_analysis(engine, review_id)

    # Get validation results
    with engine.connect() as conn:
        val_rows = conn.execute(text("""
            SELECT t.id, t.tenant_name, t.suite, v.field_name, v.source_type,
                   v.seller_value, v.lease_value, v.status, v.notes
            FROM lease_validation v
            JOIN lease_tenants t ON t.id = v.tenant_id
            WHERE t.review_id = :rid
            ORDER BY t.tenant_name, v.source_type, v.field_name
        """), {'rid': review_id}).fetchall()

    validation = [{
        'tenant_id': r[0], 'tenant': r[1], 'suite': r[2],
        'field': r[3], 'source_type': r[4],
        'seller_value': r[5], 'lease_value': r[6],
        'status': r[7], 'notes': r[8],
    } for r in val_rows]

    # Build cotenancy clauses for display
    clauses = []
    if cotenancy.get('details'):
        for tenant_name, detail in cotenancy['details'].items():
            clauses.append({
                'tenant_name': tenant_name,
                **detail,
                'trigger_description': detail.get('trigger'),
                'alt_rent_formula': detail.get('alt_rent'),
                'cure_period_days': detail.get('cure_days'),
                'named_cotenants': cotenancy['forward'].get(tenant_name, []),
            })

    # Get exclusive use data
    with engine.connect() as conn:
        exc_rows = conn.execute(text("""
            SELECT t.tenant_name, t.suite, e.restriction_text, e.restricted_use,
                   e.clause_role, e.carve_outs, e.radius_feet, e.source_doc
            FROM lease_exclusive_use e
            JOIN lease_tenants t ON t.id = e.tenant_id
            WHERE t.review_id = :rid
            ORDER BY t.tenant_name
        """), {'rid': review_id}).fetchall()

    exclusive_use = [{
        'tenant_name': r[0], 'suite': r[1],
        'restriction_text': r[2], 'restricted_use': r[3],
        'clause_role': r[4], 'carve_outs': r[5],
        'radius_feet': r[6], 'source_doc': r[7],
    } for r in exc_rows]

    # Get options
    with engine.connect() as conn:
        opt_rows = conn.execute(text("""
            SELECT o.id, t.tenant_name, t.suite, o.option_type,
                   o.option_number, o.total_options,
                   o.option_start, o.option_end, o.term_years,
                   o.notice_days, o.notice_deadline, o.rent_terms,
                   o.auto_renewal, o.exercised, o.source_doc
            FROM lease_options o
            JOIN lease_tenants t ON t.id = o.tenant_id
            WHERE t.review_id = :rid
            ORDER BY t.tenant_name, o.option_type, o.option_number
        """), {'rid': review_id}).fetchall()

    options = [{
        'id': r[0], 'tenant_name': r[1], 'suite': r[2], 'option_type': r[3],
        'option_number': r[4], 'total_options': r[5],
        'option_start': r[6], 'option_end': r[7], 'term_years': r[8],
        'notice_days': r[9], 'notice_deadline': r[10], 'rent_terms': r[11],
        'auto_renewal': bool(r[12]), 'exercised': bool(r[13]),
        'source_doc': r[14],
    } for r in opt_rows]

    # Get documents grouped by tenant_id for linking
    with engine.connect() as conn:
        doc_rows = conn.execute(text("""
            SELECT d.id, d.tenant_id, d.filename, d.doc_type, d.doc_date,
                   CASE WHEN d.file_data IS NOT NULL THEN 1 ELSE 0 END as has_file
            FROM lease_documents d
            JOIN lease_tenants t ON t.id = d.tenant_id
            WHERE t.review_id = :rid
            ORDER BY d.tenant_id, d.doc_type, d.doc_date
        """), {'rid': review_id}).fetchall()

    documents = {}
    for r in doc_rows:
        tid = r[1]
        if tid not in documents:
            documents[tid] = []
        documents[tid].append({
            'id': r[0], 'filename': r[2], 'doc_type': r[3],
            'doc_date': r[4], 'has_file': bool(r[5]),
        })

    return {
        'tenants': resolved,
        'validation': validation,
        'expirations': expirations,
        'cotenancy': {**cotenancy, 'clauses': clauses},
        'scenarios': scenarios,
        'exclusive_use': exclusive_use,
        'options': options,
        'documents': documents,
    }


# ---------------------------------------------------------------------------
# Lease Abstract — assemble & persist per-tenant abstracts
# ---------------------------------------------------------------------------

# Standard abstract sections in display order (matches Word template)
ABSTRACT_SECTIONS = [
    ('tenant', 'Tenant', 1),
    ('documents_reviewed', 'Documents Reviewed', 2),
    ('unit_address', 'Unit/Address', 3),
    ('square_feet', 'Square Feet', 4),
    ('prs', 'PRS (Proportionate Share)', 5),
    ('term', 'Term', 6),
    ('percentage_rent', 'Percentage Rent', 7),
    ('renewal_options', 'Renewal Options', 8),
    ('termination_options', 'Termination Options', 9),
    ('use', 'Use', 10),
    ('exclusive_use', 'Exclusive Use', 11),
    ('cam', 'CAM', 12),
    ('insurance', 'Insurance', 13),
    ('real_estate_taxes', 'Real Estate Taxes', 14),
    ('parking', 'Parking', 15),
    ('roof', 'Roof', 16),
    ('structural', 'Structural', 17),
    ('signage', 'Signage', 18),
    ('utilities', 'Utilities', 19),
    ('hvac', 'HVAC', 20),
    ('sales_reporting', 'Sales Reporting', 21),
    ('estoppel', 'Estoppel', 22),
    ('security_deposit', 'Security Deposit', 23),
    ('operating_covenants', 'Operating Covenants', 24),
    ('go_dark_termination', 'Go-Dark Termination', 25),
    ('business_termination', 'Landlord/Tenant Business Termination Rights', 26),
    ('casualty_termination', 'Landlord/Tenant Casualty/Default/Eminent Domain Termination Rights', 27),
    ('relocation_rights', 'Landlord/Tenant Relocation Rights', 28),
    ('rofr_rofo', 'ROFR/ROFO Options', 29),
    ('cotenancy', 'Co-Tenancy Provisions', 30),
    ('radius_restriction', 'Radius Restriction', 31),
    ('ti_allowance', 'Tenant Improvement Allowance', 32),
    ('lease_inducements', 'Lease Inducements', 33),
    ('assignment', 'Assignment and Subletting', 34),
    ('merchants_association', "Merchants' Association", 35),
    ('redevelopment', 'Redevelopment', 36),
    ('guarantor', 'Guarantor(s)', 37),
    ('notes', 'Notes', 38),
    ('date_of_review', 'Date of Review', 39),
]


def _fmt_money(val) -> str:
    """Format a number as currency."""
    if val is None:
        return ''
    try:
        v = float(val)
        if v == int(v):
            return f"${int(v):,}"
        return f"${v:,.2f}"
    except (ValueError, TypeError):
        return str(val)


def _fmt_date(val) -> str:
    """Format YYYY-MM-DD to M/D/YYYY."""
    if not val:
        return ''
    try:
        m = re.match(r'(\d{4})-(\d{2})-(\d{2})', str(val))
        if m:
            return f"{int(m.group(2))}/{int(m.group(3))}/{m.group(1)}"
    except Exception:
        pass
    return str(val)


def _assemble_abstract_from_data(
    engine, tenant_id: int, review_id: int
) -> Dict[str, Dict[str, str]]:
    """Build abstract section content from existing extracted data.

    Returns {section_key: {'content': str, 'lease_ref': str}}.
    Sections with no data return empty content (user can fill in manually).
    """
    from sqlalchemy import text

    sections: Dict[str, Dict[str, str]] = {}

    with engine.connect() as conn:
        # Tenant info
        t = conn.execute(text("""
            SELECT tenant_name, suite, square_feet, lease_type,
                   lease_start, lease_end, term_months,
                   monthly_rent, annual_rent, rent_per_sf,
                   security_deposit, extraction_json
            FROM lease_tenants WHERE id = :tid
        """), {'tid': tenant_id}).fetchone()
        if not t:
            return sections

        tenant_name = t[0] or ''
        suite = t[1] or ''
        sq_ft = t[2]
        lease_start = t[4]
        lease_end = t[5]
        term_months = t[6]
        monthly_rent = t[7]
        annual_rent = t[8]
        rent_per_sf = t[9]
        security_deposit = t[10]
        extraction_raw = t[11]

        # Parse extraction JSON (may be a list of extractions from multiple docs)
        ext = {}
        if extraction_raw:
            try:
                parsed = json.loads(extraction_raw)
                if isinstance(parsed, list) and parsed:
                    ext = parsed[0] if isinstance(parsed[0], dict) else {}
                elif isinstance(parsed, dict):
                    ext = parsed
            except (json.JSONDecodeError, IndexError):
                pass

        # Documents reviewed
        docs = conn.execute(text("""
            SELECT filename, doc_type, doc_date
            FROM lease_documents
            WHERE tenant_id = :tid
            ORDER BY doc_date, filename
        """), {'tid': tenant_id}).fetchall()

        doc_lines = []
        for d in docs:
            line = d[1] or d[0]
            if d[2]:
                line += f" (dated {_fmt_date(d[2])})"
            elif d[0] != d[1]:
                line += f" — {d[0]}"
            doc_lines.append(line)

        # Rent steps
        rent_steps = conn.execute(text("""
            SELECT effective_date, monthly_rent, annual_rent, rent_per_sf
            FROM lease_rent_steps WHERE tenant_id = :tid
            ORDER BY effective_date
        """), {'tid': tenant_id}).fetchall()

        # Renewal options
        renewals = conn.execute(text("""
            SELECT option_number, total_options, term_years,
                   notice_days, notice_deadline, rent_terms,
                   auto_renewal, exercised, option_start, option_end
            FROM lease_options
            WHERE tenant_id = :tid AND option_type = 'renewal'
            ORDER BY option_number
        """), {'tid': tenant_id}).fetchall()

        # Termination options
        terminations = conn.execute(text("""
            SELECT option_number, total_options, term_years,
                   notice_days, notice_deadline, rent_terms,
                   exercised, option_start, option_end
            FROM lease_options
            WHERE tenant_id = :tid AND option_type = 'termination'
            ORDER BY option_number
        """), {'tid': tenant_id}).fetchall()

        # Cotenancy
        cot_rows = conn.execute(text("""
            SELECT c.trigger_description, c.trigger_threshold,
                   c.cure_period_days, c.alt_rent_formula,
                   c.termination_right, c.termination_notice_days,
                   c.sunset_provision
            FROM lease_cotenancy c
            WHERE c.tenant_id = :tid
        """), {'tid': tenant_id}).fetchall()

        cot_refs = conn.execute(text("""
            SELECT cr.referenced_tenant_name
            FROM lease_cotenancy_refs cr
            JOIN lease_cotenancy c ON c.id = cr.cotenancy_id
            WHERE c.tenant_id = :tid
        """), {'tid': tenant_id}).fetchall()

        # Exclusive use
        exc_rows = conn.execute(text("""
            SELECT restriction_text, restricted_use
            FROM lease_exclusive_use WHERE tenant_id = :tid
        """), {'tid': tenant_id}).fetchall()

    # -- Build section content --

    # Tenant
    sections['tenant'] = {
        'content': ext.get('tenant_legal_name') or tenant_name,
        'lease_ref': '',
    }

    # Documents Reviewed
    sections['documents_reviewed'] = {
        'content': '\n'.join(doc_lines) if doc_lines else '',
        'lease_ref': '',
    }

    # Unit/Address
    sections['unit_address'] = {
        'content': suite,
        'lease_ref': '',
    }

    # Square Feet
    sections['square_feet'] = {
        'content': f"{int(sq_ft):,} Square Feet" if sq_ft else '',
        'lease_ref': '',
    }

    # PRS
    sections['prs'] = {'content': '', 'lease_ref': ''}

    # Term
    term_parts = []
    if term_months:
        years = term_months // 12
        months = term_months % 12
        if years and months:
            term_parts.append(f"{years} year(s), {months} month(s)")
        elif years:
            term_parts.append(f"{years} year(s)")
        else:
            term_parts.append(f"{months} month(s)")
    if lease_start:
        term_parts.append(f"Commencing {_fmt_date(lease_start)}")
    if lease_end:
        term_parts.append(f"Expiring {_fmt_date(lease_end)}")
    # Rent schedule
    if rent_steps:
        term_parts.append('')
        term_parts.append('Rent Schedule:')
        for idx, rs in enumerate(rent_steps):
            date_label = _fmt_date(rs[0])
            if not date_label:
                date_label = f"Step {idx + 1}"
            line = f"  {date_label}: {_fmt_money(rs[1])}/mo"
            if rs[2]:
                line += f" ({_fmt_money(rs[2])}/yr)"
            if rs[3]:
                line += f" — {_fmt_money(rs[3])}/SF"
            term_parts.append(line)
    elif monthly_rent or annual_rent:
        term_parts.append(
            f"Base Rent: {_fmt_money(monthly_rent)}/mo "
            f"({_fmt_money(annual_rent)}/yr)"
        )
        if rent_per_sf:
            term_parts.append(f"Rent/SF: {_fmt_money(rent_per_sf)}")

    sections['term'] = {
        'content': '\n'.join(term_parts),
        'lease_ref': '',
    }

    # Sales Provisions (percentage rent, reporting, performance clauses)
    sales = ext.get('sales_provisions', {})
    # Backwards compat: fall back to top-level percentage_rent from older extractions
    pct = sales.get('percentage_rent', ext.get('percentage_rent', {})) if sales else ext.get('percentage_rent', {})
    if pct and pct.get('has_clause'):
        pct_text = f"Rate: {pct.get('rate_pct')}%"
        if pct.get('breakpoint'):
            pct_text += f", Breakpoint: {_fmt_money(pct['breakpoint'])}"
        sections['percentage_rent'] = {'content': pct_text, 'lease_ref': ''}
    else:
        sections['percentage_rent'] = {
            'content': 'No language noted.' if ext else '',
            'lease_ref': '',
        }

    # Sales Reporting
    if sales and sales.get('sales_reporting_required'):
        rpt_parts = ['Sales reporting required.']
        if sales.get('reporting_frequency'):
            rpt_parts.append(f"Frequency: {sales['reporting_frequency']}.")
        if sales.get('reporting_deadline'):
            rpt_parts.append(f"Deadline: {sales['reporting_deadline']}.")
        if sales.get('audit_right'):
            rpt_parts.append('Landlord has audit right.')
        perf_clauses = sales.get('sales_performance_clauses', [])
        for pc in perf_clauses:
            clause_desc = pc.get('clause_type', 'clause').title()
            trigger = pc.get('trigger', '')
            consequence = pc.get('consequence', '')
            beneficiary = pc.get('beneficiary', '')
            parts = [f"{clause_desc}:"]
            if trigger:
                parts.append(f"Trigger — {trigger}.")
            if consequence:
                parts.append(f"Consequence — {consequence}.")
            if beneficiary:
                parts.append(f"({beneficiary.title()} right)")
            rpt_parts.append(' '.join(parts))
        sections['sales_reporting'] = {'content': ' '.join(rpt_parts), 'lease_ref': ''}
    elif sales and sales.get('sales_reporting_required') is False:
        sections['sales_reporting'] = {'content': 'No sales reporting required.', 'lease_ref': ''}
    else:
        sections['sales_reporting'] = {'content': '', 'lease_ref': ''}

    # Renewal Options
    if renewals:
        lines = []
        for r in renewals:
            line = f"Option {r[0]} of {r[1]}: {r[2]} year(s)"
            if r[3]:
                line += f", {r[3]} days notice"
            if r[5]:
                line += f" at {r[5]}"
            if r[6]:
                line += ' (auto-renewal)'
            if r[7]:
                line += ' [EXERCISED]'
            if r[8] or r[9]:
                line += f" ({_fmt_date(r[8])} — {_fmt_date(r[9])})"
            lines.append(line)
        sections['renewal_options'] = {
            'content': '\n'.join(lines), 'lease_ref': '',
        }
    else:
        sections['renewal_options'] = {
            'content': 'No language noted.' if ext else '',
            'lease_ref': '',
        }

    # Termination Options
    if terminations:
        lines = []
        for r in terminations:
            line = f"Option {r[0]} of {r[1]}"
            if r[3]:
                line += f", {r[3]} days notice"
            if r[5]:
                line += f" — {r[5]}"
            if r[6]:
                line += ' [EXERCISED]'
            if r[7]:
                line += f", earliest: {_fmt_date(r[7])}"
            lines.append(line)
        sections['termination_options'] = {
            'content': '\n'.join(lines), 'lease_ref': '',
        }
    else:
        sections['termination_options'] = {
            'content': 'No language noted.' if ext else '',
            'lease_ref': '',
        }

    # Use
    sections['use'] = {
        'content': ext.get('permitted_use') or '',
        'lease_ref': '',
    }

    # Exclusive Use
    if exc_rows:
        lines = [f"{r[1]}: {r[0]}" if r[1] else r[0] for r in exc_rows]
        sections['exclusive_use'] = {
            'content': '\n'.join(lines), 'lease_ref': '',
        }
    else:
        sections['exclusive_use'] = {
            'content': 'No language noted.' if ext else '',
            'lease_ref': '',
        }

    # CAM
    cam_parts = []
    if ext.get('cam_structure'):
        cam_parts.append(f"Structure: {ext['cam_structure']}")
    if ext.get('cam_cap_pct'):
        cam_parts.append(f"Cap: {ext['cam_cap_pct']}%")
    if ext.get('admin_fee_pct'):
        cam_parts.append(f"Admin Fee: {ext['admin_fee_pct']}%")
    sections['cam'] = {
        'content': ', '.join(cam_parts) if cam_parts else (
            'No language noted.' if ext else ''
        ),
        'lease_ref': '',
    }

    # Insurance
    ins_parts = []
    if ext.get('insurance_pass_through') is True:
        ins_parts.append('Insurance pass-through to tenant.')
    elif ext.get('insurance_pass_through') is False:
        ins_parts.append('No insurance pass-through.')
    sections['insurance'] = {
        'content': ' '.join(ins_parts) if ins_parts else (
            'No language noted.' if ext else ''
        ),
        'lease_ref': '',
    }

    # Real Estate Taxes
    tax_parts = []
    if ext.get('tax_pass_through') is True:
        tax_parts.append('Real estate tax pass-through to tenant.')
    elif ext.get('tax_pass_through') is False:
        tax_parts.append('No real estate tax pass-through.')
    sections['real_estate_taxes'] = {
        'content': ' '.join(tax_parts) if tax_parts else (
            'No language noted.' if ext else ''
        ),
        'lease_ref': '',
    }

    # Sections that are typically blank until manually populated
    for key in [
        'parking', 'roof', 'structural', 'signage', 'utilities',
        'hvac', 'estoppel', 'operating_covenants',
        'business_termination', 'casualty_termination',
        'relocation_rights', 'rofr_rofo', 'radius_restriction',
        'lease_inducements', 'merchants_association', 'redevelopment',
        'guarantor', 'notes',
    ]:
        sections[key] = {'content': '', 'lease_ref': ''}

    # Security Deposit
    sections['security_deposit'] = {
        'content': _fmt_money(security_deposit) if security_deposit else (
            'No language noted.' if ext else ''
        ),
        'lease_ref': '',
    }

    # Go-Dark Termination
    sections['go_dark_termination'] = {
        'content': ext.get('go_dark_provision') or (
            'No language noted.' if ext else ''
        ),
        'lease_ref': '',
    }

    # Co-Tenancy
    if cot_rows:
        lines = []
        ref_names = [r[0] for r in cot_refs]
        if ref_names:
            lines.append(f"Named Co-Tenants: {', '.join(ref_names)}")
        for c in cot_rows:
            if c[0]:
                lines.append(f"Trigger: {c[0]}")
            if c[2]:
                lines.append(f"Cure Period: {c[2]} days")
            if c[3]:
                lines.append(f"Alt Rent: {c[3]}")
            if c[4]:
                lines.append(
                    f"Termination Right: Yes"
                    + (f" ({c[5]} days notice)" if c[5] else '')
                )
            if c[6]:
                lines.append(f"Sunset: {c[6]}")
        sections['cotenancy'] = {
            'content': '\n'.join(lines), 'lease_ref': '',
        }
    else:
        sections['cotenancy'] = {
            'content': 'No language noted.' if ext else '',
            'lease_ref': '',
        }

    # TI Allowance
    sections['ti_allowance'] = {
        'content': _fmt_money(ext.get('ti_allowance')) if ext.get('ti_allowance') else (
            'No language noted.' if ext else ''
        ),
        'lease_ref': '',
    }

    # Assignment
    asgn = ext.get('assignment', {})
    if asgn:
        parts = []
        if asgn.get('consent_required') is True:
            parts.append('Landlord consent required.')
        elif asgn.get('consent_required') is False:
            parts.append('No landlord consent required.')
        if asgn.get('tenant_released') is True:
            parts.append('Tenant released upon assignment.')
        elif asgn.get('tenant_released') is False:
            parts.append('Tenant NOT released upon assignment.')
        sections['assignment'] = {
            'content': ' '.join(parts) if parts else 'No language noted.',
            'lease_ref': '',
        }
    else:
        sections['assignment'] = {
            'content': 'No language noted.' if ext else '',
            'lease_ref': '',
        }

    # Date of Review
    sections['date_of_review'] = {
        'content': datetime.now().strftime('%m/%d/%Y'),
        'lease_ref': '',
    }

    return sections


def get_tenant_abstract(
    engine, review_id: int, tenant_id: int
) -> Dict[str, Any]:
    """Get the abstract for a tenant, assembling from data if needed.

    Returns {tenant_name, suite, review_name, property_name, sections: [...]}.
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        # Tenant + review info
        info = conn.execute(text("""
            SELECT t.tenant_name, t.suite, r.property_name,
                   t.review_id
            FROM lease_tenants t
            JOIN lease_reviews r ON r.id = t.review_id
            WHERE t.id = :tid AND t.review_id = :rid
        """), {'tid': tenant_id, 'rid': review_id}).fetchone()
        if not info:
            return {'error': 'Tenant not found'}

        tenant_name = info[0]
        suite = info[1]
        property_name = info[2]

        # Check for saved abstract sections
        saved = conn.execute(text("""
            SELECT section_key, section_title, content, lease_ref, sort_order
            FROM lease_abstract_sections
            WHERE tenant_id = :tid
            ORDER BY sort_order
        """), {'tid': tenant_id}).fetchall()

    # If we have saved sections, use them (merging with template for any missing)
    saved_map = {r[0]: {
        'section_key': r[0], 'section_title': r[1],
        'content': r[2] or '', 'lease_ref': r[3] or '',
        'sort_order': r[4],
    } for r in saved}

    # Assemble from data for any section not yet saved
    if not saved_map:
        assembled = _assemble_abstract_from_data(engine, tenant_id, review_id)
    else:
        assembled = {}

    # Build final section list in template order
    result_sections = []
    for key, title, order in ABSTRACT_SECTIONS:
        if key in saved_map:
            result_sections.append(saved_map[key])
        elif key in assembled:
            result_sections.append({
                'section_key': key,
                'section_title': title,
                'content': assembled[key].get('content', ''),
                'lease_ref': assembled[key].get('lease_ref', ''),
                'sort_order': order,
            })
        else:
            result_sections.append({
                'section_key': key,
                'section_title': title,
                'content': '',
                'lease_ref': '',
                'sort_order': order,
            })

    return {
        'tenant_name': tenant_name,
        'suite': suite,
        'property_name': property_name,
        'tenant_id': tenant_id,
        'review_id': review_id,
        'sections': result_sections,
    }


def save_abstract_sections(
    engine, tenant_id: int, sections: List[Dict], username: str = ''
) -> None:
    """Save (upsert) abstract sections for a tenant."""
    from sqlalchemy import text

    with engine.begin() as conn:
        for s in sections:
            key = s.get('section_key')
            if not key:
                continue
            title = s.get('section_title', key)
            content = s.get('content', '')
            lease_ref = s.get('lease_ref', '')
            sort_order = s.get('sort_order', 0)

            if engine.dialect.name == 'postgresql':
                conn.execute(text("""
                    INSERT INTO lease_abstract_sections
                        (tenant_id, section_key, section_title, content,
                         lease_ref, sort_order, updated_by, updated_at)
                    VALUES (:tid, :sk, :st, :c, :lr, :so, :ub, CURRENT_TIMESTAMP)
                    ON CONFLICT (tenant_id, section_key)
                    DO UPDATE SET section_title = :st, content = :c,
                        lease_ref = :lr, sort_order = :so,
                        updated_by = :ub, updated_at = CURRENT_TIMESTAMP
                """), {
                    'tid': tenant_id, 'sk': key, 'st': title,
                    'c': content, 'lr': lease_ref, 'so': sort_order,
                    'ub': username,
                })
            else:
                conn.execute(text("""
                    INSERT OR REPLACE INTO lease_abstract_sections
                        (tenant_id, section_key, section_title, content,
                         lease_ref, sort_order, updated_by, updated_at)
                    VALUES (:tid, :sk, :st, :c, :lr, :so, :ub, CURRENT_TIMESTAMP)
                """), {
                    'tid': tenant_id, 'sk': key, 'st': title,
                    'c': content, 'lr': lease_ref, 'so': sort_order,
                    'ub': username,
                })


def get_review_abstracts_list(engine, review_id: int) -> List[Dict]:
    """Get a list of tenants and whether they have abstract data."""
    from sqlalchemy import text

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT t.id, t.tenant_name, t.suite, t.is_vacant,
                   (SELECT COUNT(*) FROM lease_abstract_sections a
                    WHERE a.tenant_id = t.id) as section_count
            FROM lease_tenants t
            WHERE t.review_id = :rid
            ORDER BY t.tenant_name
        """), {'rid': review_id}).fetchall()

    return [{
        'tenant_id': r[0], 'tenant_name': r[1], 'suite': r[2],
        'is_vacant': bool(r[3]),
        'has_abstract': r[4] > 0,
        'section_count': r[4],
    } for r in rows]


# ---------------------------------------------------------------------------
# Tenant Sales — AI extraction, import, TTM computation
# ---------------------------------------------------------------------------

_MONTH_MAP = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'apri': 4, 'april': 4,
    'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
    'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'oct': 10,
    'nov': 11, 'dec': 12, 'december': 12,
    'january': 1, 'february': 2, 'march': 3,
    'september': 9, 'october': 10, 'november': 11,
}


def _parse_month_range(comment: str) -> Tuple[int, int, int]:
    """Parse a month range from a sales comment string.

    Returns (month_start, month_end, months_covered).
    Examples:
        "Jan - May" → (1, 5, 5)
        "Apr - Dec" → (4, 12, 9)
        "Jan" → (1, 1, 1)
        "" or None → (1, 12, 12)  (full year)
        "REQ 8/7" → (0, 0, 0)  (no data)
    """
    if not comment:
        return (1, 12, 12)

    # Strip "REQ" follow-up dates and other non-month text
    cleaned = re.sub(r'\*.*$', '', comment).strip()
    cleaned = re.sub(r'REQ\s+[\d/\s]+', '', cleaned).strip()
    if cleaned.startswith('REQ') or not cleaned:
        # Only a request note, no actual month data
        if 'REQ' in comment and not any(
                m in comment.lower().split('req')[0].lower()
                for m in _MONTH_MAP):
            return (0, 0, 0)

    # Find month names in the comment
    months_found = []
    for word in re.split(r'[\s\-–—,]+', cleaned):
        w = word.strip().lower().rstrip('.')
        if w in _MONTH_MAP:
            months_found.append(_MONTH_MAP[w])

    if len(months_found) == 0:
        # No months found — check if pure "REQ" note
        if 'REQ' in comment.upper():
            return (0, 0, 0)
        return (1, 12, 12)
    elif len(months_found) == 1:
        return (months_found[0], months_found[0], 1)
    else:
        ms, me = months_found[0], months_found[-1]
        if me >= ms:
            return (ms, me, me - ms + 1)
        else:
            # Wrap around (unlikely but handle)
            return (ms, me, (12 - ms + 1) + me)


def extract_sales_from_pdf(
    file_obj,
    api_key: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Extract tenant sales data from a PDF using Claude AI.

    Returns a list of dicts:
        [{tenant_name, area, year, sales_amount, month_start, month_end,
          months_covered, comment}, ...]
    """
    import anthropic
    import pdfplumber
    import io

    key = api_key or os.environ.get('ANTHROPIC_API_KEY')
    if not key:
        raise ValueError("ANTHROPIC_API_KEY not set")

    # Extract text from PDF
    if isinstance(file_obj, bytes):
        file_obj = io.BytesIO(file_obj)
    text = ''
    with pdfplumber.open(file_obj) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or '') + '\n'

    if not text.strip():
        raise ValueError("No text extracted from PDF")

    client = anthropic.Anthropic(api_key=key)

    prompt = f"""Extract all tenant sales data from this property sales report.
Return a JSON array where each element has:
- "tenant_name": string (the tenant/business name)
- "area": number (square feet)
- "year": number (the reporting year)
- "sales_amount": number (dollar amount, 0 if shown as "-" or blank)
- "comment": string (the comment field, e.g. "Jan - May", "REQ 8/7", etc.)

Important:
- Include every row from every tenant, even if sales_amount is 0
- Keep tenant names exactly as shown in the report
- Parse dollar amounts removing $ and commas
- If the amount is shown as "-" or is blank, set sales_amount to 0
- Include the full comment text as-is

Return ONLY the JSON array, no other text.

REPORT TEXT:
{text}"""

    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=8192,
        messages=[{"role": "user", "content": prompt}],
    )

    response_text = message.content[0].text

    # Parse JSON from response
    try:
        json_match = re.search(r'\[[\s\S]*\]', response_text)
        if json_match:
            raw_entries = json.loads(json_match.group())
        else:
            raise ValueError("No JSON array found in AI response")
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse sales JSON: {e}")
        raise ValueError(f"AI extraction returned invalid JSON: {e}")

    # Enrich with parsed month ranges
    results = []
    for entry in raw_entries:
        comment = str(entry.get('comment', '') or '')
        ms, me, mc = _parse_month_range(comment)
        results.append({
            'tenant_name': str(entry.get('tenant_name', '')).strip(),
            'area': float(entry.get('area', 0) or 0),
            'year': int(entry.get('year', 0)),
            'sales_amount': float(entry.get('sales_amount', 0) or 0),
            'month_start': ms,
            'month_end': me,
            'months_covered': mc,
            'comment': comment,
        })

    logger.info(f"Extracted {len(results)} sales entries from PDF")
    return results


def import_sales_to_review(
    engine,
    review_id: int,
    sales_entries: List[Dict[str, Any]],
    source: str = 'ai_extract',
) -> Dict[str, Any]:
    """Import extracted sales data into lease_tenant_sales.

    Fuzzy-matches tenant names from the sales report to existing tenants.
    Uses UPSERT on (tenant_id, year) — newer imports overwrite older ones.

    Returns {matched, unmatched, total, unmatched_tenants}.
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        # Load existing tenants
        tenants = conn.execute(text("""
            SELECT id, tenant_name, suite, square_feet
            FROM lease_tenants
            WHERE review_id = :rid AND is_vacant = false
        """), {'rid': review_id}).fetchall()

        tenant_list = [{'id': r[0], 'name': r[1], 'suite': r[2],
                         'sf': r[3]} for r in tenants]

        # Group sales entries by tenant name
        by_tenant: Dict[str, List[Dict]] = {}
        for entry in sales_entries:
            tn = entry['tenant_name']
            by_tenant.setdefault(tn, []).append(entry)

        matched = 0
        unmatched = 0
        unmatched_names = []

        for sales_name, entries in by_tenant.items():
            # Find matching tenant by fuzzy name match
            match = None
            for t in tenant_list:
                if _fuzzy_match_tenant('', sales_name, '', t['name']):
                    match = t
                    break

            if not match:
                unmatched += 1
                unmatched_names.append(sales_name)
                logger.warning(f"Sales import: no match for '{sales_name}'")
                continue

            matched += 1
            for entry in entries:
                if entry['year'] == 0:
                    continue
                if engine.dialect.name == 'postgresql':
                    conn.execute(text("""
                        INSERT INTO lease_tenant_sales
                            (tenant_id, review_id, year, sales_amount,
                             month_start, month_end, months_covered,
                             comment, source, updated_at)
                        VALUES (:tid, :rid, :yr, :amt,
                                :ms, :me, :mc,
                                :cmt, :src, CURRENT_TIMESTAMP)
                        ON CONFLICT (tenant_id, year)
                        DO UPDATE SET sales_amount = :amt,
                                      month_start = :ms, month_end = :me,
                                      months_covered = :mc,
                                      comment = :cmt, source = :src,
                                      updated_at = CURRENT_TIMESTAMP
                    """), {
                        'tid': match['id'], 'rid': review_id,
                        'yr': entry['year'], 'amt': entry['sales_amount'],
                        'ms': entry['month_start'], 'me': entry['month_end'],
                        'mc': entry['months_covered'],
                        'cmt': entry.get('comment', ''), 'src': source,
                    })
                else:
                    conn.execute(text("""
                        INSERT OR REPLACE INTO lease_tenant_sales
                            (tenant_id, review_id, year, sales_amount,
                             month_start, month_end, months_covered,
                             comment, source, updated_at)
                        VALUES (:tid, :rid, :yr, :amt,
                                :ms, :me, :mc,
                                :cmt, :src, CURRENT_TIMESTAMP)
                    """), {
                        'tid': match['id'], 'rid': review_id,
                        'yr': entry['year'], 'amt': entry['sales_amount'],
                        'ms': entry['month_start'], 'me': entry['month_end'],
                        'mc': entry['months_covered'],
                        'cmt': entry.get('comment', ''), 'src': source,
                    })

        conn.commit()

    report = {
        'matched': matched,
        'unmatched': unmatched,
        'total': len(by_tenant),
        'unmatched_tenants': unmatched_names,
    }
    logger.info(f"Sales import review {review_id}: {matched} matched, "
                f"{unmatched} unmatched of {len(by_tenant)} tenants")
    return report


def compute_tenant_ttm_sales(
    engine,
    review_id: int,
) -> Dict[int, Dict[str, Any]]:
    """Compute trailing 12-month (TTM) sales per tenant.

    Logic:
    - Find the most recent year with data for each tenant
    - If full year (12 months), use it directly
    - If partial year (e.g. Jan-May = 5 months), impute remaining months
      from the prior year: prior_annual / 12 * (12 - current_months) + current_amount
    - Also returns raw sales history for display

    Returns {tenant_id: {ttm_sales, sales_per_sf, history: [{year, amount, months, comment}]}}.
    """
    from sqlalchemy import text

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT s.tenant_id, s.year, s.sales_amount,
                   s.month_start, s.month_end, s.months_covered,
                   s.comment, t.square_feet, t.annual_sales_override
            FROM lease_tenant_sales s
            JOIN lease_tenants t ON t.id = s.tenant_id
            WHERE s.review_id = :rid
            ORDER BY s.tenant_id, s.year DESC
        """), {'rid': review_id}).fetchall()

    # Group by tenant
    by_tenant: Dict[int, List] = {}
    tenant_sf: Dict[int, float] = {}
    tenant_override: Dict[int, Optional[float]] = {}
    for r in rows:
        tid = r[0]
        by_tenant.setdefault(tid, []).append({
            'year': r[1], 'amount': r[2],
            'month_start': r[3], 'month_end': r[4],
            'months': r[5], 'comment': r[6],
        })
        tenant_sf[tid] = r[7] or 0
        tenant_override[tid] = r[8]

    result = {}
    for tid, entries in by_tenant.items():
        # entries are sorted by year DESC
        history = sorted(entries, key=lambda e: e['year'])

        override = tenant_override.get(tid)
        if override is not None and override > 0:
            ttm = override
        else:
            # Find the most recent entry
            latest = entries[0]  # already sorted DESC
            if latest['months'] >= 12:
                ttm = latest['amount']
            elif latest['months'] > 0 and len(entries) > 1:
                # Partial current year — impute from prior year
                prior = entries[1]
                if prior['months'] >= 12 and prior['amount'] > 0:
                    # prior_annual / 12 * remaining_months + current_partial
                    remaining = 12 - latest['months']
                    ttm = (prior['amount'] / 12 * remaining) + latest['amount']
                elif prior['months'] > 0 and prior['amount'] > 0:
                    # Prior year also partial — annualize it then impute
                    prior_annualized = prior['amount'] / prior['months'] * 12
                    remaining = 12 - latest['months']
                    ttm = (prior_annualized / 12 * remaining) + latest['amount']
                else:
                    # No usable prior year — annualize current
                    ttm = latest['amount'] / latest['months'] * 12 if latest['months'] > 0 else 0
            elif latest['months'] > 0:
                # Only one entry — annualize it
                ttm = latest['amount'] / latest['months'] * 12
            else:
                ttm = 0

        sf = tenant_sf.get(tid, 0)
        result[tid] = {
            'ttm_sales': round(ttm, 2),
            'sales_per_sf': round(ttm / sf, 2) if sf > 0 else 0,
            'has_override': override is not None and override > 0,
            'history': [{
                'year': e['year'],
                'amount': e['amount'],
                'months': e['months'],
                'comment': e['comment'],
            } for e in history],
        }

    return result


def update_tenant_sales_override(
    engine,
    review_id: int,
    tenant_id: int,
    annual_sales: Optional[float],
) -> None:
    """Set or clear the annual sales override for a tenant."""
    from sqlalchemy import text

    with engine.begin() as conn:
        # Verify tenant belongs to this review
        t = conn.execute(text("""
            SELECT id FROM lease_tenants
            WHERE id = :tid AND review_id = :rid
        """), {'tid': tenant_id, 'rid': review_id}).fetchone()
        if not t:
            raise ValueError(f"Tenant {tenant_id} not found in review {review_id}")

        conn.execute(text("""
            UPDATE lease_tenants
            SET annual_sales_override = :val, updated_at = CURRENT_TIMESTAMP
            WHERE id = :tid
        """), {'val': annual_sales, 'tid': tenant_id})


def get_tenant_sales(engine, review_id: int) -> Dict[str, Any]:
    """Get all sales data for a review including TTM computation.

    Returns {tenants: {tid: {ttm_sales, sales_per_sf, history}}, has_sales: bool}.
    """
    ttm = compute_tenant_ttm_sales(engine, review_id)
    return {
        'tenants': {str(k): v for k, v in ttm.items()},
        'has_sales': len(ttm) > 0,
    }


# ---------------------------------------------------------------------------
# Phase 1: Tenant CRUD + Editable Rent Roll
# ---------------------------------------------------------------------------

def _verify_tenant_in_review(conn, review_id: int, tenant_id: int):
    """Verify tenant belongs to review, raise ValueError if not."""
    from sqlalchemy import text
    row = conn.execute(text(
        "SELECT id FROM lease_tenants WHERE id = :tid AND review_id = :rid"
    ), {'tid': tenant_id, 'rid': review_id}).fetchone()
    if not row:
        raise ValueError(f"Tenant {tenant_id} not found in review {review_id}")


def add_tenant(engine, review_id: int, data: Dict[str, Any]) -> Dict[str, Any]:
    """Add a new tenant to a review."""
    from sqlalchemy import text
    with engine.begin() as conn:
        # Verify review exists
        r = conn.execute(text(
            "SELECT id FROM lease_reviews WHERE id = :rid"
        ), {'rid': review_id}).fetchone()
        if not r:
            raise ValueError(f"Review {review_id} not found")

        result = conn.execute(text("""
            INSERT INTO lease_tenants (
                review_id, tenant_name, suite, square_feet, lease_type,
                lease_start, lease_end, monthly_rent, annual_rent, rent_per_sf,
                security_deposit, is_vacant, tenant_status
            ) VALUES (
                :rid, :name, :suite, :sf, :lt,
                :ls, :le, :mr, :ar, :rpsf,
                :sd, :vac, 'active'
            )
            RETURNING id
        """) if engine.dialect.name == 'postgresql' else text("""
            INSERT INTO lease_tenants (
                review_id, tenant_name, suite, square_feet, lease_type,
                lease_start, lease_end, monthly_rent, annual_rent, rent_per_sf,
                security_deposit, is_vacant, tenant_status
            ) VALUES (
                :rid, :name, :suite, :sf, :lt,
                :ls, :le, :mr, :ar, :rpsf,
                :sd, :vac, 'active'
            )
        """), {
            'rid': review_id,
            'name': data.get('tenant_name', 'New Tenant'),
            'suite': data.get('suite', ''),
            'sf': data.get('square_feet'),
            'lt': data.get('lease_type', ''),
            'ls': data.get('lease_start'),
            'le': data.get('lease_end'),
            'mr': data.get('monthly_rent'),
            'ar': data.get('annual_rent'),
            'rpsf': data.get('rent_per_sf'),
            'sd': data.get('security_deposit'),
            'vac': data.get('is_vacant', False),
        })

        if engine.dialect.name == 'postgresql':
            new_id = result.fetchone()[0]
        else:
            new_id = result.lastrowid

    return {'id': new_id, 'status': 'created'}


def update_tenant_fields(
    engine, review_id: int, tenant_id: int, fields: Dict[str, Any],
) -> Dict[str, Any]:
    """Update tenant fields directly (lighter than field resolution)."""
    from sqlalchemy import text
    ALLOWED_FIELDS = {
        'tenant_name', 'suite', 'square_feet', 'lease_type',
        'lease_start', 'lease_end', 'monthly_rent', 'annual_rent',
        'rent_per_sf', 'security_deposit', 'is_vacant', 'is_material',
        'has_cotenancy', 'has_exclusive_use', 'term_months',
    }
    update_fields = {k: v for k, v in fields.items() if k in ALLOWED_FIELDS}
    if not update_fields:
        return {'status': 'no_changes'}

    with engine.begin() as conn:
        _verify_tenant_in_review(conn, review_id, tenant_id)
        set_clause = ', '.join(f"{k} = :{k}" for k in update_fields)
        update_fields['tid'] = tenant_id
        conn.execute(text(
            f"UPDATE lease_tenants SET {set_clause}, updated_at = CURRENT_TIMESTAMP WHERE id = :tid"
        ), update_fields)

    return {'status': 'updated', 'fields': list(update_fields.keys())}


def delete_tenant(engine, review_id: int, tenant_id: int) -> Dict[str, Any]:
    """Soft-delete a tenant (set tenant_status='deleted')."""
    from sqlalchemy import text
    with engine.begin() as conn:
        _verify_tenant_in_review(conn, review_id, tenant_id)
        conn.execute(text("""
            UPDATE lease_tenants
            SET tenant_status = 'deleted', updated_at = CURRENT_TIMESTAMP
            WHERE id = :tid
        """), {'tid': tenant_id})
    return {'status': 'deleted', 'tenant_id': tenant_id}


def mark_tenant_vacant(
    engine, review_id: int, tenant_id: int, vacant: bool = True,
) -> Dict[str, Any]:
    """Mark a tenant as vacant/occupied, optionally zeroing rent."""
    from sqlalchemy import text
    with engine.begin() as conn:
        _verify_tenant_in_review(conn, review_id, tenant_id)
        if vacant:
            conn.execute(text("""
                UPDATE lease_tenants
                SET is_vacant = :vac, monthly_rent = 0, annual_rent = 0,
                    rent_per_sf = 0, updated_at = CURRENT_TIMESTAMP
                WHERE id = :tid
            """), {'vac': True, 'tid': tenant_id})
        else:
            conn.execute(text("""
                UPDATE lease_tenants
                SET is_vacant = :vac, updated_at = CURRENT_TIMESTAMP
                WHERE id = :tid
            """), {'vac': False, 'tid': tenant_id})
    return {'status': 'updated', 'is_vacant': vacant}


# ---------------------------------------------------------------------------
# Phase 2: Space Mutations (merge, split, resize)
# ---------------------------------------------------------------------------

def _create_space_event(
    conn, engine, review_id: int, event_type: str, effective_date: str,
    source_tenant_ids: List[int], description: str = '',
    status: str = 'applied', created_by: str = '',
) -> int:
    """Insert a lease_space_events row, return new id."""
    from sqlalchemy import text
    ids_str = ','.join(str(i) for i in source_tenant_ids)
    if engine.dialect.name == 'postgresql':
        row = conn.execute(text("""
            INSERT INTO lease_space_events
                (review_id, event_type, effective_date, source_tenant_ids,
                 description, status, created_by)
            VALUES (:rid, :et, :ed, :sti, :desc, :st, :cb)
            RETURNING id
        """), {
            'rid': review_id, 'et': event_type, 'ed': effective_date,
            'sti': ids_str, 'desc': description, 'st': status, 'cb': created_by,
        }).fetchone()
        return row[0]
    else:
        result = conn.execute(text("""
            INSERT INTO lease_space_events
                (review_id, event_type, effective_date, source_tenant_ids,
                 description, status, created_by)
            VALUES (:rid, :et, :ed, :sti, :desc, :st, :cb)
        """), {
            'rid': review_id, 'et': event_type, 'ed': effective_date,
            'sti': ids_str, 'desc': description, 'st': status, 'cb': created_by,
        })
        return result.lastrowid


def _insert_event_result(conn, engine, event_id: int, data: Dict) -> int:
    """Insert a lease_space_event_results row, return new id."""
    from sqlalchemy import text
    if engine.dialect.name == 'postgresql':
        row = conn.execute(text("""
            INSERT INTO lease_space_event_results
                (event_id, result_tenant_id, tenant_name, suite, square_feet,
                 monthly_rent, annual_rent, rent_per_sf, lease_start, lease_end,
                 is_vacant, notes)
            VALUES (:eid, :rtid, :tn, :s, :sf, :mr, :ar, :rpsf, :ls, :le, :iv, :n)
            RETURNING id
        """), {
            'eid': event_id, 'rtid': data.get('result_tenant_id'),
            'tn': data.get('tenant_name', ''), 's': data.get('suite', ''),
            'sf': data.get('square_feet'), 'mr': data.get('monthly_rent'),
            'ar': data.get('annual_rent'), 'rpsf': data.get('rent_per_sf'),
            'ls': data.get('lease_start'), 'le': data.get('lease_end'),
            'iv': data.get('is_vacant', False), 'n': data.get('notes'),
        }).fetchone()
        return row[0]
    else:
        result = conn.execute(text("""
            INSERT INTO lease_space_event_results
                (event_id, result_tenant_id, tenant_name, suite, square_feet,
                 monthly_rent, annual_rent, rent_per_sf, lease_start, lease_end,
                 is_vacant, notes)
            VALUES (:eid, :rtid, :tn, :s, :sf, :mr, :ar, :rpsf, :ls, :le, :iv, :n)
        """), {
            'eid': event_id, 'rtid': data.get('result_tenant_id'),
            'tn': data.get('tenant_name', ''), 's': data.get('suite', ''),
            'sf': data.get('square_feet'), 'mr': data.get('monthly_rent'),
            'ar': data.get('annual_rent'), 'rpsf': data.get('rent_per_sf'),
            'ls': data.get('lease_start'), 'le': data.get('lease_end'),
            'iv': data.get('is_vacant', False), 'n': data.get('notes'),
        })
        return result.lastrowid


def merge_suites(
    engine, review_id: int, source_ids: List[int],
    merged_suite: str, merged_name: str,
    effective_date: str = '', created_by: str = '',
) -> Dict[str, Any]:
    """Merge 2+ tenants into one. Sources marked replaced, new tenant created."""
    from sqlalchemy import text
    if len(source_ids) < 2:
        raise ValueError("Merge requires at least 2 source tenants")

    with engine.begin() as conn:
        # Load source tenants
        placeholders = ', '.join(f':t{i}' for i in range(len(source_ids)))
        params = {f't{i}': tid for i, tid in enumerate(source_ids)}
        params['rid'] = review_id
        sources = conn.execute(text(f"""
            SELECT id, tenant_name, suite, square_feet, monthly_rent,
                   annual_rent, lease_start, lease_end
            FROM lease_tenants
            WHERE id IN ({placeholders}) AND review_id = :rid
        """), params).fetchall()
        if len(sources) != len(source_ids):
            raise ValueError("Some source tenants not found in this review")

        # Aggregate
        total_sf = sum(s[3] or 0 for s in sources)
        total_monthly = sum(s[4] or 0 for s in sources)
        total_annual = sum(s[5] or 0 for s in sources)
        rpsf = total_annual / total_sf if total_sf > 0 else 0
        # Use latest lease_end, earliest lease_start
        starts = [s[6] for s in sources if s[6]]
        ends = [s[7] for s in sources if s[7]]
        ls = min(starts) if starts else None
        le = max(ends) if ends else None

        if not effective_date:
            effective_date = datetime.now().strftime('%Y-%m-%d')

        # Create the event
        event_id = _create_space_event(
            conn, engine, review_id, 'merge', effective_date,
            source_ids, f"Merged {len(source_ids)} tenants into {merged_name}",
            'applied', created_by,
        )

        # Create merged tenant
        merged_data = {
            'rid': review_id, 'name': merged_name, 'suite': merged_suite,
            'sf': total_sf, 'mr': total_monthly, 'ar': total_annual,
            'rpsf': rpsf, 'ls': ls, 'le': le,
        }
        if engine.dialect.name == 'postgresql':
            new_row = conn.execute(text("""
                INSERT INTO lease_tenants
                    (review_id, tenant_name, suite, square_feet, monthly_rent,
                     annual_rent, rent_per_sf, lease_start, lease_end, tenant_status)
                VALUES (:rid, :name, :suite, :sf, :mr, :ar, :rpsf, :ls, :le, 'active')
                RETURNING id
            """), merged_data).fetchone()
            new_id = new_row[0]
        else:
            result = conn.execute(text("""
                INSERT INTO lease_tenants
                    (review_id, tenant_name, suite, square_feet, monthly_rent,
                     annual_rent, rent_per_sf, lease_start, lease_end, tenant_status)
                VALUES (:rid, :name, :suite, :sf, :mr, :ar, :rpsf, :ls, :le, 'active')
            """), merged_data)
            new_id = result.lastrowid

        # Record event result
        _insert_event_result(conn, engine, event_id, {
            'result_tenant_id': new_id, 'tenant_name': merged_name,
            'suite': merged_suite, 'square_feet': total_sf,
            'monthly_rent': total_monthly, 'annual_rent': total_annual,
            'rent_per_sf': rpsf, 'lease_start': ls, 'lease_end': le,
        })

        # Mark sources as replaced
        for sid in source_ids:
            conn.execute(text("""
                UPDATE lease_tenants
                SET tenant_status = 'replaced', replaced_by_event_id = :eid,
                    successor_tenant_id = :nid, updated_at = CURRENT_TIMESTAMP
                WHERE id = :tid
            """), {'eid': event_id, 'nid': new_id, 'tid': sid})

    return {'event_id': event_id, 'new_tenant_id': new_id, 'status': 'merged'}


def split_suite(
    engine, review_id: int, source_id: int,
    splits: List[Dict[str, Any]],
    effective_date: str = '', created_by: str = '',
) -> Dict[str, Any]:
    """Split 1 tenant into N new tenants. Validates SF sum within tolerance."""
    from sqlalchemy import text
    if len(splits) < 2:
        raise ValueError("Split requires at least 2 result tenants")

    with engine.begin() as conn:
        _verify_tenant_in_review(conn, review_id, source_id)
        source = conn.execute(text(
            "SELECT square_feet, tenant_name FROM lease_tenants WHERE id = :tid"
        ), {'tid': source_id}).fetchone()
        source_sf = source[0] or 0
        source_name = source[1]

        # Validate SF sum
        split_sf = sum(s.get('square_feet', 0) or 0 for s in splits)
        if source_sf > 0 and abs(split_sf - source_sf) > 1:
            raise ValueError(
                f"Split SF total ({split_sf:,.0f}) must match source "
                f"({source_sf:,.0f}) within 1 SF"
            )

        if not effective_date:
            effective_date = datetime.now().strftime('%Y-%m-%d')

        event_id = _create_space_event(
            conn, engine, review_id, 'split', effective_date,
            [source_id], f"Split {source_name} into {len(splits)} units",
            'applied', created_by,
        )

        new_ids = []
        for s in splits:
            data = {
                'rid': review_id, 'name': s.get('tenant_name', 'TBD'),
                'suite': s.get('suite', ''), 'sf': s.get('square_feet'),
                'mr': s.get('monthly_rent'), 'ar': s.get('annual_rent'),
                'rpsf': s.get('rent_per_sf'),
                'ls': s.get('lease_start'), 'le': s.get('lease_end'),
                'vac': s.get('is_vacant', False),
            }
            if engine.dialect.name == 'postgresql':
                row = conn.execute(text("""
                    INSERT INTO lease_tenants
                        (review_id, tenant_name, suite, square_feet, monthly_rent,
                         annual_rent, rent_per_sf, lease_start, lease_end,
                         is_vacant, tenant_status)
                    VALUES (:rid, :name, :suite, :sf, :mr, :ar, :rpsf,
                            :ls, :le, :vac, 'active')
                    RETURNING id
                """), data).fetchone()
                nid = row[0]
            else:
                result = conn.execute(text("""
                    INSERT INTO lease_tenants
                        (review_id, tenant_name, suite, square_feet, monthly_rent,
                         annual_rent, rent_per_sf, lease_start, lease_end,
                         is_vacant, tenant_status)
                    VALUES (:rid, :name, :suite, :sf, :mr, :ar, :rpsf,
                            :ls, :le, :vac, 'active')
                """), data)
                nid = result.lastrowid

            new_ids.append(nid)
            _insert_event_result(conn, engine, event_id, {
                'result_tenant_id': nid, **s,
            })

        # Mark source as replaced
        conn.execute(text("""
            UPDATE lease_tenants
            SET tenant_status = 'replaced', replaced_by_event_id = :eid,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = :tid
        """), {'eid': event_id, 'tid': source_id})

    return {'event_id': event_id, 'new_tenant_ids': new_ids, 'status': 'split'}


def resize_tenant(
    engine, review_id: int, tenant_id: int,
    new_sf: float, new_rent: Optional[float] = None,
    effective_date: str = '', created_by: str = '',
) -> Dict[str, Any]:
    """Resize a tenant in-place with audit trail."""
    from sqlalchemy import text
    with engine.begin() as conn:
        _verify_tenant_in_review(conn, review_id, tenant_id)
        old = conn.execute(text(
            "SELECT square_feet, annual_rent, tenant_name FROM lease_tenants WHERE id = :tid"
        ), {'tid': tenant_id}).fetchone()

        if not effective_date:
            effective_date = datetime.now().strftime('%Y-%m-%d')

        event_id = _create_space_event(
            conn, engine, review_id, 'resize', effective_date,
            [tenant_id],
            f"Resize {old[2]}: {old[0]:,.0f}→{new_sf:,.0f} SF",
            'applied', created_by,
        )

        update_params = {'tid': tenant_id, 'sf': new_sf}
        set_parts = ['square_feet = :sf']
        if new_rent is not None:
            update_params['ar'] = new_rent
            update_params['rpsf'] = new_rent / new_sf if new_sf > 0 else 0
            update_params['mr'] = new_rent / 12
            set_parts.extend([
                'annual_rent = :ar', 'rent_per_sf = :rpsf', 'monthly_rent = :mr',
            ])
        conn.execute(text(
            f"UPDATE lease_tenants SET {', '.join(set_parts)}, "
            f"updated_at = CURRENT_TIMESTAMP WHERE id = :tid"
        ), update_params)

        _insert_event_result(conn, engine, event_id, {
            'result_tenant_id': tenant_id, 'tenant_name': old[2],
            'square_feet': new_sf, 'annual_rent': new_rent,
            'notes': f"Resized from {old[0]:,.0f} SF",
        })

    return {'event_id': event_id, 'status': 'resized'}


# ---------------------------------------------------------------------------
# Phase 3: Future Space Plans (events system)
# ---------------------------------------------------------------------------

def create_space_event(
    engine, review_id: int, event_data: Dict[str, Any],
) -> Dict[str, Any]:
    """Create a planned future space event with result templates."""
    from sqlalchemy import text
    with engine.begin() as conn:
        event_id = _create_space_event(
            conn, engine, review_id,
            event_data['event_type'],
            event_data['effective_date'],
            event_data.get('source_tenant_ids', []),
            event_data.get('description', ''),
            'planned',
            event_data.get('created_by', ''),
        )
        for r in event_data.get('results', []):
            _insert_event_result(conn, engine, event_id, r)

    return {'event_id': event_id, 'status': 'planned'}


def update_space_event(
    engine, event_id: int, data: Dict[str, Any],
) -> Dict[str, Any]:
    """Update a planned space event."""
    from sqlalchemy import text
    with engine.begin() as conn:
        row = conn.execute(text(
            "SELECT status FROM lease_space_events WHERE id = :eid"
        ), {'eid': event_id}).fetchone()
        if not row:
            raise ValueError(f"Event {event_id} not found")
        if row[0] != 'planned':
            raise ValueError("Only planned events can be edited")

        updates = {}
        for field in ('event_type', 'effective_date', 'description'):
            if field in data:
                updates[field] = data[field]
        if 'source_tenant_ids' in data:
            updates['source_tenant_ids'] = ','.join(
                str(i) for i in data['source_tenant_ids']
            )

        if updates:
            set_clause = ', '.join(f"{k} = :{k}" for k in updates)
            updates['eid'] = event_id
            conn.execute(text(
                f"UPDATE lease_space_events SET {set_clause}, "
                f"updated_at = CURRENT_TIMESTAMP WHERE id = :eid"
            ), updates)

        # Replace results if provided
        if 'results' in data:
            conn.execute(text(
                "DELETE FROM lease_space_event_results WHERE event_id = :eid"
            ), {'eid': event_id})
            for r in data['results']:
                _insert_event_result(conn, engine, event_id, r)

    return {'status': 'updated'}


def cancel_space_event(engine, event_id: int) -> Dict[str, Any]:
    """Cancel a planned event. If it was applied, revert tenant changes."""
    from sqlalchemy import text
    with engine.begin() as conn:
        row = conn.execute(text(
            "SELECT status, review_id FROM lease_space_events WHERE id = :eid"
        ), {'eid': event_id}).fetchone()
        if not row:
            raise ValueError(f"Event {event_id} not found")

        if row[0] == 'applied':
            # Revert: un-replace source tenants, delete result tenants
            conn.execute(text("""
                UPDATE lease_tenants
                SET tenant_status = 'active', replaced_by_event_id = NULL,
                    successor_tenant_id = NULL, updated_at = CURRENT_TIMESTAMP
                WHERE replaced_by_event_id = :eid
            """), {'eid': event_id})
            # Delete tenants that were created by this event
            result_ids = conn.execute(text(
                "SELECT result_tenant_id FROM lease_space_event_results "
                "WHERE event_id = :eid AND result_tenant_id IS NOT NULL"
            ), {'eid': event_id}).fetchall()
            for r in result_ids:
                # Only delete if the tenant was created by this event
                # (result_tenant_id != source_tenant_id)
                sources = conn.execute(text(
                    "SELECT source_tenant_ids FROM lease_space_events WHERE id = :eid"
                ), {'eid': event_id}).fetchone()
                source_ids = [int(x) for x in sources[0].split(',') if x.strip()]
                if r[0] not in source_ids:
                    conn.execute(text(
                        "UPDATE lease_tenants SET tenant_status = 'deleted' WHERE id = :tid"
                    ), {'tid': r[0]})

        conn.execute(text("""
            UPDATE lease_space_events
            SET status = 'cancelled', updated_at = CURRENT_TIMESTAMP
            WHERE id = :eid
        """), {'eid': event_id})

    return {'status': 'cancelled'}


def apply_space_event(engine, event_id: int) -> Dict[str, Any]:
    """Materialize a planned event into the tenant roster."""
    from sqlalchemy import text
    with engine.begin() as conn:
        evt = conn.execute(text("""
            SELECT id, review_id, event_type, source_tenant_ids, status
            FROM lease_space_events WHERE id = :eid
        """), {'eid': event_id}).fetchone()
        if not evt:
            raise ValueError(f"Event {event_id} not found")
        if evt[4] != 'planned':
            raise ValueError("Only planned events can be applied")

        review_id = evt[1]
        event_type = evt[2]
        source_ids = [int(x) for x in evt[3].split(',') if x.strip()]

        # Get event results (templates for new tenants)
        results = conn.execute(text("""
            SELECT id, tenant_name, suite, square_feet, monthly_rent,
                   annual_rent, rent_per_sf, lease_start, lease_end, is_vacant, notes
            FROM lease_space_event_results WHERE event_id = :eid
        """), {'eid': event_id}).fetchall()

        new_ids = []
        for r in results:
            data = {
                'rid': review_id, 'name': r[1] or 'TBD', 'suite': r[2] or '',
                'sf': r[3], 'mr': r[4], 'ar': r[5], 'rpsf': r[6],
                'ls': r[7], 'le': r[8], 'vac': bool(r[9]),
            }
            if engine.dialect.name == 'postgresql':
                row = conn.execute(text("""
                    INSERT INTO lease_tenants
                        (review_id, tenant_name, suite, square_feet, monthly_rent,
                         annual_rent, rent_per_sf, lease_start, lease_end,
                         is_vacant, tenant_status)
                    VALUES (:rid, :name, :suite, :sf, :mr, :ar, :rpsf,
                            :ls, :le, :vac, 'active')
                    RETURNING id
                """), data).fetchone()
                nid = row[0]
            else:
                result = conn.execute(text("""
                    INSERT INTO lease_tenants
                        (review_id, tenant_name, suite, square_feet, monthly_rent,
                         annual_rent, rent_per_sf, lease_start, lease_end,
                         is_vacant, tenant_status)
                    VALUES (:rid, :name, :suite, :sf, :mr, :ar, :rpsf,
                            :ls, :le, :vac, 'active')
                """), data)
                nid = result.lastrowid

            new_ids.append(nid)
            # Update result row with actual tenant id
            conn.execute(text(
                "UPDATE lease_space_event_results "
                "SET result_tenant_id = :nid WHERE id = :rid"
            ), {'nid': nid, 'rid': r[0]})

        # Mark source tenants as replaced (for non-vacate event types)
        if event_type != 'vacate':
            first_new = new_ids[0] if new_ids else None
            for sid in source_ids:
                conn.execute(text("""
                    UPDATE lease_tenants
                    SET tenant_status = 'replaced', replaced_by_event_id = :eid,
                        successor_tenant_id = :nid, updated_at = CURRENT_TIMESTAMP
                    WHERE id = :tid
                """), {'eid': event_id, 'nid': first_new, 'tid': sid})
        else:
            # Vacate: mark sources as vacant
            for sid in source_ids:
                conn.execute(text("""
                    UPDATE lease_tenants
                    SET is_vacant = 1, monthly_rent = 0, annual_rent = 0,
                        rent_per_sf = 0, updated_at = CURRENT_TIMESTAMP
                    WHERE id = :tid
                """), {'tid': sid})

        conn.execute(text("""
            UPDATE lease_space_events
            SET status = 'applied', updated_at = CURRENT_TIMESTAMP
            WHERE id = :eid
        """), {'eid': event_id})

    return {'status': 'applied', 'new_tenant_ids': new_ids}


def get_space_events(engine, review_id: int) -> List[Dict[str, Any]]:
    """Get all space events for a review with nested results."""
    from sqlalchemy import text
    with engine.connect() as conn:
        events = conn.execute(text("""
            SELECT id, event_type, effective_date, source_tenant_ids,
                   description, status, created_by, created_at
            FROM lease_space_events
            WHERE review_id = :rid
            ORDER BY effective_date, created_at
        """), {'rid': review_id}).fetchall()

        result = []
        for e in events:
            eid = e[0]
            results = conn.execute(text("""
                SELECT id, result_tenant_id, tenant_name, suite, square_feet,
                       monthly_rent, annual_rent, rent_per_sf, lease_start,
                       lease_end, is_vacant, notes
                FROM lease_space_event_results WHERE event_id = :eid
            """), {'eid': eid}).fetchall()

            # Get source tenant names
            source_ids = [int(x) for x in e[3].split(',') if x.strip()]
            source_names = []
            if source_ids:
                ph = ', '.join(f':s{i}' for i in range(len(source_ids)))
                params = {f's{i}': sid for i, sid in enumerate(source_ids)}
                names = conn.execute(text(f"""
                    SELECT id, tenant_name, suite FROM lease_tenants
                    WHERE id IN ({ph})
                """), params).fetchall()
                source_names = [
                    {'id': n[0], 'name': n[1], 'suite': n[2]} for n in names
                ]

            result.append({
                'id': eid, 'event_type': e[1], 'effective_date': e[2],
                'source_tenant_ids': source_ids,
                'source_tenants': source_names,
                'description': e[4], 'status': e[5],
                'created_by': e[6], 'created_at': str(e[7]),
                'results': [{
                    'id': r[0], 'result_tenant_id': r[1],
                    'tenant_name': r[2], 'suite': r[3],
                    'square_feet': r[4], 'monthly_rent': r[5],
                    'annual_rent': r[6], 'rent_per_sf': r[7],
                    'lease_start': r[8], 'lease_end': r[9],
                    'is_vacant': bool(r[10]), 'notes': r[11],
                } for r in results],
            })

    return result


def get_space_timeline(engine, review_id: int) -> Dict[str, Any]:
    """Project the tenant roster forward through planned events.

    Returns the current roster plus a chronological list of future transitions.
    """
    from sqlalchemy import text
    resolved = get_resolved_tenants(engine, review_id)
    active = [t for t in resolved if t.get('tenant_status', 'active') != 'deleted']
    events = get_space_events(engine, review_id)

    # Separate applied/planned/cancelled
    timeline = []
    for e in events:
        timeline.append({
            'event': e,
            'is_future': e['status'] == 'planned',
        })

    return {
        'current_roster': active,
        'timeline': timeline,
        'total_events': len(events),
        'planned_count': sum(1 for e in events if e['status'] == 'planned'),
        'applied_count': sum(1 for e in events if e['status'] == 'applied'),
    }


# ---------------------------------------------------------------------------
# Phase 4: Tenant Succession
# ---------------------------------------------------------------------------

def create_succession(
    engine, review_id: int, source_id: int,
    new_tenant_data: Dict[str, Any], effective_date: str,
    created_by: str = '',
) -> Dict[str, Any]:
    """Create a succession event: one tenant replaces another."""
    from sqlalchemy import text
    with engine.begin() as conn:
        _verify_tenant_in_review(conn, review_id, source_id)
        source = conn.execute(text(
            "SELECT tenant_name, suite, square_feet FROM lease_tenants WHERE id = :tid"
        ), {'tid': source_id}).fetchone()

        # Default suite/SF from source if not specified
        new_tenant_data.setdefault('suite', source[1])
        new_tenant_data.setdefault('square_feet', source[2])

        event_id = _create_space_event(
            conn, engine, review_id, 'succession', effective_date,
            [source_id],
            f"{source[0]} → {new_tenant_data.get('tenant_name', 'TBD')}",
            'applied', created_by,
        )

        # Create new tenant
        data = {
            'rid': review_id,
            'name': new_tenant_data.get('tenant_name', 'TBD'),
            'suite': new_tenant_data.get('suite', ''),
            'sf': new_tenant_data.get('square_feet'),
            'mr': new_tenant_data.get('monthly_rent'),
            'ar': new_tenant_data.get('annual_rent'),
            'rpsf': new_tenant_data.get('rent_per_sf'),
            'ls': new_tenant_data.get('lease_start', effective_date),
            'le': new_tenant_data.get('lease_end'),
            'vac': new_tenant_data.get('is_vacant', False),
        }
        if engine.dialect.name == 'postgresql':
            row = conn.execute(text("""
                INSERT INTO lease_tenants
                    (review_id, tenant_name, suite, square_feet, monthly_rent,
                     annual_rent, rent_per_sf, lease_start, lease_end,
                     is_vacant, tenant_status)
                VALUES (:rid, :name, :suite, :sf, :mr, :ar, :rpsf,
                        :ls, :le, :vac, 'active')
                RETURNING id
            """), data).fetchone()
            new_id = row[0]
        else:
            result = conn.execute(text("""
                INSERT INTO lease_tenants
                    (review_id, tenant_name, suite, square_feet, monthly_rent,
                     annual_rent, rent_per_sf, lease_start, lease_end,
                     is_vacant, tenant_status)
                VALUES (:rid, :name, :suite, :sf, :mr, :ar, :rpsf,
                        :ls, :le, :vac, 'active')
            """), data)
            new_id = result.lastrowid

        _insert_event_result(conn, engine, event_id, {
            'result_tenant_id': new_id, **new_tenant_data,
        })

        # Mark source as replaced
        conn.execute(text("""
            UPDATE lease_tenants
            SET tenant_status = 'replaced', replaced_by_event_id = :eid,
                successor_tenant_id = :nid, updated_at = CURRENT_TIMESTAMP
            WHERE id = :tid
        """), {'eid': event_id, 'nid': new_id, 'tid': source_id})

    return {'event_id': event_id, 'new_tenant_id': new_id, 'status': 'succession'}


def get_succession_chain(engine, tenant_id: int) -> List[Dict[str, Any]]:
    """Follow successor_tenant_id links to build a succession chain."""
    from sqlalchemy import text
    chain = []
    visited = set()
    current_id = tenant_id

    with engine.connect() as conn:
        # Walk backwards to find the original tenant
        while current_id and current_id not in visited:
            visited.add(current_id)
            row = conn.execute(text("""
                SELECT id, tenant_name, suite, square_feet, annual_rent,
                       lease_start, lease_end, tenant_status, successor_tenant_id,
                       replaced_by_event_id
                FROM lease_tenants WHERE id = :tid
            """), {'tid': current_id}).fetchone()
            if not row:
                break
            chain.append({
                'id': row[0], 'tenant_name': row[1], 'suite': row[2],
                'square_feet': row[3], 'annual_rent': row[4],
                'lease_start': row[5], 'lease_end': row[6],
                'tenant_status': row[7] or 'active',
                'successor_id': row[8], 'event_id': row[9],
            })
            # Follow successor forward
            if row[8] and row[8] not in visited:
                current_id = row[8]
            else:
                break

        # Also walk backwards from the original tenant_id
        # to find predecessors
        current_id = tenant_id
        visited_back = {tenant_id}
        predecessors = []
        while True:
            row = conn.execute(text("""
                SELECT id, tenant_name, suite, square_feet, annual_rent,
                       lease_start, lease_end, tenant_status, successor_tenant_id
                FROM lease_tenants WHERE successor_tenant_id = :tid
            """), {'tid': current_id}).fetchone()
            if not row or row[0] in visited_back:
                break
            visited_back.add(row[0])
            predecessors.insert(0, {
                'id': row[0], 'tenant_name': row[1], 'suite': row[2],
                'square_feet': row[3], 'annual_rent': row[4],
                'lease_start': row[5], 'lease_end': row[6],
                'tenant_status': row[7] or 'active',
                'successor_id': row[8],
            })
            current_id = row[0]

    # Combine: predecessors + chain (avoiding duplicates)
    seen = set()
    full_chain = []
    for item in predecessors + chain:
        if item['id'] not in seen:
            seen.add(item['id'])
            full_chain.append(item)

    return full_chain


# ---------------------------------------------------------------------------
# Phase 5: Leasing Assumptions & Projected Cash Flow
# ---------------------------------------------------------------------------

def save_market_assumptions(
    engine, review_id: int, assumptions_list: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """Upsert market assumptions by (review_id, lease_type)."""
    from sqlalchemy import text
    saved = 0
    with engine.begin() as conn:
        for a in assumptions_list:
            lt = a.get('lease_type', '')
            if not lt:
                continue
            params = {
                'rid': review_id, 'lt': lt,
                'mrpsf': a.get('market_rent_psf'),
                'arg': a.get('annual_rent_growth'),
                'rp': a.get('renewal_probability'),
                'rdm': a.get('renewal_downtime_months', 0),
                'rtipsf': a.get('renewal_ti_psf'),
                'rlcpct': a.get('renewal_lc_pct'),
                'rrs': a.get('renewal_rent_spread'),
                'rty': a.get('renewal_term_years', 5),
                'ndm': a.get('new_downtime_months', 6),
                'ntipsf': a.get('new_ti_psf'),
                'nlcpct': a.get('new_lc_pct'),
                'nrs': a.get('new_rent_spread'),
                'nty': a.get('new_term_years', 10),
                'frm': a.get('free_rent_months', 0),
                'aeg': a.get('annual_expense_growth'),
                'cb': a.get('created_by', ''),
            }
            if engine.dialect.name == 'postgresql':
                conn.execute(text("""
                    INSERT INTO lease_market_assumptions
                        (review_id, lease_type, market_rent_psf, annual_rent_growth,
                         renewal_probability, renewal_downtime_months, renewal_ti_psf,
                         renewal_lc_pct, renewal_rent_spread, renewal_term_years,
                         new_downtime_months, new_ti_psf, new_lc_pct,
                         new_rent_spread, new_term_years, free_rent_months,
                         annual_expense_growth, created_by)
                    VALUES (:rid, :lt, :mrpsf, :arg, :rp, :rdm, :rtipsf, :rlcpct,
                            :rrs, :rty, :ndm, :ntipsf, :nlcpct, :nrs, :nty,
                            :frm, :aeg, :cb)
                    ON CONFLICT (review_id, lease_type) DO UPDATE SET
                        market_rent_psf = EXCLUDED.market_rent_psf,
                        annual_rent_growth = EXCLUDED.annual_rent_growth,
                        renewal_probability = EXCLUDED.renewal_probability,
                        renewal_downtime_months = EXCLUDED.renewal_downtime_months,
                        renewal_ti_psf = EXCLUDED.renewal_ti_psf,
                        renewal_lc_pct = EXCLUDED.renewal_lc_pct,
                        renewal_rent_spread = EXCLUDED.renewal_rent_spread,
                        renewal_term_years = EXCLUDED.renewal_term_years,
                        new_downtime_months = EXCLUDED.new_downtime_months,
                        new_ti_psf = EXCLUDED.new_ti_psf,
                        new_lc_pct = EXCLUDED.new_lc_pct,
                        new_rent_spread = EXCLUDED.new_rent_spread,
                        new_term_years = EXCLUDED.new_term_years,
                        free_rent_months = EXCLUDED.free_rent_months,
                        annual_expense_growth = EXCLUDED.annual_expense_growth,
                        updated_at = CURRENT_TIMESTAMP
                """), params)
            else:
                conn.execute(text("""
                    INSERT OR REPLACE INTO lease_market_assumptions
                        (review_id, lease_type, market_rent_psf, annual_rent_growth,
                         renewal_probability, renewal_downtime_months, renewal_ti_psf,
                         renewal_lc_pct, renewal_rent_spread, renewal_term_years,
                         new_downtime_months, new_ti_psf, new_lc_pct,
                         new_rent_spread, new_term_years, free_rent_months,
                         annual_expense_growth, created_by)
                    VALUES (:rid, :lt, :mrpsf, :arg, :rp, :rdm, :rtipsf, :rlcpct,
                            :rrs, :rty, :ndm, :ntipsf, :nlcpct, :nrs, :nty,
                            :frm, :aeg, :cb)
                """), params)
            saved += 1

    return {'saved': saved}


def get_market_assumptions(engine, review_id: int) -> List[Dict[str, Any]]:
    """Get all market assumption sets for a review."""
    from sqlalchemy import text
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT id, lease_type, market_rent_psf, annual_rent_growth,
                   renewal_probability, renewal_downtime_months, renewal_ti_psf,
                   renewal_lc_pct, renewal_rent_spread, renewal_term_years,
                   new_downtime_months, new_ti_psf, new_lc_pct,
                   new_rent_spread, new_term_years, free_rent_months,
                   annual_expense_growth
            FROM lease_market_assumptions
            WHERE review_id = :rid
            ORDER BY lease_type
        """), {'rid': review_id}).fetchall()

    return [{
        'id': r[0], 'lease_type': r[1], 'market_rent_psf': r[2],
        'annual_rent_growth': r[3], 'renewal_probability': r[4],
        'renewal_downtime_months': r[5], 'renewal_ti_psf': r[6],
        'renewal_lc_pct': r[7], 'renewal_rent_spread': r[8],
        'renewal_term_years': r[9], 'new_downtime_months': r[10],
        'new_ti_psf': r[11], 'new_lc_pct': r[12],
        'new_rent_spread': r[13], 'new_term_years': r[14],
        'free_rent_months': r[15], 'annual_expense_growth': r[16],
    } for r in rows]


def generate_projected_cash_flow(
    engine, review_id: int, start_date: str, end_date: str,
) -> Dict[str, Any]:
    """Generate Argus-style projected cash flow for all suites.

    Walks each suite through time: in-place lease -> expiry ->
    renewal/new tenant decision -> next term. Returns monthly projections
    for three scenarios: renewal, new_tenant, and probability_weighted.
    """
    from sqlalchemy import text
    from dateutil.relativedelta import relativedelta

    resolved = get_resolved_tenants(engine, review_id)
    active = [t for t in resolved
              if t.get('tenant_status', 'active') in ('active', None)]
    assumptions = {a['lease_type']: a
                   for a in get_market_assumptions(engine, review_id)}

    # Get rent steps keyed by tenant_id
    with engine.connect() as conn:
        step_rows = conn.execute(text("""
            SELECT tenant_id, effective_date, monthly_rent, annual_rent, rent_per_sf
            FROM lease_rent_steps
            WHERE tenant_id IN (
                SELECT id FROM lease_tenants WHERE review_id = :rid
            )
            ORDER BY tenant_id, effective_date
        """), {'rid': review_id}).fetchall()

    rent_steps = {}
    for r in step_rows:
        rent_steps.setdefault(r[0], []).append({
            'effective_date': r[1], 'monthly_rent': r[2],
            'annual_rent': r[3], 'rent_per_sf': r[4],
        })

    start_dt = pd.to_datetime(start_date)
    end_dt = pd.to_datetime(end_date)

    # Generate month list
    months = []
    cur = start_dt.replace(day=1)
    while cur <= end_dt:
        months.append(cur)
        cur += relativedelta(months=1)

    # Default assumption for types without explicit assumptions
    default_assum = {
        'market_rent_psf': 0, 'annual_rent_growth': 0.03,
        'renewal_probability': 0.70, 'renewal_downtime_months': 0,
        'renewal_ti_psf': 5, 'renewal_lc_pct': 0.04,
        'renewal_rent_spread': 0, 'renewal_term_years': 5,
        'new_downtime_months': 6, 'new_ti_psf': 15,
        'new_lc_pct': 0.06, 'new_rent_spread': 0,
        'new_term_years': 10, 'free_rent_months': 0,
    }

    suite_projections = {}

    for tenant in active:
        tid = tenant['id']
        suite = tenant.get('suite', f'T{tid}')
        sf = tenant.get('square_feet', 0) or 0
        lt = tenant.get('lease_type', '') or ''
        assum = assumptions.get(lt, default_assum)

        # Get market rent — default to current rent/SF if not specified
        base_market_rent = assum.get('market_rent_psf') or 0
        if base_market_rent == 0 and sf > 0:
            base_market_rent = (tenant.get('annual_rent', 0) or 0) / sf

        growth = assum.get('annual_rent_growth', 0.03) or 0.03
        renewal_prob = assum.get('renewal_probability', 0.70) or 0.70

        lease_end_str = tenant.get('lease_end')
        lease_start_str = tenant.get('lease_start')
        current_monthly = tenant.get('monthly_rent', 0) or 0

        try:
            lease_end_dt = pd.to_datetime(lease_end_str) if lease_end_str else None
        except Exception:
            lease_end_dt = None
        try:
            lease_start_dt = pd.to_datetime(lease_start_str) if lease_start_str else start_dt
        except Exception:
            lease_start_dt = start_dt

        # Get rent steps for escalation during in-place period
        steps = rent_steps.get(tid, [])
        step_schedule = []
        for s in steps:
            try:
                sd = pd.to_datetime(s['effective_date'])
                step_schedule.append((sd, s.get('monthly_rent') or 0))
            except Exception:
                pass
        step_schedule.sort()

        def get_contracted_rent(dt):
            """Get contracted monthly rent at a given date from rent steps."""
            rent = current_monthly
            for sd, mr in step_schedule:
                if dt >= sd:
                    rent = mr
                else:
                    break
            return rent

        def escalated_market_rent(target_dt):
            """Escalate market rent to target date."""
            years = max(0, (target_dt - start_dt).days / 365.25)
            return base_market_rent * (1 + growth) ** years

        # Build monthly schedule for three scenarios
        renewal_months = []
        new_tenant_months = []
        weighted_months = []

        for m in months:
            m_end = m + relativedelta(months=1) - relativedelta(days=1)

            # During in-place period (before lease end)
            if lease_end_dt is None or m <= lease_end_dt:
                contracted = get_contracted_rent(m)
                entry = {
                    'month': m.strftime('%Y-%m'),
                    'suite': suite, 'tenant': tenant['tenant_name'],
                    'sf': sf, 'base_rent': contracted,
                    'effective_rent': contracted,
                    'vacancy_loss': 0, 'ti_cost': 0, 'lc_cost': 0,
                    'net_effective_rent': contracted,
                    'phase': 'in_place',
                }
                renewal_months.append(entry.copy())
                new_tenant_months.append(entry.copy())
                weighted_months.append(entry.copy())
                continue

            # Post-expiry: calculate months since lease end
            months_since_expiry = (
                (m.year - lease_end_dt.year) * 12 +
                (m.month - lease_end_dt.month)
            )

            # --- Renewal scenario ---
            r_downtime = assum.get('renewal_downtime_months', 0) or 0
            r_term = (assum.get('renewal_term_years', 5) or 5) * 12
            r_spread = assum.get('renewal_rent_spread', 0) or 0

            cycle_month = months_since_expiry % (r_downtime + r_term)
            if cycle_month < r_downtime:
                # Downtime
                renewal_months.append({
                    'month': m.strftime('%Y-%m'), 'suite': suite,
                    'tenant': 'Vacant (renewal turnover)', 'sf': sf,
                    'base_rent': 0, 'effective_rent': 0,
                    'vacancy_loss': escalated_market_rent(m) * sf / 12,
                    'ti_cost': 0, 'lc_cost': 0, 'net_effective_rent': 0,
                    'phase': 'vacancy',
                })
            else:
                mkt = escalated_market_rent(m)
                renewal_rent = (mkt + r_spread) * sf / 12
                ti = lc = 0
                if cycle_month == r_downtime:  # First month of new lease
                    ti = (assum.get('renewal_ti_psf', 0) or 0) * sf
                    lease_val = renewal_rent * 12 * (assum.get('renewal_term_years', 5) or 5)
                    lc = (assum.get('renewal_lc_pct', 0) or 0) * lease_val
                renewal_months.append({
                    'month': m.strftime('%Y-%m'), 'suite': suite,
                    'tenant': tenant['tenant_name'] + ' (renewed)', 'sf': sf,
                    'base_rent': renewal_rent, 'effective_rent': renewal_rent,
                    'vacancy_loss': 0, 'ti_cost': ti, 'lc_cost': lc,
                    'net_effective_rent': renewal_rent - ti - lc,
                    'phase': 'renewal',
                })

            # --- New tenant scenario ---
            n_downtime = assum.get('new_downtime_months', 6) or 6
            n_free = assum.get('free_rent_months', 0) or 0
            n_term = (assum.get('new_term_years', 10) or 10) * 12
            n_spread = assum.get('new_rent_spread', 0) or 0

            n_cycle = months_since_expiry % (n_downtime + n_free + n_term)
            if n_cycle < n_downtime:
                new_tenant_months.append({
                    'month': m.strftime('%Y-%m'), 'suite': suite,
                    'tenant': 'Vacant (new tenant turnover)', 'sf': sf,
                    'base_rent': 0, 'effective_rent': 0,
                    'vacancy_loss': escalated_market_rent(m) * sf / 12,
                    'ti_cost': 0, 'lc_cost': 0, 'net_effective_rent': 0,
                    'phase': 'vacancy',
                })
            elif n_cycle < n_downtime + n_free:
                # Free rent period
                mkt = escalated_market_rent(m)
                base = (mkt + n_spread) * sf / 12
                new_tenant_months.append({
                    'month': m.strftime('%Y-%m'), 'suite': suite,
                    'tenant': 'New Tenant (free rent)', 'sf': sf,
                    'base_rent': base, 'effective_rent': 0,
                    'vacancy_loss': base, 'ti_cost': 0, 'lc_cost': 0,
                    'net_effective_rent': 0, 'phase': 'free_rent',
                })
            else:
                mkt = escalated_market_rent(m)
                new_rent = (mkt + n_spread) * sf / 12
                ti = lc = 0
                if n_cycle == n_downtime + n_free:
                    ti = (assum.get('new_ti_psf', 0) or 0) * sf
                    lease_val = new_rent * 12 * (assum.get('new_term_years', 10) or 10)
                    lc = (assum.get('new_lc_pct', 0) or 0) * lease_val
                new_tenant_months.append({
                    'month': m.strftime('%Y-%m'), 'suite': suite,
                    'tenant': 'New Tenant', 'sf': sf,
                    'base_rent': new_rent, 'effective_rent': new_rent,
                    'vacancy_loss': 0, 'ti_cost': ti, 'lc_cost': lc,
                    'net_effective_rent': new_rent - ti - lc,
                    'phase': 'new_tenant',
                })

            # --- Weighted scenario ---
            r_entry = renewal_months[-1]
            n_entry = new_tenant_months[-1]
            w_eff = r_entry['effective_rent'] * renewal_prob + \
                    n_entry['effective_rent'] * (1 - renewal_prob)
            w_vac = r_entry['vacancy_loss'] * renewal_prob + \
                    n_entry['vacancy_loss'] * (1 - renewal_prob)
            w_ti = r_entry['ti_cost'] * renewal_prob + \
                   n_entry['ti_cost'] * (1 - renewal_prob)
            w_lc = r_entry['lc_cost'] * renewal_prob + \
                   n_entry['lc_cost'] * (1 - renewal_prob)
            weighted_months.append({
                'month': m.strftime('%Y-%m'), 'suite': suite,
                'tenant': r_entry['tenant'] if r_entry['effective_rent'] > 0
                          else n_entry['tenant'],
                'sf': sf,
                'base_rent': r_entry['base_rent'] * renewal_prob +
                             n_entry['base_rent'] * (1 - renewal_prob),
                'effective_rent': w_eff,
                'vacancy_loss': w_vac, 'ti_cost': w_ti, 'lc_cost': w_lc,
                'net_effective_rent': w_eff - w_ti - w_lc,
                'phase': r_entry['phase'] if renewal_prob >= 0.5
                         else n_entry['phase'],
            })

        suite_projections[suite] = {
            'tenant_id': tid, 'tenant_name': tenant['tenant_name'],
            'suite': suite, 'sf': sf, 'lease_type': lt,
            'lease_end': lease_end_str,
            'renewal': renewal_months,
            'new_tenant': new_tenant_months,
            'weighted': weighted_months,
        }

    return {
        'start_date': start_date,
        'end_date': end_date,
        'suites': suite_projections,
        'months': [m.strftime('%Y-%m') for m in months],
    }


def summarize_projected_revenue(
    engine, review_id: int, start_date: str, end_date: str,
) -> Dict[str, Any]:
    """Aggregate projected cash flow into annual totals for each scenario."""
    projection = generate_projected_cash_flow(engine, review_id, start_date, end_date)

    summaries = {}
    for scenario in ('renewal', 'new_tenant', 'weighted'):
        annual = {}
        total_sf = 0
        for suite_key, suite_data in projection['suites'].items():
            total_sf += suite_data.get('sf', 0)
            for entry in suite_data.get(scenario, []):
                year = entry['month'][:4]
                if year not in annual:
                    annual[year] = {
                        'year': int(year), 'gross_rent': 0,
                        'effective_rent': 0, 'vacancy_loss': 0,
                        'ti_costs': 0, 'lc_costs': 0, 'net_effective': 0,
                        'occupied_months': 0, 'total_months': 0,
                    }
                a = annual[year]
                a['gross_rent'] += entry.get('base_rent', 0)
                a['effective_rent'] += entry.get('effective_rent', 0)
                a['vacancy_loss'] += entry.get('vacancy_loss', 0)
                a['ti_costs'] += entry.get('ti_cost', 0)
                a['lc_costs'] += entry.get('lc_cost', 0)
                a['net_effective'] += entry.get('net_effective_rent', 0)
                a['total_months'] += 1
                if entry.get('effective_rent', 0) > 0:
                    a['occupied_months'] += 1

        # Compute rates
        for year_data in annual.values():
            tm = year_data['total_months']
            year_data['vacancy_rate'] = (
                1 - year_data['occupied_months'] / tm if tm > 0 else 0
            )
            year_data['avg_effective_rent_psf'] = (
                year_data['effective_rent'] / total_sf
                if total_sf > 0 else 0
            )

        summaries[scenario] = sorted(annual.values(), key=lambda x: x['year'])

    return {
        'start_date': start_date,
        'end_date': end_date,
        'total_sf': total_sf,
        'summaries': summaries,
    }
