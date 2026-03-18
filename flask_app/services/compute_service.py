"""Compute service — wraps compute.py without Streamlit dependency.

Replaces get_cached_deal_result() which uses st.session_state + st.toast.
Also contains ROE/MOIC audit builders extracted from app.py.
"""

import copy
from io import BytesIO
from typing import Optional

import pandas as pd
import numpy as np

from compute import compute_deal_analysis, build_partner_results
from loaders import load_waterfalls, load_mri_loans

# Module-level deal result cache (replaces st.session_state['_deal_results'])
_deal_cache: dict = {}


def get_cached_deal_result(
    vcode: str,
    start_year: int,
    horizon_years: int,
    pro_yr_base: int,
    data: dict,
    force: bool = False,
    actuals_through=None,
) -> dict:
    """Compute or retrieve cached deal result.

    Args:
        vcode: Deal vcode.
        start_year: Forecast start year.
        horizon_years: Number of years to model.
        pro_yr_base: Pro-forma base year.
        data: Dict from data_service.load_all().
        force: If True, bypass cache.
        actuals_through: Date cutoff for actuals (None = full forecast).

    Returns:
        Full result dict from compute_deal_analysis().
    """
    at_str = str(actuals_through) if actuals_through else "none"
    cache_key = f"{vcode}|{start_year}|{horizon_years}|{pro_yr_base}|{at_str}"
    if not force and cache_key in _deal_cache:
        return _deal_cache[cache_key]

    inv = data["inv"]
    wf = data["wf"]
    acct = data["acct"]
    fc = data["fc"]
    coa = data["coa"]
    mri_loans_raw = data["mri_loans_raw"]
    mri_supp = data["mri_supp"]
    mri_val = data["mri_val"]
    relationships_raw = data["relationships_raw"]
    capital_calls_raw = data["capital_calls_raw"]
    isbs_raw = data["isbs_raw"]
    prospective_loans_raw = data.get("prospective_loans_raw")

    # Resolve InvestmentID for this vcode
    deal_row = inv[inv["vcode"] == vcode]
    if deal_row.empty:
        raise ValueError(f"Deal not found: {vcode}")
    deal_investment_id = deal_row.iloc[0].get("InvestmentID", vcode)

    # Sale date
    sale_date_raw = deal_row.iloc[0].get("Sale_Date", None)

    result = compute_deal_analysis(
        deal_vcode=vcode,
        deal_investment_id=deal_investment_id,
        sale_date_raw=sale_date_raw,
        inv=inv,
        wf=wf,
        acct=acct,
        fc=fc,
        coa=coa,
        mri_loans_raw=mri_loans_raw,
        mri_supp=mri_supp,
        mri_val=mri_val,
        relationships_raw=relationships_raw,
        capital_calls_raw=capital_calls_raw,
        isbs_raw=isbs_raw,
        start_year=start_year,
        horizon_years=horizon_years,
        pro_yr_base=pro_yr_base,
        actuals_through=actuals_through,
        prospective_loans_raw=prospective_loans_raw,
    )

    _deal_cache[cache_key] = result
    return result


def clear_cache(vcode: Optional[str] = None):
    """Clear deal computation cache.

    Args:
        vcode: If provided, clear only entries for this deal. Otherwise clear all.
    """
    if vcode:
        keys_to_remove = [k for k in _deal_cache if k.startswith(f"{vcode}|")]
        for k in keys_to_remove:
            del _deal_cache[k]
    else:
        _deal_cache.clear()


def get_all_cached_vcodes() -> list[str]:
    """Return list of vcodes currently cached."""
    return list({k.split("|")[0] for k in _deal_cache})


# ============================================================
# ROE / MOIC Audit Builders  (extracted from app.py)
# ============================================================

def build_roe_timeline(combined_cashflows, cf_only_distributions, end_date):
    """Replay calculate_roe logic to produce an audit trail.

    Returns (timeline_rows, cf_dist_rows, summary_dict) — all JSON-safe.
    """
    cf_by_date = {}
    for d, a in cf_only_distributions:
        if a > 0:
            cf_by_date[d] = cf_by_date.get(d, 0.0) + a

    events = []
    for d, amt in combined_cashflows:
        if amt < 0:
            events.append((d, -amt))
        elif amt > 0:
            cf_at_date = cf_by_date.get(d, 0.0)
            if cf_at_date > 0:
                consumed = min(cf_at_date, amt)
                cf_by_date[d] -= consumed
                cap_return = amt - consumed
            else:
                cap_return = amt
            if cap_return > 0.005:
                events.append((d, -cap_return))
    events = sorted(events, key=lambda x: x[0])

    if not events:
        return [], [], {}

    inception = events[0][0]
    rows = []
    current_balance = 0.0
    prev_date = inception
    total_weighted = 0.0

    for evt_date, change in events:
        days = (evt_date - prev_date).days
        weighted = current_balance * days
        if days > 0 and current_balance > 0:
            rows.append({
                'Date': prev_date.isoformat(), 'Event': '(holding period)',
                'Amount': None, 'Capital Balance': current_balance,
                'Days at Balance': days, 'Weighted Capital': weighted,
            })
        total_weighted += weighted
        current_balance = max(0, current_balance + change)
        orig_amt = -change
        if orig_amt < 0:
            event_label = 'Contribution'
        elif orig_amt > 0:
            event_label = 'Capital Return'
        else:
            event_label = 'Event'
        rows.append({
            'Date': evt_date.isoformat(), 'Event': event_label,
            'Amount': orig_amt, 'Capital Balance': current_balance,
            'Days at Balance': 0, 'Weighted Capital': 0.0,
        })
        prev_date = evt_date

    if end_date and prev_date < end_date:
        days = (end_date - prev_date).days
        weighted = current_balance * days
        if days > 0:
            rows.append({
                'Date': prev_date.isoformat(), 'Event': '(holding period)',
                'Amount': None, 'Capital Balance': current_balance,
                'Days at Balance': days, 'Weighted Capital': weighted,
            })
            total_weighted += weighted

    total_days = (end_date - inception).days if end_date else 0
    wac = total_weighted / total_days if total_days > 0 else 0.0
    years = total_days / 365.0 if total_days > 0 else 0.0

    cf_rows = [{'Date': d.isoformat(), 'Amount': a} for d, a in cf_only_distributions if a > 0]
    total_cf_dist = sum(r['Amount'] for r in cf_rows)
    roe = (total_cf_dist / wac) / years if wac > 0 and years > 0 else 0.0

    summary = {
        'inception': inception.isoformat(), 'end': end_date.isoformat() if end_date else None,
        'total_days': total_days, 'years': years,
        'total_cf_dist': total_cf_dist, 'wac': wac, 'roe': roe,
    }
    return rows, cf_rows, summary


def build_roe_audit(partner_results, deal_summary, sale_me):
    """Build full ROE audit data for all partners + deal level.

    Returns dict with 'partners' list and 'deal_level' dict.
    """
    partners = []
    for pr in partner_results:
        tl, cf, summary = build_roe_timeline(
            pr['combined_cashflows'], pr['cf_only_distributions'], sale_me)
        partners.append({
            'partner': pr['partner'],
            'timeline': tl,
            'cf_distributions': cf,
            'summary': summary,
        })

    # Deal-level
    all_cfs = deal_summary.get('all_combined_cashflows', [])
    all_cf_dist = []
    for pr in partner_results:
        all_cf_dist.extend(pr['cf_only_distributions'])
    tl, cf, summary = build_roe_timeline(all_cfs, all_cf_dist, sale_me)

    return {
        'partners': partners,
        'deal_level': {'timeline': tl, 'cf_distributions': cf, 'summary': summary},
    }


def build_moic_breakdown(cashflow_details, pr_dict):
    """Build MOIC audit breakdown from cashflow_details.

    Returns (breakdown_rows, summary_dict) — JSON-safe.
    """
    rows = []
    for cf in cashflow_details:
        amt = cf['Amount']
        desc = cf.get('Description', '')
        if amt < 0:
            cf_type = 'Contribution'
        elif 'Unrealized' in desc or 'NAV' in desc:
            cf_type = 'Terminal Value'
        elif 'Capital' in desc or 'Cap ' in desc or 'Refi' in desc or 'Sale' in desc:
            cf_type = 'Cap Distribution'
        else:
            cf_type = 'CF Distribution'
        rows.append({
            'Date': cf['Date'].isoformat() if hasattr(cf['Date'], 'isoformat') else str(cf['Date']),
            'Description': desc, 'Type': cf_type, 'Amount': amt,
        })

    summary = {
        'contributions': pr_dict.get('contributions', 0.0),
        'cf_dist': pr_dict.get('cf_distributions', 0.0),
        'cap_dist': pr_dict.get('cap_distributions', 0.0),
        'total_dist': pr_dict.get('total_distributions', 0.0),
        'unrealized': pr_dict.get('unrealized_nav', 0.0),
        'moic': pr_dict.get('moic', 0.0),
    }
    return rows, summary


def build_moic_audit(partner_results, deal_summary, sale_me):
    """Build full MOIC audit data for all partners + deal level.

    Returns dict with 'partners' list and 'deal_level' dict.
    """
    partners = []
    for pr in partner_results:
        breakdown, summary = build_moic_breakdown(pr.get('cashflow_details', []), pr)
        partners.append({
            'partner': pr['partner'],
            'breakdown': breakdown,
            'summary': summary,
        })

    return {
        'partners': partners,
        'deal_level': {
            'total_contributions': deal_summary.get('total_contributions', 0),
            'total_cf_distributions': deal_summary.get('total_cf_distributions', 0),
            'total_cap_distributions': deal_summary.get('total_cap_distributions', 0),
            'total_distributions': deal_summary.get('total_distributions', 0),
            'deal_moic': deal_summary.get('deal_moic', 0),
            'note': 'Deal MOIC uses realized distributions only (no unrealized NAV).',
        },
    }


# ============================================================
# Excel generators  (extracted from app.py)
# ============================================================

def generate_roe_audit_excel(partner_results, deal_summary, sale_me) -> bytes:
    """Generate formatted Excel workbook for ROE audit."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "ROE Audit"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    bold_font = Font(bold=True)
    top_border = Border(top=Side(style='medium'))
    curr_fmt = '$#,##0'
    pct_fmt = '0.00%'
    date_fmt = 'MM/DD/YYYY'
    num_fmt = '#,##0'

    row = 1

    def write_header(ws, r, cols):
        for ci, name in enumerate(cols, 1):
            c = ws.cell(row=r, column=ci, value=name)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        return r + 1

    def _write_roe_section(ws, row, tl_rows, cf_rows, summary, label):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=12)
        row += 1

        ws.cell(row=row, column=1, value="Capital Balance Timeline").font = bold_font
        row += 1
        tl_cols = ['Date', 'Event', 'Amount', 'Capital Balance', 'Days at Balance', 'Weighted Capital']
        row = write_header(ws, row, tl_cols)
        for tr in tl_rows:
            ws.cell(row=row, column=1, value=tr['Date']).number_format = date_fmt
            ws.cell(row=row, column=2, value=tr['Event'])
            if tr['Amount'] is not None:
                ws.cell(row=row, column=3, value=tr['Amount']).number_format = curr_fmt
            ws.cell(row=row, column=4, value=tr['Capital Balance']).number_format = curr_fmt
            ws.cell(row=row, column=5, value=tr['Days at Balance']).number_format = num_fmt
            ws.cell(row=row, column=6, value=tr['Weighted Capital']).number_format = curr_fmt
            row += 1
        ws.cell(row=row, column=1, value="Totals").font = bold_font
        ws.cell(row=row, column=1).border = top_border
        ws.cell(row=row, column=5, value=summary.get('total_days', 0)).font = bold_font
        ws.cell(row=row, column=5).number_format = num_fmt
        ws.cell(row=row, column=5).border = top_border
        ws.cell(row=row, column=6, value=summary.get('wac', 0)).font = bold_font
        ws.cell(row=row, column=6).number_format = curr_fmt
        ws.cell(row=row, column=6).border = top_border
        row += 2

        ws.cell(row=row, column=1, value="CF Distributions").font = bold_font
        row += 1
        row = write_header(ws, row, ['Date', 'Amount'])
        for cr in cf_rows:
            ws.cell(row=row, column=1, value=cr['Date']).number_format = date_fmt
            ws.cell(row=row, column=2, value=cr['Amount']).number_format = curr_fmt
            row += 1
        ws.cell(row=row, column=1, value="Total").font = bold_font
        ws.cell(row=row, column=1).border = top_border
        ws.cell(row=row, column=2, value=summary.get('total_cf_dist', 0)).font = bold_font
        ws.cell(row=row, column=2).number_format = curr_fmt
        ws.cell(row=row, column=2).border = top_border
        row += 2

        ws.cell(row=row, column=1, value="ROE Summary").font = bold_font
        row += 1
        row = write_header(ws, row, ['Inception', 'End Date', 'Total Days', 'Years',
                                      'CF Distributions', 'Wtd Avg Capital', 'ROE'])
        ws.cell(row=row, column=1, value=summary.get('inception')).number_format = date_fmt
        ws.cell(row=row, column=2, value=summary.get('end')).number_format = date_fmt
        ws.cell(row=row, column=3, value=summary.get('total_days', 0)).number_format = num_fmt
        ws.cell(row=row, column=4, value=round(summary.get('years', 0), 2))
        ws.cell(row=row, column=5, value=summary.get('total_cf_dist', 0)).number_format = curr_fmt
        ws.cell(row=row, column=6, value=summary.get('wac', 0)).number_format = curr_fmt
        ws.cell(row=row, column=7, value=summary.get('roe', 0)).number_format = pct_fmt
        row += 2
        return row

    # Per-partner sections
    for pr in partner_results:
        tl, cf, summary = build_roe_timeline(
            pr['combined_cashflows'], pr['cf_only_distributions'], sale_me)
        row = _write_roe_section(ws, row, tl, cf, summary, f"Partner: {pr['partner']}")

    # Deal-level
    all_cfs = deal_summary.get('all_combined_cashflows', [])
    all_cf_dist = []
    for pr in partner_results:
        all_cf_dist.extend(pr['cf_only_distributions'])
    tl, cf, summary = build_roe_timeline(all_cfs, all_cf_dist, sale_me)
    _write_roe_section(ws, row, tl, cf, summary, "Deal-Level ROE")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_moic_audit_excel(partner_results, deal_summary, sale_me) -> bytes:
    """Generate formatted Excel workbook for MOIC audit."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "MOIC Audit"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    bold_font = Font(bold=True)
    top_border = Border(top=Side(style='medium'))
    curr_fmt = '$#,##0'
    mult_fmt = '0.00"x"'
    date_fmt = 'MM/DD/YYYY'

    row = 1

    def write_header(ws, r, cols):
        for ci, name in enumerate(cols, 1):
            c = ws.cell(row=r, column=ci, value=name)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        return r + 1

    for pr in partner_results:
        breakdown, summary = build_moic_breakdown(pr.get('cashflow_details', []), pr)
        ws.cell(row=row, column=1, value=f"Partner: {pr['partner']}").font = Font(bold=True, size=12)
        row += 1
        row = write_header(ws, row, ['Date', 'Description', 'Type', 'Amount'])
        for br in breakdown:
            ws.cell(row=row, column=1, value=br['Date']).number_format = date_fmt
            ws.cell(row=row, column=2, value=br['Description'])
            ws.cell(row=row, column=3, value=br['Type'])
            ws.cell(row=row, column=4, value=br['Amount']).number_format = curr_fmt
            row += 1
        row += 1

        ws.cell(row=row, column=1, value="MOIC Summary").font = bold_font
        row += 1
        row = write_header(ws, row, ['Contributions', 'CF Distributions', 'Cap Distributions',
                                      'Total Distributions', 'Unrealized NAV', 'MOIC'])
        ws.cell(row=row, column=1, value=summary['contributions']).number_format = curr_fmt
        ws.cell(row=row, column=2, value=summary['cf_dist']).number_format = curr_fmt
        ws.cell(row=row, column=3, value=summary['cap_dist']).number_format = curr_fmt
        ws.cell(row=row, column=4, value=summary['total_dist']).number_format = curr_fmt
        ws.cell(row=row, column=5, value=summary['unrealized']).number_format = curr_fmt
        ws.cell(row=row, column=6, value=summary['moic']).number_format = mult_fmt
        for ci in range(1, 7):
            ws.cell(row=row, column=ci).font = bold_font
        row += 2

    # Deal-level
    ws.cell(row=row, column=1, value="Deal-Level MOIC").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1,
            value="Note: Deal MOIC uses realized distributions only (no unrealized NAV)."
            ).font = Font(italic=True)
    row += 1
    ds = deal_summary
    row = write_header(ws, row, ['Total Contributions', 'CF Distributions', 'Cap Distributions',
                                  'Total Distributions', 'MOIC'])
    ws.cell(row=row, column=1, value=ds.get('total_contributions', 0)).number_format = curr_fmt
    ws.cell(row=row, column=2, value=ds.get('total_cf_distributions', 0)).number_format = curr_fmt
    ws.cell(row=row, column=3, value=ds.get('total_cap_distributions', 0)).number_format = curr_fmt
    ws.cell(row=row, column=4, value=ds.get('total_distributions', 0)).number_format = curr_fmt
    ws.cell(row=row, column=5, value=ds.get('deal_moic', 0)).number_format = mult_fmt
    for ci in range(1, 6):
        ws.cell(row=row, column=ci).font = bold_font

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Shared Excel helpers
# ============================================================

def _excel_styles():
    """Return common openpyxl styles dict."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        'header_font': Font(bold=True, color="FFFFFF"),
        'header_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'bold': Font(bold=True),
        'top_border': Border(top=Side(style='medium')),
        'curr': '$#,##0',
        'pct': '0.00%',
        'mult': '0.00"x"',
        'date': 'MM/DD/YYYY',
        'num': '#,##0',
        'dec2': '#,##0.00',
    }


def _write_header_row(ws, row, cols, styles):
    """Write a styled header row and return next row number."""
    from openpyxl.styles import Alignment
    for ci, name in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=name)
        c.font = styles['header_font']
        c.fill = styles['header_fill']
        c.alignment = Alignment(horizontal="center")
    return row + 1


def _autosize_columns(ws):
    """Auto-size all columns in worksheet."""
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


# ============================================================
# Per-section Excel generators
# ============================================================

def generate_partner_returns_excel(result: dict) -> bytes:
    """Excel for Partner Returns section."""
    from openpyxl import Workbook
    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Partner Returns"

    pr_list = result.get("partner_results", [])
    ds = result.get("deal_summary", {})

    cols = ['Partner', 'Contributions', 'CF Distributions', 'Cap Distributions',
            'Total Distributions', 'IRR', 'ROE', 'MOIC']
    row = _write_header_row(ws, 1, cols, s)

    for pr in pr_list:
        ws.cell(row=row, column=1, value=pr['partner'])
        ws.cell(row=row, column=2, value=pr.get('contributions', 0)).number_format = s['curr']
        ws.cell(row=row, column=3, value=pr.get('cf_distributions', 0)).number_format = s['curr']
        ws.cell(row=row, column=4, value=pr.get('cap_distributions', 0)).number_format = s['curr']
        ws.cell(row=row, column=5, value=pr.get('total_distributions', 0)).number_format = s['curr']
        ws.cell(row=row, column=6, value=pr.get('irr', 0)).number_format = s['pct']
        ws.cell(row=row, column=7, value=pr.get('roe', 0)).number_format = s['pct']
        ws.cell(row=row, column=8, value=pr.get('moic', 0)).number_format = s['mult']
        row += 1

    for ci in range(1, 9):
        ws.cell(row=row, column=ci).border = s['top_border']
    ws.cell(row=row, column=1, value="Deal Total").font = s['bold']
    ws.cell(row=row, column=2, value=ds.get('total_contributions', 0)).number_format = s['curr']
    ws.cell(row=row, column=3, value=ds.get('total_cf_distributions', 0)).number_format = s['curr']
    ws.cell(row=row, column=4, value=ds.get('total_cap_distributions', 0)).number_format = s['curr']
    ws.cell(row=row, column=5, value=ds.get('total_distributions', 0)).number_format = s['curr']
    ws.cell(row=row, column=6, value=ds.get('deal_irr', 0)).number_format = s['pct']
    ws.cell(row=row, column=7, value=ds.get('deal_roe', 0)).number_format = s['pct']
    ws.cell(row=row, column=8, value=ds.get('deal_moic', 0)).number_format = s['mult']
    for ci in range(1, 9):
        ws.cell(row=row, column=ci).font = s['bold']

    _autosize_columns(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_forecast_excel(result: dict, start_year: int, horizon_years: int) -> bytes:
    """Excel for Annual Forecast section."""
    from openpyxl import Workbook
    from reporting import annual_aggregation_table
    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Annual Forecast"

    fc_display = result.get("fc_deal_display")
    if fc_display is None or fc_display.empty:
        ws.cell(row=1, column=1, value="No forecast data available")
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    cap_events = result.get("cap_events_df")
    proceeds_by_year = None
    if cap_events is not None and not cap_events.empty and "Year" in cap_events.columns:
        proceeds_by_year = cap_events.groupby("Year")["amount"].sum()

    table = annual_aggregation_table(
        fc_display, start_year, horizon_years,
        proceeds_by_year=proceeds_by_year,
        cf_alloc=result.get("cf_alloc"),
        cap_alloc=result.get("cap_alloc"),
    )

    if "Year" not in table.columns:
        ws.cell(row=1, column=1, value="No forecast data available")
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    wide = table.set_index("Year").T
    years = [int(y) for y in wide.columns]

    fcols = ['Line Item'] + [str(y) for y in years]
    frow = _write_header_row(ws, 1, fcols, s)

    for label, row_data in wide.iterrows():
        label_str = str(label)
        ws.cell(row=frow, column=1, value=label_str).font = s['bold'] if label_str.strip() in {
            'Revenue', 'Expenses', 'NOI', 'Capital Expenditures', 'FAD',
            'Total Distributions', 'DSCR'
        } else None
        for ci, y in enumerate(years, 2):
            val = row_data.get(y)
            if val is not None and not (isinstance(val, float) and val != val):
                cell = ws.cell(row=frow, column=ci, value=float(val))
                if 'DSCR' in label_str:
                    cell.number_format = s['dec2']
                else:
                    cell.number_format = s['curr']
        frow += 1

    _autosize_columns(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_debt_service_excel(result: dict) -> bytes:
    """Excel for Debt Service section."""
    from openpyxl import Workbook
    s = _excel_styles()
    wb = Workbook()

    ws = wb.active
    ws.title = "Loan Summary"
    loans = result.get("loans", [])
    cols = ['Loan ID', 'Original Amount', 'Origination Date', 'Maturity Date',
            'Type', 'Rate', 'Loan Term (mo)', 'Amort Term (mo)', 'IO Months']
    row = _write_header_row(ws, 1, cols, s)
    for l in (loans or []):
        ws.cell(row=row, column=1, value=l.loan_id)
        ws.cell(row=row, column=2, value=l.orig_amount).number_format = s['curr']
        ws.cell(row=row, column=3, value=l.orig_date.isoformat() if l.orig_date else "")
        ws.cell(row=row, column=4, value=l.maturity_date.isoformat() if l.maturity_date else "")
        ws.cell(row=row, column=5, value=l.int_type or "")
        ws.cell(row=row, column=6, value=l.fixed_rate or 0).number_format = s['pct']
        ws.cell(row=row, column=7, value=l.loan_term_m or 0)
        ws.cell(row=row, column=8, value=l.amort_term_m or 0)
        ws.cell(row=row, column=9, value=l.io_months or 0)
        row += 1
    _autosize_columns(ws)

    loan_sched = result.get("loan_sched")
    if loan_sched is not None and not loan_sched.empty:
        ws2 = wb.create_sheet("Amortization Schedule")
        sched_cols = list(loan_sched.columns)
        row = _write_header_row(ws2, 1, sched_cols, s)
        for _, r in loan_sched.iterrows():
            for ci, col in enumerate(sched_cols, 1):
                val = r[col]
                cell = ws2.cell(row=row, column=ci, value=val)
                if isinstance(val, (int, float)) and col != 'Year':
                    cell.number_format = s['curr']
            row += 1
        _autosize_columns(ws2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_cash_schedule_excel(result: dict) -> bytes:
    """Excel for Cash Management section."""
    from openpyxl import Workbook
    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Schedule"

    sched = result.get("cash_schedule")
    if sched is None or (hasattr(sched, 'empty') and sched.empty):
        ws.cell(row=1, column=1, value="No cash schedule data available")
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    if isinstance(sched, pd.DataFrame):
        cols = list(sched.columns)
        row = _write_header_row(ws, 1, cols, s)
        for _, r in sched.iterrows():
            for ci, col in enumerate(cols, 1):
                val = r[col]
                cell = ws.cell(row=row, column=ci, value=val)
                if isinstance(val, (int, float)):
                    cell.number_format = s['curr']
            row += 1
    _autosize_columns(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_capital_calls_excel(result: dict) -> bytes:
    """Excel for Capital Calls section."""
    from openpyxl import Workbook
    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Capital Calls"

    calls = result.get("capital_calls", [])
    if not calls:
        ws.cell(row=1, column=1, value="No capital call data available")
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    if isinstance(calls, list) and len(calls) > 0:
        cols = list(calls[0].keys()) if isinstance(calls[0], dict) else ['Data']
        row = _write_header_row(ws, 1, cols, s)
        for item in calls:
            if isinstance(item, dict):
                for ci, col in enumerate(cols, 1):
                    val = item.get(col)
                    cell = ws.cell(row=row, column=ci, value=val)
                    if isinstance(val, (int, float)):
                        cell.number_format = s['curr']
            row += 1
    _autosize_columns(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_xirr_cashflows_excel(result: dict) -> bytes:
    """Excel for XIRR Cash Flows section."""
    from openpyxl import Workbook
    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "XIRR Cash Flows"

    pr_list = result.get("partner_results", [])
    if not pr_list:
        ws.cell(row=1, column=1, value="No cashflow data available")
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    partners = [pr['partner'] for pr in pr_list]
    cols = ['Date', 'Description'] + partners + ['Deal Total']
    row = _write_header_row(ws, 1, cols, s)

    all_rows = []
    for pr in pr_list:
        for cf in pr.get('cashflow_details', []):
            d = cf['Date']
            date_str = d.isoformat() if hasattr(d, 'isoformat') else str(d)
            desc = cf.get('Description', '')
            key = (date_str, desc)
            found = False
            for ar in all_rows:
                if ar['key'] == key:
                    ar[pr['partner']] = cf['Amount']
                    found = True
                    break
            if not found:
                entry = {'key': key, 'date': date_str, 'desc': desc}
                entry[pr['partner']] = cf['Amount']
                all_rows.append(entry)

    for ar in all_rows:
        ws.cell(row=row, column=1, value=ar['date'])
        ws.cell(row=row, column=2, value=ar['desc'])
        total = 0.0
        for ci, p in enumerate(partners, 3):
            val = ar.get(p, 0.0)
            ws.cell(row=row, column=ci, value=val).number_format = s['curr']
            total += (val or 0.0)
        ws.cell(row=row, column=len(partners) + 3, value=total).number_format = s['curr']
        row += 1

    _autosize_columns(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_full_deal_excel(result: dict, start_year: int, horizon_years: int) -> bytes:
    """Generate comprehensive multi-sheet Deal Analysis workbook."""
    from openpyxl import Workbook
    s = _excel_styles()
    wb = Workbook()

    pr_list = result.get("partner_results", [])
    ds = result.get("deal_summary", {})
    sale_me = result.get("sale_me")

    # --- Sheet 1: Partner Returns ---
    ws1 = wb.active
    ws1.title = "Partner Returns"
    cols = ['Partner', 'Contributions', 'CF Distributions', 'Cap Distributions',
            'Total Distributions', 'IRR', 'ROE', 'MOIC']
    row = _write_header_row(ws1, 1, cols, s)
    for pr in pr_list:
        ws1.cell(row=row, column=1, value=pr['partner'])
        ws1.cell(row=row, column=2, value=pr.get('contributions', 0)).number_format = s['curr']
        ws1.cell(row=row, column=3, value=pr.get('cf_distributions', 0)).number_format = s['curr']
        ws1.cell(row=row, column=4, value=pr.get('cap_distributions', 0)).number_format = s['curr']
        ws1.cell(row=row, column=5, value=pr.get('total_distributions', 0)).number_format = s['curr']
        ws1.cell(row=row, column=6, value=pr.get('irr', 0)).number_format = s['pct']
        ws1.cell(row=row, column=7, value=pr.get('roe', 0)).number_format = s['pct']
        ws1.cell(row=row, column=8, value=pr.get('moic', 0)).number_format = s['mult']
        row += 1
    for ci in range(1, 9):
        ws1.cell(row=row, column=ci).border = s['top_border']
    ws1.cell(row=row, column=1, value="Deal Total").font = s['bold']
    ws1.cell(row=row, column=2, value=ds.get('total_contributions', 0)).number_format = s['curr']
    ws1.cell(row=row, column=3, value=ds.get('total_cf_distributions', 0)).number_format = s['curr']
    ws1.cell(row=row, column=4, value=ds.get('total_cap_distributions', 0)).number_format = s['curr']
    ws1.cell(row=row, column=5, value=ds.get('total_distributions', 0)).number_format = s['curr']
    ws1.cell(row=row, column=6, value=ds.get('deal_irr', 0)).number_format = s['pct']
    ws1.cell(row=row, column=7, value=ds.get('deal_roe', 0)).number_format = s['pct']
    ws1.cell(row=row, column=8, value=ds.get('deal_moic', 0)).number_format = s['mult']
    for ci in range(1, 9):
        ws1.cell(row=row, column=ci).font = s['bold']
    _autosize_columns(ws1)

    # --- Sheet 2: Annual Forecast ---
    ws2 = wb.create_sheet("Annual Forecast")
    fc_display = result.get("fc_deal_display")
    if fc_display is not None and not fc_display.empty:
        from reporting import annual_aggregation_table
        cap_events = result.get("cap_events_df")
        proceeds_by_year = None
        if cap_events is not None and not cap_events.empty and "Year" in cap_events.columns:
            proceeds_by_year = cap_events.groupby("Year")["amount"].sum()
        try:
            table = annual_aggregation_table(
                fc_display, start_year, horizon_years,
                proceeds_by_year=proceeds_by_year,
                cf_alloc=result.get("cf_alloc"),
                cap_alloc=result.get("cap_alloc"),
            )
            if "Year" in table.columns:
                wide = table.set_index("Year").T
                years = [int(y) for y in wide.columns]
                fcols = ['Line Item'] + [str(y) for y in years]
                frow = _write_header_row(ws2, 1, fcols, s)
                for label, row_data in wide.iterrows():
                    ws2.cell(row=frow, column=1, value=str(label))
                    for ci, y in enumerate(years, 2):
                        val = row_data.get(y)
                        if val is not None and not (isinstance(val, float) and val != val):
                            cell = ws2.cell(row=frow, column=ci, value=float(val))
                            cell.number_format = s['dec2'] if 'DSCR' in str(label) else s['curr']
                    frow += 1
                _autosize_columns(ws2)
        except Exception:
            ws2.cell(row=1, column=1, value="Error generating forecast")
    else:
        ws2.cell(row=1, column=1, value="No forecast data")

    # --- Sheet 3: Debt Service ---
    ws3 = wb.create_sheet("Debt Service")
    loans = result.get("loans", [])
    dcols = ['Loan ID', 'Original Amount', 'Origination', 'Maturity', 'Type', 'Rate']
    drow = _write_header_row(ws3, 1, dcols, s)
    for l in (loans or []):
        ws3.cell(row=drow, column=1, value=l.loan_id)
        ws3.cell(row=drow, column=2, value=l.orig_amount).number_format = s['curr']
        ws3.cell(row=drow, column=3, value=l.orig_date.isoformat() if l.orig_date else "")
        ws3.cell(row=drow, column=4, value=l.maturity_date.isoformat() if l.maturity_date else "")
        ws3.cell(row=drow, column=5, value=l.int_type or "")
        ws3.cell(row=drow, column=6, value=l.fixed_rate or 0).number_format = s['pct']
        drow += 1
    _autosize_columns(ws3)

    # --- Sheet 4: Cash Schedule ---
    ws4 = wb.create_sheet("Cash Schedule")
    sched = result.get("cash_schedule")
    if sched is not None and isinstance(sched, pd.DataFrame) and not sched.empty:
        ccols = list(sched.columns)
        crow = _write_header_row(ws4, 1, ccols, s)
        for _, r in sched.iterrows():
            for ci, col in enumerate(ccols, 1):
                val = r[col]
                cell = ws4.cell(row=crow, column=ci, value=val)
                if isinstance(val, (int, float)):
                    cell.number_format = s['curr']
            crow += 1
        _autosize_columns(ws4)
    else:
        ws4.cell(row=1, column=1, value="No cash schedule data")

    # --- Sheet 5: XIRR Cash Flows ---
    ws5 = wb.create_sheet("XIRR Cash Flows")
    if pr_list:
        partners = [pr['partner'] for pr in pr_list]
        xcols = ['Date', 'Description'] + partners + ['Deal Total']
        xrow = _write_header_row(ws5, 1, xcols, s)
        all_rows = []
        for pr in pr_list:
            for cf in pr.get('cashflow_details', []):
                d = cf['Date']
                date_str = d.isoformat() if hasattr(d, 'isoformat') else str(d)
                desc = cf.get('Description', '')
                key = (date_str, desc)
                found = False
                for ar in all_rows:
                    if ar['key'] == key:
                        ar[pr['partner']] = cf['Amount']
                        found = True
                        break
                if not found:
                    entry = {'key': key, 'date': date_str, 'desc': desc}
                    entry[pr['partner']] = cf['Amount']
                    all_rows.append(entry)
        for ar in all_rows:
            ws5.cell(row=xrow, column=1, value=ar['date'])
            ws5.cell(row=xrow, column=2, value=ar['desc'])
            total = 0.0
            for ci, p in enumerate(partners, 3):
                val = ar.get(p, 0.0)
                ws5.cell(row=xrow, column=ci, value=val).number_format = s['curr']
                total += (val or 0.0)
            ws5.cell(row=xrow, column=len(partners) + 3, value=total).number_format = s['curr']
            xrow += 1
        _autosize_columns(ws5)

    # --- Sheet 6: ROE Audit ---
    try:
        roe_bytes = generate_roe_audit_excel(pr_list, ds, sale_me)
        from openpyxl import load_workbook
        roe_wb = load_workbook(BytesIO(roe_bytes))
        roe_src = roe_wb.active
        ws6 = wb.create_sheet("ROE Audit")
        for r in roe_src.iter_rows():
            for cell in r:
                ws6.cell(row=cell.row, column=cell.column, value=cell.value).number_format = cell.number_format
        _autosize_columns(ws6)
    except Exception:
        ws6 = wb.create_sheet("ROE Audit")
        ws6.cell(row=1, column=1, value="Error generating ROE audit")

    # --- Sheet 7: MOIC Audit ---
    try:
        moic_bytes = generate_moic_audit_excel(pr_list, ds, sale_me)
        from openpyxl import load_workbook
        moic_wb = load_workbook(BytesIO(moic_bytes))
        moic_src = moic_wb.active
        ws7 = wb.create_sheet("MOIC Audit")
        for r in moic_src.iter_rows():
            for cell in r:
                ws7.cell(row=cell.row, column=cell.column, value=cell.value).number_format = cell.number_format
        _autosize_columns(ws7)
    except Exception:
        ws7 = wb.create_sheet("MOIC Audit")
        ws7.cell(row=1, column=1, value="Error generating MOIC audit")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
