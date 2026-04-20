"""Property Financials service — performance chart, IS, BS, tenants, one pager.

Extracts pure data-shaping logic from property_financials_ui.py and one_pager.py.
All functions are pure Python (no Streamlit). They accept DataFrames and return
dicts suitable for JSON serialization.
"""

import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, date
from typing import Optional

from config import IS_ACCOUNTS, BS_ACCOUNTS


# ============================================================
# Shared ISBS helpers
# ============================================================

def _prepare_isbs(isbs_raw, vcode):
    """Filter and parse ISBS data for a deal."""
    if isbs_raw is None or isbs_raw.empty:
        return pd.DataFrame()
    isbs = isbs_raw.copy()
    isbs.columns = [str(c).strip() for c in isbs.columns]
    if 'vcode' in isbs.columns:
        isbs['vcode'] = isbs['vcode'].astype(str).str.strip().str.lower()
        isbs = isbs[isbs['vcode'] == str(vcode).strip().lower()]
    if isbs.empty or 'dtEntry' not in isbs.columns:
        return pd.DataFrame()
    try:
        isbs['dtEntry_parsed'] = pd.to_datetime(isbs['dtEntry'], unit='D', origin='1899-12-30', errors='coerce')
    except Exception:
        isbs['dtEntry_parsed'] = pd.to_datetime(isbs['dtEntry'], errors='coerce')
    null_dates = isbs['dtEntry_parsed'].isna()
    if null_dates.any():
        isbs.loc[null_dates, 'dtEntry_parsed'] = pd.to_datetime(
            isbs.loc[null_dates, 'dtEntry'], errors='coerce')
    if 'vSource' in isbs.columns:
        isbs['vSource'] = isbs['vSource'].astype(str).str.strip()
    if 'vAccount' in isbs.columns:
        isbs['vAccount'] = isbs['vAccount'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    if 'mAmount' in isbs.columns:
        isbs['mAmount'] = pd.to_numeric(isbs['mAmount'], errors='coerce').fillna(0)
    return isbs


def _parse_occupancy(occupancy_raw, vcode):
    """Parse and filter occupancy data. Returns dict {month_end: occ_value}."""
    if occupancy_raw is None or occupancy_raw.empty:
        return {}
    occ = occupancy_raw.copy()
    occ.columns = [str(c).strip() for c in occ.columns]
    if 'vCode' not in occ.columns:
        return {}
    occ['vCode'] = occ['vCode'].astype(str).str.strip().str.lower()
    occ = occ[occ['vCode'] == str(vcode).strip().lower()]
    occ_col = 'Occ%' if 'Occ%' in occ.columns else (
        'OccupancyPercent' if 'OccupancyPercent' in occ.columns else None)
    if not occ_col or 'dtReported' not in occ.columns:
        return {}
    occ['occ_val'] = pd.to_numeric(occ[occ_col], errors='coerce')
    try:
        occ['date_parsed'] = pd.to_datetime(occ['dtReported'], unit='D', origin='1899-12-30', errors='coerce')
    except Exception:
        occ['date_parsed'] = pd.to_datetime(occ['dtReported'], errors='coerce')
    occ = occ.dropna(subset=['date_parsed', 'occ_val'])
    monthly = {}
    for _, row in occ.iterrows():
        me = pd.Timestamp(row['date_parsed']) + pd.offsets.MonthEnd(0)
        monthly[me] = row['occ_val']
    return monthly


# ============================================================
# Performance Chart
# ============================================================

def get_performance_chart_data(
    isbs_raw: pd.DataFrame,
    occupancy_raw: pd.DataFrame,
    vcode: str,
    freq: str = "Quarterly",
    periods: int = 12,
    period_end: Optional[str] = None,
) -> dict:
    """Build performance chart data (NOI + occupancy) for a single deal.

    Full pipeline: cumulative → periodic monthly → aggregate to frequency.
    """
    isbs = _prepare_isbs(isbs_raw, vcode)
    if isbs.empty:
        return {"periods": [], "actual_noi": [], "uw_noi": [], "occupancy": [],
                "frequency": freq, "available_period_ends": []}

    actual_data = isbs[isbs['vSource'] == 'Interim IS']
    uw_data = isbs[isbs['vSource'] == 'Projected IS']

    # Flatten account codes
    rev_accounts = [a for lst in IS_ACCOUNTS['REVENUES'].values() for a in lst]
    exp_accounts = [a for lst in IS_ACCOUNTS['EXPENSES'].values() for a in lst]

    def _compute_cumulative_noi(data, dates):
        noi_by_date = {}
        for dt in dates:
            period = data[data['dtEntry_parsed'] == dt]
            rev = period[period['vAccount'].isin(rev_accounts)]['mAmount'].sum()
            exp = period[period['vAccount'].isin(exp_accounts)]['mAmount'].sum()
            noi_by_date[dt] = (-rev) - exp
        return noi_by_date

    actual_dates = sorted(actual_data['dtEntry_parsed'].dropna().unique())
    uw_dates = sorted(uw_data['dtEntry_parsed'].dropna().unique())

    actual_cum = _compute_cumulative_noi(actual_data, actual_dates) if actual_dates else {}
    uw_cum = _compute_cumulative_noi(uw_data, uw_dates) if uw_dates else {}

    def _cumulative_to_periodic(cum_dict, sorted_dates):
        periodic = {}
        for i, dt in enumerate(sorted_dates):
            dt_ts = pd.Timestamp(dt)
            if dt_ts.month == 1:
                periodic[dt_ts] = cum_dict[dt]
            else:
                prior = None
                for j in range(i - 1, -1, -1):
                    p = pd.Timestamp(sorted_dates[j])
                    if p.year == dt_ts.year:
                        prior = sorted_dates[j]
                        break
                periodic[dt_ts] = cum_dict[dt] - cum_dict.get(prior, 0) if prior else cum_dict[dt]
        return periodic

    actual_periodic = _cumulative_to_periodic(actual_cum, actual_dates)
    uw_periodic = _cumulative_to_periodic(uw_cum, uw_dates)

    def _aggregate(periodic_dict, freq):
        if not periodic_dict:
            return {}
        if freq == "Monthly":
            return periodic_dict
        elif freq == "Quarterly":
            quarterly = {}
            month_counts = {}
            for dt, val in sorted(periodic_dict.items()):
                dt_ts = pd.Timestamp(dt)
                q_month = ((dt_ts.month - 1) // 3 + 1) * 3
                q_end = pd.Timestamp(year=dt_ts.year, month=q_month, day=1) + pd.offsets.MonthEnd(0)
                quarterly[q_end] = quarterly.get(q_end, 0) + val
                month_counts[q_end] = month_counts.get(q_end, 0) + 1
            return {k: v for k, v in quarterly.items() if month_counts.get(k, 0) == 3}
        else:  # Annually
            annual = {}
            month_counts = {}
            for dt, val in sorted(periodic_dict.items()):
                yr_end = pd.Timestamp(year=pd.Timestamp(dt).year, month=12, day=31)
                annual[yr_end] = annual.get(yr_end, 0) + val
                month_counts[yr_end] = month_counts.get(yr_end, 0) + 1
            return {k: v for k, v in annual.items() if month_counts.get(k, 0) == 12}

    actual_agg = _aggregate(actual_periodic, freq)
    uw_agg = _aggregate(uw_periodic, freq)

    all_period_ends = sorted(set(actual_agg.keys()) | set(uw_agg.keys()))
    if not all_period_ends:
        return {"periods": [], "actual_noi": [], "uw_noi": [], "occupancy": [],
                "frequency": freq, "available_period_ends": []}

    available_labels = [pd.Timestamp(d).strftime('%Y-%m-%d') for d in all_period_ends]

    # Determine display range — default to last completed actual period
    if period_end:
        try:
            sel_ts = pd.Timestamp(period_end)
            sel_idx = min(range(len(all_period_ends)), key=lambda i: abs(pd.Timestamp(all_period_ends[i]) - sel_ts))
        except Exception:
            sel_idx = len(all_period_ends) - 1
    else:
        # Default: most recent period with actual NOI data
        actual_end_set = set(actual_agg.keys())
        sel_idx = len(all_period_ends) - 1  # fallback
        for i in range(len(all_period_ends) - 1, -1, -1):
            if all_period_ends[i] in actual_end_set:
                sel_idx = i
                break

    start_idx = max(0, sel_idx - periods + 1)
    display_dates = all_period_ends[start_idx:sel_idx + 1]

    # Occupancy
    monthly_occ = _parse_occupancy(occupancy_raw, vcode)
    occ_by_date = {}
    for dt in display_dates:
        dt_ts = pd.Timestamp(dt)
        if freq == "Monthly":
            me = dt_ts + pd.offsets.MonthEnd(0)
            if me in monthly_occ:
                occ_by_date[dt_ts] = monthly_occ[me]
        elif freq == "Quarterly":
            q_start = ((dt_ts.month - 1) // 3) * 3 + 1
            vals = []
            for m in range(q_start, q_start + 3):
                me = pd.Timestamp(year=dt_ts.year, month=m, day=1) + pd.offsets.MonthEnd(0)
                if me in monthly_occ:
                    vals.append(monthly_occ[me])
            if vals:
                occ_by_date[dt_ts] = sum(vals) / len(vals)
        else:
            yr_vals = [v for k, v in monthly_occ.items() if pd.Timestamp(k).year == dt_ts.year]
            if yr_vals:
                occ_by_date[dt_ts] = sum(yr_vals) / len(yr_vals)

    # Build output
    period_labels = []
    actual_noi = []
    uw_noi = []
    occupancy = []
    for dt in display_dates:
        dt_ts = pd.Timestamp(dt)
        if freq == "Monthly":
            label = dt_ts.strftime('%b %Y')
        elif freq == "Quarterly":
            label = f"Q{(dt_ts.month - 1) // 3 + 1} {dt_ts.year}"
        else:
            label = str(dt_ts.year)
        period_labels.append(label)
        actual_noi.append(actual_agg.get(dt_ts))
        uw_noi.append(uw_agg.get(dt_ts))
        occupancy.append(occ_by_date.get(dt_ts))

    return {
        "periods": period_labels,
        "actual_noi": [float(v) if v is not None else None for v in actual_noi],
        "uw_noi": [float(v) if v is not None else None for v in uw_noi],
        "occupancy": [float(v) if v is not None else None for v in occupancy],
        "frequency": freq,
        "available_period_ends": available_labels,
    }


# ============================================================
# Income Statement helpers
# ============================================================

def _get_cumulative_balances(data_df, as_of_date, accounts_dict):
    period_data = data_df[data_df['dtEntry_parsed'] == as_of_date]
    if period_data.empty:
        return {}
    balances = {}
    for section, categories in accounts_dict.items():
        balances[section] = {}
        for category, acct_list in categories.items():
            total = period_data[period_data['vAccount'].isin(acct_list)]['mAmount'].sum()
            balances[section][category] = float(total)
    return balances


def _get_budget_sum(data_df, start_date, end_date, accounts_dict):
    period_data = data_df[
        (data_df['dtEntry_parsed'] > start_date) & (data_df['dtEntry_parsed'] <= end_date)
    ]
    if period_data.empty:
        return {}
    balances = {}
    for section, categories in accounts_dict.items():
        balances[section] = {}
        for category, acct_list in categories.items():
            total = period_data[period_data['vAccount'].isin(acct_list)]['mAmount'].sum()
            balances[section][category] = float(total)
    return balances


def _get_valuation_sum(fc_df, start_date, end_date, accounts_dict):
    """Get valuation amounts from forecast_feed.

    mAmount_norm uses forecast convention (revenue positive, expenses negative).
    Actuals use MRI accounting convention (revenue negative/credit, expenses
    positive/debit).  Negate mAmount_norm so returned balances match the MRI
    sign convention expected by the display layer.
    """
    if fc_df is None or fc_df.empty:
        return {}
    period_data = fc_df[
        (fc_df['event_date'] > start_date) & (fc_df['event_date'] <= end_date)
    ].copy()
    if period_data.empty:
        return {}
    # Normalize vAccount to string for matching against IS_ACCOUNTS keys
    period_data['vAccount'] = period_data['vAccount'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    balances = {}
    for section, categories in accounts_dict.items():
        balances[section] = {}
        for category, acct_list in categories.items():
            total = period_data[period_data['vAccount'].isin(acct_list)]['mAmount_norm'].sum()
            # Negate to convert forecast sign convention → MRI accounting convention
            balances[section][category] = float(-total)
    return balances


def _add_balances(bal1, bal2, accounts_dict):
    result = {}
    for section, categories in accounts_dict.items():
        result[section] = {}
        for category in categories.keys():
            v1 = bal1.get(section, {}).get(category, 0)
            v2 = bal2.get(section, {}).get(category, 0)
            result[section][category] = v1 + v2
    return result


def _subtract_balances(bal1, bal2, accounts_dict):
    result = {}
    for section, categories in accounts_dict.items():
        result[section] = {}
        for category in categories.keys():
            v1 = bal1.get(section, {}).get(category, 0)
            v2 = bal2.get(section, {}).get(category, 0)
            result[section][category] = v1 - v2
    return result


def _calculate_is_amounts(period_type, source, ref_date, year, accounts_dict,
                          actual_data, actual_periods, budget_data, uw_data, fc_deal_modeled):
    ref_date = pd.Timestamp(ref_date)

    if source == "Actual":
        if period_type == "TTM":
            current_bal = _get_cumulative_balances(actual_data, ref_date, accounts_dict)
            dec_prior = next((pd.Timestamp(p) for p in actual_periods
                              if pd.Timestamp(p).year == ref_date.year - 1 and pd.Timestamp(p).month == 12), None)
            same_month_ly = next((pd.Timestamp(p) for p in actual_periods
                                  if pd.Timestamp(p).year == ref_date.year - 1 and pd.Timestamp(p).month == ref_date.month), None)
            if dec_prior and same_month_ly:
                dec_bal = _get_cumulative_balances(actual_data, dec_prior, accounts_dict)
                ly_bal = _get_cumulative_balances(actual_data, same_month_ly, accounts_dict)
                return _subtract_balances(_add_balances(current_bal, dec_bal, accounts_dict), ly_bal, accounts_dict)
            elif dec_prior:
                return _add_balances(current_bal, _get_cumulative_balances(actual_data, dec_prior, accounts_dict), accounts_dict)
            return current_bal

        elif period_type == "YTD":
            return _get_cumulative_balances(actual_data, ref_date, accounts_dict)

        elif period_type == "Full Year":
            dec_date = next((pd.Timestamp(p) for p in actual_periods
                             if pd.Timestamp(p).year == year and pd.Timestamp(p).month == 12), None)
            return _get_cumulative_balances(actual_data, dec_date, accounts_dict) if dec_date else {}

        elif period_type == "Estimate":
            ytd = _get_cumulative_balances(actual_data, ref_date, accounts_dict)
            dec_end = pd.Timestamp(f"{ref_date.year}-12-31")
            remainder = _get_budget_sum(budget_data, ref_date, dec_end, accounts_dict)
            return _add_balances(ytd, remainder, accounts_dict)

        else:  # Custom
            return _get_cumulative_balances(actual_data, ref_date, accounts_dict)

    elif source == "Budget":
        if period_type == "TTM":
            start = ref_date - pd.DateOffset(months=12)
            return _get_budget_sum(budget_data, start, ref_date, accounts_dict)
        elif period_type == "YTD":
            jan1 = pd.Timestamp(f"{ref_date.year}-01-01") - pd.DateOffset(days=1)
            return _get_budget_sum(budget_data, jan1, ref_date, accounts_dict)
        elif period_type == "Full Year":
            jan1 = pd.Timestamp(f"{year}-01-01") - pd.DateOffset(days=1)
            dec31 = pd.Timestamp(f"{year}-12-31")
            return _get_budget_sum(budget_data, jan1, dec31, accounts_dict)
        elif period_type == "Estimate":
            jan1 = pd.Timestamp(f"{ref_date.year}-01-01") - pd.DateOffset(days=1)
            dec31 = pd.Timestamp(f"{ref_date.year}-12-31")
            return _get_budget_sum(budget_data, jan1, dec31, accounts_dict)
        else:
            prior = ref_date - pd.DateOffset(months=1)
            return _get_budget_sum(budget_data, prior, ref_date, accounts_dict)

    elif source == "Underwriting":
        # Underwriting (Projected IS) is stored as YTD cumulative snapshots
        # (trial balance), same as Actuals — NOT periodic like Budget.
        uw_periods_list = sorted(uw_data['dtEntry_parsed'].dropna().unique()) if not uw_data.empty else []

        if period_type == "TTM":
            current_bal = _get_cumulative_balances(uw_data, ref_date, accounts_dict)
            dec_prior = next((pd.Timestamp(p) for p in uw_periods_list
                              if pd.Timestamp(p).year == ref_date.year - 1 and pd.Timestamp(p).month == 12), None)
            same_month_ly = next((pd.Timestamp(p) for p in uw_periods_list
                                  if pd.Timestamp(p).year == ref_date.year - 1 and pd.Timestamp(p).month == ref_date.month), None)
            if dec_prior and same_month_ly:
                dec_bal = _get_cumulative_balances(uw_data, dec_prior, accounts_dict)
                ly_bal = _get_cumulative_balances(uw_data, same_month_ly, accounts_dict)
                return _subtract_balances(_add_balances(current_bal, dec_bal, accounts_dict), ly_bal, accounts_dict)
            elif dec_prior:
                return _add_balances(current_bal, _get_cumulative_balances(uw_data, dec_prior, accounts_dict), accounts_dict)
            return current_bal

        elif period_type == "YTD":
            return _get_cumulative_balances(uw_data, ref_date, accounts_dict)

        elif period_type == "Full Year":
            dec_date = next((pd.Timestamp(p) for p in uw_periods_list
                             if pd.Timestamp(p).year == year and pd.Timestamp(p).month == 12), None)
            return _get_cumulative_balances(uw_data, dec_date, accounts_dict) if dec_date else {}

        elif period_type == "Estimate":
            # Full year from underwriting
            dec_date = next((pd.Timestamp(p) for p in uw_periods_list
                             if pd.Timestamp(p).year == ref_date.year and pd.Timestamp(p).month == 12), None)
            return _get_cumulative_balances(uw_data, dec_date, accounts_dict) if dec_date else {}

        else:  # Custom — single-month periodic: current YTD minus prior month YTD
            current_bal = _get_cumulative_balances(uw_data, ref_date, accounts_dict)
            prior_date = ref_date - pd.DateOffset(months=1)
            closest_prior = next((pd.Timestamp(p) for p in reversed(uw_periods_list)
                                  if pd.Timestamp(p) <= prior_date), None)
            if closest_prior:
                prior_bal = _get_cumulative_balances(uw_data, closest_prior, accounts_dict)
                return _subtract_balances(current_bal, prior_bal, accounts_dict)
            return current_bal

    elif source == "Valuation":
        if period_type == "TTM":
            start = ref_date - pd.DateOffset(months=12)
            return _get_valuation_sum(fc_deal_modeled, start.date(), ref_date.date(), accounts_dict)
        elif period_type == "YTD":
            jan1 = pd.Timestamp(f"{ref_date.year}-01-01") - pd.DateOffset(days=1)
            return _get_valuation_sum(fc_deal_modeled, jan1.date(), ref_date.date(), accounts_dict)
        elif period_type == "Full Year":
            jan1 = pd.Timestamp(f"{year}-01-01") - pd.DateOffset(days=1)
            dec31 = pd.Timestamp(f"{year}-12-31")
            return _get_valuation_sum(fc_deal_modeled, jan1.date(), dec31.date(), accounts_dict)
        elif period_type == "Estimate":
            jan1 = pd.Timestamp(f"{ref_date.year}-01-01") - pd.DateOffset(days=1)
            dec31 = pd.Timestamp(f"{ref_date.year}-12-31")
            return _get_valuation_sum(fc_deal_modeled, jan1.date(), dec31.date(), accounts_dict)
        else:
            prior = ref_date - pd.DateOffset(months=1)
            return _get_valuation_sum(fc_deal_modeled, prior.date(), ref_date.date(), accounts_dict)

    return {}


# ============================================================
# Income Statement
# ============================================================

def get_income_statement(
    isbs_raw: pd.DataFrame,
    vcode: str,
    left_period: str = "TTM",
    left_source: str = "Actual",
    right_period: str = "YTD",
    right_source: str = "Budget",
    left_as_of_date: Optional[str] = None,
    right_as_of_date: Optional[str] = None,
    fc_deal_modeled: Optional[pd.DataFrame] = None,
) -> dict:
    """Build income statement comparison data.

    Each column has its own as-of date so callers can compare different periods.
    """
    isbs = _prepare_isbs(isbs_raw, vcode)
    if isbs.empty and (fc_deal_modeled is None or fc_deal_modeled.empty):
        return {"rows": [], "available_dates": {}, "left_label": "", "right_label": ""}

    if not isbs.empty:
        actual_data = isbs[isbs['vSource'] == 'Interim IS']
        budget_data = isbs[isbs['vSource'] == 'Budget IS']
        uw_data = isbs[isbs['vSource'] == 'Projected IS']
    else:
        actual_data = pd.DataFrame()
        budget_data = pd.DataFrame()
        uw_data = pd.DataFrame()

    # Collect available dates per source
    actual_periods = sorted(actual_data['dtEntry_parsed'].dropna().unique()) if not actual_data.empty else []
    budget_periods = sorted(budget_data['dtEntry_parsed'].dropna().unique()) if not budget_data.empty else []
    uw_periods = sorted(uw_data['dtEntry_parsed'].dropna().unique()) if not uw_data.empty else []
    val_periods = []
    if fc_deal_modeled is not None and not fc_deal_modeled.empty:
        raw_dates = fc_deal_modeled['event_date'].dropna().unique()
        val_periods = sorted(set(pd.Timestamp(d) for d in raw_dates))

    available_dates = {
        "Actual": [pd.Timestamp(d).strftime('%Y-%m-%d') for d in actual_periods],
        "Budget": [pd.Timestamp(d).strftime('%Y-%m-%d') for d in budget_periods],
        "Underwriting": [pd.Timestamp(d).strftime('%Y-%m-%d') for d in uw_periods],
        "Valuation": [pd.Timestamp(d).strftime('%Y-%m-%d') for d in val_periods],
    }

    # Resolve default ref dates per column
    def _default_ref():
        for periods in (actual_periods, budget_periods, uw_periods, val_periods):
            if periods:
                return pd.Timestamp(periods[-1])
        return None

    left_ref = pd.Timestamp(left_as_of_date) if left_as_of_date else _default_ref()
    right_ref = pd.Timestamp(right_as_of_date) if right_as_of_date else _default_ref()

    if left_ref is None or right_ref is None:
        return {"rows": [], "available_dates": available_dates, "left_label": "", "right_label": ""}

    left_year = left_ref.year
    right_year = right_ref.year

    left_amounts = _calculate_is_amounts(
        left_period, left_source, left_ref, left_year, IS_ACCOUNTS,
        actual_data, actual_periods, budget_data, uw_data, fc_deal_modeled)
    right_amounts = _calculate_is_amounts(
        right_period, right_source, right_ref, right_year, IS_ACCOUNTS,
        actual_data, actual_periods, budget_data, uw_data, fc_deal_modeled)

    # Build rows
    rows = []
    for section in ['REVENUES', 'EXPENSES']:
        rows.append({"account": section, "level": 0, "is_header": True,
                      "left": None, "right": None, "var_usd": None, "var_pct": None})
        section_left_total = 0
        section_right_total = 0
        for category in IS_ACCOUNTS[section].keys():
            left_val = left_amounts.get(section, {}).get(category, 0)
            right_val = right_amounts.get(section, {}).get(category, 0)
            # Revenue stored as credits (negative) → display positive
            if section == 'REVENUES':
                left_val = -left_val
                right_val = -right_val
            var_usd = left_val - right_val
            var_pct = (var_usd / abs(right_val)) if right_val != 0 else None
            rows.append({"account": category, "level": 1, "is_header": False,
                          "left": left_val, "right": right_val,
                          "var_usd": var_usd, "var_pct": var_pct})
            section_left_total += left_val
            section_right_total += right_val

        total_label = f"Total {section.title()}"
        var_usd = section_left_total - section_right_total
        var_pct = (var_usd / abs(section_right_total)) if section_right_total != 0 else None
        rows.append({"account": total_label, "level": 0, "is_total": True,
                      "left": section_left_total, "right": section_right_total,
                      "var_usd": var_usd, "var_pct": var_pct})

    # NOI
    rev_left = sum(r['left'] for r in rows if r.get('is_total') and 'Revenues' in r['account']) or 0
    rev_right = sum(r['right'] for r in rows if r.get('is_total') and 'Revenues' in r['account']) or 0
    exp_left = sum(r['left'] for r in rows if r.get('is_total') and 'Expenses' in r['account']) or 0
    exp_right = sum(r['right'] for r in rows if r.get('is_total') and 'Expenses' in r['account']) or 0
    noi_left = rev_left - exp_left
    noi_right = rev_right - exp_right
    noi_var = noi_left - noi_right
    rows.append({"account": "NET OPERATING INCOME", "level": 0, "is_calc": True,
                  "left": noi_left, "right": noi_right,
                  "var_usd": noi_var, "var_pct": (noi_var / abs(noi_right)) if noi_right != 0 else None})

    # Debt Service
    if 'DEBT_SERVICE' in IS_ACCOUNTS:
        rows.append({"account": "DEBT SERVICE", "level": 0, "is_header": True,
                      "left": None, "right": None, "var_usd": None, "var_pct": None})
        ds_left_total = 0
        ds_right_total = 0
        for category in IS_ACCOUNTS['DEBT_SERVICE'].keys():
            left_val = left_amounts.get('DEBT_SERVICE', {}).get(category, 0)
            right_val = right_amounts.get('DEBT_SERVICE', {}).get(category, 0)
            var_usd = left_val - right_val
            rows.append({"account": category, "level": 1, "is_header": False,
                          "left": left_val, "right": right_val,
                          "var_usd": var_usd, "var_pct": (var_usd / abs(right_val)) if right_val != 0 else None})
            ds_left_total += left_val
            ds_right_total += right_val
        rows.append({"account": "Total Debt Service", "level": 0, "is_total": True,
                      "left": ds_left_total, "right": ds_right_total,
                      "var_usd": ds_left_total - ds_right_total,
                      "var_pct": ((ds_left_total - ds_right_total) / abs(ds_right_total)) if ds_right_total != 0 else None})
        # DSCR
        dscr_left = noi_left / abs(ds_left_total) if ds_left_total != 0 else None
        dscr_right = noi_right / abs(ds_right_total) if ds_right_total != 0 else None
        rows.append({"account": "DSCR", "level": 0, "is_calc": True,
                      "left": dscr_left, "right": dscr_right, "var_usd": None, "var_pct": None})

    # Other Below the Line
    if 'OTHER_BTL' in IS_ACCOUNTS:
        rows.append({"account": "OTHER BELOW THE LINE", "level": 0, "is_header": True,
                      "left": None, "right": None, "var_usd": None, "var_pct": None})
        for category in IS_ACCOUNTS['OTHER_BTL'].keys():
            left_val = left_amounts.get('OTHER_BTL', {}).get(category, 0)
            right_val = right_amounts.get('OTHER_BTL', {}).get(category, 0)
            if category == 'Interest Income':
                left_val = -left_val
                right_val = -right_val
            var_usd = left_val - right_val
            rows.append({"account": category, "level": 1, "is_header": False,
                          "left": left_val, "right": right_val,
                          "var_usd": var_usd, "var_pct": (var_usd / abs(right_val)) if right_val != 0 else None})

    def _period_label(source, period, ref):
        date_part = ref.strftime('%Y-%m') if period in ('TTM', 'YTD', 'Custom') else str(ref.year)
        return f"{source} {period} {date_part}"

    left_label = _period_label(left_source, left_period, left_ref)
    right_label = _period_label(right_source, right_period, right_ref)

    return {
        "rows": rows,
        "available_dates": available_dates,
        "left_label": left_label,
        "right_label": right_label,
    }


# ============================================================
# Balance Sheet
# ============================================================

def _get_period_balances(isbs_df, period_date, accounts_dict):
    period_data = isbs_df[isbs_df['dtEntry_parsed'] == period_date]
    if period_data.empty:
        return {}
    balances = {}
    for section, subsections in accounts_dict.items():
        balances[section] = {}
        for subsection, categories in subsections.items():
            balances[section][subsection] = {}
            for category, acct_list in categories.items():
                total = period_data[period_data['vAccount'].isin(acct_list)]['mAmount'].sum()
                balances[section][subsection][category] = float(total)
    return balances


def get_balance_sheet(
    isbs_raw: pd.DataFrame,
    vcode: str,
    period1: Optional[str] = None,
    period2: Optional[str] = None,
) -> dict:
    """Build balance sheet comparison data."""
    isbs = _prepare_isbs(isbs_raw, vcode)
    if isbs.empty:
        return {"rows": [], "available_periods": [], "period1_label": "", "period2_label": ""}

    bs_data = isbs[isbs['vSource'] == 'Interim BS']
    available = sorted(bs_data['dtEntry_parsed'].dropna().unique())
    available_labels = [pd.Timestamp(d).strftime('%Y-%m-%d') for d in available]

    if len(available) < 1:
        return {"rows": [], "available_periods": available_labels, "period1_label": "", "period2_label": ""}

    p2_date = pd.Timestamp(period2) if period2 else (available[-1] if available else None)
    p1_date = pd.Timestamp(period1) if period1 else (available[-2] if len(available) >= 2 else available[0])

    bal1 = _get_period_balances(bs_data, p1_date, BS_ACCOUNTS)
    bal2 = _get_period_balances(bs_data, p2_date, BS_ACCOUNTS)

    rows = []
    for section in ['ASSETS', 'LIABILITIES', 'EQUITY']:
        rows.append({"account": section, "level": 0, "is_header": True,
                      "period1": None, "period2": None, "var_usd": None, "var_pct": None})
        section_p1 = 0
        section_p2 = 0
        for subsection in BS_ACCOUNTS.get(section, {}).keys():
            rows.append({"account": subsection, "level": 1, "is_header": True,
                          "period1": None, "period2": None, "var_usd": None, "var_pct": None})
            sub_p1 = 0
            sub_p2 = 0
            for category in BS_ACCOUNTS[section][subsection].keys():
                v1 = bal1.get(section, {}).get(subsection, {}).get(category, 0)
                v2 = bal2.get(section, {}).get(subsection, {}).get(category, 0)
                var = v2 - v1
                rows.append({"account": category, "level": 2, "is_header": False,
                              "period1": v1, "period2": v2,
                              "var_usd": var, "var_pct": (var / abs(v1)) if v1 != 0 else None})
                sub_p1 += v1
                sub_p2 += v2
            var = sub_p2 - sub_p1
            rows.append({"account": f"Total {subsection}", "level": 1, "is_total": True,
                          "period1": sub_p1, "period2": sub_p2,
                          "var_usd": var, "var_pct": (var / abs(sub_p1)) if sub_p1 != 0 else None})
            section_p1 += sub_p1
            section_p2 += sub_p2
        var = section_p2 - section_p1
        rows.append({"account": f"TOTAL {section}", "level": 0, "is_total": True,
                      "period1": section_p1, "period2": section_p2,
                      "var_usd": var, "var_pct": (var / abs(section_p1)) if section_p1 != 0 else None})

    return {
        "rows": rows,
        "available_periods": available_labels,
        "period1_label": p1_date.strftime('%Y-%m-%d') if p1_date else "",
        "period2_label": p2_date.strftime('%Y-%m-%d') if p2_date else "",
    }


# ============================================================
# Tenant Roster
# ============================================================

def _clean_currency(val):
    if pd.isna(val) or val == '':
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    return float(str(val).replace('$', '').replace(',', '').strip())


def _clean_number(val):
    if pd.isna(val) or val == '':
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    return float(str(val).replace(',', '').strip())


def get_tenant_roster(tenants_raw: pd.DataFrame, vcode: str, inv: Optional[pd.DataFrame] = None) -> dict:
    """Get tenant roster and lease rollover data."""
    if tenants_raw is None or tenants_raw.empty:
        return {"tenants": [], "summary": {}, "rollover": {}}

    t = tenants_raw.copy()
    t['Code'] = t['Code'].astype(str).str.strip()
    t = t[t['Code'] == str(vcode)]
    if t.empty:
        return {"tenants": [], "summary": {}, "rollover": {}}

    t['SF_Leased'] = t['SF Leased'].apply(_clean_number)
    t['Monthly_Rent'] = t['Rent'].apply(_clean_currency)
    t['Annual_Rent'] = t['Monthly_Rent'] * 12
    t['Rentable_SF'] = t['Rentable SF'].apply(_clean_number)
    t['RPSF'] = t.apply(lambda r: r['Annual_Rent'] / r['SF_Leased'] if r['SF_Leased'] > 0 else 0, axis=1)

    total_rentable = t['Rentable_SF'].iloc[0] if len(t) > 0 else 0
    total_annual_rent = t['Annual_Rent'].sum()
    t['Pct_GLA'] = t['SF_Leased'] / total_rentable if total_rentable > 0 else 0
    t['Pct_ABR'] = t['Annual_Rent'] / total_annual_rent if total_annual_rent > 0 else 0

    t['Lease_Start'] = pd.to_datetime(t['Lease Start'], errors='coerce')
    t['Lease_End'] = pd.to_datetime(t['Lease End'], errors='coerce')

    two_years = pd.Timestamp.now() + pd.DateOffset(years=2)
    t['Expiring_Soon'] = (t['Lease_End'] <= two_years) & (t['Lease_End'] >= pd.Timestamp.now())
    t['Is_Vacant'] = t['Tenant Name'].str.strip().str.lower() == 'vacant'
    t = t.sort_values('SF_Leased', ascending=False)

    occupied = t[~t['Is_Vacant']]
    total_occupied = occupied['SF_Leased'].sum()
    occupancy_pct = total_occupied / t['SF_Leased'].sum() if t['SF_Leased'].sum() > 0 else 0
    wtd_avg_rpsf = occupied['Annual_Rent'].sum() / total_occupied if total_occupied > 0 else 0

    tenants = []
    for _, r in t.iterrows():
        tenants.append({
            "tenant_name": r['Tenant Name'],
            "sf_leased": r['SF_Leased'],
            "lease_start": r['Lease_Start'].isoformat() if pd.notna(r['Lease_Start']) else None,
            "lease_end": r['Lease_End'].isoformat() if pd.notna(r['Lease_End']) else None,
            "annual_rent": r['Annual_Rent'],
            "rpsf": r['RPSF'],
            "pct_gla": r['Pct_GLA'],
            "pct_abr": r['Pct_ABR'],
            "expiring_soon": bool(r['Expiring_Soon'] and not r['Is_Vacant']),
            "is_vacant": bool(r['Is_Vacant']),
        })

    summary = {"occupied_sf": total_occupied, "occupancy_pct": occupancy_pct, "wtd_avg_rpsf": wtd_avg_rpsf}

    # Rollover
    rollover = {}
    if inv is not None:
        deal_row = inv[inv['vcode'] == str(vcode)]
        if not deal_row.empty:
            p = deal_row.iloc[0]
            rollover["property_info"] = {
                "name": p.get("Investment_Name", ""),
                "partner": p.get("Operating_Partner", ""),
                "location": f"{p.get('City', '')}, {p.get('State', '')}".strip(', '),
                "asset_type": p.get("Asset_Type", ""),
                "gla": _clean_number(p.get("Size_Sqf", 0)) or total_rentable,
            }

    # Maturity by year
    non_vacant = t[~t['Is_Vacant']].copy()
    non_vacant['End_Year'] = non_vacant['Lease_End'].apply(
        lambda x: str(x.year) if pd.notna(x) else 'Unknown')
    maturity = non_vacant.groupby('End_Year').agg({'SF_Leased': 'sum', 'Annual_Rent': 'sum'}).reset_index()
    maturity['avg_rpsf'] = maturity.apply(lambda r: r['Annual_Rent'] / r['SF_Leased'] if r['SF_Leased'] > 0 else 0, axis=1)
    maturity['pct_revenue'] = maturity['Annual_Rent'] / total_annual_rent if total_annual_rent > 0 else 0
    rollover["maturity_by_year"] = [
        {"year": r['End_Year'], "sf": r['SF_Leased'], "annual_rent": r['Annual_Rent'],
         "avg_rpsf": r['avg_rpsf'], "pct_revenue": r['pct_revenue']}
        for _, r in maturity.iterrows()
    ]

    # Exposure 2yr
    expiring = t[(t['Expiring_Soon']) & (~t['Is_Vacant'])]
    exp_gla = expiring['SF_Leased'].sum()
    exp_abr = expiring['Annual_Rent'].sum()
    rollover["exposure_2yr"] = {
        "gla": exp_gla,
        "gla_pct": exp_gla / total_rentable if total_rentable > 0 else 0,
        "abr": exp_abr,
        "abr_pct": exp_abr / total_annual_rent if total_annual_rent > 0 else 0,
        "rpsf": exp_abr / exp_gla if exp_gla > 0 else 0,
    }

    return {"tenants": tenants, "summary": summary, "rollover": rollover}


# ============================================================
# One Pager (delegates to one_pager.py)
# ============================================================

def get_one_pager_data(vcode, quarter_str, inv, isbs_raw, mri_loans, mri_val,
                       waterfalls, commitments, acct, occupancy_raw=None):
    """Aggregate all One Pager sections into a single response."""
    from one_pager import (
        get_general_information, get_capitalization_stack,
        get_property_performance, get_pe_performance,
        get_one_pager_comments, get_available_quarters,
    )

    available = get_available_quarters(isbs_raw) if isbs_raw is not None else []

    # Default to latest quarter if none provided
    if not quarter_str and available:
        quarter_str = available[0]

    general = get_general_information(inv, vcode)
    cap_stack = get_capitalization_stack(vcode, mri_loans, mri_val, waterfalls, commitments, acct, inv,
                                         isbs_raw=isbs_raw, quarter_str=quarter_str)
    prop_perf = get_property_performance(vcode, quarter_str, isbs_raw, mri_val, occupancy_raw) if quarter_str else {}
    pe_perf = get_pe_performance(vcode, quarter_str, acct, commitments, waterfalls, inv) if quarter_str else {}
    comments = get_one_pager_comments(vcode, quarter_str) if quarter_str else {}

    # Compute PE Yield on Exposure = Actual YE NOI / (Debt + PE)
    if prop_perf and cap_stack:
        noi_ye = prop_perf.get('noi', {}).get('actual_ye', 0) or prop_perf.get('noi', {}).get('ytd_actual', 0) or 0
        senior_plus_pe = cap_stack.get('debt', 0) + cap_stack.get('pref_equity', 0)
        if senior_plus_pe > 0 and noi_ye > 0:
            cap_stack['pe_yield_on_exposure'] = noi_ye / senior_plus_pe

    return {
        "available_quarters": available,
        "general": general,
        "cap_stack": cap_stack,
        "property_performance": prop_perf,
        "pe_performance": pe_perf,
        "comments": comments,
    }


def get_one_pager_chart(vcode, isbs_raw, occupancy_raw, num_quarters=12):
    """Build quarterly NOI chart data for One Pager (same pipeline as perf chart, fixed quarterly)."""
    return get_performance_chart_data(isbs_raw, occupancy_raw, vcode, freq="Quarterly", periods=num_quarters)


# ============================================================
# Excel Generators
# ============================================================

def _safe_autosize(ws):
    """Auto-size columns, skipping merged cells."""
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            max_len = max(max_len, len(str(cell.value or "")))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_performance_chart_excel(chart_data: dict, deal_name: str = "") -> bytes:
    """Create Excel workbook with performance chart data as a table."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from flask_app.services.compute_service import _excel_styles, _write_header_row

    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Data"

    # Title row
    title = f"{deal_name} — Performance Data" if deal_name else "Performance Data"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # Header row 3
    cols = ["Period", "Actual NOI ($)", "U/W NOI ($)", "Occupancy (%)"]
    r = _write_header_row(ws, 3, cols, s)

    # Data rows
    periods = chart_data.get("periods", [])
    actual_noi = chart_data.get("actual_noi", [])
    uw_noi = chart_data.get("uw_noi", [])
    occupancy = chart_data.get("occupancy", [])

    for i, period in enumerate(periods):
        ws.cell(row=r, column=1, value=period)
        c_actual = ws.cell(row=r, column=2, value=actual_noi[i] if i < len(actual_noi) else None)
        c_actual.number_format = '$#,##0'
        c_uw = ws.cell(row=r, column=3, value=uw_noi[i] if i < len(uw_noi) else None)
        c_uw.number_format = '$#,##0'
        occ_val = occupancy[i] if i < len(occupancy) else None
        c_occ = ws.cell(row=r, column=4, value=occ_val / 100.0 if occ_val is not None else None)
        c_occ.number_format = '0.0%'
        r += 1

    _safe_autosize(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_income_statement_excel(is_data: dict, deal_name: str = "") -> bytes:
    """Create Excel workbook with income statement data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask_app.services.compute_service import _excel_styles, _write_header_row

    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    # Title row
    title = f"{deal_name} — Income Statement" if deal_name else "Income Statement"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # Header row 3
    left_label = is_data.get("left_label", "Left")
    right_label = is_data.get("right_label", "Right")
    cols = ["Account", left_label, right_label, "Variance ($)", "Variance (%)"]
    r = _write_header_row(ws, 3, cols, s)

    calc_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    top_border = Border(top=Side(style='medium'))

    rows = is_data.get("rows", [])
    for row in rows:
        account = row.get("account", "")
        is_header = row.get("is_header", False)
        is_total = row.get("is_total", False)
        is_calc = row.get("is_calc", False)
        level = row.get("level", 0)
        is_dscr = account == "DSCR"

        # Indent level 1 rows
        if level == 1 and not is_header:
            account = "  " + account

        ws.cell(row=r, column=1, value=account)

        if is_header:
            ws.cell(row=r, column=1).font = Font(bold=True)
        elif is_total:
            for ci in range(1, 6):
                ws.cell(row=r, column=ci).font = Font(bold=True)
                ws.cell(row=r, column=ci).border = top_border
        elif is_calc:
            for ci in range(1, 6):
                ws.cell(row=r, column=ci).font = Font(bold=True)
                ws.cell(row=r, column=ci).fill = calc_fill

        if not is_header:
            left_val = row.get("left")
            right_val = row.get("right")
            var_usd = row.get("var_usd")
            var_pct = row.get("var_pct")

            if is_dscr:
                fmt = '0.00"x"'
                c_left = ws.cell(row=r, column=2, value=left_val)
                c_left.number_format = fmt
                c_right = ws.cell(row=r, column=3, value=right_val)
                c_right.number_format = fmt
            else:
                c_left = ws.cell(row=r, column=2, value=left_val)
                c_left.number_format = '$#,##0'
                c_right = ws.cell(row=r, column=3, value=right_val)
                c_right.number_format = '$#,##0'
                c_var = ws.cell(row=r, column=4, value=var_usd)
                c_var.number_format = '$#,##0'

            c_vpct = ws.cell(row=r, column=5, value=var_pct)
            c_vpct.number_format = '0.0%'

        r += 1

    _safe_autosize(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_balance_sheet_excel(bs_data: dict, deal_name: str = "") -> bytes:
    """Create Excel workbook with balance sheet data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from flask_app.services.compute_service import _excel_styles, _write_header_row

    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    # Title row
    title = f"{deal_name} — Balance Sheet" if deal_name else "Balance Sheet"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # Header row 3
    p1_label = bs_data.get("period1_label", "Period 1")
    p2_label = bs_data.get("period2_label", "Period 2")
    cols = ["Account", p1_label, p2_label, "Variance ($)", "Variance (%)"]
    r = _write_header_row(ws, 3, cols, s)

    top_border = Border(top=Side(style='medium'))

    rows = bs_data.get("rows", [])
    for row in rows:
        account = row.get("account", "")
        is_header = row.get("is_header", False)
        is_total = row.get("is_total", False)
        level = row.get("level", 0)

        # Indent based on level
        if level == 2 and not is_header:
            account = "    " + account
        elif level == 1 and not is_header:
            account = "  " + account

        ws.cell(row=r, column=1, value=account)

        if is_header:
            ws.cell(row=r, column=1).font = Font(bold=True)
        elif is_total:
            for ci in range(1, 6):
                ws.cell(row=r, column=ci).font = Font(bold=True)
                ws.cell(row=r, column=ci).border = top_border

        if not is_header:
            c_p1 = ws.cell(row=r, column=2, value=row.get("period1"))
            c_p1.number_format = '$#,##0'
            c_p2 = ws.cell(row=r, column=3, value=row.get("period2"))
            c_p2.number_format = '$#,##0'
            c_var = ws.cell(row=r, column=4, value=row.get("var_usd"))
            c_var.number_format = '$#,##0'
            c_vpct = ws.cell(row=r, column=5, value=row.get("var_pct"))
            c_vpct.number_format = '0.0%'

        r += 1

    _safe_autosize(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_tenant_roster_excel(tenant_data: dict, deal_name: str = "") -> bytes:
    """Create multi-sheet Excel workbook with tenant roster and lease rollover."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask_app.services.compute_service import _excel_styles, _write_header_row

    s = _excel_styles()
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    wb = Workbook()

    # ---- Sheet 1: Tenant Roster ----
    ws = wb.active
    ws.title = "Tenant Roster"

    # Title row 1
    title = f"{deal_name} — Tenant Roster" if deal_name else "Tenant Roster"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # Summary row 2
    summary = tenant_data.get("summary", {})
    occ_sf = summary.get("occupied_sf", 0)
    occ_pct = summary.get("occupancy_pct", 0)
    wtd_rpsf = summary.get("wtd_avg_rpsf", 0)
    summary_text = f"Occupied SF: {occ_sf:,.0f} | Occupancy: {occ_pct:.1%} | Wtd Avg RPSF: ${wtd_rpsf:,.2f}"
    ws.cell(row=2, column=1, value=summary_text).font = Font(bold=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Header row 4
    cols = ["Tenant", "SF Leased", "Lease Start", "Lease End", "Annual Rent", "$/SF", "% GLA", "% ABR"]
    r = _write_header_row(ws, 4, cols, s)

    tenants = tenant_data.get("tenants", [])
    for t in tenants:
        ws.cell(row=r, column=1, value=t.get("tenant_name", ""))
        c_sf = ws.cell(row=r, column=2, value=t.get("sf_leased"))
        c_sf.number_format = '#,##0'

        # Lease dates
        start_str = t.get("lease_start")
        if start_str:
            try:
                ws.cell(row=r, column=3, value=pd.Timestamp(start_str).to_pydatetime()).number_format = 'MM/DD/YYYY'
            except Exception:
                ws.cell(row=r, column=3, value=start_str)
        end_str = t.get("lease_end")
        if end_str:
            try:
                ws.cell(row=r, column=4, value=pd.Timestamp(end_str).to_pydatetime()).number_format = 'MM/DD/YYYY'
            except Exception:
                ws.cell(row=r, column=4, value=end_str)

        c_rent = ws.cell(row=r, column=5, value=t.get("annual_rent"))
        c_rent.number_format = '$#,##0'
        c_rpsf = ws.cell(row=r, column=6, value=t.get("rpsf"))
        c_rpsf.number_format = '$#,##0.00'
        c_gla = ws.cell(row=r, column=7, value=t.get("pct_gla"))
        c_gla.number_format = '0.0%'
        c_abr = ws.cell(row=r, column=8, value=t.get("pct_abr"))
        c_abr.number_format = '0.0%'

        # Vacant rows: italic
        if t.get("is_vacant"):
            for ci in range(1, 9):
                ws.cell(row=r, column=ci).font = Font(italic=True)

        # Expiring rows: yellow fill
        if t.get("expiring_soon"):
            for ci in range(1, 9):
                ws.cell(row=r, column=ci).fill = yellow_fill

        r += 1

    _safe_autosize(ws)

    # ---- Sheet 2: Lease Rollover ----
    ws2 = wb.create_sheet("Lease Rollover")

    ws2.cell(row=1, column=1, value="Lease Maturity Schedule").font = Font(bold=True, size=13)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    r2 = 3
    rollover = tenant_data.get("rollover", {})
    exposure = rollover.get("exposure_2yr")
    if exposure:
        exp_gla = exposure.get("gla", 0)
        exp_gla_pct = exposure.get("gla_pct", 0)
        exp_abr = exposure.get("abr", 0)
        exp_abr_pct = exposure.get("abr_pct", 0)
        kpi_text = (f"2-Year Exposure — Expiring GLA: {exp_gla:,.0f} ({exp_gla_pct:.1%}) "
                    f"| Expiring ABR: ${exp_abr:,.0f} ({exp_abr_pct:.1%})")
        ws2.cell(row=r2, column=1, value=kpi_text).font = Font(bold=True)
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=5)
        r2 += 2

    # Maturity table header
    mat_cols = ["Year", "SF", "Annual Rent", "Avg $/SF", "% Revenue"]
    r2 = _write_header_row(ws2, r2, mat_cols, s)

    maturity = rollover.get("maturity_by_year", [])
    for m in maturity:
        ws2.cell(row=r2, column=1, value=m.get("year", ""))
        c_sf = ws2.cell(row=r2, column=2, value=m.get("sf"))
        c_sf.number_format = '#,##0'
        c_rent = ws2.cell(row=r2, column=3, value=m.get("annual_rent"))
        c_rent.number_format = '$#,##0'
        c_rpsf = ws2.cell(row=r2, column=4, value=m.get("avg_rpsf"))
        c_rpsf.number_format = '$#,##0.00'
        c_pct = ws2.cell(row=r2, column=5, value=m.get("pct_revenue"))
        c_pct.number_format = '0.0%'
        r2 += 1

    _safe_autosize(ws2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_full_financials_excel(chart_data, is_data, bs_data, tenant_data, deal_name="") -> bytes:
    """Combine all property financials sections into a single 5-sheet workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask_app.services.compute_service import _excel_styles, _write_header_row

    s = _excel_styles()
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    calc_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    top_border = Border(top=Side(style='medium'))
    wb = Workbook()

    # ---- Sheet 1: Performance Data ----
    ws = wb.active
    ws.title = "Performance Data"
    title = f"{deal_name} — Performance Data" if deal_name else "Performance Data"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    cols = ["Period", "Actual NOI ($)", "U/W NOI ($)", "Occupancy (%)"]
    r = _write_header_row(ws, 3, cols, s)
    periods = chart_data.get("periods", [])
    actual_noi = chart_data.get("actual_noi", [])
    uw_noi = chart_data.get("uw_noi", [])
    occupancy = chart_data.get("occupancy", [])
    for i, period in enumerate(periods):
        ws.cell(row=r, column=1, value=period)
        ws.cell(row=r, column=2, value=actual_noi[i] if i < len(actual_noi) else None).number_format = '$#,##0'
        ws.cell(row=r, column=3, value=uw_noi[i] if i < len(uw_noi) else None).number_format = '$#,##0'
        occ_val = occupancy[i] if i < len(occupancy) else None
        ws.cell(row=r, column=4, value=occ_val / 100.0 if occ_val is not None else None).number_format = '0.0%'
        r += 1
    _safe_autosize(ws)

    # ---- Sheet 2: Income Statement ----
    ws2 = wb.create_sheet("Income Statement")
    title = f"{deal_name} — Income Statement" if deal_name else "Income Statement"
    ws2.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    left_label = is_data.get("left_label", "Left")
    right_label = is_data.get("right_label", "Right")
    is_cols = ["Account", left_label, right_label, "Variance ($)", "Variance (%)"]
    r = _write_header_row(ws2, 3, is_cols, s)
    for row in is_data.get("rows", []):
        account = row.get("account", "")
        is_header = row.get("is_header", False)
        is_total_row = row.get("is_total", False)
        is_calc_row = row.get("is_calc", False)
        level = row.get("level", 0)
        is_dscr = account == "DSCR"
        if level == 1 and not is_header:
            account = "  " + account
        ws2.cell(row=r, column=1, value=account)
        if is_header:
            ws2.cell(row=r, column=1).font = Font(bold=True)
        elif is_total_row:
            for ci in range(1, 6):
                ws2.cell(row=r, column=ci).font = Font(bold=True)
                ws2.cell(row=r, column=ci).border = top_border
        elif is_calc_row:
            for ci in range(1, 6):
                ws2.cell(row=r, column=ci).font = Font(bold=True)
                ws2.cell(row=r, column=ci).fill = calc_fill
        if not is_header:
            left_val = row.get("left")
            right_val = row.get("right")
            if is_dscr:
                fmt = '0.00"x"'
                ws2.cell(row=r, column=2, value=left_val).number_format = fmt
                ws2.cell(row=r, column=3, value=right_val).number_format = fmt
            else:
                ws2.cell(row=r, column=2, value=left_val).number_format = '$#,##0'
                ws2.cell(row=r, column=3, value=right_val).number_format = '$#,##0'
                ws2.cell(row=r, column=4, value=row.get("var_usd")).number_format = '$#,##0'
            ws2.cell(row=r, column=5, value=row.get("var_pct")).number_format = '0.0%'
        r += 1
    _safe_autosize(ws2)

    # ---- Sheet 3: Balance Sheet ----
    ws3 = wb.create_sheet("Balance Sheet")
    title = f"{deal_name} — Balance Sheet" if deal_name else "Balance Sheet"
    ws3.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    p1_label = bs_data.get("period1_label", "Period 1")
    p2_label = bs_data.get("period2_label", "Period 2")
    bs_cols = ["Account", p1_label, p2_label, "Variance ($)", "Variance (%)"]
    r = _write_header_row(ws3, 3, bs_cols, s)
    for row in bs_data.get("rows", []):
        account = row.get("account", "")
        is_header = row.get("is_header", False)
        is_total_row = row.get("is_total", False)
        level = row.get("level", 0)
        if level == 2 and not is_header:
            account = "    " + account
        elif level == 1 and not is_header:
            account = "  " + account
        ws3.cell(row=r, column=1, value=account)
        if is_header:
            ws3.cell(row=r, column=1).font = Font(bold=True)
        elif is_total_row:
            for ci in range(1, 6):
                ws3.cell(row=r, column=ci).font = Font(bold=True)
                ws3.cell(row=r, column=ci).border = top_border
        if not is_header:
            ws3.cell(row=r, column=2, value=row.get("period1")).number_format = '$#,##0'
            ws3.cell(row=r, column=3, value=row.get("period2")).number_format = '$#,##0'
            ws3.cell(row=r, column=4, value=row.get("var_usd")).number_format = '$#,##0'
            ws3.cell(row=r, column=5, value=row.get("var_pct")).number_format = '0.0%'
        r += 1
    _safe_autosize(ws3)

    # ---- Sheet 4: Tenant Roster ----
    ws4 = wb.create_sheet("Tenant Roster")
    title = f"{deal_name} — Tenant Roster" if deal_name else "Tenant Roster"
    ws4.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    summary = tenant_data.get("summary", {})
    occ_sf = summary.get("occupied_sf", 0)
    occ_pct = summary.get("occupancy_pct", 0)
    wtd_rpsf = summary.get("wtd_avg_rpsf", 0)
    summary_text = f"Occupied SF: {occ_sf:,.0f} | Occupancy: {occ_pct:.1%} | Wtd Avg RPSF: ${wtd_rpsf:,.2f}"
    ws4.cell(row=2, column=1, value=summary_text).font = Font(bold=True)
    ws4.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    t_cols = ["Tenant", "SF Leased", "Lease Start", "Lease End", "Annual Rent", "$/SF", "% GLA", "% ABR"]
    r = _write_header_row(ws4, 4, t_cols, s)
    for t in tenant_data.get("tenants", []):
        ws4.cell(row=r, column=1, value=t.get("tenant_name", ""))
        ws4.cell(row=r, column=2, value=t.get("sf_leased")).number_format = '#,##0'
        start_str = t.get("lease_start")
        if start_str:
            try:
                ws4.cell(row=r, column=3, value=pd.Timestamp(start_str).to_pydatetime()).number_format = 'MM/DD/YYYY'
            except Exception:
                ws4.cell(row=r, column=3, value=start_str)
        end_str = t.get("lease_end")
        if end_str:
            try:
                ws4.cell(row=r, column=4, value=pd.Timestamp(end_str).to_pydatetime()).number_format = 'MM/DD/YYYY'
            except Exception:
                ws4.cell(row=r, column=4, value=end_str)
        ws4.cell(row=r, column=5, value=t.get("annual_rent")).number_format = '$#,##0'
        ws4.cell(row=r, column=6, value=t.get("rpsf")).number_format = '$#,##0.00'
        ws4.cell(row=r, column=7, value=t.get("pct_gla")).number_format = '0.0%'
        ws4.cell(row=r, column=8, value=t.get("pct_abr")).number_format = '0.0%'
        if t.get("is_vacant"):
            for ci in range(1, 9):
                ws4.cell(row=r, column=ci).font = Font(italic=True)
        if t.get("expiring_soon"):
            for ci in range(1, 9):
                ws4.cell(row=r, column=ci).fill = yellow_fill
        r += 1
    _safe_autosize(ws4)

    # ---- Sheet 5: Lease Rollover ----
    ws5 = wb.create_sheet("Lease Rollover")
    ws5.cell(row=1, column=1, value="Lease Maturity Schedule").font = Font(bold=True, size=13)
    ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    r5 = 3
    rollover = tenant_data.get("rollover", {})
    exposure = rollover.get("exposure_2yr")
    if exposure:
        exp_gla = exposure.get("gla", 0)
        exp_gla_pct = exposure.get("gla_pct", 0)
        exp_abr = exposure.get("abr", 0)
        exp_abr_pct = exposure.get("abr_pct", 0)
        kpi_text = (f"2-Year Exposure — Expiring GLA: {exp_gla:,.0f} ({exp_gla_pct:.1%}) "
                    f"| Expiring ABR: ${exp_abr:,.0f} ({exp_abr_pct:.1%})")
        ws5.cell(row=r5, column=1, value=kpi_text).font = Font(bold=True)
        ws5.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=5)
        r5 += 2
    mat_cols = ["Year", "SF", "Annual Rent", "Avg $/SF", "% Revenue"]
    r5 = _write_header_row(ws5, r5, mat_cols, s)
    for m in rollover.get("maturity_by_year", []):
        ws5.cell(row=r5, column=1, value=m.get("year", ""))
        ws5.cell(row=r5, column=2, value=m.get("sf")).number_format = '#,##0'
        ws5.cell(row=r5, column=3, value=m.get("annual_rent")).number_format = '$#,##0'
        ws5.cell(row=r5, column=4, value=m.get("avg_rpsf")).number_format = '$#,##0.00'
        ws5.cell(row=r5, column=5, value=m.get("pct_revenue")).number_format = '0.0%'
        r5 += 1
    _safe_autosize(ws5)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
