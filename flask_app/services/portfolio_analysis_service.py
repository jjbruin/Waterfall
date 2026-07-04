"""Portfolio Analysis service — upstream entity analysis for active portfolios.

Generalized version of PSCKOC service. Traces deal cash flows through PPI
entities to a selected portfolio entity (e.g., TGA23), with two modes:

- Actual: Uses the entity's saved waterfall (CF_WF / Cap_WF)
- Proposed: Applies simplified assumptions (AM Fee, Hurdle, Promote, Expenses)
  at the selected entity level only. All lower-tier waterfalls remain actual.
"""

import pandas as pd
from io import BytesIO
from datetime import date, datetime
from typing import Optional

from ownership_tree import load_relationships, build_ownership_tree
from loaders import load_waterfalls
from waterfall import run_recursive_upstream_waterfalls, build_amfee_exclusions
from models import InvestorState
from metrics import xirr, calculate_roe


def find_portfolio_entities(wf: pd.DataFrame,
                            relationships_raw: pd.DataFrame,
                            inv: pd.DataFrame) -> list[dict]:
    """Find upstream entities that own multiple deals through intermediaries.

    Returns entities like TGA22, TGA23, TGA24 — holding companies that are
    investors in PPIs which are referenced in deal waterfalls.
    """
    if relationships_raw is None or relationships_raw.empty:
        return []
    if wf is None or wf.empty:
        return []

    rel = relationships_raw.copy()
    rel["InvestorID"] = rel["InvestorID"].astype(str).str.strip()
    rel["InvestmentID"] = rel["InvestmentID"].astype(str).str.strip()

    wf_norm = wf.copy()
    wf_norm["PropCode"] = wf_norm["PropCode"].fillna("").astype(str).str.strip()
    wf_norm["vcode"] = wf_norm["vcode"].fillna("").astype(str).str.strip()

    deal_vcodes_set = set(inv["vcode"].astype(str).str.strip()) if inv is not None else set()

    # Find entities that are investors in intermediary entities (PPIs)
    # which in turn appear as PropCode recipients in deal waterfalls.
    # Also require the entity itself to have a waterfall defined.
    entity_wf_vcodes = set(wf_norm["vcode"].unique())

    # Build: intermediary -> set of deal vcodes it receives from
    intermediary_deals = {}
    for pc in wf_norm["PropCode"].unique():
        if pc in deal_vcodes_set:
            continue  # skip deals themselves
        refs = wf_norm[wf_norm["PropCode"] == pc]["vcode"].unique()
        deals = [v for v in refs if v in deal_vcodes_set]
        if deals:
            intermediary_deals[pc] = set(deals)

    # Build: entity -> set of intermediaries it invests in
    # Only consider current relationships (EndDate is null)
    current_rel = rel[rel["EndDate"].isna()]
    entity_intermediaries = {}
    for _, r in current_rel.iterrows():
        investor = r["InvestorID"]
        investment = r["InvestmentID"]
        if investment in intermediary_deals:
            entity_intermediaries.setdefault(investor, set()).add(investment)

    # Filter to entities that:
    # 1. Have their own waterfall defined
    # 2. Own at least 2 intermediaries that link to deals
    # 3. Are not themselves deals
    results = []
    for entity_id, intermediaries in entity_intermediaries.items():
        if entity_id in deal_vcodes_set:
            continue
        if entity_id not in entity_wf_vcodes:
            continue
        # Count total deals reachable through intermediaries
        all_deals = set()
        for ppi in intermediaries:
            all_deals |= intermediary_deals.get(ppi, set())
        if len(all_deals) < 2:
            continue

        # Get entity name from relationships
        name_rows = rel[rel["InvestmentID"] == entity_id]
        name = name_rows.iloc[0]["Name"] if not name_rows.empty else entity_id

        results.append({
            "entity_id": entity_id,
            "name": name,
            "deal_count": len(all_deals),
            "intermediary_count": len(intermediaries),
        })

    results.sort(key=lambda x: x["entity_id"])
    return results


def find_entity_deals(entity_id: str, inv: pd.DataFrame, wf: pd.DataFrame,
                      relationships_raw: pd.DataFrame) -> list[dict]:
    """Find deals linked to a specific portfolio entity through intermediaries.

    Uses recursive downward traversal to find all intermediate entities, then
    filters to only deals whose waterfall PropCode references one of those
    entities. This prevents pulling in deals from shared holding entities
    that the selected entity doesn't have direct waterfall exposure to.
    Excludes sold deals and filters ended relationships.

    Returns list of deal info dicts with vcode, name, PPI linkage, ownership.
    """
    if relationships_raw is None or relationships_raw.empty:
        return []
    if inv is None or inv.empty:
        return []

    relationships = load_relationships(relationships_raw)

    # Filter out ended relationships
    if "EndDate" in relationships.columns:
        end_col = relationships["EndDate"]
        is_empty = end_col.isna() | (end_col.astype(str).str.strip().isin(["", "NaT", "nan", "None"]))
        relationships = relationships[is_empty].copy()

    nodes = build_ownership_tree(relationships)

    # Exclude sold deals
    inv_norm = inv.copy()
    inv_norm["vcode"] = inv_norm["vcode"].astype(str).str.strip()
    inv_norm["_sale"] = inv_norm.get("Sale_Status", "").fillna("").astype(str).str.strip().str.upper()
    inv_norm["_life"] = inv_norm.get("Lifecycle", "").fillna("").astype(str).str.strip().str.upper()
    active = inv_norm[(inv_norm["_sale"] != "SOLD") & (inv_norm["_life"] != "SOLD")]
    deal_vcodes_set = set(active["vcode"])

    # Trace DOWNWARD from entity through investments recursively
    def _get_downstream_entities(eid: str, visited: set = None) -> set:
        if visited is None:
            visited = set()
        if eid in visited:
            return set()
        visited.add(eid)
        result = set()
        node = nodes.get(eid)
        if not node:
            return result
        for child in node.investments:
            result.add(child)
            result |= _get_downstream_entities(child, visited)
        return result

    downstream = _get_downstream_entities(entity_id)

    # Find deals whose waterfall PropCode references the entity itself or a
    # downstream entity. Including the entity itself catches deals that list
    # it directly as a PropCode (e.g., Life Storage lists TGA22 as PropCode).
    wf_norm = wf.copy() if wf is not None and not wf.empty else pd.DataFrame()
    deal_vcodes = set()
    ppi_for_deal = {}
    match_entities = downstream | {entity_id}

    if not wf_norm.empty:
        wf_norm["PropCode"] = wf_norm["PropCode"].fillna("").astype(str).str.strip()
        wf_norm["vcode"] = wf_norm["vcode"].fillna("").astype(str).str.strip()

        for entity in match_entities:
            refs = wf_norm[wf_norm["PropCode"] == entity]["vcode"].unique()
            for vc in refs:
                if vc in deal_vcodes_set:
                    deal_vcodes.add(vc)
                    ppi_for_deal[vc] = entity

    # Remove the entity itself from results
    deal_vcodes -= {entity_id}

    # Build entity's direct investment ownership percentages (for display)
    entity_inv_rels = relationships[relationships["InvestorID"].astype(str).str.strip() == entity_id]
    ppi_pcts = {}
    for _, r in entity_inv_rels.iterrows():
        ppi = str(r["InvestmentID"]).strip()
        pct = float(r.get("OwnershipPct", 0))
        if pct > 1:
            pct = pct / 100.0
        ppi_pcts[ppi] = pct

    # Build deal info list
    result = []
    for vc in sorted(deal_vcodes):
        row_info = {"vcode": vc, "name": vc}
        match = inv[inv["vcode"].astype(str).str.strip() == vc]
        if not match.empty:
            r = match.iloc[0]
            row_info["name"] = str(r.get("Investment_Name", vc))
            row_info["asset_type"] = str(r.get("Asset_Type", ""))
            row_info["sale_status"] = str(r.get("Sale_Status", "") or "Active")
        # Show intermediate entity that links this deal
        ppi = ppi_for_deal.get(vc)
        if ppi and ppi in ppi_pcts:
            row_info["ppi_entity"] = ppi
            row_info["entity_pct"] = ppi_pcts[ppi]
        result.append(row_info)

    return result


def get_entity_investors(entity_id: str,
                         relationships_raw: pd.DataFrame) -> list[dict]:
    """Get current investors in the entity (who the waterfall allocates to)."""
    if relationships_raw is None or relationships_raw.empty:
        return []
    rel = relationships_raw.copy()
    rel["InvestorID"] = rel["InvestorID"].astype(str).str.strip()
    rel["InvestmentID"] = rel["InvestmentID"].astype(str).str.strip()
    if "EndDate" in rel.columns:
        end_col = rel["EndDate"]
        is_empty = end_col.isna() | (end_col.astype(str).str.strip().isin(["", "NaT", "nan", "None"]))
    else:
        is_empty = True
    current = rel[(rel["InvestmentID"] == entity_id) & is_empty]
    investors = []
    for _, r in current.iterrows():
        investors.append({
            "investor_id": r["InvestorID"],
            "ownership_pct": float(r.get("OwnershipPct", 0)),
            "name": r.get("Name", ""),
        })
    return investors


def compute_portfolio_actual(entity_id: str, data: dict,
                             start_year: int, horizon_years: int,
                             pro_yr_base: int) -> dict:
    """Compute portfolio returns using actual saved waterfalls at all tiers.

    Runs deal waterfalls -> upstream through PPIs -> through entity waterfall.
    """
    inv = data["inv"]
    wf = data["wf"]
    relationships_raw = data["relationships_raw"]

    deal_list = find_entity_deals(entity_id, inv, wf, relationships_raw)
    deal_vcodes = [d["vcode"] for d in deal_list]

    if not deal_vcodes:
        return {"error": f"No deals found for entity {entity_id}"}

    return _run_upstream_computation(
        entity_id=entity_id,
        deal_vcodes=deal_vcodes,
        data=data,
        start_year=start_year,
        horizon_years=horizon_years,
        pro_yr_base=pro_yr_base,
        wf_override=None,
    )


def compute_portfolio_proposed(entity_id: str, data: dict,
                               assumptions: dict,
                               start_year: int, horizon_years: int,
                               pro_yr_base: int) -> dict:
    """Compute portfolio returns with proposed waterfall at entity level only.

    Lower tiers (deals, PPIs) use actual waterfalls. The entity-level waterfall
    is replaced with a simplified model using the provided assumptions.

    Assumptions: {am_fee_pct, hurdle_rate, promote_pct, annual_expenses}
    """
    inv = data["inv"]
    wf = data["wf"]
    relationships_raw = data["relationships_raw"]

    deal_list = find_entity_deals(entity_id, inv, wf, relationships_raw)
    deal_vcodes = [d["vcode"] for d in deal_list]

    if not deal_vcodes:
        return {"error": f"No deals found for entity {entity_id}"}

    # Build a synthetic waterfall for this entity based on assumptions.
    # Get current investors to build the allocation steps.
    investors = get_entity_investors(entity_id, relationships_raw)

    # Separate capital investors (ownership > 0) from fee/promote recipients (ownership == 0)
    capital_investors = [i for i in investors if i["ownership_pct"] > 0]
    total_ownership = sum(i["ownership_pct"] for i in capital_investors)

    if total_ownership <= 0:
        return {"error": f"No capital investors found for {entity_id}"}

    # Normalize ownership percentages
    for inv_info in capital_investors:
        inv_info["norm_pct"] = inv_info["ownership_pct"] / total_ownership

    # Build synthetic waterfall steps as a DataFrame matching waterfall format.
    # Structure: Amt (expenses) -> Pref (hurdle) -> Share (with promote deducted)
    #
    # For "Proposed" we apply the Sold Portfolio approach:
    # the entity receives cash from lower tiers, then we simulate net returns
    # using the assumptions. This is done post-hoc on the cashflows, not via
    # the waterfall engine, because the Sold Portfolio model is fundamentally
    # a cashflow-level simulation (AM Fee on capital, pref accrual, promote).
    #
    # So we run upstream waterfalls stopping BEFORE the entity level,
    # collect the cashflows arriving at the entity, then apply the
    # proposed waterfall logic to split among investors.

    return _run_upstream_with_proposed(
        entity_id=entity_id,
        deal_vcodes=deal_vcodes,
        data=data,
        assumptions=assumptions,
        capital_investors=capital_investors,
        start_year=start_year,
        horizon_years=horizon_years,
        pro_yr_base=pro_yr_base,
    )


def _build_entity_seeded_states(entity_id: str, acct) -> dict:
    """Build pre-seeded InvestorStates for entity investors from accounting.

    The upstream waterfall creates empty InvestorStates by default, so
    capital_outstanding=0 and pref never accrues. This function seeds states
    with net capital (contributions - capital returned) from entity accounting
    so that Pref/Def_Int steps in the entity waterfall work correctly.
    """
    if acct is None or acct.empty:
        return {}

    ent = acct[acct["InvestmentID"].astype(str).str.strip() == entity_id].copy()
    if ent.empty:
        return {}

    ent["_amt"] = pd.to_numeric(ent["Amt"], errors="coerce").fillna(0)
    ent["_iid"] = ent["InvestorID"].astype(str).str.strip()
    ent["_date"] = pd.to_datetime(ent["EffectiveDate"], format="mixed", dayfirst=False)

    states = {}
    for iid, grp in ent.groupby("_iid"):
        # Net capital = sum of contributions (negative) + capital returns (positive)
        net_capital = -grp["_amt"].sum()  # flip sign: contributions are negative in acct
        if net_capital <= 0:
            continue
        st = InvestorState(propcode=iid)
        st.capital_outstanding = net_capital
        # Set last_accrual_date to earliest contribution date so pref accrues from start
        dates = grp.loc[grp["_amt"] < 0, "_date"].dropna()
        if not dates.empty:
            st.get_pool("initial").last_accrual_date = dates.min().date()
        states[iid] = st

    return states


def _run_upstream_computation(entity_id: str, deal_vcodes: list[str],
                              data: dict, start_year: int,
                              horizon_years: int, pro_yr_base: int,
                              wf_override) -> dict:
    """Core computation: run deals + upstream waterfalls through entity."""
    from flask_app.services.compute_service import get_cached_deal_result

    inv = data["inv"]
    wf = data["wf"]
    wf_steps = load_waterfalls(wf)
    relationships_raw = data["relationships_raw"]
    actuals_through = data.get("actuals_through")

    all_cf_allocs = []
    all_cap_allocs = []
    deal_results = {}
    errors = []

    for vcode in deal_vcodes:
        try:
            result = get_cached_deal_result(
                vcode=vcode,
                start_year=start_year,
                horizon_years=horizon_years,
                pro_yr_base=pro_yr_base,
                data=data,
                actuals_through=actuals_through,
            )
        except Exception as e:
            errors.append(f"Deal {vcode}: {e}")
            continue

        if "error" in result:
            errors.append(f"Deal {vcode}: {result['error']}")
            continue

        deal_results[vcode] = result

        cf_alloc = result.get("cf_alloc")
        cap_alloc = result.get("cap_alloc")

        if cf_alloc is not None and not cf_alloc.empty:
            cf = cf_alloc.copy()
            cf["vcode"] = vcode
            all_cf_allocs.append(cf)

        if cap_alloc is not None and not cap_alloc.empty:
            ca = cap_alloc.copy()
            ca["vcode"] = vcode
            all_cap_allocs.append(ca)

    if not all_cf_allocs and not all_cap_allocs:
        return {"error": "No deal allocations produced", "errors": errors}

    # Build relationships
    if relationships_raw is not None and not relationships_raw.empty:
        relationships = load_relationships(relationships_raw)
    else:
        relationships = pd.DataFrame()

    combined_cf = pd.concat(all_cf_allocs, ignore_index=True) if all_cf_allocs else pd.DataFrame()
    combined_cap = pd.concat(all_cap_allocs, ignore_index=True) if all_cap_allocs else pd.DataFrame()

    # AMFee exclusion capital
    _acct = data.get("acct")
    _excl = build_amfee_exclusions(_acct, relationships) if _acct is not None else {}

    # Seed entity-level InvestorStates with capital from accounting so pref accrues
    seeded = _build_entity_seeded_states(entity_id, _acct)

    # Run CF upstream waterfalls
    cf_upstream_alloc = pd.DataFrame()
    cf_entity_states = {}
    if not combined_cf.empty:
        cf_upstream_alloc, cf_entity_states, _ = \
            run_recursive_upstream_waterfalls(
                deal_allocations=combined_cf,
                wf_steps=wf_steps,
                relationships=relationships,
                wf_type="CF_WF",
                amfee_exclusions=_excl,
                pre_seeded_states=seeded,
            )

    # Run Cap upstream waterfalls
    cap_upstream_alloc = pd.DataFrame()
    cap_entity_states = {}
    if not combined_cap.empty:
        cap_upstream_alloc, cap_entity_states, _ = \
            run_recursive_upstream_waterfalls(
                deal_allocations=combined_cap,
                wf_steps=wf_steps,
                relationships=relationships,
                wf_type="Cap_WF",
                amfee_exclusions=_excl,
                pre_seeded_states=seeded,
            )

    return _build_entity_results(
        entity_id=entity_id,
        cf_upstream_alloc=cf_upstream_alloc,
        cap_upstream_alloc=cap_upstream_alloc,
        cf_entity_states=cf_entity_states,
        cap_entity_states=cap_entity_states,
        deal_results=deal_results,
        deal_vcodes=deal_vcodes,
        inv=inv,
        relationships_raw=data["relationships_raw"],
        mode="actual",
        errors=errors,
        acct=data.get("acct"),
        wf_raw=wf,
    )


def _run_upstream_with_proposed(entity_id: str, deal_vcodes: list[str],
                                data: dict, assumptions: dict,
                                capital_investors: list[dict],
                                start_year: int, horizon_years: int,
                                pro_yr_base: int) -> dict:
    """Run deals + upstream waterfalls, but stop before entity and apply proposed model.

    Strategy: Run the full upstream computation (which goes through the entity's
    actual waterfall), then recompute the entity-level allocation using proposed
    assumptions on the cashflows that arrive at the entity.
    """
    from flask_app.services.compute_service import get_cached_deal_result

    inv = data["inv"]
    wf = data["wf"]
    wf_steps = load_waterfalls(wf)
    relationships_raw = data["relationships_raw"]
    actuals_through = data.get("actuals_through")

    all_cf_allocs = []
    all_cap_allocs = []
    deal_results = {}
    errors = []

    for vcode in deal_vcodes:
        try:
            result = get_cached_deal_result(
                vcode=vcode,
                start_year=start_year,
                horizon_years=horizon_years,
                pro_yr_base=pro_yr_base,
                data=data,
                actuals_through=actuals_through,
            )
        except Exception as e:
            errors.append(f"Deal {vcode}: {e}")
            continue

        if "error" in result:
            errors.append(f"Deal {vcode}: {result['error']}")
            continue

        deal_results[vcode] = result

        cf_alloc = result.get("cf_alloc")
        cap_alloc = result.get("cap_alloc")

        if cf_alloc is not None and not cf_alloc.empty:
            cf = cf_alloc.copy()
            cf["vcode"] = vcode
            all_cf_allocs.append(cf)

        if cap_alloc is not None and not cap_alloc.empty:
            ca = cap_alloc.copy()
            ca["vcode"] = vcode
            all_cap_allocs.append(ca)

    if not all_cf_allocs and not all_cap_allocs:
        return {"error": "No deal allocations produced", "errors": errors}

    # Build relationships
    if relationships_raw is not None and not relationships_raw.empty:
        relationships = load_relationships(relationships_raw)
    else:
        relationships = pd.DataFrame()

    combined_cf = pd.concat(all_cf_allocs, ignore_index=True) if all_cf_allocs else pd.DataFrame()
    combined_cap = pd.concat(all_cap_allocs, ignore_index=True) if all_cap_allocs else pd.DataFrame()

    _acct = data.get("acct")
    _excl = build_amfee_exclusions(_acct, relationships) if _acct is not None else {}

    # Run upstream waterfalls through ALL tiers (including entity).
    # We need the full results to extract what arrives at the entity level.
    cf_upstream_alloc = pd.DataFrame()
    cf_entity_states = {}
    if not combined_cf.empty:
        cf_upstream_alloc, cf_entity_states, _ = \
            run_recursive_upstream_waterfalls(
                deal_allocations=combined_cf,
                wf_steps=wf_steps,
                relationships=relationships,
                wf_type="CF_WF",
                amfee_exclusions=_excl,
            )

    cap_upstream_alloc = pd.DataFrame()
    cap_entity_states = {}
    if not combined_cap.empty:
        cap_upstream_alloc, cap_entity_states, _ = \
            run_recursive_upstream_waterfalls(
                deal_allocations=combined_cap,
                wf_steps=wf_steps,
                relationships=relationships,
                wf_type="Cap_WF",
                amfee_exclusions=_excl,
            )

    # Extract cashflows arriving at the entity (before entity-level waterfall).
    # These are rows where PropCode == entity_id (cash allocated TO this entity).
    entity_cf_income = _extract_entity_income(cf_upstream_alloc, entity_id)
    entity_cap_income = _extract_entity_income(cap_upstream_alloc, entity_id)

    # Apply proposed waterfall model on entity-level cashflows
    proposed_results = _apply_proposed_waterfall(
        entity_id=entity_id,
        cf_income=entity_cf_income,
        cap_income=entity_cap_income,
        assumptions=assumptions,
        capital_investors=capital_investors,
        deal_results=deal_results,
        deal_vcodes=deal_vcodes,
        inv=inv,
        relationships_raw=data["relationships_raw"],
        acct=data.get("acct"),
    )
    proposed_results["errors"] = errors
    proposed_results["deals_computed"] = len(deal_results)
    proposed_results["mode"] = "proposed"
    proposed_results["assumptions"] = assumptions
    return proposed_results


def _extract_entity_income(upstream_alloc: pd.DataFrame,
                           entity_id: str) -> list[dict]:
    """Extract cashflow events arriving at an entity from upstream allocations."""
    if upstream_alloc is None or upstream_alloc.empty:
        return []
    # Rows where PropCode == entity_id represent cash flowing TO this entity
    entity_rows = upstream_alloc[
        upstream_alloc["PropCode"].astype(str).str.strip() == entity_id
    ]
    events = []
    for _, row in entity_rows.iterrows():
        events.append({
            "date": row.get("event_date"),
            "amount": float(row.get("Allocated", 0)),
            "source_entity": str(row.get("Entity", "")),
            "vState": str(row.get("vState", "")),
            "path": str(row.get("Path", "")),
        })
    return events


def _apply_proposed_waterfall(entity_id: str,
                              cf_income: list[dict],
                              cap_income: list[dict],
                              assumptions: dict,
                              capital_investors: list[dict],
                              deal_results: dict,
                              deal_vcodes: list[str],
                              inv: pd.DataFrame,
                              relationships_raw: pd.DataFrame,
                              acct: pd.DataFrame = None) -> dict:
    """Apply Sold Portfolio-style proposed waterfall to entity-level cashflows.

    This simulates: AM Fee, Expenses, Pref accrual, Promote on the cash
    arriving at the entity, then splits among investors by ownership %.

    For the proposed model, all capital investors share pari passu (by
    ownership %) after fees and promote. The promote goes to the GP/manager.
    """
    am_fee_pct = assumptions.get("am_fee_pct", 0)
    hurdle_rate = assumptions.get("hurdle_rate", 0)
    promote_pct = assumptions.get("promote_pct", 0)
    annual_expenses = assumptions.get("annual_expenses", 0)

    # Combine CF and Cap income events chronologically
    all_events = []
    for ev in cf_income:
        all_events.append({**ev, "type": "CF"})
    for ev in cap_income:
        all_events.append({**ev, "type": "Cap"})

    # Parse dates and sort
    for ev in all_events:
        d = ev["date"]
        if isinstance(d, str):
            try:
                ev["date"] = pd.to_datetime(d).date()
            except Exception:
                ev["date"] = date.today()
        elif hasattr(d, "date"):
            ev["date"] = d.date()
        elif not isinstance(d, date):
            ev["date"] = date.today()

    all_events.sort(key=lambda x: x["date"])

    if not all_events:
        return _empty_proposed_results(entity_id, capital_investors, deal_vcodes, inv, relationships_raw)

    # Walk through events and compute proposed waterfall
    # Track: capital balance (from contributions seeded from accounting),
    # pref balance, cumulative cashflows for XNPV hurdle
    capital_balance = 0.0
    pref_balance = 0.0
    last_date = all_events[0]["date"]  # updated below after accounting seed
    waterfall_detail = []
    net_cashflows = {}  # investor_id -> [(date, amount)]
    total_net_cashflows = []  # for entity-level XIRR

    # Initialize investor net cashflow lists
    for inv_info in capital_investors:
        net_cashflows[inv_info["investor_id"]] = []

    # Seed entity-level contributions from accounting data.
    # These are the actual capital calls to the entity's investors.
    if acct is not None and not acct.empty:
        ent_acct = acct[acct["InvestmentID"].astype(str).str.strip() == entity_id].copy()
        if not ent_acct.empty:
            ent_acct["_date"] = pd.to_datetime(ent_acct["EffectiveDate"], format="mixed", dayfirst=False)
            ent_acct["_amt"] = pd.to_numeric(ent_acct["Amt"], errors="coerce").fillna(0)
            # Build per-investor contribution cashflows
            inv_id_set = {i["investor_id"] for i in capital_investors}
            for _, row in ent_acct.iterrows():
                iid = str(row["InvestorID"]).strip()
                d = row["_date"]
                amt = float(row["_amt"])
                if pd.isna(d) or amt == 0 or iid not in inv_id_set:
                    continue
                d_date = d.date()
                net_cashflows[iid].append((d_date, amt))
                total_net_cashflows.append((d_date, amt))
                if amt < 0:
                    capital_balance += abs(amt)

    # Set last_date to earliest contribution date for pref accrual
    all_dates = [d for cfs in net_cashflows.values() for d, _ in cfs]
    if all_dates:
        last_date = min(last_date, min(all_dates))

    # Walk through waterfall events (distributions from lower tiers)
    for ev in all_events:
        ev_date = ev["date"]
        amount = ev["amount"]
        ev_type = ev["type"]

        if amount <= 0:
            continue

        # Distribution: apply proposed waterfall
        days = (ev_date - last_date).days
        years_frac = days / 365.0

        # AM Fee on capital balance
        am_fee = capital_balance * am_fee_pct * years_frac if capital_balance > 0 else 0
        am_fee = min(am_fee, amount)

        # Expenses
        remaining = amount - am_fee
        expenses = annual_expenses * years_frac
        expenses = min(expenses, max(0, remaining))

        # Available after fees
        available = amount - am_fee - expenses

        # Pref accrual
        pref_accrued = pref_balance + capital_balance * hurdle_rate * years_frac
        pref_paid = min(pref_accrued, available) if available > 0 else 0

        # Capital return (for Cap events only)
        cap_returned = 0
        if ev_type == "Cap" and available > pref_paid:
            cap_returned = min(capital_balance, available - pref_paid)

        # Excess
        excess = available - pref_paid - cap_returned

        # Promote on excess
        promote = excess * promote_pct if excess > 0 else 0

        # Net to investors
        net_to_investors = pref_paid + cap_returned + excess - promote

        # Update balances
        capital_balance = max(0, capital_balance - cap_returned)
        pref_balance = pref_accrued - pref_paid

        # Allocate net to investors by ownership
        for inv_info in capital_investors:
            share = net_to_investors * inv_info["norm_pct"]
            net_cashflows[inv_info["investor_id"]].append((ev_date, share))
        total_net_cashflows.append((ev_date, net_to_investors))

        waterfall_detail.append({
            "Date": str(ev_date),
            "Event": f"{ev_type} Distribution",
            "Type": ev_type,
            "Gross": amount,
            "AM Fee": round(am_fee),
            "Expenses": round(expenses),
            "Pref Accrued": round(pref_accrued),
            "Pref Paid": round(pref_paid),
            "Capital Returned": round(cap_returned),
            "Excess": round(excess),
            "Promote": round(promote),
            "Net Available": round(net_to_investors),
            "Capital Balance": round(capital_balance),
            "Pref Balance": round(pref_balance),
        })
        last_date = ev_date

    # Build partner returns
    partner_returns = []
    for inv_info in capital_investors:
        iid = inv_info["investor_id"]
        cfs = net_cashflows[iid]
        if not cfs:
            continue
        contributions = sum(a for _, a in cfs if a < 0)
        distributions = sum(a for _, a in cfs if a > 0)
        irr_val = xirr(cfs) if len(cfs) >= 2 else None

        # CF distributions for ROE (only CF-type distributions)
        cf_dists = []
        cap_dists_total = 0.0
        for ev in waterfall_detail:
            ev_date_str = ev.get("Date", "")
            try:
                ev_d = datetime.strptime(ev_date_str, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue
            inv_share = ev.get("Net Available", 0) * inv_info["norm_pct"]
            if inv_share <= 0:
                continue
            if "CF" in ev.get("Type", ""):
                cf_dists.append((ev_d, inv_share))
            else:
                cap_dists_total += inv_share

        roe_val = calculate_roe(
            cfs,
            cf_dists,
            cfs[0][0] if cfs else date.today(),
            cfs[-1][0] if cfs else date.today(),
        ) if cfs else 0.0

        moic_val = distributions / abs(contributions) if contributions < 0 else 0.0

        cashflows_serial = [{"date": str(d), "amount": float(a)} for d, a in cfs]

        partner_returns.append({
            "partner": iid,
            "ownership_pct": inv_info["ownership_pct"],
            "contributions": contributions,
            "cf_distributions": sum(a for _, a in cf_dists),
            "cap_distributions": cap_dists_total,
            "total_distributions": distributions,
            "irr": irr_val,
            "roe": roe_val,
            "moic": moic_val,
            "combined_cashflows": cashflows_serial,
        })

    # Entity-level totals
    total_contribs = sum(a for _, a in total_net_cashflows if a < 0)
    total_dists = sum(a for _, a in total_net_cashflows if a > 0)
    entity_irr = xirr(total_net_cashflows) if len(total_net_cashflows) >= 2 else None
    entity_moic = total_dists / abs(total_contribs) if total_contribs < 0 else 0.0

    deal_returns_list, pref_eq_summary = _build_deal_returns(deal_vcodes, deal_results, inv)

    return {
        "entity_id": entity_id,
        "entity_name": _get_entity_name(entity_id, relationships_raw),
        "mode": "proposed",
        "assumptions": assumptions,
        "partner_results": partner_returns,
        "waterfall_detail": waterfall_detail,
        "deal_summary": {
            "deal_irr": entity_irr,
            "deal_moic": entity_moic,
            "deal_roe": None,
            "total_contributions": total_contribs,
            "total_distributions": total_dists,
        },
        "deal_returns": deal_returns_list,
        "pref_equity_summary": pref_eq_summary,
        "deal_info": _build_deal_info(deal_vcodes, deal_results, inv),
        "investors": [{"investor_id": i["investor_id"],
                       "ownership_pct": i["ownership_pct"]}
                      for i in capital_investors],
        "deals_computed": len(deal_results),
    }


def _build_entity_results(entity_id, cf_upstream_alloc, cap_upstream_alloc,
                          cf_entity_states, cap_entity_states,
                          deal_results, deal_vcodes, inv,
                          relationships_raw, mode, errors,
                          acct=None, wf_raw=None):
    """Build structured results from upstream waterfall outputs."""
    # Get investors in this entity
    investors = get_entity_investors(entity_id, relationships_raw)
    investor_ids = [i["investor_id"] for i in investors if i["ownership_pct"] > 0]

    # Also include 0% investors (fee/promote recipients) that appear in entity states
    all_recipient_ids = set(investor_ids)
    # Check waterfall PropCodes for this entity
    for alloc_df in [cf_upstream_alloc, cap_upstream_alloc]:
        if isinstance(alloc_df, pd.DataFrame) and not alloc_df.empty:
            entity_dists = alloc_df[alloc_df["Entity"].astype(str) == entity_id]
            all_recipient_ids |= set(entity_dists["PropCode"].astype(str).unique())
    # Remove expense sinks and entity itself
    all_recipient_ids -= {entity_id}
    all_recipient_ids = {r for r in all_recipient_ids if not r.endswith("_EXP")}

    # Income schedule: cash arriving at entity
    income_rows = []
    for wf_type, alloc_df in [("CF", cf_upstream_alloc), ("Cap", cap_upstream_alloc)]:
        if isinstance(alloc_df, pd.DataFrame) and alloc_df.empty:
            continue
        entity_income = alloc_df[alloc_df["PropCode"].astype(str) == entity_id].copy()
        for _, row in entity_income.iterrows():
            path_str = str(row.get("Path", ""))
            income_rows.append({
                "Date": str(row.get("event_date", "")),
                "Source Entity": str(row.get("Entity", "")),
                "Source Deal": path_str.split("->")[0] if path_str else "",
                "Type": wf_type,
                "vState": str(row.get("vState", "")),
                "Amount": float(row.get("Allocated", 0)),
            })

    # Member allocations: cash leaving entity to investors
    member_alloc_rows = []
    for wf_type, alloc_df in [("CF", cf_upstream_alloc), ("Cap", cap_upstream_alloc)]:
        if isinstance(alloc_df, pd.DataFrame) and alloc_df.empty:
            continue
        entity_dists = alloc_df[alloc_df["Entity"].astype(str) == entity_id].copy()
        for _, row in entity_dists.iterrows():
            # Extract source deal from Path (first segment before "->")
            path_str = str(row.get("Path", ""))
            source_deal = path_str.split("->")[0] if path_str else ""
            member_alloc_rows.append({
                "Date": str(row.get("event_date", "")),
                "Member": str(row.get("PropCode", "")),
                "Type": wf_type,
                "iOrder": int(row.get("iOrder", 0)),
                "vState": str(row.get("vState", "")),
                "FXRate": float(row.get("FXRate", 0)),
                "Amount": float(row.get("Allocated", 0)),
                "source_deal": source_deal,
            })

    # Seed entity-level contributions from accounting data.
    # The upstream waterfall only produces distributions; contributions to
    # entity investors come from accounting (MajorType='Contribution').
    entity_acct_cfs = {}  # investor_id -> [(date, amount)]
    if acct is not None and not acct.empty:
        ent_acct = acct[acct["InvestmentID"].astype(str).str.strip() == entity_id].copy()
        if not ent_acct.empty:
            ent_acct["_date"] = pd.to_datetime(ent_acct["EffectiveDate"], format="mixed", dayfirst=False)
            ent_acct["_amt"] = pd.to_numeric(ent_acct["Amt"], errors="coerce").fillna(0)
            for _, row in ent_acct.iterrows():
                iid = str(row["InvestorID"]).strip()
                d = row["_date"]
                amt = float(row["_amt"])
                if pd.isna(d) or amt == 0:
                    continue
                entity_acct_cfs.setdefault(iid, []).append((d.date(), amt))

    # Include departed investors who have accounting history (e.g. transfers)
    # but are no longer in the current waterfall or relationships
    all_recipient_ids |= set(entity_acct_cfs.keys())
    all_recipient_ids -= {entity_id}
    all_recipient_ids = {r for r in all_recipient_ids if not r.endswith("_EXP")}

    # Partner returns from entity states
    partner_returns = []
    for member in sorted(all_recipient_ids):
        cf_st = cf_entity_states.get(member)
        cap_st = cap_entity_states.get(member)
        state = cf_st or cap_st
        if not state and member not in entity_acct_cfs:
            continue

        combined_cfs = []
        seen = set()

        # Add accounting cashflows (contributions + historical distributions)
        for d, amt in entity_acct_cfs.get(member, []):
            key = (d, amt)
            if key not in seen:
                combined_cfs.append((d, amt))
                seen.add(key)

        # Add projected distributions from waterfall entity states
        for st_obj in [cf_st, cap_st]:
            if st_obj and st_obj.cashflows:
                for cf in st_obj.cashflows:
                    key = (cf[0], cf[1])
                    if key not in seen:
                        combined_cfs.append(cf)
                        seen.add(key)
        combined_cfs.sort(key=lambda x: x[0])

        irr_val = xirr(combined_cfs) if combined_cfs else None
        contributions = sum(amt for _, amt in combined_cfs if amt < 0)
        distributions = sum(amt for _, amt in combined_cfs if amt > 0)

        cf_dists = []
        if cf_st and cf_st.cf_distributions:
            cf_dists = cf_st.cf_distributions

        roe_val = calculate_roe(
            combined_cfs,
            cf_dists,
            combined_cfs[0][0] if combined_cfs else date.today(),
            combined_cfs[-1][0] if combined_cfs else date.today(),
        ) if combined_cfs else 0.0

        moic_val = distributions / abs(contributions) if contributions < 0 else 0.0

        # CF vs Cap distribution totals
        member_alloc_df = pd.DataFrame(member_alloc_rows) if member_alloc_rows else pd.DataFrame()
        cf_dist_total = 0.0
        cap_dist_total = 0.0
        if not member_alloc_df.empty:
            cf_member = member_alloc_df[
                (member_alloc_df["Member"] == member) & (member_alloc_df["Type"] == "CF")
            ]
            cf_dist_total = cf_member["Amount"].sum()
            cap_member = member_alloc_df[
                (member_alloc_df["Member"] == member) & (member_alloc_df["Type"] == "Cap")
            ]
            cap_dist_total = cap_member["Amount"].sum()

        # Get ownership pct
        inv_match = [i for i in investors if i["investor_id"] == member]
        own_pct = inv_match[0]["ownership_pct"] if inv_match else 0

        cashflows_serial = [{"date": str(d), "amount": float(a)} for d, a in combined_cfs]

        partner_returns.append({
            "partner": member,
            "ownership_pct": own_pct,
            "contributions": contributions,
            "cf_distributions": cf_dist_total,
            "cap_distributions": cap_dist_total,
            "total_distributions": distributions,
            "irr": irr_val,
            "roe": roe_val,
            "moic": moic_val,
            "combined_cashflows": cashflows_serial,
        })

    # Entity-level totals
    all_combined = []
    for pr in partner_returns:
        for cf in pr["combined_cashflows"]:
            all_combined.append((cf["date"], cf["amount"]))
    all_combined.sort(key=lambda x: x[0])

    total_contributions = sum(amt for _, amt in all_combined if amt < 0)
    total_distributions = sum(amt for _, amt in all_combined if amt > 0)

    xirr_cfs = []
    for d_str, amt in all_combined:
        try:
            d = datetime.strptime(d_str, "%Y-%m-%d").date() if isinstance(d_str, str) else d_str
            xirr_cfs.append((d, amt))
        except (ValueError, TypeError):
            pass
    deal_irr = xirr(xirr_cfs) if xirr_cfs else None
    deal_moic = total_distributions / abs(total_contributions) if total_contributions < 0 else 0.0

    deal_returns_list, pref_eq_summary = _build_deal_returns(deal_vcodes, deal_results, inv)

    return {
        "entity_id": entity_id,
        "entity_name": _get_entity_name(entity_id, relationships_raw),
        "mode": mode,
        "partner_results": partner_returns,
        "income_schedule": income_rows,
        "member_allocations": member_alloc_rows,
        "allocation_table": _pivot_allocations_by_year(
            member_alloc_rows,
            deal_name_map=_build_deal_name_map(deal_vcodes, inv),
            cf_entity_states=cf_entity_states,
            cap_entity_states=cap_entity_states,
            entity_id=entity_id,
            wf_raw=wf_raw,
        ),
        "deal_summary": {
            "deal_irr": deal_irr,
            "deal_moic": deal_moic,
            "deal_roe": None,
            "total_contributions": total_contributions,
            "total_distributions": total_distributions,
        },
        "deal_returns": deal_returns_list,
        "pref_equity_summary": pref_eq_summary,
        "deal_info": _build_deal_info(deal_vcodes, deal_results, inv),
        "investors": [{"investor_id": i["investor_id"],
                       "ownership_pct": i["ownership_pct"]}
                      for i in investors],
        "deals_computed": len(deal_results),
        "errors": errors,
    }


def _empty_proposed_results(entity_id, capital_investors, deal_vcodes, inv, relationships_raw):
    """Return empty results structure for proposed mode."""
    return {
        "entity_id": entity_id,
        "entity_name": _get_entity_name(entity_id, relationships_raw),
        "mode": "proposed",
        "partner_results": [],
        "waterfall_detail": [],
        "deal_summary": {"deal_irr": None, "deal_moic": 0, "total_contributions": 0, "total_distributions": 0},
        "deal_info": [],
        "investors": [{"investor_id": i["investor_id"], "ownership_pct": i["ownership_pct"]}
                      for i in capital_investors],
        "deals_computed": 0,
        "errors": ["No cashflow events found"],
    }


def _get_entity_name(entity_id: str, relationships_raw: pd.DataFrame) -> str:
    """Get entity display name from relationships."""
    if relationships_raw is None or relationships_raw.empty:
        return entity_id
    matches = relationships_raw[
        relationships_raw["InvestmentID"].astype(str).str.strip() == entity_id
    ]
    if not matches.empty:
        return str(matches.iloc[0].get("Name", entity_id))
    return entity_id


def _build_deal_name_map(deal_vcodes: list[str], inv: pd.DataFrame) -> dict:
    """Build vcode -> display name mapping."""
    name_map = {}
    for vc in deal_vcodes:
        match = inv[inv["vcode"].astype(str).str.strip() == vc]
        if not match.empty:
            name_map[vc] = str(match.iloc[0].get("Investment_Name", vc))
        else:
            name_map[vc] = vc
    return name_map


def _pivot_allocations_by_year(member_alloc_rows: list[dict],
                                deal_name_map: dict = None,
                                cf_entity_states: dict = None,
                                cap_entity_states: dict = None,
                                entity_id: str = None,
                                wf_raw: pd.DataFrame = None) -> dict:
    """Pivot member allocations into yearly columns, matching Deal Analysis format.

    Returns { "years": [...], "cf": { "rows": [...] }, "cap": { ... } }
    Years are unified across CF and Cap for aligned display.
    Each row: { "label": "2 | 90% | Pref | TGAM | 8.0%", "values": { "2026": 100, ... } }
    """
    from collections import OrderedDict

    if deal_name_map is None:
        deal_name_map = {}

    # Build step detail lookup from raw waterfall definitions, keyed per wf_type.
    # Key: (wf_type, iOrder, PropCode) -> {nPercent, mAmount, FXRate (effective share)}
    # A single PropCode at a given iOrder can have multiple rows (e.g. Capital Share
    # FXRate=0.1 + GP Promote FXRate=0.2 for INV23). We collect unique FXRate values
    # and sum them for the effective share.
    # CF_WF steps have vAmtType starting with "6.02" or "EXP";
    # Cap_WF steps have vAmtType starting with "6.03" or "EXP".
    step_detail = {}  # (wf_type, iOrder, PropCode) -> {nPercent, mAmount, fx_values: set}
    if wf_raw is not None and entity_id and not wf_raw.empty:
        ent_wf = wf_raw[wf_raw["vcode"].astype(str).str.strip() == entity_id]

        def _float(v):
            try:
                return float(v) if v is not None and str(v).strip() != "" else None
            except (ValueError, TypeError):
                return None

        for _, s in ent_wf.iterrows():
            order = int(s.get("iOrder", 0))
            pc = str(s.get("PropCode", "")).strip()
            n_pct = _float(s.get("nPercent"))
            m_amt = _float(s.get("mAmount"))
            fx = _float(s.get("FXRate")) or 0.0
            vamt = str(s.get("vAmtType", "")).strip()

            # Determine which wf_type(s) this step belongs to
            wf_types = []
            if vamt.startswith("6.02") or vamt == "EXP":
                wf_types.append("CF")
            if vamt.startswith("6.03") or vamt == "EXP":
                wf_types.append("Cap")
            if not wf_types:
                wf_types = ["CF", "Cap"]  # fallback: include in both

            for wft in wf_types:
                detail_key = (wft, order, pc)
                if detail_key not in step_detail:
                    step_detail[detail_key] = {
                        "nPercent": n_pct, "mAmount": m_amt, "fx_values": {fx},
                    }
                else:
                    step_detail[detail_key]["fx_values"].add(fx)
                    if n_pct is not None:
                        step_detail[detail_key]["nPercent"] = n_pct
                    if m_amt is not None:
                        step_detail[detail_key]["mAmount"] = m_amt

        # Convert fx_values sets to summed FXRate
        for detail in step_detail.values():
            detail["FXRate"] = sum(detail.pop("fx_values"))

    # current_wf_type is set per iteration in the outer loop below
    current_wf_type = "CF"

    def _step_label(order, state, member, fx_from_row):
        """Build descriptive step label with rates and sharing percentages."""
        detail = step_detail.get((current_wf_type, order, member), {})
        # FXRate: use summed value from step definition, fall back to allocation row
        fx = detail.get("FXRate")
        if fx is None or fx == 0:
            fx = fx_from_row
        fx_str = f"{fx * 100:g}%" if fx is not None and fx > 0 else ""
        # nPercent (rate for Pref, Def_Int, etc.)
        n_pct = detail.get("nPercent")
        rate_str = f"{n_pct:g}%" if n_pct is not None and n_pct > 0 else ""
        # mAmount (for Amt/fixed amount steps)
        m_amt = detail.get("mAmount")
        amt_str = f"${m_amt:,.0f}" if m_amt is not None and m_amt > 0 else ""

        parts = [f"{order}"]
        if fx_str:
            parts.append(fx_str)
        parts.append(state)
        parts.append(member)
        if rate_str:
            parts.append(rate_str)
        elif amt_str:
            parts.append(amt_str)
        return "  " + " | ".join(parts)

    def _step_sort_key(label):
        """Sort steps: by iOrder, then leads before tags."""
        # Parse iOrder from label (first segment before |)
        parts = label.strip().split(" | ")
        try:
            order = int(parts[0])
        except (ValueError, IndexError):
            order = 999
        # Determine if this is a Tag row
        is_tag = " | Tag | " in label
        return (order, 1 if is_tag else 0, label)

    # Collect ALL years across both waterfall types for unified columns
    all_years_set = set()
    for r in member_alloc_rows:
        d = r.get("Date", "")
        if d:
            all_years_set.add(int(d[:4]))
    all_years = sorted(all_years_set)

    def _zero_values():
        return {str(y): 0 for y in all_years}

    def _none_values():
        return {str(y): None for y in all_years}

    result = {"years": all_years}
    for wf_type in ("CF", "Cap"):
        current_wf_type = wf_type
        type_rows = [r for r in member_alloc_rows if r.get("Type") == wf_type]
        if not type_rows:
            result[wf_type.lower()] = {"rows": [], "years": all_years}
            continue

        # Build step keys and aggregate by year
        step_totals = OrderedDict()  # step_key -> { year -> amount }
        step_order_map = {}
        for r in type_rows:
            order = r.get("iOrder", 0)
            state = r.get("vState", "")
            member = r.get("Member", "")
            fx_from_row = r.get("FXRate", 0)
            key = _step_label(order, state, member, fx_from_row)
            year = int(r["Date"][:4]) if r.get("Date") else None
            if year is None:
                continue
            if key not in step_totals:
                step_totals[key] = {}
                step_order_map[key] = order
            step_totals[key][year] = step_totals[key].get(year, 0) + r.get("Amount", 0)

        sorted_keys = sorted(step_totals.keys(), key=_step_sort_key)

        # Build partner totals
        partner_totals = {}
        for r in type_rows:
            member = r.get("Member", "")
            year = int(r["Date"][:4]) if r.get("Date") else None
            if year is None:
                continue
            partner_totals.setdefault(member, {})[year] = \
                partner_totals.get(member, {}).get(year, 0) + r.get("Amount", 0)

        rows = []

        # Capital event sources (Cap only): show which deals produced events
        if wf_type == "Cap" and deal_name_map:
            # Group source deals by year
            deal_by_year = {}  # year -> set of deal names
            deal_amt_by_year = {}  # year -> { deal_name: total_amount }
            for r in type_rows:
                sd = r.get("source_deal", "")
                if not sd:
                    continue
                year = int(r["Date"][:4]) if r.get("Date") else None
                if year is None:
                    continue
                dname = deal_name_map.get(sd, sd)
                deal_by_year.setdefault(year, set()).add(dname)
                deal_amt_by_year.setdefault(year, {})
                deal_amt_by_year[year][dname] = \
                    deal_amt_by_year[year].get(dname, 0) + r.get("Amount", 0)

            # Add source deals header row
            rows.append({"label": "Capital Event Sources:", "values": _none_values()})
            # Collect all deal names across years
            all_deal_names = sorted({
                n for names in deal_by_year.values() for n in names
            })
            for dname in all_deal_names:
                values = {}
                for y in all_years:
                    amt = deal_amt_by_year.get(y, {}).get(dname, 0)
                    values[str(y)] = amt
                rows.append({"label": f"  {dname}", "values": values})
            rows.append({"label": "", "values": _none_values()})

        # Step rows — with 0 placeholders for all unified years
        for key in sorted_keys:
            values = {}
            for y in all_years:
                values[str(y)] = step_totals[key].get(y, 0)
            rows.append({"label": key, "values": values})

        # Blank separator
        rows.append({"label": "", "values": _none_values()})

        # Partner totals section
        rows.append({"label": "Partner Totals:", "values": _none_values()})
        grand_total = {y: 0.0 for y in all_years}
        for member in sorted(partner_totals.keys()):
            values = {}
            for y in all_years:
                val = partner_totals[member].get(y, 0)
                values[str(y)] = val
                grand_total[y] += val
            rows.append({"label": f"  {member} Total", "values": values})

        rows.append({
            "label": "  Total Distributions",
            "values": {str(y): grand_total[y] for y in all_years},
        })

        # Entity investor state diagnostics (why steps 1-N are zero)
        entity_states = cap_entity_states if wf_type == "Cap" else cf_entity_states
        if entity_states:
            rows.append({"label": "", "values": _none_values()})
            rows.append({"label": "Entity Investor Balances (End of Projection):", "values": _none_values()})
            for member in sorted(partner_totals.keys()):
                if member.endswith("_EXP"):
                    continue  # skip expense sinks
                st = entity_states.get(member)
                if not st:
                    continue
                cap_out = st.capital_outstanding
                pref_unpaid = st.pref_unpaid_compounded + st.pref_accrued_current_year
                rows.append({
                    "label": f"  {member} Capital Outstanding",
                    "values": _none_values(),
                    "end_value": round(cap_out),
                })
                rows.append({
                    "label": f"  {member} Accrued Pref",
                    "values": _none_values(),
                    "end_value": round(pref_unpaid),
                })
            # Show note about why steps are zero
            if wf_type == "Cap":
                rows.append({
                    "label": "  Note: Steps with $0 indicate pref/capital was fully satisfied by prior CF distributions",
                    "values": _none_values(),
                })

        result[wf_type.lower()] = {"rows": rows, "years": all_years}

    return result


def _build_deal_info(deal_vcodes, deal_results, inv):
    """Build deal info summary for display."""
    deal_info = []
    for vc in deal_vcodes:
        match = inv[inv["vcode"].astype(str).str.strip() == vc]
        info = {"vcode": vc}
        if not match.empty:
            r = match.iloc[0]
            info["name"] = str(r.get("Investment_Name", vc))
            info["asset_type"] = str(r.get("Asset_Type", ""))
            info["sale_status"] = str(r.get("Sale_Status", "") or "Active")
        info["computed"] = vc in deal_results
        deal_info.append(info)
    return deal_info


def _build_deal_returns(deal_vcodes, deal_results, inv):
    """Build per-deal pref equity returns for the summary table.

    Shows the preferred equity partner's returns for each deal, since upstream
    investors only participate in the pref equity component. Also computes
    a portfolio-level aggregate of all pref equity cashflows.

    Returns (deal_returns_list, pref_equity_summary).
    """
    deal_returns = []
    all_pref_cfs = []  # for portfolio-level IRR
    total_pref_contribs = 0.0
    total_pref_cf_dist = 0.0
    total_pref_cap_dist = 0.0

    for vc in deal_vcodes:
        match = inv[inv["vcode"].astype(str).str.strip() == vc]
        name = str(match.iloc[0].get("Investment_Name", vc)) if not match.empty else vc
        asset_type = str(match.iloc[0].get("Asset_Type", "")) if not match.empty else ""

        if vc not in deal_results:
            deal_returns.append({
                "vcode": vc, "name": name, "asset_type": asset_type,
                "computed": False, "pref_partner": "",
                "contributions": 0, "distributions": 0,
                "cf_distributions": 0, "cap_distributions": 0,
                "irr": None, "roe": None, "moic": 0,
            })
            continue

        partner_results = deal_results[vc].get("partner_results", [])
        pref_partners = [pr for pr in partner_results if pr.get("is_pref_equity", False)]

        if pref_partners:
            contribs = sum(pr["contributions"] for pr in pref_partners)
            cf_dist = sum(pr["cf_distributions"] for pr in pref_partners)
            cap_dist = sum(pr["cap_distributions"] for pr in pref_partners)
            total_dist = sum(pr["total_distributions"] for pr in pref_partners)

            # Aggregate pref cashflows for portfolio-level metrics
            for pr in pref_partners:
                all_pref_cfs.extend(pr.get("combined_cashflows", []))
            total_pref_contribs += contribs
            total_pref_cf_dist += cf_dist
            total_pref_cap_dist += cap_dist

            if len(pref_partners) == 1:
                pr = pref_partners[0]
                irr_val = pr["irr"]
                roe_val = pr.get("roe")
                moic_val = pr["moic"]
                pref_name = pr["partner"]
            else:
                # Multiple pref partners — combine cashflows for IRR
                deal_pref_cfs = []
                for pr in pref_partners:
                    deal_pref_cfs.extend(pr.get("combined_cashflows", []))
                deal_pref_cfs.sort(key=lambda x: x[0])
                irr_val = xirr(deal_pref_cfs) if len(deal_pref_cfs) >= 2 else None
                moic_val = total_dist / abs(contribs) if contribs else 0
                roe_val = None
                pref_name = ", ".join(pr["partner"] for pr in pref_partners)

            deal_returns.append({
                "vcode": vc, "name": name, "asset_type": asset_type,
                "computed": True, "pref_partner": pref_name,
                "contributions": contribs,
                "distributions": total_dist,
                "cf_distributions": cf_dist,
                "cap_distributions": cap_dist,
                "irr": irr_val, "roe": roe_val, "moic": moic_val,
            })
        else:
            # No pref equity partner found — use deal summary as fallback
            ds = deal_results[vc].get("deal_summary", {})
            deal_returns.append({
                "vcode": vc, "name": name, "asset_type": asset_type,
                "computed": True, "pref_partner": "",
                "contributions": ds.get("total_contributions", 0),
                "distributions": ds.get("total_distributions", 0),
                "cf_distributions": ds.get("total_cf_distributions", 0),
                "cap_distributions": ds.get("total_cap_distributions", 0),
                "irr": ds.get("deal_irr"), "roe": ds.get("deal_roe"),
                "moic": ds.get("deal_moic", 0),
            })

    # Portfolio-level pref equity aggregate
    all_pref_cfs.sort(key=lambda x: x[0])
    total_pref_dist = total_pref_cf_dist + total_pref_cap_dist
    pref_summary = {
        "total_contributions": total_pref_contribs,
        "total_distributions": total_pref_dist,
        "cf_distributions": total_pref_cf_dist,
        "cap_distributions": total_pref_cap_dist,
        "irr": xirr(all_pref_cfs) if len(all_pref_cfs) >= 2 else None,
        "moic": total_pref_dist / abs(total_pref_contribs) if total_pref_contribs else 0,
    }

    return deal_returns, pref_summary


def get_deal_detail(vcode, data, start_year, horizon_years, pro_yr_base):
    """Get detailed partner results for a specific deal (for drill-down)."""
    from flask_app.services.compute_service import get_cached_deal_result

    result = get_cached_deal_result(
        vcode=vcode,
        start_year=start_year,
        horizon_years=horizon_years,
        pro_yr_base=pro_yr_base,
        data=data,
        actuals_through=data.get("actuals_through"),
    )
    if "error" in result:
        return {"error": result["error"]}

    partner_results = result.get("partner_results", [])
    deal_summary = result.get("deal_summary", {})

    inv = data["inv"]
    match = inv[inv["vcode"].astype(str).str.strip() == vcode]
    name = str(match.iloc[0].get("Investment_Name", vcode)) if not match.empty else vcode

    serialized = []
    for pr in partner_results:
        serialized.append({
            "partner": pr["partner"],
            "is_pref_equity": pr.get("is_pref_equity", False),
            "contributions": pr["contributions"],
            "cf_distributions": pr["cf_distributions"],
            "cap_distributions": pr["cap_distributions"],
            "total_distributions": pr["total_distributions"],
            "irr": pr["irr"],
            "roe": pr.get("roe"),
            "moic": pr["moic"],
            "cashflow_details": [
                {"date": str(d["Date"]), "description": d.get("Description", ""),
                 "amount": d.get("Amount", 0)}
                for d in pr.get("cashflow_details", [])
            ],
        })

    return {
        "vcode": vcode,
        "name": name,
        "partner_results": serialized,
        "deal_summary": {
            "total_contributions": deal_summary.get("total_contributions", 0),
            "total_distributions": deal_summary.get("total_distributions", 0),
            "cf_distributions": deal_summary.get("total_cf_distributions", 0),
            "cap_distributions": deal_summary.get("total_cap_distributions", 0),
            "deal_irr": deal_summary.get("deal_irr"),
            "deal_roe": deal_summary.get("deal_roe"),
            "deal_moic": deal_summary.get("deal_moic", 0),
        },
    }


def generate_portfolio_excel(results: dict) -> bytes:
    """Generate Excel workbook for portfolio analysis results.

    Sheets:
    1. Deal Returns — per-deal summary (like Sold Portfolio summary)
    2. Partner Returns — entity-level partner splits
    3. Entity Allocations — CF_WF + Cap_WF member allocations (actual mode)
    4. Waterfall Detail — proposed waterfall simulation (proposed mode)
    5. XIRR Cash Flows — per-partner combined cashflows
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    bold_font = Font(bold=True)
    top_border = Border(top=Side(style="medium"))
    CUR = "$#,##0"
    PCT = "0.00%"
    MULT = '0.00"x"'

    entity_name = results.get("entity_name", results.get("entity_id", ""))
    mode = results.get("mode", "actual")
    deal_summary = results.get("deal_summary", {})

    def _write_headers(ws, headers, row=1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")

    def _autosize(ws, headers, min_width=12):
        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = max(
                len(headers[ci - 1]) + 4, min_width)

    # Sheet 1: Deal Returns Summary
    ws = wb.active
    ws.title = "Deal Returns"
    dr_headers = ["Investment Name", "Asset Type", "Contributions",
                  "CF Distributions", "Cap Distributions", "Total Distributions",
                  "IRR", "ROE", "MOIC"]
    _write_headers(ws, dr_headers)
    ws.cell(row=1, column=11, value=f"{entity_name} ({mode.title()})").font = bold_font

    deal_returns = results.get("deal_returns", [])
    ri = 2
    for dr in deal_returns:
        ws.cell(row=ri, column=1, value=dr.get("name", dr.get("vcode", "")))
        ws.cell(row=ri, column=2, value=dr.get("asset_type", ""))
        ws.cell(row=ri, column=3, value=dr["contributions"]).number_format = CUR
        ws.cell(row=ri, column=4, value=dr.get("cf_distributions", 0)).number_format = CUR
        ws.cell(row=ri, column=5, value=dr.get("cap_distributions", 0)).number_format = CUR
        ws.cell(row=ri, column=6, value=dr["distributions"]).number_format = CUR
        if dr.get("irr") is not None:
            ws.cell(row=ri, column=7, value=dr["irr"]).number_format = PCT
        if dr.get("roe") is not None:
            ws.cell(row=ri, column=8, value=dr["roe"]).number_format = PCT
        ws.cell(row=ri, column=9, value=dr.get("moic", 0)).number_format = MULT
        ri += 1

    # Portfolio total row
    for ci in range(1, 10):
        ws.cell(row=ri, column=ci).font = bold_font
        ws.cell(row=ri, column=ci).border = top_border
    ws.cell(row=ri, column=1, value=f"{entity_name} Total")
    ws.cell(row=ri, column=3, value=deal_summary.get("total_contributions", 0)).number_format = CUR
    ws.cell(row=ri, column=6, value=deal_summary.get("total_distributions", 0)).number_format = CUR
    if deal_summary.get("deal_irr") is not None:
        ws.cell(row=ri, column=7, value=deal_summary["deal_irr"]).number_format = PCT
    if deal_summary.get("deal_roe") is not None:
        ws.cell(row=ri, column=8, value=deal_summary["deal_roe"]).number_format = PCT
    ws.cell(row=ri, column=9, value=deal_summary.get("deal_moic", 0)).number_format = MULT
    _autosize(ws, dr_headers, 14)

    # Sheet 2: Partner Returns
    partner_results = results.get("partner_results", [])
    ws2 = wb.create_sheet("Partner Returns")
    pr_headers = ["Partner", "Ownership %", "Contributions", "CF Distributions",
                  "Cap Distributions", "Total Distributions", "IRR", "ROE", "MOIC"]
    _write_headers(ws2, pr_headers)
    ri = 2
    for pr in partner_results:
        ws2.cell(row=ri, column=1, value=pr["partner"])
        own_pct = pr.get("ownership_pct", 0)
        ws2.cell(row=ri, column=2, value=own_pct / 100.0 if own_pct > 1 else own_pct).number_format = PCT
        ws2.cell(row=ri, column=3, value=pr["contributions"]).number_format = CUR
        ws2.cell(row=ri, column=4, value=pr.get("cf_distributions", 0)).number_format = CUR
        ws2.cell(row=ri, column=5, value=pr.get("cap_distributions", 0)).number_format = CUR
        ws2.cell(row=ri, column=6, value=pr["total_distributions"]).number_format = CUR
        if pr.get("irr") is not None:
            ws2.cell(row=ri, column=7, value=pr["irr"]).number_format = PCT
        if pr.get("roe") is not None:
            ws2.cell(row=ri, column=8, value=pr["roe"]).number_format = PCT
        ws2.cell(row=ri, column=9, value=pr.get("moic", 0)).number_format = MULT
        ri += 1
    # Total row
    for ci in range(1, 10):
        ws2.cell(row=ri, column=ci).font = bold_font
        ws2.cell(row=ri, column=ci).border = top_border
    ws2.cell(row=ri, column=1, value=f"{entity_name} Total")
    ws2.cell(row=ri, column=3, value=deal_summary.get("total_contributions", 0)).number_format = CUR
    ws2.cell(row=ri, column=6, value=deal_summary.get("total_distributions", 0)).number_format = CUR
    if deal_summary.get("deal_irr") is not None:
        ws2.cell(row=ri, column=7, value=deal_summary["deal_irr"]).number_format = PCT
    ws2.cell(row=ri, column=9, value=deal_summary.get("deal_moic", 0)).number_format = MULT
    _autosize(ws2, pr_headers, 14)

    # Sheet 3: Entity Allocations (actual mode — CF + Cap member allocations)
    member_allocs = results.get("member_allocations", [])
    if member_allocs:
        ws3 = wb.create_sheet("Entity Allocations")
        alloc_headers = ["Date", "Type", "Order", "Member", "vState", "FXRate", "Amount"]
        _write_headers(ws3, alloc_headers)
        ri = 2
        for row in sorted(member_allocs, key=lambda r: (r.get("Date", ""), r.get("Type", ""), r.get("iOrder", 0))):
            ws3.cell(row=ri, column=1, value=row.get("Date", ""))
            ws3.cell(row=ri, column=2, value=row.get("Type", ""))
            ws3.cell(row=ri, column=3, value=row.get("iOrder", 0))
            ws3.cell(row=ri, column=4, value=row.get("Member", ""))
            ws3.cell(row=ri, column=5, value=row.get("vState", ""))
            ws3.cell(row=ri, column=6, value=row.get("FXRate", 0))
            ws3.cell(row=ri, column=7, value=row.get("Amount", 0)).number_format = CUR
            ri += 1
        _autosize(ws3, alloc_headers, 12)

    # Sheet 4: Waterfall Detail (proposed mode)
    wf_detail = results.get("waterfall_detail", [])
    if wf_detail:
        ws4 = wb.create_sheet("Waterfall Detail")
        wf_headers = list(wf_detail[0].keys())
        _write_headers(ws4, wf_headers)
        for ri, row in enumerate(wf_detail, 2):
            for ci, h in enumerate(wf_headers, 1):
                val = row.get(h, "")
                cell = ws4.cell(row=ri, column=ci, value=val)
                if isinstance(val, (int, float)) and h != "Date":
                    cell.number_format = CUR

    # Sheet 5: XIRR Cash Flows
    ws5 = wb.create_sheet("XIRR Cash Flows")
    cf_headers = ["Date", "Partner", "Amount"]
    _write_headers(ws5, cf_headers)
    ri = 2
    for pr in partner_results:
        for cf in pr.get("combined_cashflows", []):
            ws5.cell(row=ri, column=1, value=cf["date"])
            ws5.cell(row=ri, column=2, value=pr["partner"])
            ws5.cell(row=ri, column=3, value=cf["amount"]).number_format = CUR
            ri += 1
    _autosize(ws5, cf_headers, 14)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
