"""
prospect_excel.py
Audit workbook for a Prospect Deal Analysis.

Built to be handed to a third party: every tab shows the inputs or the
supporting calculation behind the results, and the partner cash flow tab
carries live =XIRR formulas so an auditor's Excel recomputes the returns
from the same cash flows the model used.
"""

import logging
from io import BytesIO
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def _fmt_cell(ws, row, col, value, styles, kind):
    c = ws.cell(row=row, column=col, value=value)
    if kind == "curr":
        c.number_format = styles["curr"]
    elif kind == "pct":
        c.number_format = styles["pct"]
    elif kind == "num":
        c.number_format = styles["num"]
    return c


def _section(ws, row, title, styles):
    c = ws.cell(row=row, column=1, value=title)
    c.font = styles["bold"]
    return row + 1


def _waterfall_narrative(budget: Dict[str, Any],
                         wf_steps: List[Dict[str, Any]]) -> List[str]:
    """The deal's economics in the words a term sheet would use.

    Built from the declared capital split and the saved waterfall steps:
    prefs, full-pool IRR lookbacks, gated tiers ("X% until a Y% IRR, partner
    tagging the rest"), and the terminal split.
    """
    lines: List[str] = []
    if budget and budget.get("equity"):
        lines.append(
            f"PSC Equity — ${budget['pe_amount']:,.0f} ({budget['pe_pct']:.0%});  "
            f"OP Equity — ${budget['op_amount']:,.0f} ({1 - budget['pe_pct']:.0%})"
        )

    cf = [st for st in (wf_steps or []) if st.get("vmisc") == "CF_WF"]
    cap = [st for st in (wf_steps or []) if st.get("vmisc") == "Cap_WF"]

    prefs = sorted({float(st.get("nPercent") or 0) for st in cf
                    if str(st.get("vState")) == "Pref"})
    if prefs:
        split = " / ".join(
            f"{float(st.get('FXRate') or 0):.0%} {st.get('PropCode')}"
            for st in cf if str(st.get("vState")) in ("Share", "Tag"))
        lines.append(
            f"Operating cash: {', '.join(f'{p:g}%' for p in prefs)} preferred "
            f"return, excess split {split}"
        )

    # Capital waterfall: group by iOrder (Tie #) to read the tiers in sequence
    by_order: Dict[int, List[Dict]] = {}
    for st in cap:
        by_order.setdefault(int(st.get("iOrder") or 0), []).append(st)
    for order in sorted(by_order):
        group = by_order[order]
        irr = [st for st in group if str(st.get("vState")) == "IRR"]
        tags = [st for st in group if str(st.get("vState")) == "Tag"]
        shares = [st for st in group if str(st.get("vState")) == "Share"]
        if irr and not tags:
            g = irr[0]
            lines.append(
                f"Capital events: {g.get('PropCode')} to a "
                f"{float(g.get('nPercent') or 0):g}% IRR lookback"
            )
        elif irr and tags:
            g = irr[0]
            lines.append(
                f"then {float(g.get('FXRate') or 0):.0%} to {g.get('PropCode')} "
                f"until a {float(g.get('nPercent') or 0):g}% IRR ("
                + ", ".join(f"{float(t.get('FXRate') or 0):.0%} {t.get('PropCode')}"
                            for t in tags) + " alongside)"
            )
        elif shares:
            lines.append(
                "thereafter "
                + " / ".join(
                    f"{float(st.get('FXRate') or 0):.0%} {st.get('PropCode')}"
                    for st in shares + tags)
            )
    return lines


def generate_prospect_analysis_excel(
    result: Dict[str, Any],
    deal: Dict[str, Any],
    assumptions: Dict[str, Any],
    wf_steps: Optional[List[Dict[str, Any]]] = None,
    scenario: Optional[Dict[str, Any]] = None,
    annual_forecast: Optional[Dict[str, Any]] = None,
    ppi: Optional[Dict[str, Any]] = None,
) -> bytes:
    """Build the audit workbook. Returns xlsx bytes.

    annual_forecast is the same anniversary-year table the app renders
    (from _continue_analyze), so the sheet ties to the screen by construction.
    """
    import json
    import pandas as pd
    from openpyxl import Workbook
    from flask_app.services.compute_service import (
        _excel_styles, _write_header_row, _autosize_columns,
    )

    s = _excel_styles()
    wb = Workbook()
    budget = result.get("capital_budget") or {}
    forecast_source = result.get("forecast_source") or ""
    src_label = {"argus": "Argus projection", "cashflows": "Uploaded cash flows",
                 "noi_growth": "NOI growth assumptions"}.get(forecast_source,
                                                             forecast_source)
    sd = result.get("sale_dbg") or {}

    # ------------------------------------------------------------------
    # 1. Summary
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    r = 1
    ws.cell(row=r, column=1, value=f"Deal Analysis — {deal.get('deal_name', '')}").font = s["bold"]
    r += 1
    for label, val in [
        ("vCode", deal.get("vcode")),
        ("Stage", deal.get("stage")),
        ("Scenario", (scenario or {}).get("name") or "Live assumptions"),
        ("Cash Flow Source", src_label),
        ("Generated", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")),
    ]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1

    r = _section(ws, r, "Partner Returns", s)
    _write_header_row(ws, r, ["Partner", "Contributions", "CF Distributions",
                              "Capital Distributions", "Total Distributions",
                              "IRR", "ROE", "MOIC"], s)
    r += 1
    for p in (result.get("partner_results") or []):
        ws.cell(row=r, column=1, value=p.get("partner"))
        _fmt_cell(ws, r, 2, p.get("contributions"), s, "curr")
        _fmt_cell(ws, r, 3, p.get("cf_distributions"), s, "curr")
        _fmt_cell(ws, r, 4, p.get("cap_distributions"), s, "curr")
        _fmt_cell(ws, r, 5, p.get("total_distributions"), s, "curr")
        _fmt_cell(ws, r, 6, p.get("irr"), s, "pct")
        _fmt_cell(ws, r, 7, p.get("roe"), s, "pct")
        c = ws.cell(row=r, column=8, value=p.get("moic"))
        c.number_format = s["mult"]
        r += 1
    ds = result.get("deal_summary") or {}
    ws.cell(row=r, column=1, value="Deal Total").font = s["bold"]
    _fmt_cell(ws, r, 2, ds.get("total_contributions"), s, "curr").font = s["bold"]
    _fmt_cell(ws, r, 3, ds.get("total_cf_distributions"), s, "curr").font = s["bold"]
    _fmt_cell(ws, r, 4, ds.get("total_cap_distributions"), s, "curr").font = s["bold"]
    _fmt_cell(ws, r, 5, ds.get("total_distributions"), s, "curr").font = s["bold"]
    _fmt_cell(ws, r, 6, ds.get("deal_irr"), s, "pct").font = s["bold"]
    _fmt_cell(ws, r, 7, ds.get("deal_roe"), s, "pct").font = s["bold"]
    c = ws.cell(row=r, column=8, value=ds.get("deal_moic"))
    c.number_format = s["mult"]
    c.font = s["bold"]
    _autosize_columns(ws)

    # ------------------------------------------------------------------
    # 2. Assumptions — only inputs the projection actually uses
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Assumptions")
    r = 1

    r = _section(ws, r, "Debt", s)
    for key, label, kind in [
        ("debt_amount", "Debt Amount", "curr"),
        ("debt_rate", "Interest Rate", "pct"),
        ("debt_term_months", "Term (months)", "num"),
        ("io_months", "IO (months)", "num"),
        ("amort_months", "Amortization (months)", "num"),
        ("lender", "Lender", "text"),
        ("rate_type", "Rate Type", "text"),
    ]:
        val = (assumptions or {}).get(key)
        if val in (None, ""):
            continue
        ws.cell(row=r, column=1, value=label)
        _fmt_cell(ws, r, 2, val, s, kind)
        r += 1
    r += 1

    # -- Equity & Waterfall: the arrangement in words --
    r = _section(ws, r, "Equity & Waterfall", s)
    narrative = _waterfall_narrative(budget, wf_steps or [])
    if narrative:
        for line in narrative:
            ws.cell(row=r, column=1, value=line)
            r += 1
    else:
        ws.cell(row=r, column=1, value="No waterfall steps saved")
        r += 1
    r += 1

    # -- Operations & Exit: skip growth-model inputs when a real projection
    #    (Argus / uploaded cash flows) drives the forecast --
    r = _section(ws, r, "Operations & Exit", s)
    ws.cell(row=r, column=1, value="Cash Flow Source")
    ws.cell(row=r, column=2, value=src_label)
    r += 1
    ops_fields = [
        ("hold_years", "Hold (years)", "num"),
        ("mgmt_fee_pct", "Management Fee %", "pct"),
        ("replacement_reserve_psf", "Replacement Reserve /SF", "num"),
    ]
    if forecast_source == "noi_growth":
        ops_fields = [("noi_year1", "NOI Year 1", "curr"),
                      ("noi_growth_rate", "NOI Growth", "pct"),
                      ("capex_reserve_psf", "CapEx Reserve /SF", "num")] + ops_fields
    for key, label, kind in ops_fields:
        val = (assumptions or {}).get(key)
        if val in (None, "", 0):
            continue
        ws.cell(row=r, column=1, value=label)
        _fmt_cell(ws, r, 2, val, s, kind)
        r += 1
    # The exit, end to end: cap, cost, terminal NOI, and the proceeds math
    ws.cell(row=r, column=1, value="Exit Cap Rate")
    _fmt_cell(ws, r, 2, sd.get("CapRate_Sale"), s, "pct")
    r += 1
    ws.cell(row=r, column=1, value="Cost of Sale %")
    _fmt_cell(ws, r, 2, (assumptions or {}).get("selling_cost_pct"), s, "pct")
    r += 1
    for label, key, bold in [
        ("Terminal NOI (fwd 12 mo)", "NOI_12m_After_Sale", False),
        ("Sale Price (NOI / Cap)", "Implied_Value", False),
        ("Less: Cost of Sale", "Selling_Cost_Amount", False),
        ("Less: Loan Payoff", "Less_Loan_Balances", False),
        ("Net Sale Proceeds", "Net_Sale_Proceeds", True),
    ]:
        ws.cell(row=r, column=1, value=label)
        _fmt_cell(ws, r, 2, sd.get(key), s, "curr")
        if bold:
            ws.cell(row=r, column=1).font = s["bold"]
            ws.cell(row=r, column=2).font = s["bold"]
        r += 1
    r += 1

    def _json(blob):
        if not blob:
            return None
        try:
            return json.loads(blob) if isinstance(blob, str) else blob
        except (TypeError, ValueError):
            return None

    refi = _json((assumptions or {}).get("planned_refi_json"))
    if refi and refi.get("enabled"):
        r = _section(ws, r, "Planned Refinancing", s)
        for label, key, kind in [
            ("Refi Date", "refi_date", "text"), ("New Loan Amount", "loan_amount", "curr"),
            ("Rate", "rate", "pct"), ("Term (years)", "term_years", "num"),
            ("Amort (years)", "amort_years", "num"), ("IO (years)", "io_years", "num"),
            ("Closing Costs", "closing_costs", "curr"), ("Reserve Holdback", "holdback", "curr"),
        ]:
            if refi.get(key) in (None, "", 0) and key != "refi_date":
                continue
            ws.cell(row=r, column=1, value=label)
            _fmt_cell(ws, r, 2, refi.get(key), s, kind)
            r += 1
        r += 1

    if scenario:
        r = _section(ws, r, f"Scenario — {scenario.get('name')}", s)
        for k, v in (scenario.get("assumption_overrides") or {}).items():
            ws.cell(row=r, column=1, value=f"Override: {k}")
            ws.cell(row=r, column=2, value=v)
            r += 1
        for adj in (scenario.get("adjustments") or []):
            span = f"from {adj.get('start_date')}" + (f" to {adj.get('end_date')}" if adj.get("end_date") else "")
            rev = sum(float(v or 0) for v in (adj.get("revenue") or {}).values())
            exp = sum(float(v or 0) for v in (adj.get("expense") or {}).values())
            ws.cell(row=r, column=1, value=f"Adjustment: {adj.get('label')} ({span})")
            ws.cell(row=r, column=2, value=f"revenue {rev:+,.0f}/yr, expense {exp:+,.0f}/yr")
            r += 1
    _autosize_columns(ws)

    # ------------------------------------------------------------------
    # 3. Capital Budget — the app's Sources & Uses, tying to the penny
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Capital Budget")
    if budget.get("total_uses"):
        r = _section(ws, 1, "Uses", s)
        for _id, label, amt in (budget.get("uses") or []):
            ws.cell(row=r, column=1, value=label)
            _fmt_cell(ws, r, 2, amt, s, "curr")
            r += 1
        ws.cell(row=r, column=1, value="Total Uses").font = s["bold"]
        _fmt_cell(ws, r, 2, budget.get("total_uses"), s, "curr").font = s["bold"]
        r += 2

        r = _section(ws, r, "Sources", s)
        for d in (budget.get("debt_rows") or []):
            amt = float(d.get("amount") or 0)
            if not amt:
                continue
            label = d.get("label") or d.get("id")
            if d.get("rate"):
                label += f"  ({float(d['rate']) * 100:.2f}%, own loan)"
            ws.cell(row=r, column=1, value=label)
            _fmt_cell(ws, r, 2, amt, s, "curr")
            r += 1
        pe_pct = budget.get("pe_pct") or 0
        ws.cell(row=r, column=1, value=f"PSC Preferred Equity ({pe_pct:.0%})")
        _fmt_cell(ws, r, 2, budget.get("pe_amount"), s, "curr")
        r += 1
        ws.cell(row=r, column=1, value=f"OP Equity ({1 - pe_pct:.0%})")
        _fmt_cell(ws, r, 2, budget.get("op_amount"), s, "curr")
        r += 1
        for e_row in (budget.get("extra_equity_rows") or []):
            amt = float(e_row.get("amount") or 0)
            if not amt:
                continue
            ws.cell(row=r, column=1, value=e_row.get("label") or e_row.get("id"))
            _fmt_cell(ws, r, 2, amt, s, "curr")
            r += 1
        total_sources = (budget.get("total_debt") or 0) + (budget.get("equity") or 0)
        ws.cell(row=r, column=1, value="Total Sources").font = s["bold"]
        _fmt_cell(ws, r, 2, total_sources, s, "curr").font = s["bold"]
        r += 1
        ws.cell(row=r, column=1, value="Sources − Uses")
        _fmt_cell(ws, r, 2, total_sources - (budget.get("total_uses") or 0), s, "curr")
    else:
        ws.cell(row=1, column=1, value="No capital budget saved")
    _autosize_columns(ws)

    # ------------------------------------------------------------------
    # 4. Annual Forecast — the app's anniversary-year table, verbatim
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Annual Forecast")
    if annual_forecast and annual_forecast.get("rows"):
        cols = annual_forecast.get("columns") or []
        years = annual_forecast.get("years") or [c.get("year") for c in cols]
        header = ["Line Item"]
        for c in cols:
            lbl = str(c.get("label") or c.get("year") or "")
            if c.get("sublabel"):
                lbl = f"{lbl} ({c['sublabel']})"
            header.append(lbl)
        _write_header_row(ws, 1, header, s)
        r = 2
        for row in annual_forecast["rows"]:
            label = str(row.get("label") or "")
            c = ws.cell(row=r, column=1, value=label)
            if row.get("isBold") or row.get("is_header"):
                c.font = s["bold"]
            if not row.get("is_header"):
                vals = row.get("values") or {}
                ratio = "DSCR" in label or "Occup" in label
                for ci, yr in enumerate(years, 2):
                    v = vals.get(yr, vals.get(str(yr)))
                    if v is None:
                        continue
                    cell = _fmt_cell(ws, r, ci, v, s, "curr")
                    if ratio:
                        cell.number_format = "0.00"
                    if row.get("isBold"):
                        cell.font = s["bold"]
            r += 1
    else:
        ws.cell(row=1, column=1, value="No forecast data")
    _autosize_columns(ws)

    # ------------------------------------------------------------------
    # 5. Debt Service (loan summary + monthly amortization)
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Debt Service")
    ls = result.get("loan_sched")
    if ls is not None and not ls.empty:
        r = _section(ws, 1, "Loan Summary", s)
        _write_header_row(ws, r, ["Loan", "First Period", "Last Period", "Rate",
                                  "Total Interest", "Total Principal",
                                  "Curtailments", "Ending Balance"], s)
        r += 1
        for lid, grp in ls.groupby("LoanID"):
            g = grp.sort_values("event_date")
            ws.cell(row=r, column=1, value=str(lid))
            ws.cell(row=r, column=2, value=str(g.iloc[0]["event_date"])[:10])
            ws.cell(row=r, column=3, value=str(g.iloc[-1]["event_date"])[:10])
            _fmt_cell(ws, r, 4, float(g.iloc[0]["rate"]), s, "pct")
            _fmt_cell(ws, r, 5, float(g["interest"].sum()), s, "curr")
            _fmt_cell(ws, r, 6, float(g["principal"].sum()), s, "curr")
            cur = float(g["curtailment"].sum()) if "curtailment" in g.columns else 0.0
            _fmt_cell(ws, r, 7, cur, s, "curr")
            _fmt_cell(ws, r, 8, float(g.iloc[-1]["ending_balance"]), s, "curr")
            r += 1
        r += 1
        r = _section(ws, r, "Amortization Schedule (monthly)", s)
        cols = ["LoanID", "Period", "Rate", "Interest", "Principal", "Payment",
                "Curtailment", "Ending Balance"]
        _write_header_row(ws, r, cols, s)
        r += 1
        for _, x in ls.sort_values(["LoanID", "event_date"]).iterrows():
            ws.cell(row=r, column=1, value=str(x["LoanID"]))
            ws.cell(row=r, column=2, value=str(x["event_date"])[:10])
            _fmt_cell(ws, r, 3, float(x["rate"]), s, "pct")
            _fmt_cell(ws, r, 4, float(x["interest"]), s, "curr")
            _fmt_cell(ws, r, 5, float(x["principal"]), s, "curr")
            _fmt_cell(ws, r, 6, float(x["payment"]), s, "curr")
            _fmt_cell(ws, r, 7, float(x.get("curtailment", 0) or 0), s, "curr")
            _fmt_cell(ws, r, 8, float(x["ending_balance"]), s, "curr")
            r += 1
    else:
        ws.cell(row=1, column=1, value="No debt modelled")
    _autosize_columns(ws)

    # ------------------------------------------------------------------
    # 6. Cash Management
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Cash Management")
    cs = result.get("cash_schedule")
    if cs is not None and not cs.empty:
        cols = [c for c in ["event_date", "beginning_cash", "capital_call",
                            "reserve_deposit", "operating_cf", "capex_paid",
                            "capex_unpaid", "deficit_covered", "distributable",
                            "ending_cash"] if c in cs.columns]
        _write_header_row(ws, 1, [c.replace("_", " ").title() for c in cols], s)
        r = 2
        for _, x in cs.iterrows():
            for ci, c in enumerate(cols, 1):
                v = x[c]
                if c == "event_date":
                    ws.cell(row=r, column=ci, value=str(v)[:10])
                else:
                    _fmt_cell(ws, r, ci, float(v or 0), s, "curr")
            r += 1
    else:
        ws.cell(row=1, column=1, value="No cash schedule")
    _autosize_columns(ws)

    # ------------------------------------------------------------------
    # 7. Waterfall Steps (the rules the money followed)
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Waterfall Steps")
    _write_header_row(ws, 1, ["Waterfall", "Tie #", "Partner", "Step", "FX Rate",
                              "Rate %", "Amount", "Description"], s)
    r = 2
    for st in (wf_steps or []):
        ws.cell(row=r, column=1, value=st.get("vmisc"))
        ws.cell(row=r, column=2, value=st.get("iOrder"))
        ws.cell(row=r, column=3, value=st.get("PropCode"))
        ws.cell(row=r, column=4, value=st.get("vState"))
        ws.cell(row=r, column=5, value=st.get("FXRate"))
        ws.cell(row=r, column=6, value=st.get("nPercent"))
        _fmt_cell(ws, r, 7, st.get("mAmount"), s, "curr")
        ws.cell(row=r, column=8, value=st.get("vtranstype"))
        r += 1
    _autosize_columns(ws)

    # ------------------------------------------------------------------
    # 7b. PPI Ownership — the stack upstream of the PE vehicle
    # ------------------------------------------------------------------
    if ppi and (ppi.get("participants") or ppi.get("notes")):
        ws = wb.create_sheet("PPI Ownership")
        r = 1
        ws.cell(row=r, column=1, value=f"PE Vehicle: {ppi.get('vehicle', '')}"
                ).font = s["bold"]
        r += 1
        for note in (ppi.get("notes") or []):
            ws.cell(row=r, column=1, value=str(note))
            r += 1
        r += 1
        if ppi.get("participants"):
            r = _section(ws, r, "Participants", s)
            _write_header_row(ws, r, ["Participant", "Type", "Relationships",
                                      "Contributions", "Distributions",
                                      "AM Fees Paid", "AM Fees Received",
                                      "Net", "IRR", "MOIC"], s)
            r += 1
            for p_row in ppi["participants"]:
                ws.cell(row=r, column=1,
                        value=f"{p_row.get('name')} ({p_row.get('investor_id')})")
                ws.cell(row=r, column=2, value=p_row.get("type"))
                ws.cell(row=r, column=3,
                        value=", ".join(p_row.get("relationships") or []))
                _fmt_cell(ws, r, 4, p_row.get("contributions"), s, "curr")
                _fmt_cell(ws, r, 5, p_row.get("distributions"), s, "curr")
                _fmt_cell(ws, r, 6, p_row.get("am_fees_paid"), s, "curr")
                _fmt_cell(ws, r, 7, p_row.get("am_fees_received"), s, "curr")
                _fmt_cell(ws, r, 8, p_row.get("net_total"), s, "curr")
                _fmt_cell(ws, r, 9, p_row.get("irr"), s, "pct")
                c = ws.cell(row=r, column=10, value=p_row.get("moic"))
                c.number_format = s["mult"]
                r += 1
            r += 1
        for rel in (ppi.get("relationships") or []):
            r = _section(ws, r, f"{rel.get('name')} ({rel.get('entity_id')}, "
                                f"{rel.get('slice_pct')}%)", s)
            for b in (rel.get("breakdown") or []):
                ws.cell(row=r, column=1, value=b.get("category"))
                ws.cell(row=r, column=2, value=b.get("participant"))
                _fmt_cell(ws, r, 3, b.get("amount"), s, "curr")
                r += 1
            fs = rel.get("fee_schedule") or []
            if fs:
                ws.cell(row=r, column=1, value="AM fee schedule").font = s["bold"]
                r += 1
                _write_header_row(ws, r, ["Date", "Recipient", "Waterfall", "Fee"], s)
                r += 1
                for f in fs:
                    ws.cell(row=r, column=1, value=f.get("date"))
                    ws.cell(row=r, column=2, value=f.get("recipient"))
                    ws.cell(row=r, column=3, value=f.get("waterfall"))
                    _fmt_cell(ws, r, 4, f.get("fee"), s, "curr")
                    r += 1
            r += 1
        _autosize_columns(ws)

    # ------------------------------------------------------------------
    # 8. Partner Cash Flows — with live =XIRR so Excel audits the IRR
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Partner Cash Flows")
    col = 1
    for p in (result.get("partner_results") or []):
        cfs = p.get("combined_cashflows") or []
        ws.cell(row=1, column=col, value=p.get("partner")).font = s["bold"]
        for ci, h in ((col, "Date"), (col + 1, "Amount")):
            hc = ws.cell(row=2, column=ci, value=h)
            hc.font = s["header_font"]
            hc.fill = s["header_fill"]
        r = 3
        for d, amt in cfs:
            c = ws.cell(row=r, column=col, value=d)
            c.number_format = s["date"]
            _fmt_cell(ws, r, col + 1, float(amt), s, "curr")
            r += 1
        from openpyxl.utils import get_column_letter
        dcol, acol = get_column_letter(col), get_column_letter(col + 1)
        ws.cell(row=r + 1, column=col, value="Excel XIRR:").font = s["bold"]
        fc = ws.cell(row=r + 1, column=col + 1,
                     value=f"=XIRR({acol}3:{acol}{r - 1},{dcol}3:{dcol}{r - 1})")
        fc.number_format = s["pct"]
        ws.cell(row=r + 2, column=col, value="Model IRR:").font = s["bold"]
        _fmt_cell(ws, r + 2, col + 1, p.get("irr"), s, "pct")
        col += 3
    _autosize_columns(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
