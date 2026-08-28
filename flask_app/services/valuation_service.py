"""Annual property valuation cycle management (Phase 1).

Owns four app-managed tables (all in PROTECTED_TABLES — never overwritten by
CSV import or MRI refresh):

    valuation_cycles     — one row per annual valuation cycle (year, as-of date)
    valuation_records    — one row per deal per cycle: policy classification,
                           appraisal assumptions, linked Argus import, status
    valuation_documents  — evidence store (appraisal PDFs, LLC excerpts, BS
                           support) as DB blobs with SHA-256 dedup per record
    valuation_comments   — analyst commentary blocks per record + section

The module deliberately does NOT own the `valuations` table (that name is the
MRI_Val feed, loaded as data["mri_val"]). Publishing approved results into it
is Phase 3.

Policy classification (Valuation Policy, 9/18/25), computed as a default and
overridable per record with a note:
  - cost:        held < 12 months at the valuation date, or a development /
                 new-construction strategy (value derived from accounting
                 balances + accrued pref — Phase 3)
  - third_party: invested preferred equity >= $5M and held >= 12 months
  - internal:    everything else (< $5M PE — income cap / DCF in-house)

Child properties of a portfolio parent inherit the parent's classification:
appraisals are evaluated at the property level, NAV runs at the parent.
"""

from __future__ import annotations

import hashlib
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from sqlalchemy import text

from config import IS_ACCOUNTS, is_dev_deal
from flask_app.db import get_engine

logger = logging.getLogger(__name__)

PE_THRESHOLD = 5_000_000.0
HOLD_MONTHS_MIN = 12

CLASSIFICATIONS = ("third_party", "internal", "cost")
RECORD_STATUSES = ("open", "signed_off", "approved", "excluded")

# Valuation Committee (parallel unanimous — decisions Aug 28, 2026).
# Roles live in the shared review_roles table; 'cio' was added to
# REVIEW_ROLE_NAMES for this. The CCO records but does not approve.
COMMITTEE_ROLES = ("president", "ceo", "cio")
RECORDER_ROLE = "cco"
DOC_TYPES = ("appraisal", "argus", "llc_excerpt", "bs_support", "other")
COMMENT_SECTIONS = ("budget_review", "balance_sheet", "general")

_VALUATION_DDL = [
    """
    CREATE TABLE IF NOT EXISTS valuation_cycles (
        id {pk},
        year INTEGER NOT NULL UNIQUE,
        as_of_date TEXT NOT NULL,
        status TEXT NOT NULL DEFAULT 'open',
        opened_by TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        closed_at TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS valuation_records (
        id {pk},
        cycle_id INTEGER NOT NULL,
        vcode TEXT NOT NULL,
        parent_vcode TEXT,
        classification TEXT,
        classification_reason TEXT,
        classification_override TEXT,
        override_note TEXT,
        method TEXT,
        concluded_value DOUBLE PRECISION,
        cap_rate DOUBLE PRECISION,
        term_cap_rate DOUBLE PRECISION,
        discount_rate DOUBLE PRECISION,
        direct_cap_noi DOUBLE PRECISION,
        cost_of_sale_pct DOUBLE PRECISION,
        appraiser TEXT,
        appraisal_date TEXT,
        argus_import_id INTEGER,
        status TEXT NOT NULL DEFAULT 'open',
        signed_off_by TEXT,
        signed_off_at TIMESTAMP,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(cycle_id, vcode)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS valuation_documents (
        id {pk},
        record_id INTEGER NOT NULL,
        doc_type TEXT,
        filename TEXT NOT NULL,
        file_hash TEXT,
        file_data {blob},
        uploaded_by TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS valuation_comments (
        id {pk},
        record_id INTEGER NOT NULL,
        section TEXT NOT NULL,
        comment_text TEXT,
        updated_by TEXT,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(record_id, section)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS valuation_questions (
        id {pk},
        record_id INTEGER NOT NULL,
        question_text TEXT NOT NULL,
        asked_by TEXT,
        asked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        answer_text TEXT,
        answered_by TEXT,
        answered_at TIMESTAMP,
        status TEXT NOT NULL DEFAULT 'open'
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS valuation_approvals (
        id {pk},
        record_id INTEGER NOT NULL,
        member_role TEXT NOT NULL,
        username TEXT,
        action TEXT NOT NULL,
        note TEXT,
        active INTEGER NOT NULL DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS valuation_snapshots (
        id {pk},
        record_id INTEGER NOT NULL UNIQUE,
        snapshot_json TEXT,
        approved_by TEXT,
        approved_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS valuation_ai_summaries (
        id {pk},
        record_id INTEGER NOT NULL UNIQUE,
        doc_id INTEGER,
        summary_json TEXT,
        model TEXT,
        created_by TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS valuation_nav_results (
        id {pk},
        record_id INTEGER NOT NULL UNIQUE,
        inputs_json TEXT,
        walk_json TEXT,
        net_proceeds DOUBLE PRECISION,
        psc_nav DOUBLE PRECISION,
        op_nav DOUBLE PRECISION,
        computed_by TEXT,
        computed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS valuation_bs_selections (
        id {pk},
        record_id INTEGER NOT NULL,
        account TEXT NOT NULL,
        included INTEGER NOT NULL,
        updated_by TEXT,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(record_id, account)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS valuation_step_refs (
        id {pk},
        vcode TEXT NOT NULL,
        wf_type TEXT NOT NULL,
        iorder INTEGER NOT NULL,
        agreement_ref TEXT,
        UNIQUE(vcode, wf_type, iorder)
    )
    """,
    "CREATE INDEX IF NOT EXISTS idx_valuation_records_cycle ON valuation_records(cycle_id)",
    "CREATE INDEX IF NOT EXISTS idx_valuation_documents_record ON valuation_documents(record_id)",
    "CREATE INDEX IF NOT EXISTS idx_valuation_questions_record ON valuation_questions(record_id)",
    "CREATE INDEX IF NOT EXISTS idx_valuation_approvals_record ON valuation_approvals(record_id)",
]


def ensure_valuation_tables(engine=None):
    """Create valuation tables if missing. Idempotent, both dialects."""
    if engine is None:
        engine = get_engine()
    is_pg = engine.dialect.name == "postgresql"
    pk = "SERIAL PRIMARY KEY" if is_pg else "INTEGER PRIMARY KEY AUTOINCREMENT"
    blob = "BYTEA" if is_pg else "BLOB"
    with engine.begin() as conn:
        for ddl in _VALUATION_DDL:
            conn.execute(text(ddl.format(pk=pk, blob=blob)))
    # Idempotent column migrations (records created before Phase 3)
    for col, col_type in (("published_at", "TIMESTAMP"), ("published_by", "TEXT")):
        try:
            with engine.begin() as conn:
                conn.execute(text(f"ALTER TABLE valuation_records ADD COLUMN {col} {col_type}"))
        except Exception:
            pass  # column already exists


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _serialize(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def _row_dict(row) -> Dict[str, Any]:
    return {k: _serialize(v) for k, v in dict(row._mapping).items()}


def _insert_id(conn, engine, result, table: str, where_sql: str, params: dict) -> int:
    """Portable lastrowid (feedback_service pattern)."""
    if engine.dialect.name == "postgresql":
        row = conn.execute(text(f"SELECT max(id) FROM {table} WHERE {where_sql}"), params).fetchone()
        return int(row[0])
    return int(result.lastrowid)


# ============================================================
# Classification
# ============================================================

def _pe_funded(acct: pd.DataFrame, investment_id: str) -> float:
    """Net preferred-equity contributions for a deal (signs respected —
    reversal rows net out rather than double-count)."""
    if acct is None or acct.empty or not investment_id:
        return 0.0
    rows = acct[
        (acct["InvestmentID"] == str(investment_id).strip().upper())
        & (acct["is_contribution"])
        & (~acct["InvestorID"].str.upper().str.startswith("OP"))
    ]
    if rows.empty:
        return 0.0
    # Contributions are stored negative; net and flip sign.
    return float(-rows["Amt"].sum())


def _classify(strategy, acq_date, pe_funded: float, as_of: date) -> Tuple[str, str]:
    """Default policy classification for one deal."""
    if is_dev_deal(strategy):
        return "cost", "Development / new-construction strategy — carried at cost"
    acq = None
    if acq_date is not None and not pd.isna(acq_date):
        acq = pd.Timestamp(acq_date).date()
    if acq is None:
        return "internal", "No acquisition date on file — review classification"
    months_held = (as_of.year - acq.year) * 12 + (as_of.month - acq.month)
    if months_held < HOLD_MONTHS_MIN:
        return "cost", f"Held {months_held} months (< 12) at the valuation date — carried at cost"
    if pe_funded >= PE_THRESHOLD:
        return "third_party", (
            f"${pe_funded:,.0f} preferred equity (>= $5M) held {months_held} months — "
            "third-party appraisal required"
        )
    return "internal", f"${pe_funded:,.0f} preferred equity (< $5M) — internal valuation eligible"


def _child_parent_map(inv: pd.DataFrame) -> Dict[str, str]:
    """{child_vcode: parent_vcode} using the app-wide Portfolio_Name pairing.

    Parents have Property_Count >= 1; genuine children carry 0. A child's
    Portfolio_Name matches the parent's Investment_Name or (Burton exception)
    the parent's own Portfolio_Name.
    """
    out: Dict[str, str] = {}
    if inv is None or inv.empty:
        return out
    df = inv.copy()
    for col in ("Portfolio_Name", "Investment_Name"):
        if col not in df.columns:
            return out
        df[col] = df[col].fillna("").astype(str).str.strip()
    pc = pd.to_numeric(df.get("Property_Count"), errors="coerce").fillna(0)
    parents = df[pc >= 1]
    children = df[(pc == 0) & (df["Portfolio_Name"] != "")]
    for _, ch in children.iterrows():
        grp = ch["Portfolio_Name"]
        match = parents[
            (parents["Investment_Name"] == grp) | (parents["Portfolio_Name"] == grp)
        ]
        match = match[match["vcode"] != ch["vcode"]]
        if not match.empty:
            out[str(ch["vcode"])] = str(match.iloc[0]["vcode"])
    return out


# ============================================================
# Cycles
# ============================================================

def list_cycles(engine) -> List[Dict[str, Any]]:
    ensure_valuation_tables(engine)
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT c.*,
                   (SELECT COUNT(*) FROM valuation_records r WHERE r.cycle_id = c.id) AS record_count
            FROM valuation_cycles c ORDER BY c.year DESC
        """)).fetchall()
    return [_row_dict(r) for r in rows]


def create_cycle(engine, year: int, username: str, data: dict,
                 as_of_date: Optional[str] = None) -> Dict[str, Any]:
    """Create a cycle and seed a record per active deal (children included —
    appraisals are evaluated per property; NAV rolls up to the parent)."""
    ensure_valuation_tables(engine)
    year = int(year)
    as_of = as_of_date or f"{year}-12-31"

    with engine.connect() as conn:
        existing = conn.execute(
            text("SELECT id FROM valuation_cycles WHERE year = :y"), {"y": year}
        ).fetchone()
    if existing:
        raise ValueError(f"A {year} valuation cycle already exists")

    with engine.begin() as conn:
        result = conn.execute(text("""
            INSERT INTO valuation_cycles (year, as_of_date, status, opened_by)
            VALUES (:y, :d, 'open', :u)
        """), {"y": year, "d": as_of, "u": username})
        cycle_id = _insert_id(conn, engine, result, "valuation_cycles", "year = :y", {"y": year})

    seeded = seed_cycle_records(engine, cycle_id, data, username)
    return {"id": cycle_id, "year": year, "as_of_date": as_of, **seeded}


def seed_cycle_records(engine, cycle_id: int, data: dict, username: str) -> Dict[str, Any]:
    """Insert a record for every active deal missing one. Idempotent — safe to
    re-run mid-cycle to pick up newly added deals."""
    from flask_app.services import data_service

    with engine.connect() as conn:
        cyc = conn.execute(
            text("SELECT * FROM valuation_cycles WHERE id = :i"), {"i": cycle_id}
        ).fetchone()
        if not cyc:
            raise ValueError(f"Cycle {cycle_id} not found")
        have = {r[0] for r in conn.execute(
            text("SELECT vcode FROM valuation_records WHERE cycle_id = :i"), {"i": cycle_id}
        ).fetchall()}

    as_of = pd.Timestamp(cyc._mapping["as_of_date"]).date()
    inv = data_service.exclude_sold(data.get("inv"))
    acct = data.get("acct")
    if inv is None or inv.empty:
        return {"seeded": 0, "skipped": 0}

    child_of = _child_parent_map(inv)

    # First pass — classify every deal on its own accounting
    rows_by_vcode: Dict[str, dict] = {}
    for _, row in inv.iterrows():
        vcode = str(row["vcode"])
        pe = _pe_funded(acct, row.get("InvestmentID", ""))
        cls, reason = _classify(row.get("Investment_Strategy"), row.get("Acquisition_Date"), pe, as_of)
        rows_by_vcode[vcode] = {"classification": cls, "reason": reason}

    # Second pass — children inherit the parent's classification
    for child, parent in child_of.items():
        if child in rows_by_vcode and parent in rows_by_vcode:
            pcls = rows_by_vcode[parent]["classification"]
            rows_by_vcode[child] = {
                "classification": pcls,
                "reason": f"Child property — inherits {parent} classification; "
                          "appraised at property level, NAV rolls up to the parent",
            }

    seeded = 0
    with engine.begin() as conn:
        for vcode, info in rows_by_vcode.items():
            if vcode in have:
                continue
            conn.execute(text("""
                INSERT INTO valuation_records
                    (cycle_id, vcode, parent_vcode, classification, classification_reason,
                     status, updated_at)
                VALUES (:c, :v, :p, :cls, :r, 'open', :now)
            """), {
                "c": cycle_id, "v": vcode, "p": child_of.get(vcode),
                "cls": info["classification"], "r": info["reason"], "now": _now(),
            })
            seeded += 1
    return {"seeded": seeded, "skipped": len(have)}


# ============================================================
# Dashboard
# ============================================================

def get_cycle_dashboard(engine, cycle_id: int, data: dict) -> Dict[str, Any]:
    ensure_valuation_tables(engine)
    with engine.connect() as conn:
        cyc = conn.execute(
            text("SELECT * FROM valuation_cycles WHERE id = :i"), {"i": cycle_id}
        ).fetchone()
        if not cyc:
            raise ValueError(f"Cycle {cycle_id} not found")
        recs = conn.execute(text("""
            SELECT r.*,
                   (SELECT COUNT(*) FROM valuation_documents d
                     WHERE d.record_id = r.id) AS doc_count,
                   (SELECT COUNT(*) FROM valuation_documents d
                     WHERE d.record_id = r.id AND d.doc_type = 'appraisal') AS appraisal_count,
                   (SELECT COUNT(*) FROM valuation_questions q
                     WHERE q.record_id = r.id AND q.status = 'open') AS open_questions,
                   (SELECT COUNT(*) FROM valuation_approvals a
                     WHERE a.record_id = r.id AND a.active = 1 AND a.action = 'approve') AS approval_count
            FROM valuation_records r WHERE r.cycle_id = :i
        """), {"i": cycle_id}).fetchall()

    cycle = _row_dict(cyc)
    year = int(cycle["year"])

    inv = data.get("inv")
    inv_by_vcode = {}
    if inv is not None and not inv.empty:
        inv_by_vcode = {str(r["vcode"]): r for _, r in inv.iterrows()}

    prior_vals = _prior_values(data.get("mri_val"), year - 1)

    items = []
    for r in recs:
        d = _row_dict(r)
        deal = inv_by_vcode.get(d["vcode"])
        d["deal_name"] = str(deal["Investment_Name"]) if deal is not None else d["vcode"]
        d["asset_type"] = str(deal["Asset_Type"]) if deal is not None and pd.notna(deal.get("Asset_Type")) else ""
        d["is_child"] = bool(d.get("parent_vcode"))
        d["effective_classification"] = d.get("classification_override") or d.get("classification")
        d["prior_value"] = prior_vals.get(d["vcode"])
        cv = d.get("concluded_value")
        d["value_change"] = (cv - d["prior_value"]) if cv is not None and d.get("prior_value") else None
        d["has_appraisal"] = (d.pop("appraisal_count", 0) or 0) > 0
        d["has_argus"] = d.get("argus_import_id") is not None
        items.append(d)

    # Parents first (alpha), children directly under their parent
    parents = sorted([i for i in items if not i["is_child"]], key=lambda x: x["deal_name"].lower())
    by_parent: Dict[str, list] = {}
    for i in items:
        if i["is_child"]:
            by_parent.setdefault(i["parent_vcode"], []).append(i)
    ordered = []
    for p in parents:
        ordered.append(p)
        ordered.extend(sorted(by_parent.get(p["vcode"], []), key=lambda x: x["deal_name"].lower()))
    # Orphan children (parent not in cycle) at the end
    seen = {id(i) for i in ordered}
    ordered.extend(i for i in items if id(i) not in seen)

    return {"cycle": cycle, "records": ordered}


def _prior_values(mri_val: Optional[pd.DataFrame], prior_year: int) -> Dict[str, float]:
    """{vcode: concluded value} from the valuations feed for a given year."""
    out: Dict[str, float] = {}
    if mri_val is None or mri_val.empty:
        return out
    df = mri_val.copy()
    vcode_col = "vCode" if "vCode" in df.columns else "vcode"
    if vcode_col not in df.columns or "dtValuation" not in df.columns:
        return out
    df["_dt"] = pd.to_datetime(df["dtValuation"], errors="coerce")
    df = df[df["_dt"].dt.year == prior_year]
    for _, r in df.iterrows():
        val = pd.to_numeric(r.get("mIncomeCapConcludedValue"), errors="coerce")
        if pd.notna(val):
            out[str(r[vcode_col]).strip()] = float(val)
    return out


# ============================================================
# Records
# ============================================================

_RECORD_FIELDS = (
    "method", "concluded_value", "cap_rate", "term_cap_rate", "discount_rate",
    "direct_cap_noi", "cost_of_sale_pct", "appraiser", "appraisal_date",
    "classification_override", "override_note",
)
_NUMERIC_FIELDS = {
    "concluded_value", "cap_rate", "term_cap_rate", "discount_rate",
    "direct_cap_noi", "cost_of_sale_pct",
}


def get_record(engine, record_id: int, data: dict) -> Dict[str, Any]:
    ensure_valuation_tables(engine)
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT * FROM valuation_records WHERE id = :i"), {"i": record_id}
        ).fetchone()
        if not row:
            raise ValueError(f"Valuation record {record_id} not found")
        cyc = conn.execute(
            text("SELECT * FROM valuation_cycles WHERE id = :i"),
            {"i": row._mapping["cycle_id"]},
        ).fetchone()
        docs = conn.execute(text("""
            SELECT id, doc_type, filename, file_hash, uploaded_by, created_at,
                   CASE WHEN file_data IS NOT NULL THEN 1 ELSE 0 END AS has_file
            FROM valuation_documents WHERE record_id = :i ORDER BY created_at
        """), {"i": record_id}).fetchall()
        comments = conn.execute(text("""
            SELECT section, comment_text, updated_by, updated_at
            FROM valuation_comments WHERE record_id = :i
        """), {"i": record_id}).fetchall()
        questions = conn.execute(text("""
            SELECT * FROM valuation_questions WHERE record_id = :i ORDER BY asked_at, id
        """), {"i": record_id}).fetchall()
        approvals = conn.execute(text("""
            SELECT member_role, username, action, note, created_at
            FROM valuation_approvals WHERE record_id = :i AND active = 1
            ORDER BY created_at
        """), {"i": record_id}).fetchall()
        ai_meta = conn.execute(text("""
            SELECT id, doc_id, model, created_by, created_at
            FROM valuation_ai_summaries WHERE record_id = :i
        """), {"i": record_id}).fetchone()

    rec = _row_dict(row)
    rec["cycle"] = _row_dict(cyc) if cyc else None
    rec["documents"] = [_row_dict(d) for d in docs]
    rec["comments"] = {c._mapping["section"]: _row_dict(c) for c in comments}
    rec["questions"] = [_row_dict(q) for q in questions]
    rec["approvals"] = [_row_dict(a) for a in approvals]
    approved_roles = {a._mapping["member_role"] for a in approvals if a._mapping["action"] == "approve"}
    rec["approval_state"] = {
        "approved_roles": sorted(approved_roles),
        "missing_roles": [r for r in COMMITTEE_ROLES if r not in approved_roles],
        "is_approved": rec.get("status") == "approved",
    }
    rec["ai_summary_meta"] = _row_dict(ai_meta) if ai_meta else None
    rec["effective_classification"] = rec.get("classification_override") or rec.get("classification")

    # Deal header + valuation history from the MRI_Val feed
    inv = data.get("inv")
    if inv is not None and not inv.empty:
        match = inv[inv["vcode"].astype(str) == rec["vcode"]]
        if not match.empty:
            deal = match.iloc[0]
            rec["deal"] = {
                "name": _s(deal.get("Investment_Name")),
                "investment_id": _s(deal.get("InvestmentID")),
                "asset_type": _s(deal.get("Asset_Type")),
                "city": _s(deal.get("City")),
                "state": _s(deal.get("State")),
                "operating_partner": _s(deal.get("Operating_Partner")),
                "strategy": _s(deal.get("Investment_Strategy")),
                "acquisition_date": _s(deal.get("Acquisition_Date")),
                "size_sqf": _num(deal.get("Size_Sqf")),
                "total_units": _num(deal.get("Total_Units")),
            }
            rec["pe_funded"] = _pe_funded(data.get("acct"), deal.get("InvestmentID", ""))

    rec["valuation_history"] = _valuation_history(data.get("mri_val"), rec["vcode"])
    return rec


def _s(v) -> str:
    return "" if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v) else str(v)


def _num(v):
    n = pd.to_numeric(v, errors="coerce")
    return None if pd.isna(n) else float(n)


def _valuation_history(mri_val: Optional[pd.DataFrame], vcode: str) -> List[Dict[str, Any]]:
    if mri_val is None or mri_val.empty:
        return []
    df = mri_val.copy()
    vcode_col = "vCode" if "vCode" in df.columns else "vcode"
    df = df[df[vcode_col].astype(str).str.strip().str.lower() == str(vcode).strip().lower()]
    if df.empty:
        return []
    df["_dt"] = pd.to_datetime(df["dtValuation"], errors="coerce")
    df = df.sort_values("_dt", ascending=False)
    out = []
    for _, r in df.iterrows():
        out.append({
            "date": r["_dt"].strftime("%Y-%m-%d") if pd.notna(r["_dt"]) else _s(r.get("dtValuation")),
            "method": _s(r.get("vMethod")),
            "noi": _num(r.get("mAnnualNOI")),
            "cap_rate": _num(r.get("fCapRate")),
            "term_cap_rate": _num(r.get("nTermCapRate")),
            "discount_rate": _num(r.get("nDiscountRateForEquityInterest")),
            "value": _num(r.get("mIncomeCapConcludedValue")),
            "debt": _num(r.get("mDebtValue")),
            "pe_nav": _num(r.get("mMezzanineValue")),
            "cost_of_sale": _num(r.get("nCostSaleRate")),
        })
    return out


def _require_not_approved(engine, record_id: int):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT status FROM valuation_records WHERE id = :i"), {"i": record_id}
        ).fetchone()
    if not row:
        raise ValueError(f"Valuation record {record_id} not found")
    if row[0] == "approved":
        raise ValueError("This valuation is approved and locked — the committee must return it before edits")


def update_record(engine, record_id: int, fields: Dict[str, Any], username: str) -> Dict[str, Any]:
    ensure_valuation_tables(engine)
    _require_not_approved(engine, record_id)
    sets, params = [], {"i": record_id, "now": _now()}
    for k in _RECORD_FIELDS:
        if k not in fields:
            continue
        v = fields[k]
        if v == "" or v is None:
            v = None
        elif k in _NUMERIC_FIELDS:
            v = float(v)
        elif k == "classification_override" and v not in CLASSIFICATIONS:
            raise ValueError(f"Invalid classification: {v}")
        sets.append(f"{k} = :{k}")
        params[k] = v
    if not sets:
        raise ValueError("No recognized fields to update")
    if "classification_override" in params and params["classification_override"] and not (
        fields.get("override_note") or ""
    ).strip():
        raise ValueError("A classification override requires a note")
    with engine.begin() as conn:
        conn.execute(text(f"""
            UPDATE valuation_records SET {', '.join(sets)}, updated_at = :now WHERE id = :i
        """), params)
    return {"status": "updated", "id": record_id}


def record_action(engine, record_id: int, action: str, username: str) -> Dict[str, Any]:
    """Phase 1 status transitions: sign_off, reopen, exclude."""
    ensure_valuation_tables(engine)
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT status FROM valuation_records WHERE id = :i"), {"i": record_id}
        ).fetchone()
    if not row:
        raise ValueError(f"Valuation record {record_id} not found")
    status = row[0]

    if action == "sign_off":
        if status != "open":
            raise ValueError(f"Cannot sign off a record in status '{status}'")
        with engine.begin() as conn:
            conn.execute(text("""
                UPDATE valuation_records
                SET status = 'signed_off', signed_off_by = :u, signed_off_at = :now,
                    updated_at = :now
                WHERE id = :i
            """), {"i": record_id, "u": username, "now": _now()})
        return {"status": "signed_off"}

    if action == "reopen":
        if status == "approved":
            raise ValueError("Approved valuations must be returned by the Valuation Committee")
        with engine.begin() as conn:
            conn.execute(text("""
                UPDATE valuation_records
                SET status = 'open', signed_off_by = NULL, signed_off_at = NULL,
                    updated_at = :now
                WHERE id = :i
            """), {"i": record_id, "now": _now()})
        return {"status": "open"}

    if action == "exclude":
        with engine.begin() as conn:
            conn.execute(text("""
                UPDATE valuation_records SET status = 'excluded', updated_at = :now
                WHERE id = :i
            """), {"i": record_id, "now": _now()})
        return {"status": "excluded"}

    raise ValueError(f"Unknown action: {action}")


# ============================================================
# Documents
# ============================================================

def upload_documents(engine, record_id: int, files: List[Tuple[str, bytes]],
                     doc_type: str, username: str) -> Dict[str, Any]:
    ensure_valuation_tables(engine)
    _require_not_approved(engine, record_id)
    if doc_type not in DOC_TYPES:
        doc_type = "other"
    with engine.connect() as conn:
        exists = conn.execute(
            text("SELECT id FROM valuation_records WHERE id = :i"), {"i": record_id}
        ).fetchone()
        if not exists:
            raise ValueError(f"Valuation record {record_id} not found")
        existing_hashes = {r[0] for r in conn.execute(text("""
            SELECT file_hash FROM valuation_documents
            WHERE record_id = :i AND file_hash IS NOT NULL
        """), {"i": record_id}).fetchall()}

    added, skipped, details = 0, 0, []
    with engine.begin() as conn:
        for filename, file_bytes in files:
            file_hash = hashlib.sha256(file_bytes).hexdigest()
            if file_hash in existing_hashes:
                skipped += 1
                details.append({"filename": filename, "action": "skipped_duplicate"})
                continue
            conn.execute(text("""
                INSERT INTO valuation_documents
                    (record_id, doc_type, filename, file_hash, file_data, uploaded_by)
                VALUES (:r, :t, :f, :h, :d, :u)
            """), {"r": record_id, "t": doc_type, "f": filename,
                   "h": file_hash, "d": file_bytes, "u": username})
            existing_hashes.add(file_hash)
            added += 1
            details.append({"filename": filename, "action": "added"})
    return {"added": added, "skipped_duplicate": skipped, "details": details}


def get_document(engine, record_id: int, doc_id: int) -> Tuple[str, bytes]:
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT filename, file_data FROM valuation_documents
            WHERE id = :d AND record_id = :r
        """), {"d": doc_id, "r": record_id}).fetchone()
    if not row or not row[1]:
        raise ValueError("Document not found or no file data")
    return row[0], bytes(row[1])


def delete_document(engine, record_id: int, doc_id: int) -> bool:
    _require_not_approved(engine, record_id)
    with engine.begin() as conn:
        result = conn.execute(text("""
            DELETE FROM valuation_documents WHERE id = :d AND record_id = :r
        """), {"d": doc_id, "r": record_id})
    return result.rowcount > 0


# ============================================================
# Comments
# ============================================================

def save_comment(engine, record_id: int, section: str, comment_text: str,
                 username: str) -> Dict[str, Any]:
    ensure_valuation_tables(engine)
    _require_not_approved(engine, record_id)
    if section not in COMMENT_SECTIONS:
        raise ValueError(f"Unknown comment section: {section}")
    with engine.begin() as conn:
        result = conn.execute(text("""
            UPDATE valuation_comments
            SET comment_text = :t, updated_by = :u, updated_at = :now
            WHERE record_id = :r AND section = :s
        """), {"t": comment_text, "u": username, "now": _now(),
               "r": record_id, "s": section})
        if result.rowcount == 0:
            conn.execute(text("""
                INSERT INTO valuation_comments (record_id, section, comment_text, updated_by)
                VALUES (:r, :s, :t, :u)
            """), {"r": record_id, "s": section, "t": comment_text, "u": username})
    return {"status": "saved"}


# ============================================================
# Argus import (staged to the cycle)
# ============================================================

def import_argus(engine, record_id: int, file_bytes: bytes, filename: str,
                 username: str) -> Dict[str, Any]:
    """Import the appraisal's Argus export against the deal's vcode and link
    it to this valuation record. The import is visible in the Deal Analysis
    projection dropdown but does NOT change the default forecast — publishing
    into forecast_feed is the Phase 3 approval step."""
    from flask_app.services import argus_service

    _require_not_approved(engine, record_id)
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT r.vcode, c.year FROM valuation_records r
            JOIN valuation_cycles c ON c.id = r.cycle_id
            WHERE r.id = :i
        """), {"i": record_id}).fetchone()
    if not row:
        raise ValueError(f"Valuation record {record_id} not found")
    vcode, year = row[0], row[1]

    result = argus_service.import_argus_cashflow(
        engine, vcode, file_bytes, filename,
        import_label=f"{year} Valuation", import_type="valuation",
        username=username,
    )
    if result.get("status") in ("success", "duplicate") and result.get("import_id"):
        with engine.begin() as conn:
            conn.execute(text("""
                UPDATE valuation_records SET argus_import_id = :a, updated_at = :now
                WHERE id = :i
            """), {"a": int(result["import_id"]), "i": record_id, "now": _now()})
    return result


# ============================================================
# Budget review (Review Form page 1)
# ============================================================

def get_budget_review(engine, record_id: int, data: dict) -> Dict[str, Any]:
    """Three-column comparison: cycle-year Estimate (YTD actual + budget
    remainder), next-year Budget, and valuation year 1 (the linked Argus
    import) — with Estimate-vs-Budget and Budget-vs-Valuation variances.
    Mirrors the printed Budget Review form."""
    from flask_app.services import argus_service
    from flask_app.services.financials_service import (
        _prepare_isbs, _calculate_is_amounts, _get_bs_principal,
        _get_budget_principal, _get_valuation_sum,
    )

    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT r.vcode, r.argus_import_id, c.year FROM valuation_records r
            JOIN valuation_cycles c ON c.id = r.cycle_id
            WHERE r.id = :i
        """), {"i": record_id}).fetchone()
    if not row:
        raise ValueError(f"Valuation record {record_id} not found")
    vcode, argus_import_id, year = row[0], row[1], int(row[2])
    budget_year = year + 1

    isbs = _prepare_isbs(data.get("isbs_raw"), vcode)
    if isbs.empty:
        actual_data = budget_data = bs_data = pd.DataFrame()
        actual_periods = []
    else:
        actual_data = isbs[isbs["vSource"] == "Interim IS"]
        budget_data = isbs[isbs["vSource"] == "Budget IS"]
        bs_data = isbs[isbs["vSource"] == "Interim BS"]
        actual_periods = sorted(actual_data["dtEntry_parsed"].dropna().unique()) if not actual_data.empty else []

    argus_fc = None
    if argus_import_id:
        try:
            from flask import current_app
            pro_yr_base = current_app.config.get("PRO_YR_BASE_DEFAULT", year)
        except Exception:
            pro_yr_base = year
        argus_fc = argus_service.get_forecast_df_by_id(engine, vcode, int(argus_import_id), pro_yr_base)

    est_ref = pd.Timestamp(f"{year}-12-31")
    bud_ref = pd.Timestamp(f"{budget_year}-12-31")

    est = _calculate_is_amounts("Estimate", "Actual", est_ref, year, IS_ACCOUNTS,
                                actual_data, actual_periods, budget_data, pd.DataFrame(), None)
    bud = _calculate_is_amounts("Full Year", "Budget", bud_ref, budget_year, IS_ACCOUNTS,
                                actual_data, actual_periods, budget_data, pd.DataFrame(), None)
    val = {}
    if argus_fc is not None and not argus_fc.empty:
        val = _calculate_is_amounts("Full Year", "Valuation", bud_ref, budget_year, IS_ACCOUNTS,
                                    actual_data, actual_periods, budget_data, pd.DataFrame(), argus_fc)

    def _get(bal, section, category):
        return float(bal.get(section, {}).get(category, 0) or 0)

    rows = []

    def _add_row(label, e, b, v, level=1, **flags):
        rows.append({
            "account": label, "level": level,
            "estimate": e, "budget": b, "valuation": v,
            "var_est_bud": (e - b) if e is not None and b is not None else None,
            "var_bud_val": (b - v) if b is not None and v is not None else None,
            **flags,
        })

    totals = {}
    for section, display in (("REVENUES", "Total Revenues"), ("EXPENSES", "Total Expenses")):
        sec_e = sec_b = sec_v = 0.0
        for category in IS_ACCOUNTS[section].keys():
            e, b, v = _get(est, section, category), _get(bud, section, category), _get(val, section, category)
            if section == "REVENUES":
                e, b, v = -e, -b, -v  # credits → display positive
            _add_row(category, e, b, v)
            sec_e += e; sec_b += b; sec_v += v
        _add_row(display, sec_e, sec_b, sec_v, level=0, is_total=True)
        totals[section] = (sec_e, sec_b, sec_v)

    noi = tuple(totals["REVENUES"][i] - totals["EXPENSES"][i] for i in range(3))
    _add_row("Net Operating Income", *noi, level=0, is_calc=True)

    # Debt service — interest from 5190; principal per-source
    int_e = _get(est, "DEBT_SERVICE", "Interest")
    int_b = _get(bud, "DEBT_SERVICE", "Interest")
    int_v = _get(val, "DEBT_SERVICE", "Interest")

    jan1 = pd.Timestamp(f"{year}-01-01") - pd.DateOffset(days=1)
    last_actual = next(
        (pd.Timestamp(p) for p in reversed(actual_periods)
         if pd.Timestamp(p).year == year and pd.Timestamp(p) <= est_ref), None)
    if last_actual is not None:
        prin_e = _get_bs_principal(bs_data, jan1, last_actual)
        if last_actual < est_ref:
            prin_e += _get_budget_principal(budget_data, last_actual, est_ref)
    else:
        prin_e = _get_budget_principal(budget_data, jan1, est_ref)

    bud_jan1 = pd.Timestamp(f"{budget_year}-01-01") - pd.DateOffset(days=1)
    prin_b = _get_budget_principal(budget_data, bud_jan1, bud_ref)

    prin_v = 0.0
    if argus_fc is not None and not argus_fc.empty:
        pv = _get_valuation_sum(argus_fc, bud_jan1.date(), bud_ref.date(), {"_": {"P": ["7060"]}})
        prin_v = abs(pv.get("_", {}).get("P", 0) or 0)

    _add_row("Interest Expense", int_e, int_b, int_v)
    _add_row("Principal Payments", float(prin_e or 0), float(prin_b or 0), float(prin_v or 0))
    ds = (int_e + (prin_e or 0), int_b + (prin_b or 0), int_v + (prin_v or 0))
    _add_row("Total Debt Service", *ds, level=0, is_total=True)
    rows.append({
        "account": "DSCR", "level": 0, "is_calc": True,
        "estimate": (noi[0] / ds[0]) if ds[0] else None,
        "budget": (noi[1] / ds[1]) if ds[1] else None,
        "valuation": (noi[2] / ds[2]) if ds[2] else None,
        "var_est_bud": None, "var_bud_val": None, "is_ratio": True,
    })

    # Below the line
    btl_e = btl_b = btl_v = 0.0
    for category in IS_ACCOUNTS.get("OTHER_BTL", {}).keys():
        e, b, v = _get(est, "OTHER_BTL", category), _get(bud, "OTHER_BTL", category), _get(val, "OTHER_BTL", category)
        _add_row(category, e, b, v)
        btl_e += e; btl_b += b; btl_v += v
    _add_row("Other Below the Line", btl_e, btl_b, btl_v, level=0, is_total=True)

    return {
        "vcode": vcode,
        "estimate_year": year,
        "budget_year": budget_year,
        "has_argus": argus_fc is not None and not argus_fc.empty,
        "last_actual_month": (pd.Timestamp(last_actual).strftime("%Y-%m-%d") if last_actual is not None else None),
        "rows": rows,
        "occupancy_trend": _occupancy_trend(data.get("occupancy_raw"), vcode, year),
    }


def _occupancy_trend(occ_raw: Optional[pd.DataFrame], vcode: str, year: int) -> List[Dict[str, Any]]:
    """Quarterly average occupancy, trailing 12 quarters through the cycle
    year Q4 plus any reported quarters of the following year."""
    if occ_raw is None or occ_raw.empty:
        return []
    occ = occ_raw.copy()
    if "vCode" not in occ.columns or "Qtr" not in occ.columns:
        return []
    occ_col = "Occ%" if "Occ%" in occ.columns else ("OccupancyPercent" if "OccupancyPercent" in occ.columns else None)
    if occ_col is None:
        return []
    occ = occ[occ["vCode"].astype(str).str.strip().str.lower() == str(vcode).strip().lower()]
    if occ.empty:
        return []
    occ["_v"] = pd.to_numeric(occ[occ_col], errors="coerce").clip(lower=0, upper=100)
    grouped = occ.dropna(subset=["_v"]).groupby("Qtr")["_v"].mean()
    lo, hi = f"{year - 2}-Q1", f"{year + 1}-Q4"
    out = []
    for q in sorted(grouped.index):
        if lo <= str(q) <= hi:
            out.append({"quarter": str(q), "occupancy": float(grouped[q])})
    return out


# ============================================================
# Balance sheet analysis (Review Form page 2)
# ============================================================

def get_balance_sheet(engine, record_id: int, data: dict) -> Dict[str, Any]:
    """Prior year-end vs latest Interim BS, grouped by vAccountType and
    vDescription — the raw material for the Review Form balance sheet page
    and, in Phase 3, the current asset/liability curation."""
    from flask_app.services.financials_service import _prepare_isbs

    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT r.vcode, c.year, c.as_of_date FROM valuation_records r
            JOIN valuation_cycles c ON c.id = r.cycle_id
            WHERE r.id = :i
        """), {"i": record_id}).fetchone()
    if not row:
        raise ValueError(f"Valuation record {record_id} not found")
    vcode, year = row[0], int(row[1])
    as_of = pd.Timestamp(row[2])

    isbs = _prepare_isbs(data.get("isbs_raw"), vcode)
    if isbs.empty:
        return {"vcode": vcode, "rows": [], "prior_date": None, "current_date": None}
    bs = isbs[isbs["vSource"] == "Interim BS"]
    if bs.empty:
        return {"vcode": vcode, "rows": [], "prior_date": None, "current_date": None}

    periods = sorted(bs["dtEntry_parsed"].dropna().unique())
    prior_target = pd.Timestamp(f"{year - 1}-12-31")
    prior_date = next((pd.Timestamp(p) for p in reversed(periods) if pd.Timestamp(p) <= prior_target), None)
    current_date = next((pd.Timestamp(p) for p in reversed(periods) if pd.Timestamp(p) <= as_of), None)
    if current_date is None:
        current_date = pd.Timestamp(periods[-1])

    def _grouped(dt):
        if dt is None:
            return pd.DataFrame(columns=["vAccountType", "vAccount", "vDescription", "mAmount"])
        sub = bs[bs["dtEntry_parsed"] == dt]
        return sub.groupby(["vAccountType", "vAccount", "vDescription"], dropna=False)["mAmount"].sum().reset_index()

    prior_g = _grouped(prior_date)
    cur_g = _grouped(current_date)
    merged = pd.merge(
        prior_g, cur_g, how="outer",
        on=["vAccountType", "vAccount", "vDescription"], suffixes=("_prior", "_current"),
    ).fillna({"mAmount_prior": 0.0, "mAmount_current": 0.0})

    type_order = {"Assets": 0, "Liabilities": 1, "Equity": 2}
    merged["_ord"] = merged["vAccountType"].map(lambda t: type_order.get(str(t), 3))
    merged = merged.sort_values(["_ord", "vAccount"])

    rows = []
    for _, r in merged.iterrows():
        prior_v, cur_v = float(r["mAmount_prior"]), float(r["mAmount_current"])
        if abs(prior_v) < 0.005 and abs(cur_v) < 0.005:
            continue
        rows.append({
            "account_type": _s(r["vAccountType"]),
            "account": _s(r["vAccount"]),
            "description": _s(r["vDescription"]),
            "prior": prior_v,
            "current": cur_v,
            "variance": cur_v - prior_v,
        })

    return {
        "vcode": vcode,
        "prior_date": prior_date.strftime("%Y-%m-%d") if prior_date is not None else None,
        "current_date": current_date.strftime("%Y-%m-%d") if current_date is not None else None,
        "rows": rows,
    }


# ============================================================
# Phase 2 — Reviewer Q&A
# ============================================================

def ask_question(engine, record_id: int, question_text: str, username: str,
                 role_label: str = "") -> Dict[str, Any]:
    """Any reviewer can ask; the asking user's review role is kept for display."""
    ensure_valuation_tables(engine)
    if not (question_text or "").strip():
        raise ValueError("Question text is required")
    with engine.connect() as conn:
        exists = conn.execute(
            text("SELECT id FROM valuation_records WHERE id = :i"), {"i": record_id}
        ).fetchone()
    if not exists:
        raise ValueError(f"Valuation record {record_id} not found")
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO valuation_questions (record_id, question_text, asked_by, asked_at, status)
            VALUES (:r, :q, :u, :now, 'open')
        """), {"r": record_id, "q": question_text.strip(),
               "u": f"{username} ({role_label})" if role_label else username, "now": _now()})
    return {"status": "asked"}


def answer_question(engine, question_id: int, answer_text: str, username: str) -> Dict[str, Any]:
    """Asset managers (analyst/admin) answer; re-answering updates the answer."""
    if not (answer_text or "").strip():
        raise ValueError("Answer text is required")
    with engine.begin() as conn:
        result = conn.execute(text("""
            UPDATE valuation_questions
            SET answer_text = :a, answered_by = :u, answered_at = :now,
                status = CASE WHEN status = 'resolved' THEN status ELSE 'answered' END
            WHERE id = :i
        """), {"a": answer_text.strip(), "u": username, "now": _now(), "i": question_id})
    if result.rowcount == 0:
        raise ValueError(f"Question {question_id} not found")
    return {"status": "answered"}


def resolve_question(engine, question_id: int, username: str) -> Dict[str, Any]:
    with engine.begin() as conn:
        result = conn.execute(text("""
            UPDATE valuation_questions SET status = 'resolved' WHERE id = :i
        """), {"i": question_id})
    if result.rowcount == 0:
        raise ValueError(f"Question {question_id} not found")
    return {"status": "resolved"}


# ============================================================
# Phase 2 — Committee approvals (parallel unanimous) + snapshot
# ============================================================

def committee_approve(engine, record_id: int, member_roles: List[str], username: str,
                      data: dict, note: str = "") -> Dict[str, Any]:
    """Record an approval for each committee role the user holds. When all of
    COMMITTEE_ROLES carry an active approval, the record is Approved and a
    snapshot freezes it."""
    ensure_valuation_tables(engine)
    roles = [r for r in member_roles if r in COMMITTEE_ROLES]
    if not roles:
        raise PermissionError("Only Valuation Committee members (President, CEO, CIO) can approve")

    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT status FROM valuation_records WHERE id = :i"), {"i": record_id}
        ).fetchone()
    if not row:
        raise ValueError(f"Valuation record {record_id} not found")
    if row[0] not in ("signed_off", "approved"):
        raise ValueError("The analyst must sign off before the committee approves")

    with engine.begin() as conn:
        for role in roles:
            conn.execute(text("""
                UPDATE valuation_approvals SET active = 0
                WHERE record_id = :r AND member_role = :m AND active = 1
            """), {"r": record_id, "m": role})
            conn.execute(text("""
                INSERT INTO valuation_approvals
                    (record_id, member_role, username, action, note, active, created_at)
                VALUES (:r, :m, :u, 'approve', :n, 1, :now)
            """), {"r": record_id, "m": role, "u": username, "n": note or None, "now": _now()})
        approved_roles = {r[0] for r in conn.execute(text("""
            SELECT member_role FROM valuation_approvals
            WHERE record_id = :r AND active = 1 AND action = 'approve'
        """), {"r": record_id}).fetchall()}

    fully_approved = all(r in approved_roles for r in COMMITTEE_ROLES)
    if fully_approved and row[0] != "approved":
        with engine.begin() as conn:
            conn.execute(text("""
                UPDATE valuation_records SET status = 'approved', updated_at = :now
                WHERE id = :i
            """), {"i": record_id, "now": _now()})
        _save_snapshot(engine, record_id, data, username)

    return {
        "status": "approved" if fully_approved else "pending",
        "approved_roles": sorted(approved_roles),
        "missing_roles": [r for r in COMMITTEE_ROLES if r not in approved_roles],
    }


def committee_return(engine, record_id: int, member_roles: List[str], username: str,
                     note: str) -> Dict[str, Any]:
    """Return a valuation to the asset manager. Requires a note (the reason is
    part of the committee record). Clears all active approvals; a snapshot from
    a prior approval is removed — it no longer represents an approved state."""
    ensure_valuation_tables(engine)
    allowed = [r for r in member_roles if r in COMMITTEE_ROLES or r == RECORDER_ROLE]
    if not allowed:
        raise PermissionError("Only committee members or the recorder (CCO) can return a valuation")
    if not (note or "").strip():
        raise ValueError("A return requires a note explaining what needs to change")

    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT status FROM valuation_records WHERE id = :i"), {"i": record_id}
        ).fetchone()
    if not row:
        raise ValueError(f"Valuation record {record_id} not found")

    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE valuation_approvals SET active = 0 WHERE record_id = :r AND active = 1
        """), {"r": record_id})
        conn.execute(text("""
            INSERT INTO valuation_approvals
                (record_id, member_role, username, action, note, active, created_at)
            VALUES (:r, :m, :u, 'return', :n, 1, :now)
        """), {"r": record_id, "m": allowed[0], "u": username, "n": note.strip(), "now": _now()})
        conn.execute(text("""
            UPDATE valuation_records
            SET status = 'open', signed_off_by = NULL, signed_off_at = NULL, updated_at = :now
            WHERE id = :i
        """), {"i": record_id, "now": _now()})
        conn.execute(text("DELETE FROM valuation_snapshots WHERE record_id = :r"), {"r": record_id})
    return {"status": "returned"}


def _save_snapshot(engine, record_id: int, data: dict, username: str):
    """Freeze the record, review-form data, Q&A and approvals as approved."""
    import json as _json
    from flask_app.serializers import safe_json

    payload: Dict[str, Any] = {"record": get_record(engine, record_id, data)}
    try:
        payload["budget_review"] = get_budget_review(engine, record_id, data)
    except Exception as e:
        payload["budget_review"] = {"error": str(e)}
    try:
        payload["balance_sheet"] = get_balance_sheet(engine, record_id, data)
    except Exception as e:
        payload["balance_sheet"] = {"error": str(e)}
    try:
        from flask_app.services.valuation_nav_service import get_nav
        payload["nav"] = get_nav(engine, record_id)
    except Exception as e:
        payload["nav"] = {"error": str(e)}

    snapshot_json = _json.dumps(safe_json(payload), default=str)
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM valuation_snapshots WHERE record_id = :r"), {"r": record_id})
        conn.execute(text("""
            INSERT INTO valuation_snapshots (record_id, snapshot_json, approved_by, approved_at)
            VALUES (:r, :j, :u, :now)
        """), {"r": record_id, "j": snapshot_json, "u": username, "now": _now()})


def get_snapshot(engine, record_id: int) -> Optional[Dict[str, Any]]:
    import json as _json
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT snapshot_json, approved_by, approved_at
            FROM valuation_snapshots WHERE record_id = :r
        """), {"r": record_id}).fetchone()
    if not row:
        return None
    return {
        "approved_by": row[1],
        "approved_at": _serialize(row[2]),
        "data": _json.loads(row[0]) if row[0] else None,
    }


# ============================================================
# Phase 2 — Committee summary (Analyses 1 & 2)
# ============================================================

def get_committee_summary(engine, cycle_id: int, data: dict) -> Dict[str, Any]:
    """The two analyses the Valuation Committee reviews, live.

    Analysis 1 (Pref NAV): pref balance, accrued pref (accounting + waterfall
    rates through the cycle as-of date), prior-year Pref NAV from the
    valuations feed. Current-cycle Pref NAV lands with the Phase 3 NAV engine.

    Analysis 2 (Methods & values): method / cap / terminal cap / discount /
    NOI / value, this cycle vs prior year, with a simple net-proceeds estimate
    (value less current ISBS debt) pending the full Phase 3 NAV walk.
    """
    from flask_app.services.reports_service import build_roe_summary_row
    from flask_app.services.valuation_nav_service import nav_results_for_cycle
    from compute import get_isbs_debt_balance

    dash = get_cycle_dashboard(engine, cycle_id, data)
    nav_by_vcode = nav_results_for_cycle(engine, cycle_id)
    cycle = dash["cycle"]
    year = int(cycle["year"])
    as_of = pd.Timestamp(cycle["as_of_date"]).date()

    mri_val = data.get("mri_val")
    prior = _prior_rows(mri_val, year - 1)
    acct = data.get("acct")
    inv = data.get("inv")
    wf = data.get("wf")
    isbs_raw = data.get("isbs_raw")

    analysis1, analysis2 = [], []
    for rec in dash["records"]:
        if rec["status"] == "excluded":
            continue
        vcode = rec["vcode"]
        p = prior.get(vcode, {})

        cur_value = rec.get("concluded_value")
        cur_debt = None
        try:
            if isbs_raw is not None:
                cur_debt = float(get_isbs_debt_balance(isbs_raw, vcode) or 0) or None
        except Exception:
            cur_debt = None

        prior_value = p.get("value")
        cap_delta = _delta(rec.get("cap_rate"), p.get("cap_rate"))
        disc_delta = _delta(rec.get("discount_rate"), p.get("discount_rate"))
        value_var = (cur_value - prior_value) if cur_value is not None and prior_value else None
        analysis2.append({
            "vcode": vcode,
            "deal_name": rec["deal_name"],
            "is_child": rec["is_child"],
            "status": rec["status"],
            "open_questions": rec.get("open_questions", 0),
            "approval_count": rec.get("approval_count", 0),
            "prior_method": p.get("method"),
            "method": rec.get("method"),
            "prior_cap_rate": p.get("cap_rate"), "cap_rate": rec.get("cap_rate"),
            "prior_term_cap": p.get("term_cap_rate"), "term_cap": rec.get("term_cap_rate"),
            "prior_discount": p.get("discount_rate"), "discount": rec.get("discount_rate"),
            "prior_noi": p.get("noi"), "noi": rec.get("direct_cap_noi"),
            "prior_value": prior_value, "value": cur_value,
            "value_var": value_var,
            "value_var_pct": (value_var / prior_value) if value_var is not None and prior_value else None,
            "prior_debt": p.get("debt"), "debt": cur_debt,
            "prior_net_proceeds": (p.get("value") - p.get("debt")) if p.get("value") and p.get("debt") is not None else None,
            # Full NAV net proceeds when computed; simple value-less-debt estimate otherwise
            "net_proceeds": nav_by_vcode.get(vcode, {}).get("net_proceeds")
                if vcode in nav_by_vcode
                else ((cur_value - cur_debt) if cur_value is not None and cur_debt is not None else None),
            "has_nav": vcode in nav_by_vcode,
            "direction": ("Up" if value_var > 0 else "Down") if value_var else None,
            "cap_delta": cap_delta,
            "discount_delta": disc_delta,
        })

        # Analysis 1 — parent/standalone deals only (pref lives at the deal)
        if rec["is_child"]:
            continue
        pref_balance = accrued = None
        try:
            roe_row = build_roe_summary_row(
                vcode, rec["deal_name"], acct, inv, as_of, wf_steps=wf, isbs_raw=None)
            if roe_row:
                pref_balance = roe_row["Current Balance"]
                accrued = roe_row["Accrued Pref"]
        except Exception:
            logger.warning(f"Analysis 1 pref computation failed for {vcode}", exc_info=True)
        analysis1.append({
            "vcode": vcode,
            "deal_name": rec["deal_name"],
            "status": rec["status"],
            "pref_balance": pref_balance,
            "accrued_pref": accrued,
            "balance_with_accrual": (pref_balance + accrued) if pref_balance is not None and accrued is not None else None,
            "pref_nav": nav_by_vcode.get(vcode, {}).get("psc_nav"),
            "prior_pref_nav": p.get("pe_nav"),
        })

    return {"cycle": cycle, "analysis1": analysis1, "analysis2": analysis2}


def _prior_rows(mri_val: Optional[pd.DataFrame], prior_year: int) -> Dict[str, dict]:
    out: Dict[str, dict] = {}
    if mri_val is None or mri_val.empty:
        return out
    df = mri_val.copy()
    vcode_col = "vCode" if "vCode" in df.columns else "vcode"
    if vcode_col not in df.columns or "dtValuation" not in df.columns:
        return out
    df["_dt"] = pd.to_datetime(df["dtValuation"], errors="coerce")
    df = df[df["_dt"].dt.year == prior_year]
    for _, r in df.iterrows():
        out[str(r[vcode_col]).strip()] = {
            "method": _s(r.get("vMethod")) or None,
            "cap_rate": _num(r.get("fCapRate")),
            "term_cap_rate": _num(r.get("nTermCapRate")),
            "discount_rate": _num(r.get("nDiscountRateForEquityInterest")),
            "noi": _num(r.get("mAnnualNOI")),
            "value": _num(r.get("mIncomeCapConcludedValue")),
            "debt": _num(r.get("mDebtValue")),
            "pe_nav": _num(r.get("mMezzanineValue")),
        }
    return out


def _delta(cur, prior):
    if cur is None or prior is None:
        return None
    return cur - prior


def generate_committee_workbook(engine, cycle_id: int, data: dict) -> bytes:
    """Two-sheet Excel replacing the manual Valuation Summary workbook."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    summary = get_committee_summary(engine, cycle_id, data)
    year = summary["cycle"]["year"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    curr = "#,##0"
    pct = "0.00%"

    wb = Workbook()

    def _sheet(ws, title_text, cols, rows):
        ws.cell(row=1, column=1, value=title_text).font = Font(bold=True, size=13)
        for ci, (label, _, _) in enumerate(cols, 1):
            c = ws.cell(row=3, column=ci, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        r = 4
        for row in rows:
            for ci, (_, key, fmt) in enumerate(cols, 1):
                v = row.get(key)
                cell = ws.cell(row=r, column=ci, value=v)
                if fmt and isinstance(v, (int, float)):
                    cell.number_format = fmt
            r += 1
        for ci, (label, _, _) in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=3, column=ci).column_letter].width = max(12, min(34, len(label) + 4))
        ws.freeze_panes = "A4"

    yy, py = str(year)[2:], str(year - 1)[2:]
    ws1 = wb.active
    ws1.title = "Pref NAV"
    cols1 = [
        ("Vcode", "vcode", None), ("Deal", "deal_name", None), ("Status", "status", None),
        (f"'{yy} Pref Balance", "pref_balance", curr),
        (f"'{yy} Accrued Pref", "accrued_pref", curr),
        ("Balance w/ Accrual", "balance_with_accrual", curr),
        (f"'{yy} Pref NAV", "pref_nav", curr),
        (f"'{py} Pref NAV", "prior_pref_nav", curr),
    ]
    _sheet(ws1, f"PSC - Property Valuation Analysis (As of 12/31/{year}) - Pref NAV",
           cols1, summary["analysis1"])

    ws2 = wb.create_sheet("Methods & Values")
    cols2 = [
        ("Vcode", "vcode", None), ("Deal", "deal_name", None), ("Status", "status", None),
        (f"'{py} Method", "prior_method", None), (f"'{yy} Method", "method", None),
        (f"'{py} Cap", "prior_cap_rate", pct), (f"'{yy} Cap", "cap_rate", pct),
        (f"'{py} Exit Cap", "prior_term_cap", pct), (f"'{yy} Exit Cap", "term_cap", pct),
        (f"'{py} Disc", "prior_discount", pct), (f"'{yy} Disc", "discount", pct),
        (f"'{py} NOI", "prior_noi", curr), (f"'{yy} NOI", "noi", curr),
        (f"'{py} Value", "prior_value", curr), (f"'{yy} Value", "value", curr),
        ("Var to Prior", "value_var", curr), ("Var %", "value_var_pct", pct),
        (f"'{py} Debt", "prior_debt", curr), (f"'{yy} Debt", "debt", curr),
        ("Net Proceeds (est)", "net_proceeds", curr),
        ("Up/Down", "direction", None),
        ("Cap Chg", "cap_delta", pct), ("Disc Chg", "discount_delta", pct),
    ]
    _sheet(ws2, f"PSC - Property Valuation Analysis (As of 12/31/{year}) - Methods & Values",
           cols2, summary["analysis2"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

