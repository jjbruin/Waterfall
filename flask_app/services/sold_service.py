"""Sold Portfolio service — historical returns from accounting.

Extracts pure logic from sold_portfolio_ui.py (no Streamlit dependency).
"""

import pandas as pd
import numpy as np
import io
from typing import Optional

from loaders import normalize_accounting_feed, build_investmentid_to_vcode
from metrics import xirr, xnpv, calculate_roe


def compute_all_sold_returns(inv_sold: pd.DataFrame, acct: pd.DataFrame,
                             inv: pd.DataFrame) -> pd.DataFrame:
    """Compute pref-equity returns for all sold deals from accounting.

    Extracted from sold_portfolio_ui.py::_compute_all_sold_returns().
    Filters out OP partners, case-insensitive InvestorID grouping,
    computes IRR/ROE/MOIC per deal, plus portfolio total from combined pool.

    Returns DataFrame with columns: Investment Name, Acquisition Date, Sale Date,
    Total Contributions, Total Distributions, IRR, ROE, MOIC, _is_deal_total, vcode.
    """
    if "is_contribution" not in acct.columns:
        acct = normalize_accounting_feed(acct)

    inv_to_vcode = build_investmentid_to_vcode(inv)
    vcode_to_iids = {}
    for iid, vc in inv_to_vcode.items():
        vcode_to_iids.setdefault(vc, []).append(iid)

    all_rows = []
    portfolio_cashflows = []
    portfolio_capital_events = []
    portfolio_cf_distributions = []

    for _, deal_row in inv_sold.iterrows():
        deal_name = str(deal_row.get("Investment_Name", "")).strip()
        deal_vcode = str(deal_row.get("vcode", "")).strip()

        acq_raw = deal_row.get("Acquisition_Date", None)
        sale_raw = deal_row.get("Sale_Date", None)
        acq_date = pd.to_datetime(acq_raw, errors="coerce")
        sale_date = pd.to_datetime(sale_raw, errors="coerce")
        acq_str = acq_date.strftime("%m/%d/%Y") if pd.notna(acq_date) else ""
        sale_str = sale_date.strftime("%m/%d/%Y") if pd.notna(sale_date) else ""

        deal_iids = set(vcode_to_iids.get(deal_vcode, []))
        direct_iid = str(deal_row.get("InvestmentID", "")).strip()
        if direct_iid:
            deal_iids.add(direct_iid)

        if not deal_iids or acct is None or acct.empty:
            continue

        deal_acct = acct[acct["InvestmentID"].isin(deal_iids)].copy()
        if deal_acct.empty:
            continue

        deal_acct = deal_acct[~deal_acct["InvestorID"].str.upper().str.startswith("OP")].copy()
        if deal_acct.empty:
            continue

        deal_acct["_investor_key"] = deal_acct["InvestorID"].str.upper()

        pref_cashflows = []
        pref_capital_events = []
        pref_cf_distributions = []

        for _, grp in deal_acct.groupby("_investor_key"):
            _, _, _, _, _, cashflows, capital_events, cf_dists = (
                _compute_partner_metrics(grp)
            )
            pref_cashflows.extend(cashflows)
            pref_capital_events.extend(capital_events)
            pref_cf_distributions.extend(cf_dists)

        if not pref_cashflows:
            continue

        contribs = sum(abs(a) for _, a in pref_cashflows if a < 0)
        distribs = sum(a for _, a in pref_cashflows if a > 0)
        irr_val = xirr(pref_cashflows) if len(pref_cashflows) >= 2 else None

        start = min(d for d, _ in pref_cashflows)
        end = max(d for d, _ in pref_cashflows)
        roe_val = calculate_roe(pref_capital_events, pref_cf_distributions, start, end)
        moic_val = distribs / contribs if contribs > 0 else 0.0

        all_rows.append({
            "Investment Name": deal_name,
            "vcode": deal_vcode,
            "Acquisition Date": acq_str,
            "Sale Date": sale_str,
            "Total Contributions": contribs,
            "Total Distributions": distribs,
            "IRR": irr_val,
            "ROE": roe_val,
            "MOIC": moic_val,
            "_is_deal_total": False,
        })

        portfolio_cashflows.extend(pref_cashflows)
        portfolio_capital_events.extend(pref_capital_events)
        portfolio_cf_distributions.extend(pref_cf_distributions)

    if not all_rows:
        return pd.DataFrame()

    # Portfolio total row
    port_contribs = sum(abs(a) for _, a in portfolio_cashflows if a < 0)
    port_distribs = sum(a for _, a in portfolio_cashflows if a > 0)
    port_irr = xirr(portfolio_cashflows) if len(portfolio_cashflows) >= 2 else None

    port_start = min(d for d, _ in portfolio_cashflows)
    port_end = max(d for d, _ in portfolio_cashflows)
    port_roe = calculate_roe(portfolio_capital_events, portfolio_cf_distributions, port_start, port_end)
    port_moic = port_distribs / port_contribs if port_contribs > 0 else 0.0

    all_rows.append({
        "Investment Name": "Portfolio Total",
        "vcode": "",
        "Acquisition Date": "",
        "Sale Date": "",
        "Total Contributions": port_contribs,
        "Total Distributions": port_distribs,
        "IRR": port_irr,
        "ROE": port_roe,
        "MOIC": port_moic,
        "_is_deal_total": True,
    })

    return pd.DataFrame(all_rows)


def build_deal_detail(vcode: str, inv_sold: pd.DataFrame, acct: pd.DataFrame,
                      inv: pd.DataFrame) -> dict:
    """Build cashflow detail table for a single sold deal.

    Returns dict with:
      - rows: list of dicts (Date, InvestorID, MajorType, Typename, Capital, Amount,
              Cashflow (XIRR), Capital Balance)
      - summary: dict with IRR, ROE, MOIC, Total Contributions, Total Distributions
    """
    if "is_contribution" not in acct.columns:
        acct = normalize_accounting_feed(acct)

    inv_to_vcode = build_investmentid_to_vcode(inv)
    vcode_to_iids = {}
    for iid, vc in inv_to_vcode.items():
        vcode_to_iids.setdefault(vc, []).append(iid)

    # Find deal by vcode
    match = inv_sold[inv_sold["vcode"].astype(str).str.strip() == str(vcode)]
    if match.empty:
        return {"rows": [], "summary": {}}

    deal_row = match.iloc[0]
    deal_vcode = str(deal_row.get("vcode", "")).strip()

    deal_iids = set(vcode_to_iids.get(deal_vcode, []))
    direct_iid = str(deal_row.get("InvestmentID", "")).strip()
    if direct_iid:
        deal_iids.add(direct_iid)

    if not deal_iids:
        return {"rows": [], "summary": {}}

    deal_acct = acct[acct["InvestmentID"].isin(deal_iids)].copy()
    if deal_acct.empty:
        return {"rows": [], "summary": {}}

    # Pref equity only
    deal_acct = deal_acct[~deal_acct["InvestorID"].str.upper().str.startswith("OP")].copy()
    if deal_acct.empty:
        return {"rows": [], "summary": {}}

    # Build detail rows
    rows = []
    for _, r in deal_acct.iterrows():
        amt = float(r["Amt"])
        if r["is_contribution"]:
            cashflow = -abs(amt)
        elif r["is_distribution"]:
            cashflow = abs(amt)
        else:
            cashflow = 0.0

        rows.append({
            "Date": str(r["EffectiveDate"]),
            "InvestorID": r["InvestorID"],
            "MajorType": r["MajorType"],
            "Typename": r.get("Typename", ""),
            "Capital": r["Capital"],
            "Amount": amt,
            "Cashflow (XIRR)": cashflow,
        })

    if not rows:
        return {"rows": [], "summary": {}}

    df = pd.DataFrame(rows).sort_values("Date").reset_index(drop=True)

    # Running capital balance
    balance = 0.0
    balances = []
    for _, r in df.iterrows():
        cf = r["Cashflow (XIRR)"]
        if cf < 0:
            balance += abs(cf)
        elif r["Capital"] == "Y" and cf > 0:
            balance = max(0.0, balance - cf)
        balances.append(balance)
    df["Capital Balance"] = balances

    # Compute summary metrics — use original EffectiveDate (date objects) for XIRR
    # The df "Date" column is stringified for JSON; rebuild cashflows from deal_acct
    cashflows = []
    for _, r in deal_acct.iterrows():
        amt = float(r["Amt"])
        if r["is_contribution"]:
            cashflows.append((r["EffectiveDate"], -abs(amt)))
        elif r["is_distribution"]:
            cashflows.append((r["EffectiveDate"], abs(amt)))
    contribs = sum(abs(cf) for _, cf in cashflows if cf < 0)
    distribs = sum(cf for _, cf in cashflows if cf > 0)
    irr_val = xirr(cashflows) if len(cashflows) >= 2 else None
    moic_val = distribs / contribs if contribs > 0 else 0.0

    # ROE from capital events and CF distributions (using date objects from deal_acct)
    capital_events = []
    cf_dists = []
    for _, r in deal_acct.iterrows():
        amt = float(r["Amt"])
        if r["is_contribution"]:
            cf = -abs(amt)
            capital_events.append((r["EffectiveDate"], cf))
        elif r["is_distribution"]:
            cf = abs(amt)
            if r["is_capital"]:
                capital_events.append((r["EffectiveDate"], cf))
            else:
                cf_dists.append((r["EffectiveDate"], cf))

    if cashflows:
        start = min(d for d, _ in cashflows)
        end = max(d for d, _ in cashflows)
        roe_val = calculate_roe(capital_events, cf_dists, start, end)
    else:
        roe_val = 0.0

    detail_rows = df.to_dict(orient="records")
    summary = {
        "Total Contributions": contribs,
        "Total Distributions": distribs,
        "IRR": irr_val,
        "ROE": roe_val,
        "MOIC": moic_val,
    }

    return {"rows": detail_rows, "summary": summary}


def generate_sold_excel(df: pd.DataFrame) -> bytes:
    """Generate formatted Excel for sold portfolio summary.

    Extracted from sold_portfolio_ui.py::_generate_sold_excel().
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Sold Portfolio Returns"

    display_cols = [c for c in df.columns if c not in ("_is_deal_total", "vcode")]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    bold_font = Font(bold=True)
    top_border = Border(top=Side(style="medium"))

    currency_cols = {"Total Contributions", "Total Distributions"}
    pct_cols = {"IRR", "ROE"}
    moic_cols = {"MOIC"}

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_total = bool(row.get("_is_deal_total", False))
        for col_idx, col_name in enumerate(display_cols, 1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name in currency_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = "$#,##0"
            elif col_name in pct_cols:
                cell.value = float(val) if pd.notna(val) else None
                cell.number_format = "0.00%"
            elif col_name in moic_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = '0.00"x"'
            else:
                cell.value = val

            if is_total:
                cell.font = bold_font
                cell.border = top_border

    for col_idx, col_name in enumerate(display_cols, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, len(df) + 2):
            cv = ws.cell(row=row_idx, column=col_idx).value
            if cv is not None:
                max_len = max(max_len, len(str(cv)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_detail_excel(detail_rows: list[dict], deal_name: str,
                          summary: dict) -> bytes:
    """Generate formatted Excel for deal activity detail.

    Extracted from sold_portfolio_ui.py::_generate_detail_excel().
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Detail"

    if not detail_rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    df = pd.DataFrame(detail_rows)
    cols = list(df.columns)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    currency_cols = {"Amount", "Cashflow (XIRR)", "Capital Balance"}
    date_cols = {"Date"}

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci)
            if col in date_cols:
                cell.value = val
                cell.number_format = "MM/DD/YYYY"
            elif col in currency_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = "$#,##0"
            else:
                cell.value = val

    # Summary section below
    summary_row = len(df) + 3
    bold = Font(bold=True)
    ws.cell(row=summary_row, column=1, value="Computed Returns").font = bold

    if summary:
        labels = [
            ("Total Contributions", summary.get("Total Contributions", 0), "$#,##0"),
            ("Total Distributions", summary.get("Total Distributions", 0), "$#,##0"),
            ("IRR", summary.get("IRR"), "0.00%"),
            ("ROE", summary.get("ROE"), "0.00%"),
            ("MOIC", summary.get("MOIC", 0), '0.00"x"'),
        ]
        for i, (label, val, fmt) in enumerate(labels):
            r = summary_row + 1 + i
            ws.cell(row=r, column=1, value=label).font = bold
            cell = ws.cell(row=r, column=2)
            cell.value = float(val) if val is not None and pd.notna(val) else None
            cell.number_format = fmt

    for ci, col in enumerate(cols, 1):
        max_len = len(str(col))
        for ri in range(2, len(df) + 2):
            cv = ws.cell(row=ri, column=ci).value
            if cv is not None:
                max_len = max(max_len, len(str(cv)))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _compute_partner_metrics(partner_acct):
    """Compute metrics for a single partner from accounting data.

    Returns:
        (contributions, distributions, irr, roe, moic,
         cashflows, capital_events, cf_distributions)

    Capital event identification uses Typename (not the Capital flag from
    accounting, which is unreliable for sale events).  Typenames containing
    "return of capital" or "realized gain" are capital events that reduce
    capital outstanding; everything else is a CF distribution (operating income).
    """
    CAPITAL_TYPENAMES = {"return of capital", "realized gain"}

    cashflows = []
    capital_events = []
    cf_distributions = []

    for _, r in partner_acct.iterrows():
        d = r["EffectiveDate"]
        amt = float(r["Amt"])

        if r["is_contribution"]:
            cf = -abs(amt)
            cashflows.append((d, cf))
            capital_events.append((d, cf))
        elif r["is_distribution"]:
            cf = abs(amt)
            cashflows.append((d, cf))
            tn = str(r.get("Typename", "")).lower()
            is_cap = any(kw in tn for kw in CAPITAL_TYPENAMES)
            if is_cap:
                capital_events.append((d, cf))
            else:
                cf_distributions.append((d, cf))

    contributions = sum(abs(a) for _, a in cashflows if a < 0)
    distributions = sum(a for _, a in cashflows if a > 0)

    irr_val = xirr(cashflows) if len(cashflows) >= 2 else None

    if cashflows:
        start = min(d for d, _ in cashflows)
        end = max(d for d, _ in cashflows)
        roe_val = calculate_roe(capital_events, cf_distributions, start, end)
    else:
        roe_val = 0.0

    moic_val = distributions / contributions if contributions > 0 else 0.0

    return contributions, distributions, irr_val, roe_val, moic_val, cashflows, capital_events, cf_distributions


def compute_net_waterfall_for_deal(deal_acct: pd.DataFrame, inv: pd.DataFrame,
                                   deal_vcode: str, assumptions: dict) -> dict:
    """Run simplified net waterfall for a single deal from investor perspective.

    Walks through accounting events chronologically, applying AM fee, expenses,
    pref accrual, and promote split to compute net-of-fees cashflows.
    """
    if "is_contribution" not in deal_acct.columns:
        deal_acct = normalize_accounting_feed(deal_acct)

    ownership_pct = assumptions["ownership_pct"]
    am_fee_pct = assumptions["am_fee_pct"]
    hurdle_rate = assumptions["hurdle_rate"]
    promote_pct = assumptions["promote_pct"]
    annual_expenses = assumptions["annual_expenses"]

    # Filter to pref equity (no OP)
    pref = deal_acct[~deal_acct["InvestorID"].str.upper().str.startswith("OP")].copy()
    if pref.empty:
        return {"net_irr": None, "net_roe": 0.0, "net_moic": 0.0,
                "net_contributions": 0.0, "net_distributions": 0.0,
                "waterfall_detail": [], "net_cashflows": []}

    # Build sorted event list
    # Capital events identified by Typename (not the Capital flag from accounting):
    #   "Distribution: Return of Capital" or "Distribution: Realized Gain"
    # These are the events that trigger the promote XIRR hurdle test.
    CAPITAL_TYPENAMES = {"return of capital", "realized gain"}
    events = []
    for _, r in pref.iterrows():
        amt = float(r["Amt"])
        tn = str(r.get("Typename", "")).lower()
        is_cap_event = any(kw in tn for kw in CAPITAL_TYPENAMES)
        events.append({
            "date": r["EffectiveDate"],
            "is_contribution": bool(r["is_contribution"]),
            "is_distribution": bool(r["is_distribution"]),
            "is_capital_event": is_cap_event,
            "gross_amount": amt,
            "typename": str(r.get("Typename", "")),
        })
    events.sort(key=lambda e: e["date"])

    capital_balance = 0.0
    pref_accrued = 0.0
    last_date = None
    net_cashflows = []
    waterfall_detail = []

    for ev in events:
        d = ev["date"]
        gross = ev["gross_amount"]
        scaled = gross * ownership_pct

        if ev["is_contribution"]:
            contrib = abs(scaled)
            capital_balance += contrib
            net_cashflows.append((d, -contrib))
            waterfall_detail.append({
                "Date": str(d),
                "Event": "Contribution",
                "Gross Amount": -abs(gross),
                "Ownership %": ownership_pct,
                "Scaled Amount": -contrib,
                "Acq Fee Paid": 0.0,
                "AM Fee": 0.0,
                "Expenses": 0.0,
                "Available": 0.0,
                "Pref Accrued": 0.0,
                "Pref Paid": 0.0,
                "Capital Returned": 0.0,
                "Excess": 0.0,
                "Hurdle Amount": 0.0,
                "Promote (GP)": 0.0,
                "Net to Investor": -contrib,
                "Capital Balance": capital_balance,
                "Pref Balance": pref_accrued,
            })
            last_date = d
            continue

        if ev["is_distribution"]:
            scaled_dist = abs(scaled)
            is_acq_fee = "acquisition fee" in ev["typename"].lower()

            # Days since last event
            if last_date is not None:
                days_period = max((d - last_date).days, 0)
            else:
                days_period = 0

            # Acquisition Fee: entire distribution consumed by the fee.
            # Investor gets $0. Pref still accrues. AM fees and expenses
            # for this period are foregone (matching Excel formula behavior).
            if is_acq_fee:
                # Accrue pref for elapsed time (time still passes)
                if capital_balance > 0 and days_period > 0:
                    pref_accrued += capital_balance * hurdle_rate * days_period / 365.0

                # Append $0 cashflow so Python XIRR uses same inputs as Excel
                net_cashflows.append((d, 0.0))
                waterfall_detail.append({
                    "Date": str(d),
                    "Event": "Acquisition Fee",
                    "Gross Amount": abs(gross),
                    "Ownership %": ownership_pct,
                    "Scaled Amount": scaled_dist,
                    "Acq Fee Paid": scaled_dist,
                    "AM Fee": 0.0,
                    "Expenses": 0.0,
                    "Available": 0.0,
                    "Pref Accrued": pref_accrued,
                    "Pref Paid": 0.0,
                    "Capital Returned": 0.0,
                    "Excess": 0.0,
                    "Hurdle Amount": 0.0,
                    "Promote (GP)": 0.0,
                    "Net to Investor": 0.0,
                    "Capital Balance": capital_balance,
                    "Pref Balance": pref_accrued,
                })
                last_date = d
                continue

            # Accrue pref since last event
            if capital_balance > 0 and days_period > 0:
                pref_accrued += capital_balance * hurdle_rate * days_period / 365.0

            # Compute fees (independent, then cap sequentially)
            am_fee = capital_balance * am_fee_pct * days_period / 365.0 if capital_balance > 0 else 0.0
            am_fee = min(am_fee, scaled_dist)  # cap at distribution
            expenses = annual_expenses * days_period / 365.0
            expenses = min(expenses, max(0.0, scaled_dist - am_fee))  # cap at remainder
            available = scaled_dist - am_fee - expenses

            # Pay accrued pref
            pref_paid = min(pref_accrued, available)
            pref_accrued -= pref_paid
            remaining = available - pref_paid

            # Return capital if capital event (identified by Typename)
            capital_returned = 0.0
            if ev["is_capital_event"] and remaining > 0:
                capital_returned = min(capital_balance, remaining)
                capital_balance -= capital_returned
                remaining -= capital_returned

            # Promote split — different logic for capital vs CF distributions:
            #
            # Capital events (sale/refi): xnpv hurdle test.
            #   hurdle_amount = -xnpv(hurdle_rate, prior_net_cfs) * (1+hurdle)^years
            #   promote = max(0, total_distribution - hurdle_amount) * promote_pct
            #   Promote is deducted from total distribution (can reduce pref/cap portions).
            #
            # CF distributions: promote on Excess only.
            #   Pref paid IS the hurdle return (accrued at hurdle rate on capital).
            #   Promote = remaining (Excess after pref) * promote_pct.
            net_before_promote = pref_paid + capital_returned + remaining
            hurdle_amount = 0.0

            if ev["is_capital_event"]:
                # Capital event: xnpv hurdle determines how much exceeds hurdle IRR
                if net_cashflows and hurdle_rate > 0 and net_before_promote > 0:
                    npv_prior = xnpv(hurdle_rate, net_cashflows)
                    if npv_prior < 0:
                        t0 = net_cashflows[0][0]
                        years = (d - t0).days / 365.0
                        hurdle_amount = -npv_prior * ((1 + hurdle_rate) ** years)

                excess_over_hurdle = max(0.0, net_before_promote - hurdle_amount)
                gp_promote = excess_over_hurdle * promote_pct
                net_to_investor = net_before_promote - gp_promote
            else:
                # CF distribution: pref paid IS the hurdle return for this period.
                # Promote only on Excess (profit after pref).
                gp_promote = remaining * promote_pct
                net_to_investor = pref_paid + capital_returned + remaining - gp_promote
            net_cashflows.append((d, net_to_investor))

            waterfall_detail.append({
                "Date": str(d),
                "Event": "Capital Distribution" if ev["is_capital_event"] else "CF Distribution",
                "Gross Amount": abs(gross),
                "Ownership %": ownership_pct,
                "Scaled Amount": scaled_dist,
                "Acq Fee Paid": 0.0,
                "AM Fee": am_fee,
                "Expenses": expenses,
                "Available": available,
                "Pref Accrued": pref_accrued + pref_paid,  # show pre-payment balance
                "Pref Paid": pref_paid,
                "Capital Returned": capital_returned,
                "Excess": remaining,
                "Hurdle Amount": hurdle_amount,
                "Promote (GP)": gp_promote,
                "Net to Investor": net_to_investor,
                "Capital Balance": capital_balance,
                "Pref Balance": pref_accrued,
            })
            last_date = d

    # Compute net metrics
    if len(net_cashflows) >= 2:
        net_irr = xirr(net_cashflows)
    else:
        net_irr = None

    net_contribs = sum(abs(a) for _, a in net_cashflows if a < 0)
    net_distribs = sum(a for _, a in net_cashflows if a > 0)
    net_moic = net_distribs / net_contribs if net_contribs > 0 else 0.0

    # ROE: Only "CF Distribution" events count as operating income (numerator).
    # Capital events (sale/refi) go entirely to capital_events — pref paid at
    # sale and gain on sale are NOT operating income.
    # Net capital returned is capped at Net to Investor since promote reduces
    # what the investor actually gets back.
    net_capital_events = []
    net_cf_dists = []
    for row in waterfall_detail:
        d = pd.to_datetime(row["Date"]).date()
        net_inv = row["Net to Investor"]
        if row["Event"] == "Contribution":
            net_capital_events.append((d, -abs(row["Scaled Amount"])))
        elif row["Event"] == "CF Distribution":
            # Operating income — all goes to ROE numerator
            if net_inv > 0.005:
                net_cf_dists.append((d, net_inv))
        elif row["Event"] == "Capital Distribution":
            # Sale/refi — capital event only (reduces weighted avg capital)
            net_cap_ret = min(row["Capital Returned"], max(0.0, net_inv))
            if net_cap_ret > 0.005:
                net_capital_events.append((d, net_cap_ret))

    if net_cashflows:
        start = min(d for d, _ in net_cashflows)
        end = max(d for d, _ in net_cashflows)
        net_roe = calculate_roe(net_capital_events, net_cf_dists, start, end)
    else:
        net_roe = 0.0

    return {
        "net_irr": net_irr,
        "net_roe": net_roe,
        "net_moic": net_moic,
        "net_contributions": net_contribs,
        "net_distributions": net_distribs,
        "waterfall_detail": waterfall_detail,
        "net_cashflows": net_cashflows,
    }


def compute_all_net_returns(inv_sold: pd.DataFrame, acct: pd.DataFrame,
                            inv: pd.DataFrame, assumptions: dict) -> dict:
    """Compute net returns for all sold deals with fee/promote waterfall.

    Returns dict with deal_results (list), portfolio metrics, and assumptions.
    """
    if "is_contribution" not in acct.columns:
        acct = normalize_accounting_feed(acct)

    inv_to_vcode = build_investmentid_to_vcode(inv)
    vcode_to_iids = {}
    for iid, vc in inv_to_vcode.items():
        vcode_to_iids.setdefault(vc, []).append(iid)

    deal_results = []

    for _, deal_row in inv_sold.iterrows():
        deal_name = str(deal_row.get("Investment_Name", "")).strip()
        deal_vcode = str(deal_row.get("vcode", "")).strip()

        deal_iids = set(vcode_to_iids.get(deal_vcode, []))
        direct_iid = str(deal_row.get("InvestmentID", "")).strip()
        if direct_iid:
            deal_iids.add(direct_iid)

        if not deal_iids or acct.empty:
            continue

        deal_acct = acct[acct["InvestmentID"].isin(deal_iids)].copy()
        if deal_acct.empty:
            continue

        result = compute_net_waterfall_for_deal(deal_acct, inv, deal_vcode, assumptions)

        if not result["net_cashflows"]:
            continue

        deal_results.append({
            "vcode": deal_vcode,
            "Investment Name": deal_name,
            "Net IRR": result["net_irr"],
            "Net ROE": result["net_roe"],
            "Net MOIC": result["net_moic"],
            "Net Contributions": result["net_contributions"],
            "Net Distributions": result["net_distributions"],
            "waterfall_detail": result["waterfall_detail"],
        })

    # Portfolio-level metrics — full waterfall recomputation at portfolio level.
    # Matches the Excel Portfolio Detail sheet: portfolio-level capital balance,
    # pref accrual, AM fee, and dynamic expenses based on active deal count.
    #
    # Portfolio annual expenses = (3 + 0.25 × active_deals) × per_deal_expenses
    # Active deals = deals with capital balance > 0.
    hurdle_rate = assumptions.get("hurdle_rate", 0)
    promote_pct = assumptions.get("promote_pct", 0)
    am_fee_pct = assumptions.get("am_fee_pct", 0)
    annual_expenses = assumptions.get("annual_expenses", 0)

    # Collect all events across deals, tagged with deal name, sort chronologically
    all_port_events = []
    for dr in deal_results:
        deal_name = dr["Investment Name"]
        for row in dr["waterfall_detail"]:
            all_port_events.append((deal_name, row))
    all_port_events.sort(key=lambda x: pd.to_datetime(x[1]["Date"]))

    if all_port_events:
        port_net_cfs = []       # (date, net_amount) for XIRR
        port_net_capital = []   # for ROE
        port_net_cf_dists = []  # for ROE

        port_cap_balance = 0.0
        port_pref_balance = 0.0
        last_port_date = None
        deal_last_cap = {}      # deal_name -> last Capital Balance (for active count)

        for deal_name, row in all_port_events:
            d = pd.to_datetime(row["Date"]).date()
            event = row["Event"]
            scaled = float(row["Scaled Amount"])

            # Count active deals BEFORE updating this event's balance
            active_deals = sum(1 for b in deal_last_cap.values() if b > 0.005)
            deal_last_cap[deal_name] = float(row.get("Capital Balance", 0))

            if event == "Contribution":
                port_cap_balance += abs(scaled)
                port_net_cfs.append((d, scaled))  # negative
                port_net_capital.append((d, scaled))
                last_port_date = d
                continue

            if event == "Acquisition Fee":
                # Pref accrues during acquisition fee periods (deferred to next dist)
                port_net_cfs.append((d, 0.0))
                last_port_date = d
                continue

            # ── Distribution: full portfolio-level waterfall ──
            scaled_dist = abs(scaled)
            days_p = (d - last_port_date).days if last_port_date else 0

            # AM Fee on portfolio capital balance
            am_fee = port_cap_balance * am_fee_pct * days_p / 365.0 if port_cap_balance > 0 else 0.0
            am_fee = min(am_fee, scaled_dist)

            # Dynamic expenses: (3 + 0.25 × active_deals) × annual_expenses
            port_annual_exp = (3.0 + 0.25 * active_deals) * annual_expenses
            expenses = port_annual_exp * days_p / 365.0
            expenses = min(expenses, max(0.0, scaled_dist - am_fee))

            available = scaled_dist - am_fee - expenses

            # Pref accrual on portfolio capital
            if port_cap_balance > 0 and days_p > 0:
                port_pref_balance += port_cap_balance * hurdle_rate * days_p / 365.0
            pref_paid = min(port_pref_balance, available)
            port_pref_balance -= pref_paid
            remaining = available - pref_paid

            # Capital return (capital events only)
            is_cap_event = (event == "Capital Distribution")
            cap_returned = 0.0
            if is_cap_event and remaining > 0:
                cap_returned = min(port_cap_balance, remaining)
                port_cap_balance -= cap_returned
                remaining -= cap_returned

            # Promote (hybrid: capital events use xnpv hurdle, CF uses Excess)
            net_before_promote = pref_paid + cap_returned + remaining

            if is_cap_event:
                hurdle_amount = 0.0
                if port_net_cfs and hurdle_rate > 0 and net_before_promote > 0:
                    npv_prior = xnpv(hurdle_rate, port_net_cfs)
                    if npv_prior < 0:
                        t0 = port_net_cfs[0][0]
                        years = (d - t0).days / 365.0
                        hurdle_amount = -npv_prior * ((1 + hurdle_rate) ** years)
                excess_over_hurdle = max(0.0, net_before_promote - hurdle_amount)
                gp_promote = excess_over_hurdle * promote_pct
                net_to_investor = net_before_promote - gp_promote
            else:
                gp_promote = remaining * promote_pct
                net_to_investor = net_before_promote - gp_promote
            port_net_cfs.append((d, net_to_investor))

            # ROE: only CF Distribution events count as operating income
            if is_cap_event:
                net_cap_ret = min(cap_returned, max(0.0, net_to_investor))
                if net_cap_ret > 0.005:
                    port_net_capital.append((d, net_cap_ret))
            else:
                if net_to_investor > 0.005:
                    port_net_cf_dists.append((d, net_to_investor))

            last_port_date = d

        port_irr = xirr(port_net_cfs) if len(port_net_cfs) >= 2 else None
        port_contribs = sum(abs(a) for _, a in port_net_cfs if a < 0)
        port_distribs = sum(a for _, a in port_net_cfs if a > 0)
        port_moic = port_distribs / port_contribs if port_contribs > 0 else 0.0

        port_start = min(d for d, _ in port_net_cfs)
        port_end = max(d for d, _ in port_net_cfs)
        port_roe = calculate_roe(port_net_capital, port_net_cf_dists, port_start, port_end)
    else:
        port_irr = None
        port_contribs = 0.0
        port_distribs = 0.0
        port_moic = 0.0
        port_roe = 0.0

    return {
        "deal_results": deal_results,
        "portfolio": {
            "Net IRR": port_irr,
            "Net ROE": port_roe,
            "Net MOIC": port_moic,
            "Net Contributions": port_contribs,
            "Net Distributions": port_distribs,
        },
        "assumptions": assumptions,
    }


def generate_net_returns_excel(gross_df: pd.DataFrame, net_result: dict) -> bytes:
    """Generate multi-sheet Excel with gross+net summary and per-deal waterfall detail.

    Detail sheets are built first so the Summary can reference their formula cells
    (Net IRR, Net MOIC, Net Contributions, Net Distributions), ensuring consistency.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    bold_font = Font(bold=True)
    top_border = Border(top=Side(style="medium"))

    # ── Phase 1: Build per-deal detail sheets first ──
    # Track cell references for each deal so Summary can link to them.
    # detail_refs[vcode] = {"sheet": sheet_name, "irr_cell": "B23", "moic_cell": ...}
    detail_refs = {}

    # Per-deal waterfall sheets — formula-driven for auditability
    # Users can change assumptions in B2/D2/F2/H2/J2 and the entire sheet recalculates.
    #
    # Column map (row 4 header, data from row 5):
    #   A=Date  B=Event  C=Gross  D=Own%  E=Scaled  F=AcqFeePaid
    #   G=AMFee  H=Expenses  I=Available  J=PrefAccrued  K=PrefPaid
    #   L=CapReturned  M=Excess  N=HurdleAmount  O=Promote  P=NetToInvestor
    #   Q=CapBalance  R=PrefBalance
    #
    # Assumption cells: B2=Ownership%, D2=AM Fee%, F2=Hurdle%, H2=Promote%, J2=AnnualExp

    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    note_font = Font(italic=True, color="808080")
    assumptions = net_result.get("assumptions", {})

    used_names = set()
    for dr in net_result["deal_results"]:
        detail = dr.get("waterfall_detail", [])
        if not detail:
            continue

        sheet_name = dr["Investment Name"][:31]
        base = sheet_name
        counter = 2
        while sheet_name in used_names:
            sheet_name = base[:28] + f" ({counter})"
            counter += 1
        used_names.add(sheet_name)

        dws = wb.create_sheet(title=sheet_name)

        # ── Assumptions block (rows 1-2) ──
        dws.cell(row=1, column=1, value="Assumptions (editable — formulas update automatically)").font = bold_font
        assumption_defs = [
            # (col, label, value, format)
            (1, "Ownership %", assumptions.get("ownership_pct", 0), "0.00%"),
            (3, "AM Fee %", assumptions.get("am_fee_pct", 0), "0.00%"),
            (5, "Hurdle Rate %", assumptions.get("hurdle_rate", 0), "0.00%"),
            (7, "Promote %", assumptions.get("promote_pct", 0), "0.00%"),
            (9, "Annual Expenses", assumptions.get("annual_expenses", 0), "$#,##0"),
        ]
        for col, label, val, fmt in assumption_defs:
            dws.cell(row=2, column=col, value=label).font = bold_font
            c = dws.cell(row=2, column=col + 1, value=val)
            c.number_format = fmt

        # Absolute references to assumption cells
        OWN = "$B$2"   # Ownership %
        AMF = "$D$2"   # AM Fee %
        HUR = "$F$2"   # Hurdle Rate %
        PRO = "$H$2"   # Promote %
        EXP = "$J$2"   # Annual Expenses

        # ── Column headers (row 4) ──
        HDR = 4
        detail_cols = [
            "Date", "Event", "Gross Amount", "Ownership %", "Scaled Amount",
            "Acq Fee Paid", "AM Fee", "Expenses", "Available", "Pref Accrued",
            "Pref Paid", "Capital Returned", "Excess", "Hurdle Amount",
            "Promote (GP)", "Net to Investor", "Capital Balance", "Pref Balance",
        ]
        for ci, col in enumerate(detail_cols, 1):
            cell = dws.cell(row=HDR, column=ci, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Add formula-explanation comments on headers
        dws.cell(row=HDR, column=5).comment = Comment("= Gross Amount × Ownership %", "Formula")
        dws.cell(row=HDR, column=6).comment = Comment(
            "= Scaled Amount for Acquisition Fee rows; 0 otherwise.\n"
            "Consumes the entire distribution — nothing left for the waterfall.", "Formula")
        dws.cell(row=HDR, column=7).comment = Comment(
            "= prior Capital Balance × AM Fee % × Days / 365\n"
            "Capped at Scaled Amount − Acq Fee Paid.", "Formula")
        dws.cell(row=HDR, column=8).comment = Comment(
            "= Annual Expenses × Days / 365\n"
            "Capped at Scaled − Acq Fee − AM Fee.", "Formula")
        dws.cell(row=HDR, column=9).comment = Comment(
            "= Scaled Amount − Acq Fee Paid − AM Fee − Expenses", "Formula")
        dws.cell(row=HDR, column=10).comment = Comment(
            "= prior Pref Balance + prior Capital Balance × Hurdle Rate % × Days / 365", "Formula")
        dws.cell(row=HDR, column=11).comment = Comment("= MIN(Pref Accrued, Available)", "Formula")
        dws.cell(row=HDR, column=12).comment = Comment(
            "= MIN(prior Capital Balance, Available − Pref Paid)\n"
            "Only on Capital Distribution events; 0 for CF Distributions.", "Formula")
        dws.cell(row=HDR, column=13).comment = Comment("= Available − Pref Paid − Capital Returned", "Formula")
        dws.cell(row=HDR, column=14).comment = Comment(
            "Amount needed to deliver hurdle rate IRR to investor.\n"
            "= -XNPV(hurdle_rate, prior net cashflows) × (1+hurdle)^years\n"
            "Promote only applies to the excess above this amount.", "Formula")
        dws.cell(row=HDR, column=15).comment = Comment(
            "Capital Distribution: MAX(0, (K+L+M) − Hurdle Amount) × Promote %\n"
            "  → xnpv hurdle test on total distribution above cumulative hurdle IRR.\n"
            "CF Distribution: Excess × Promote %\n"
            "  → Pref Paid IS the hurdle return; promote only on profit.\n"
            "Acquisition Fee: always 0.", "Formula")
        dws.cell(row=HDR, column=16).comment = Comment(
            "= Pref Paid + Capital Returned + Excess − Promote (GP)", "Formula")
        dws.cell(row=HDR, column=17).comment = Comment(
            "Contributions: prior + |Scaled Amount|\n"
            "Distributions: prior − Capital Returned", "Formula")
        dws.cell(row=HDR, column=18).comment = Comment("= Pref Accrued − Pref Paid", "Formula")

        # ── Data rows with formulas (row 5+) ──
        FDR = HDR + 1  # first data row
        CUR = "$#,##0"

        for idx, row_data in enumerate(detail):
            r = FDR + idx
            p = r - 1  # previous row
            is_first = (idx == 0)
            is_contrib = (row_data["Event"] == "Contribution")

            # A: Date (actual Excel date for date arithmetic)
            dt_val = pd.to_datetime(row_data["Date"])
            dws.cell(row=r, column=1, value=dt_val).number_format = "MM/DD/YYYY"

            # B: Event (text value)
            dws.cell(row=r, column=2, value=row_data["Event"])

            # C: Gross Amount (input value)
            dws.cell(row=r, column=3, value=float(row_data["Gross Amount"])).number_format = CUR

            # D: Ownership % → assumption cell
            c = dws.cell(row=r, column=4)
            c.value = f"={OWN}"
            c.number_format = "0.00%"

            # E: Scaled Amount = Gross × Ownership%
            c = dws.cell(row=r, column=5)
            c.value = f"=C{r}*{OWN}"
            c.number_format = CUR

            if is_contrib:
                # F-N: zeros for contributions (no fees, no waterfall)
                for ci in range(6, 16):
                    dws.cell(row=r, column=ci, value=0).number_format = CUR

                # P: Net to Investor = Scaled (negative contribution)
                c = dws.cell(row=r, column=16)
                c.value = f"=E{r}"
                c.number_format = CUR

                # Q: Capital Balance = prior + |contribution|
                c = dws.cell(row=r, column=17)
                c.value = f"=ABS(E{r})" if is_first else f"=Q{p}+ABS(E{r})"
                c.number_format = CUR

                # R: Pref Balance (unchanged through contributions)
                c = dws.cell(row=r, column=18)
                c.value = 0.0 if is_first else f"=R{p}"
                c.number_format = CUR

            else:
                # ── Distribution row: full waterfall formulas ──

                # F: Acq Fee Paid = Scaled Amount if Acquisition Fee, else 0
                c = dws.cell(row=r, column=6)
                c.value = f'=IF(B{r}="Acquisition Fee",E{r},0)'
                c.number_format = CUR

                # G: AM Fee = prior Capital Balance × AM Fee % × Days / 365
                # Capped at Scaled − Acq Fee Paid
                c = dws.cell(row=r, column=7)
                if is_first:
                    c.value = 0.0
                else:
                    days = f"(A{r}-A{p})/365"
                    am_raw = f"Q{p}*{AMF}*{days}"
                    c.value = f"=IF(Q{p}<=0,0,MIN({am_raw},E{r}-F{r}))"
                c.number_format = CUR

                # H: Expenses = Annual Expenses × Days / 365
                # Capped at remaining after Acq Fee + AM Fee
                c = dws.cell(row=r, column=8)
                if is_first:
                    c.value = 0.0
                else:
                    days = f"(A{r}-A{p})/365"
                    exp_raw = f"{EXP}*{days}"
                    c.value = f"=MIN({exp_raw},MAX(0,E{r}-F{r}-G{r}))"
                c.number_format = CUR

                # I: Available = Scaled - Acq Fee - AM Fee - Expenses
                c = dws.cell(row=r, column=9)
                c.value = f"=E{r}-F{r}-G{r}-H{r}"
                c.number_format = CUR

                # J: Pref Accrued = prior Pref Balance + accrual
                c = dws.cell(row=r, column=10)
                if is_first:
                    c.value = 0.0
                else:
                    c.value = f"=R{p}+Q{p}*{HUR}*(A{r}-A{p})/365"
                c.number_format = CUR

                # K: Pref Paid = MIN(Accrued, Available)
                c = dws.cell(row=r, column=11)
                c.value = f"=MIN(J{r},I{r})"
                c.number_format = CUR

                # L: Capital Returned (capital events only)
                c = dws.cell(row=r, column=12)
                if is_first:
                    c.value = 0.0
                else:
                    c.value = f'=IF(B{r}="Capital Distribution",MIN(Q{p},I{r}-K{r}),0)'
                c.number_format = CUR

                # M: Excess = Available - Pref Paid - Capital Returned
                c = dws.cell(row=r, column=13)
                c.value = f"=I{r}-K{r}-L{r}"
                c.number_format = CUR

                # N: Hurdle Amount = -XNPV(hurdle, prior net CFs) × (1+hurdle)^years
                # The amount of this distribution needed to deliver hurdle IRR.
                # Uses P column (Net to Investor) for prior cashflows.
                # For Acquisition Fee rows: 0 (no promote anyway).
                c = dws.cell(row=r, column=14)
                if is_first:
                    # No prior cashflows — hurdle amount = 0
                    c.value = 0.0
                else:
                    # XNPV of prior net cashflows at hurdle rate
                    # Hurdle amount = MAX(0, -XNPV × (1+hurdle)^years)
                    xnpv_prior = f"XNPV({HUR},P{FDR}:P{p},A{FDR}:A{p})"
                    years = f"((A{r}-A{FDR})/365)"
                    c.value = f'=IF(B{r}="Acquisition Fee",0,MAX(0,-{xnpv_prior}*(1+{HUR})^{years}))'
                c.number_format = CUR

                # O: Promote
                # Capital Distribution: MAX(0, (K+L+M) - N) × Promote% (xnpv hurdle)
                # CF Distribution: M × Promote% (pref IS the hurdle return)
                # Acquisition Fee: 0
                c = dws.cell(row=r, column=15)
                c.value = (
                    f'=IF(B{r}="Acquisition Fee",0,'
                    f'IF(B{r}="Capital Distribution",'
                    f'MAX(0,(K{r}+L{r}+M{r})-N{r})*{PRO},'
                    f'M{r}*{PRO}))'
                )
                c.number_format = CUR

                # P: Net to Investor = Pref + Capital + Excess - Promote
                c = dws.cell(row=r, column=16)
                c.value = f"=K{r}+L{r}+M{r}-O{r}"
                c.number_format = CUR

                # Q: Capital Balance = prior - Capital Returned
                c = dws.cell(row=r, column=17)
                if is_first:
                    c.value = 0.0
                else:
                    c.value = f"=Q{p}-L{r}"
                c.number_format = CUR

                # R: Pref Balance = Accrued - Paid
                c = dws.cell(row=r, column=18)
                c.value = f"=J{r}-K{r}"
                c.number_format = CUR

        # ── Summary metrics as formulas ──
        last_r = FDR + len(detail) - 1
        sr = last_r + 2
        dws.cell(row=sr, column=1, value="Net Metrics (formulas)").font = bold_font

        dws.cell(row=sr + 1, column=1, value="Net Contributions").font = bold_font
        c = dws.cell(row=sr + 1, column=2)
        c.value = f'=SUMPRODUCT((B{FDR}:B{last_r}="Contribution")*P{FDR}:P{last_r})'
        c.number_format = CUR

        dws.cell(row=sr + 2, column=1, value="Net Distributions").font = bold_font
        c = dws.cell(row=sr + 2, column=2)
        c.value = f'=SUMPRODUCT((B{FDR}:B{last_r}<>"Contribution")*P{FDR}:P{last_r})'
        c.number_format = CUR

        dws.cell(row=sr + 3, column=1, value="Net MOIC").font = bold_font
        c = dws.cell(row=sr + 3, column=2)
        c.value = f'=IF(ABS(B{sr+1})>0,B{sr+2}/ABS(B{sr+1}),0)'
        c.number_format = '0.00"x"'

        dws.cell(row=sr + 4, column=1, value="Net IRR").font = bold_font
        c = dws.cell(row=sr + 4, column=2)
        c.value = f'=IFERROR(XIRR(FILTER(P{FDR}:P{last_r},P{FDR}:P{last_r}<>0),FILTER(A{FDR}:A{last_r},P{FDR}:P{last_r}<>0)),"N/A")'
        c.number_format = "0.00%"

        dws.cell(row=sr + 5, column=1, value="Net ROE").font = bold_font
        c = dws.cell(row=sr + 5, column=2)
        net_roe_val = dr.get("Net ROE")
        c.value = float(net_roe_val) if net_roe_val is not None and pd.notna(net_roe_val) else 0.0
        c.number_format = "0.00%"

        # Record cell references for Summary cross-sheet formulas
        detail_refs[dr["vcode"]] = {
            "sheet": sheet_name,
            "contribs": f"B{sr + 1}",
            "distribs": f"B{sr + 2}",
            "moic": f"B{sr + 3}",
            "irr": f"B{sr + 4}",
            "roe": f"B{sr + 5}",
        }

        # ── Auto-size columns ──
        col_widths = [12, 20, 16, 12, 16, 14, 14, 14, 14, 14, 14, 16, 14, 16, 14, 16, 16, 14]
        for ci, w in enumerate(col_widths, 1):
            dws.column_dimensions[get_column_letter(ci)].width = w

    # ── Phase 2: Build Summary sheet (first tab) ──
    # The default sheet was created by Workbook(); rename and use it.
    ws = wb.active
    ws.title = "Summary"
    # Move Summary to be the first sheet
    wb.move_sheet(ws, offset=-(len(wb.sheetnames) - 1))

    summary_cols = [
        "Investment Name", "Acquisition Date", "Sale Date",
        "Gross Contributions", "Gross Distributions", "Gross IRR", "Gross ROE", "Gross MOIC",
        "",  # spacer
        "Net Contributions", "Net Distributions", "Net IRR", "Net ROE", "Net MOIC",
    ]

    for ci, col in enumerate(summary_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    spacer_idx = 9
    ws.column_dimensions[ws.cell(row=1, column=spacer_idx).column_letter].width = 2

    net_by_vcode = {dr["vcode"]: dr for dr in net_result["deal_results"]}
    portfolio_net = net_result["portfolio"]

    currency_cols = {4, 5, 10, 11}
    pct_cols = {6, 7, 12, 13}
    moic_cols = {8, 14}

    row_idx = 2
    # Track which summary rows have detail sheet refs (for portfolio XIRR)
    deal_summary_rows = []

    for _, grow in gross_df.iterrows():
        is_total = bool(grow.get("_is_deal_total", False))
        vcode = str(grow.get("vcode", ""))

        if is_total:
            net_data = portfolio_net
            net_roe = net_data.get("Net ROE", 0)
        elif vcode in net_by_vcode:
            nd = net_by_vcode[vcode]
            net_roe = nd.get("Net ROE", 0)
        else:
            net_roe = None

        # Gross columns (always static values)
        gross_vals = [
            grow.get("Investment Name", ""),
            grow.get("Acquisition Date", ""),
            grow.get("Sale Date", ""),
            grow.get("Total Contributions", 0),
            grow.get("Total Distributions", 0),
            grow.get("IRR"),
            grow.get("ROE"),
            grow.get("MOIC", 0),
            "",  # spacer
        ]

        for ci, val in enumerate(gross_vals, 1):
            cell = ws.cell(row=row_idx, column=ci)
            if ci == spacer_idx:
                cell.value = ""
            elif ci in currency_cols:
                cell.value = float(val) if val is not None and pd.notna(val) else 0.0
                cell.number_format = "$#,##0"
            elif ci in pct_cols:
                cell.value = float(val) if val is not None and pd.notna(val) else None
                cell.number_format = "0.00%"
            elif ci in moic_cols:
                cell.value = float(val) if val is not None and pd.notna(val) else 0.0
                cell.number_format = '0.00"x"'
            else:
                cell.value = val
            if is_total:
                cell.font = bold_font
                cell.border = top_border

        # Net columns — use cross-sheet formula references where available
        ref = detail_refs.get(vcode)

        if not is_total and ref:
            # Per-deal: reference the detail sheet's formula cells
            sn = ref["sheet"].replace("'", "''")  # escape single quotes
            quoted = f"'{sn}'"

            # Col 10: Net Contributions (formula)
            c = ws.cell(row=row_idx, column=10)
            c.value = f"={quoted}!{ref['contribs']}"
            c.number_format = "$#,##0"

            # Col 11: Net Distributions (formula)
            c = ws.cell(row=row_idx, column=11)
            c.value = f"={quoted}!{ref['distribs']}"
            c.number_format = "$#,##0"

            # Col 12: Net IRR (formula)
            c = ws.cell(row=row_idx, column=12)
            c.value = f"={quoted}!{ref['irr']}"
            c.number_format = "0.00%"

            # Col 13: Net ROE (cross-sheet reference to per-deal ROE)
            c = ws.cell(row=row_idx, column=13)
            if "roe" in ref:
                c.value = f"={quoted}!{ref['roe']}"
            else:
                c.value = float(net_roe) if net_roe is not None and pd.notna(net_roe) else None
            c.number_format = "0.00%"

            # Col 14: Net MOIC (formula)
            c = ws.cell(row=row_idx, column=14)
            c.value = f"={quoted}!{ref['moic']}"
            c.number_format = '0.00"x"'

            deal_summary_rows.append(row_idx)
        elif is_total:
            # Portfolio Total — aggregate from deal rows using SUM/XIRR formulas
            deal_rows = deal_summary_rows

            # Col 10: Net Contributions = SUM of deal contributions
            c = ws.cell(row=row_idx, column=10)
            if deal_rows:
                refs_str = "+".join(f"J{r}" for r in deal_rows)
                c.value = f"={refs_str}"
            else:
                c.value = 0.0
            c.number_format = "$#,##0"

            # Col 11: Net Distributions = SUM of deal distributions
            c = ws.cell(row=row_idx, column=11)
            if deal_rows:
                refs_str = "+".join(f"K{r}" for r in deal_rows)
                c.value = f"={refs_str}"
            else:
                c.value = 0.0
            c.number_format = "$#,##0"

            # Col 12: Net IRR — use Python value (portfolio XIRR across all deals
            # can't be computed as a simple formula from per-deal IRRs)
            c = ws.cell(row=row_idx, column=12)
            pirr = portfolio_net.get("Net IRR")
            c.value = float(pirr) if pirr is not None and pd.notna(pirr) else None
            c.number_format = "0.00%"

            # Col 13: Net ROE (Python)
            c = ws.cell(row=row_idx, column=13)
            c.value = float(net_roe) if net_roe is not None and pd.notna(net_roe) else None
            c.number_format = "0.00%"

            # Col 14: Net MOIC = Net Distributions / |Net Contributions|
            c = ws.cell(row=row_idx, column=14)
            c.value = f"=IF(ABS(J{row_idx})>0,K{row_idx}/ABS(J{row_idx}),0)"
            c.number_format = '0.00"x"'

            for ci in range(10, 15):
                ws.cell(row=row_idx, column=ci).font = bold_font
                ws.cell(row=row_idx, column=ci).border = top_border
        else:
            # No net data for this deal
            for ci in range(10, 15):
                cell = ws.cell(row=row_idx, column=ci)
                cell.value = None
                if is_total:
                    cell.font = bold_font
                    cell.border = top_border

        row_idx += 1

    # Assumptions footnote
    foot_row = row_idx + 1
    ws.cell(row=foot_row, column=1, value="Net Returns Assumptions:").font = bold_font
    labels = [
        f"Ownership: {assumptions.get('ownership_pct', 0) * 100:.1f}%",
        f"AM Fee: {assumptions.get('am_fee_pct', 0) * 100:.2f}%",
        f"Hurdle Rate: {assumptions.get('hurdle_rate', 0) * 100:.2f}%",
        f"Promote: {assumptions.get('promote_pct', 0) * 100:.1f}%",
        f"Annual Expenses: ${assumptions.get('annual_expenses', 0):,.0f}",
    ]
    ws.cell(row=foot_row + 1, column=1, value="  ".join(labels))

    # Auto-size columns
    for ci in range(1, len(summary_cols) + 1):
        if ci == spacer_idx:
            continue
        max_len = len(str(summary_cols[ci - 1]))
        for ri in range(2, row_idx):
            cv = ws.cell(row=ri, column=ci).value
            if cv is not None:
                max_len = max(max_len, len(str(cv)))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 4, 30)

    # ── Phase 3: Portfolio Detail sheet ──
    # Pools all deals' waterfall events chronologically for portfolio-level validation.
    # Formula-driven: changing assumptions recalculates the entire portfolio waterfall.
    #
    # Column map (19 cols):
    #   A=Deal  B=Date  C=Event  D=Gross  E=Scaled  F=AcqFeePaid
    #   G=AMFee  H=Expenses  I=Available  J=PrefAccrued  K=PrefPaid
    #   L=CapReturned  M=Excess  N=HurdleAmount  O=Promote  P=NetToInvestor
    #   Q=CapBalance  R=PrefBalance  S=ActiveDeals

    pws = wb.create_sheet(title="Portfolio Detail")

    # Assumption references (editable)
    pws.cell(row=1, column=1, value="Assumptions (editable — formulas update automatically)").font = bold_font
    port_assumption_defs = [
        (1, "AM Fee %", assumptions.get("am_fee_pct", 0), "0.00%"),
        (3, "Hurdle Rate %", assumptions.get("hurdle_rate", 0), "0.00%"),
        (5, "Promote %", assumptions.get("promote_pct", 0), "0.00%"),
        (7, "Annual Expenses (per deal)", assumptions.get("annual_expenses", 0), "$#,##0"),
    ]
    for col, label, val, fmt in port_assumption_defs:
        pws.cell(row=2, column=col, value=label).font = bold_font
        c = pws.cell(row=2, column=col + 1, value=val)
        c.number_format = fmt

    # Expense scaling note
    pws.cell(row=3, column=1,
             value="Portfolio Expenses = (3 + 0.25 × # Active Deals) × Annual Expenses per deal"
             ).font = note_font

    P_AMF = "$B$2"  # AM Fee %
    P_HUR = "$D$2"  # Hurdle Rate
    P_PRO = "$F$2"  # Promote %
    P_EXP = "$H$2"  # Annual Expenses (per deal)

    HDR_P = 4  # header row
    port_cols = [
        "Deal", "Date", "Event", "Gross Amount", "Scaled Amount",
        "Acq Fee Paid", "AM Fee", "Expenses", "Available", "Pref Accrued",
        "Pref Paid", "Capital Returned", "Excess", "Hurdle Amount",
        "Promote (GP)", "Net to Investor", "Capital Balance", "Pref Balance",
        "# Active Deals",
    ]
    for ci, col in enumerate(port_cols, 1):
        cell = pws.cell(row=HDR_P, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Collect all events across deals, tagged with deal name
    all_events = []
    for dr in net_result["deal_results"]:
        deal_name = dr["Investment Name"]
        for row in dr.get("waterfall_detail", []):
            all_events.append((deal_name, row))

    # Sort by date, then by deal name for stable ordering
    all_events.sort(key=lambda x: (pd.to_datetime(x[1]["Date"]), x[0]))

    CUR = "$#,##0"
    FDR_P = HDR_P + 1  # first data row

    # Track per-deal capital balances for active deal count
    excel_deal_caps = {}  # deal_name -> last Capital Balance

    for idx, (deal_name, row) in enumerate(all_events):
        r = FDR_P + idx
        p = r - 1  # previous row
        is_first = (idx == 0)
        is_contrib = (row["Event"] == "Contribution")

        # Count active deals BEFORE updating this event's balance
        active_deals = sum(1 for b in excel_deal_caps.values() if b > 0.005)
        excel_deal_caps[deal_name] = float(row.get("Capital Balance", 0))

        # A: Deal name
        pws.cell(row=r, column=1, value=deal_name)

        # B: Date
        dt_val = pd.to_datetime(row["Date"])
        pws.cell(row=r, column=2, value=dt_val).number_format = "MM/DD/YYYY"

        # C: Event
        pws.cell(row=r, column=3, value=row["Event"])

        # D: Gross Amount (static from accounting)
        pws.cell(row=r, column=4, value=float(row["Gross Amount"])).number_format = CUR

        # E: Scaled Amount (static from per-deal computation)
        pws.cell(row=r, column=5, value=float(row["Scaled Amount"])).number_format = CUR

        # S (col 19): # Active Deals (static — computed from per-deal capital balances)
        pws.cell(row=r, column=19, value=active_deals)

        if is_contrib:
            # F-O: zeros for contributions
            for ci in range(6, 16):
                pws.cell(row=r, column=ci, value=0).number_format = CUR

            # P: Net to Investor = Scaled (negative contribution)
            c = pws.cell(row=r, column=16)
            c.value = f"=E{r}"
            c.number_format = CUR

            # Q: Capital Balance = prior + |Scaled|
            c = pws.cell(row=r, column=17)
            c.value = f"=ABS(E{r})" if is_first else f"=Q{p}+ABS(E{r})"
            c.number_format = CUR

            # R: Pref Balance (unchanged through contributions)
            c = pws.cell(row=r, column=18)
            c.value = 0.0 if is_first else f"=R{p}"
            c.number_format = CUR
        else:
            # ── Distribution row: full waterfall formulas ──

            # F: Acq Fee Paid
            c = pws.cell(row=r, column=6)
            c.value = f'=IF(C{r}="Acquisition Fee",E{r},0)'
            c.number_format = CUR

            # G: AM Fee = prior CapBalance × AM Fee% × Days/365, capped
            c = pws.cell(row=r, column=7)
            if is_first:
                c.value = 0.0
            else:
                days = f"(B{r}-B{p})/365"
                am_raw = f"Q{p}*{P_AMF}*{days}"
                c.value = f"=IF(Q{p}<=0,0,MIN({am_raw},E{r}-F{r}))"
            c.number_format = CUR

            # H: Expenses = (3 + 0.25 × ActiveDeals) × AnnualExp × Days/365, capped
            c = pws.cell(row=r, column=8)
            if is_first:
                c.value = 0.0
            else:
                days = f"(B{r}-B{p})/365"
                exp_raw = f"(3+0.25*S{r})*{P_EXP}*{days}"
                c.value = f"=MIN({exp_raw},MAX(0,E{r}-F{r}-G{r}))"
            c.number_format = CUR

            # I: Available = Scaled - AcqFee - AMFee - Expenses
            c = pws.cell(row=r, column=9)
            c.value = f"=E{r}-F{r}-G{r}-H{r}"
            c.number_format = CUR

            # J: Pref Accrued = prior PrefBalance + accrual on prior CapBalance
            c = pws.cell(row=r, column=10)
            if is_first:
                c.value = 0.0
            else:
                c.value = f"=R{p}+Q{p}*{P_HUR}*(B{r}-B{p})/365"
            c.number_format = CUR

            # K: Pref Paid = MIN(Accrued, Available)
            c = pws.cell(row=r, column=11)
            c.value = f"=MIN(J{r},I{r})"
            c.number_format = CUR

            # L: Capital Returned (capital events only)
            c = pws.cell(row=r, column=12)
            if is_first:
                c.value = 0.0
            else:
                c.value = f'=IF(C{r}="Capital Distribution",MIN(Q{p},I{r}-K{r}),0)'
            c.number_format = CUR

            # M: Excess = Available - PrefPaid - CapReturned
            c = pws.cell(row=r, column=13)
            c.value = f"=I{r}-K{r}-L{r}"
            c.number_format = CUR

            # N: Hurdle Amount = -XNPV(hurdle, prior net CFs) × (1+hurdle)^years
            c = pws.cell(row=r, column=14)
            if is_first:
                c.value = 0.0
            else:
                xnpv_prior = f"XNPV({P_HUR},P{FDR_P}:P{p},B{FDR_P}:B{p})"
                years = f"((B{r}-B{FDR_P})/365)"
                c.value = f'=IF(C{r}="Acquisition Fee",0,MAX(0,-{xnpv_prior}*(1+{P_HUR})^{years}))'
            c.number_format = CUR

            # O: Promote (hybrid: capital events use xnpv hurdle, CF uses Excess)
            c = pws.cell(row=r, column=15)
            c.value = (
                f'=IF(C{r}="Acquisition Fee",0,'
                f'IF(C{r}="Capital Distribution",'
                f'MAX(0,(K{r}+L{r}+M{r})-N{r})*{P_PRO},'
                f'M{r}*{P_PRO}))'
            )
            c.number_format = CUR

            # P: Net to Investor = PrefPaid + CapReturned + Excess - Promote
            c = pws.cell(row=r, column=16)
            c.value = f"=K{r}+L{r}+M{r}-O{r}"
            c.number_format = CUR

            # Q: Capital Balance = prior - Capital Returned
            c = pws.cell(row=r, column=17)
            if is_first:
                c.value = 0.0
            else:
                c.value = f"=Q{p}-L{r}"
            c.number_format = CUR

            # R: Pref Balance = Accrued - Paid
            c = pws.cell(row=r, column=18)
            c.value = f"=J{r}-K{r}"
            c.number_format = CUR

    if all_events:
        last_r = FDR_P + len(all_events) - 1
        sr = last_r + 2

        pws.cell(row=sr, column=1, value="Portfolio Net Metrics").font = bold_font

        # Net Contributions
        pws.cell(row=sr + 1, column=1, value="Net Contributions").font = bold_font
        c = pws.cell(row=sr + 1, column=2)
        c.value = f'=SUMPRODUCT((C{FDR_P}:C{last_r}="Contribution")*P{FDR_P}:P{last_r})'
        c.number_format = CUR

        # Net Distributions
        pws.cell(row=sr + 2, column=1, value="Net Distributions").font = bold_font
        c = pws.cell(row=sr + 2, column=2)
        c.value = f'=SUMPRODUCT((C{FDR_P}:C{last_r}<>"Contribution")*P{FDR_P}:P{last_r})'
        c.number_format = CUR

        # Net MOIC
        pws.cell(row=sr + 3, column=1, value="Net MOIC").font = bold_font
        c = pws.cell(row=sr + 3, column=2)
        c.value = f'=IF(ABS(B{sr+1})>0,B{sr+2}/ABS(B{sr+1}),0)'
        c.number_format = '0.00"x"'

        # Portfolio Net IRR
        pws.cell(row=sr + 4, column=1, value="Net IRR").font = bold_font
        c = pws.cell(row=sr + 4, column=2)
        c.value = f'=IFERROR(XIRR(FILTER(P{FDR_P}:P{last_r},P{FDR_P}:P{last_r}<>0),FILTER(B{FDR_P}:B{last_r},P{FDR_P}:P{last_r}<>0)),"N/A")'
        c.number_format = "0.00%"

        # Portfolio Net ROE (Python-computed — weighted avg capital calc not feasible in Excel)
        pws.cell(row=sr + 5, column=1, value="Net ROE").font = bold_font
        c = pws.cell(row=sr + 5, column=2)
        port_roe = portfolio_net.get("Net ROE", 0)
        c.value = float(port_roe) if port_roe is not None and pd.notna(port_roe) else 0.0
        c.number_format = "0.00%"

        # Update Summary Portfolio Total IRR to reference this sheet's formula
        port_irr_cell = f"B{sr + 4}"
        for ri in range(2, row_idx):
            cv = ws.cell(row=ri, column=1).value
            if cv and "Portfolio Total" in str(cv):
                ws.cell(row=ri, column=12).value = f"='Portfolio Detail'!{port_irr_cell}"
                ws.cell(row=ri, column=12).number_format = "0.00%"
                ws.cell(row=ri, column=12).font = bold_font
                ws.cell(row=ri, column=12).border = top_border
                break

    # Auto-size Portfolio Detail columns (19 cols A-S)
    port_widths = [28, 12, 20, 16, 16, 14, 14, 14, 14, 14, 14, 16, 14, 16, 14, 16, 16, 14, 14]
    for ci, w in enumerate(port_widths, 1):
        pws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_sold_deals(inv: pd.DataFrame) -> pd.DataFrame:
    """Filter investment map to sold deals, excluding child properties."""
    if "Sale_Status" not in inv.columns:
        return pd.DataFrame()

    inv_sold = inv[inv["Sale_Status"].fillna("").str.upper() == "SOLD"].copy()

    if "Portfolio_Name" in inv_sold.columns:
        inv_sold["Investment_Name"] = inv_sold["Investment_Name"].fillna("").astype(str).str.strip()
        inv_sold["Portfolio_Name"] = inv_sold["Portfolio_Name"].fillna("").astype(str).str.strip()
        parent_names = set(inv.loc[
            inv["Sale_Status"].fillna("").str.upper() == "SOLD", "Investment_Name"
        ].fillna("").astype(str).str.strip())
        is_child = (
            inv_sold["Portfolio_Name"].isin(parent_names)
            & (inv_sold["Portfolio_Name"] != inv_sold["Investment_Name"])
            & (inv_sold["Portfolio_Name"] != "")
        )
        inv_sold = inv_sold[~is_child].copy()

    return inv_sold
