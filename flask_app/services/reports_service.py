"""Reports service — partner returns, population selectors, Excel generation.

Extracts pure logic from reports_ui.py (no Streamlit dependency).
"""

import logging
import calendar
import pandas as pd
import io
from datetime import date
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from utils import normalize_columns

log = logging.getLogger(__name__)


def build_partner_returns(deal_result: dict, deal_name: str) -> list[dict]:
    """Build partner + deal-total rows from a compute result.

    Thin wrapper around partner_results/deal_summary from compute.py.
    Returns list of dicts with columns: Deal Name, Partner, Contributions,
    CF Distributions, Capital Distributions, IRR, ROE, MOIC, _is_deal_total.
    """
    rows = []
    partner_results = deal_result.get("partner_results", [])
    for pr in partner_results:
        rows.append({
            "Deal Name": deal_name,
            "Partner": pr["partner"],
            "Contributions": pr["contributions"],
            "CF Distributions": pr["cf_distributions"],
            "Capital Distributions": pr["cap_distributions"],
            "IRR": pr["irr"],
            "ROE": pr["roe"],
            "MOIC": pr["moic"],
            "_is_deal_total": False,
        })

    ds = deal_result.get("deal_summary", {})
    rows.append({
        "Deal Name": deal_name,
        "Partner": "DEAL TOTAL",
        "Contributions": ds.get("total_contributions", 0),
        "CF Distributions": ds.get("total_cf_distributions", 0),
        "Capital Distributions": ds.get("total_cap_distributions", 0),
        "IRR": ds.get("deal_irr", None),
        "ROE": ds.get("deal_roe", None),
        "MOIC": ds.get("deal_moic", None),
        "_is_deal_total": True,
    })

    return rows


def build_deal_lookup(inv: pd.DataFrame, wf: pd.DataFrame) -> dict:
    """Build deal lookup tables for population selectors.

    Returns dict with:
      - eligible: list of {vcode, name, label} for deals with waterfall definitions
      - vcode_to_label: dict mapping vcode -> display label
    """
    inv_disp = inv.copy()

    inv_disp["Investment_Name"] = inv_disp["Investment_Name"].fillna("").astype(str)
    inv_disp["vcode"] = inv_disp["vcode"].astype(str)

    name_counts = inv_disp["Investment_Name"].value_counts()
    inv_disp["DealLabel"] = inv_disp.apply(
        lambda r: (
            f"{r['Investment_Name']} ({r['vcode']})"
            if name_counts.get(r["Investment_Name"], 0) > 1
            else r["Investment_Name"]
        ),
        axis=1,
    )

    # Exclude child properties
    if "Portfolio_Name" in inv_disp.columns:
        inv_disp["Portfolio_Name"] = inv_disp["Portfolio_Name"].fillna("").astype(str).str.strip()
        parent_names = set(inv_disp["Investment_Name"].str.strip())
        is_child = (
            inv_disp["Portfolio_Name"].isin(parent_names)
            & (inv_disp["Portfolio_Name"] != inv_disp["Investment_Name"].str.strip())
            & (inv_disp["Portfolio_Name"] != "")
        )
        inv_disp = inv_disp[~is_child].copy()

    # Waterfall normalisation
    wf_norm = wf.copy()
    normalize_columns(wf_norm)
    if "vCode" in wf_norm.columns and "vcode" not in wf_norm.columns:
        wf_norm = wf_norm.rename(columns={"vCode": "vcode"})
    wf_norm["vcode"] = wf_norm["vcode"].astype(str)

    wf_vcodes = set(wf_norm["vcode"].unique())
    eligible = inv_disp[inv_disp["vcode"].isin(wf_vcodes)]

    eligible_list = sorted(
        [
            {"vcode": r["vcode"], "name": r["Investment_Name"], "label": r["DealLabel"]}
            for _, r in eligible.iterrows()
        ],
        key=lambda x: x["label"].lower(),
    )

    vcode_to_label = {r["vcode"]: r["DealLabel"] for _, r in eligible.iterrows()}

    return {
        "eligible": eligible_list,
        "vcode_to_label": vcode_to_label,
        "wf_norm": wf_norm,
        "eligible_vcodes": wf_vcodes & set(inv_disp["vcode"]),
    }


def get_partner_deals(wf: pd.DataFrame, eligible_vcodes: set,
                      vcode_to_label: dict) -> dict[str, list[str]]:
    """Build partner -> list of deal vcodes from waterfall PropCodes.

    Returns dict mapping partner_id -> sorted list of vcodes.
    """
    partner_to_vcodes: dict[str, set[str]] = {}

    for _, r in wf.iterrows():
        vc = str(r.get("vcode", ""))
        pc = str(r.get("PropCode", "")).strip()
        if vc in eligible_vcodes and pc:
            partner_to_vcodes.setdefault(pc, set()).add(vc)

    return {
        partner: sorted(vcodes)
        for partner, vcodes in sorted(partner_to_vcodes.items(), key=lambda x: x[0].lower())
    }


def get_upstream_investor_deals(
    relationships_raw: pd.DataFrame, inv: pd.DataFrame,
    eligible_vcodes: set
) -> dict[str, dict]:
    """Build upstream_investor -> list of deal vcodes from ownership tree.

    Returns dict mapping investor_id -> {name, vcodes: [...]}.
    """
    if relationships_raw is None or relationships_raw.empty:
        return {}

    from ownership_tree import load_relationships, build_ownership_tree, get_ultimate_investors
    from loaders import build_investmentid_to_vcode

    relationships = load_relationships(relationships_raw)

    # Filter out ended relationships (matching Review Tracking behaviour)
    if "EndDate" in relationships.columns:
        end_col = relationships["EndDate"]
        is_empty = end_col.isna() | (end_col.astype(str).str.strip().isin(["", "NaT", "nan", "None"]))
        relationships = relationships[is_empty].copy()

    nodes = build_ownership_tree(relationships)
    inv_to_vcode = build_investmentid_to_vcode(inv)

    investor_to_vcodes: dict[str, set[str]] = {}
    investor_names: dict[str, str] = {}

    for inv_id, vc in inv_to_vcode.items():
        if str(vc) not in eligible_vcodes:
            continue
        if inv_id not in nodes:
            continue
        ultimate = get_ultimate_investors(inv_id, nodes, normalize=True)
        for investor_id, _ in ultimate:
            investor_to_vcodes.setdefault(investor_id, set()).add(str(vc))
            node = nodes.get(investor_id)
            if node and hasattr(node, "name") and node.name:
                investor_names[investor_id] = node.name

    # Also include direct investors from relationships
    for _, rel_row in relationships.iterrows():
        inv_id = str(rel_row.get("InvestmentID", "")).strip()
        investor_id = str(rel_row.get("InvestorID", "")).strip()
        vc = inv_to_vcode.get(inv_id)
        if vc and str(vc) in eligible_vcodes and investor_id:
            investor_to_vcodes.setdefault(investor_id, set()).add(str(vc))

    result = {}
    for iid in sorted(investor_to_vcodes.keys(), key=lambda x: x.lower()):
        name = investor_names.get(iid, "")
        result[iid] = {
            "name": name,
            "display": f"{iid} — {name}" if name and name != iid else iid,
            "vcodes": sorted(investor_to_vcodes[iid]),
        }

    return result


def generate_returns_excel(df: pd.DataFrame) -> bytes:
    """Generate formatted Excel for Projected Returns Summary.

    Extracted from reports_ui.py::_generate_excel().
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Projected Returns"

    display_cols = [c for c in df.columns if c != "_is_deal_total"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    bold_font = Font(bold=True)
    top_border = Border(top=Side(style="medium"))

    currency_cols = {"Contributions", "CF Distributions", "Capital Distributions"}
    pct_cols = {"IRR", "ROE"}
    moic_cols = {"MOIC"}

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_total = bool(row.get("_is_deal_total", False))
        for col_idx, col_name in enumerate(display_cols, 1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name in currency_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = "$#,##0"
            elif col_name in pct_cols:
                cell.value = float(val) if pd.notna(val) else None
                cell.number_format = "0.00%"
            elif col_name in moic_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = '0.00"x"'
            else:
                cell.value = val

            if is_total:
                cell.font = bold_font
                cell.border = top_border

    for col_idx, col_name in enumerate(display_cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(col_name) + 2, 14)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# ROE Summary Report
# ---------------------------------------------------------------------------

def _compute_accrued_pref(deal_acct: pd.DataFrame, report_date: date,
                          pref_rates: dict) -> float:
    """Compute total accrued (unpaid) pref across all PE investors through report_date.

    Walks accounting chronologically per investor, accrues pref daily on capital
    balance at the waterfall pref rate, reduces accrued by pref payments (TypeID 1019)
    and excess CF distributions. Compounds at year-end (always, no grace period).
    Uses Act/Act day count convention (366 for leap years).
    """
    if not pref_rates:
        return 0.0

    total_accrued = 0.0

    for investor_id, grp in deal_acct.groupby("InvestorID"):
        if investor_id.upper().startswith("OP"):
            continue
        pref_rate = pref_rates.get(investor_id, 0.0)
        if pref_rate <= 0:
            continue

        rows = grp.sort_values("EffectiveDate")
        capital = 0.0
        pref_compounded = 0.0
        pref_cy = 0.0
        prev_date = None

        # Detect TypeID column
        has_typeid = "TypeID" in rows.columns

        for _, r in rows.iterrows():
            if r.get("is_commitment", False):
                continue
            evt_date = r["EffectiveDate"].date() if pd.notna(r["EffectiveDate"]) else None
            if evt_date is None:
                continue
            amt = float(r["Amt"])
            major = r["MajorType"].lower()
            tname = r["TypeName"].lower()
            type_id = None
            if has_typeid:
                try:
                    type_id = float(r.get("TypeID", 0))
                except (ValueError, TypeError):
                    pass

            # Accrue pref from prev_date to this event (Act/Act)
            if prev_date is not None and capital > 0 and evt_date > prev_date:
                cur = prev_date
                while cur < evt_date:
                    year_end = date(cur.year, 12, 31)
                    next_stop = min(evt_date, year_end)
                    days = (next_stop - cur).days
                    if days > 0:
                        diy = _days_in_year(cur.year)
                        base = max(0.0, capital + pref_compounded)
                        pref_cy += base * pref_rate * (days / diy)
                    if next_stop == year_end and next_stop < evt_date:
                        # Always compound at year-end
                        pref_compounded += pref_cy
                        pref_cy = 0.0
                        cur = date(cur.year + 1, 1, 1)
                    else:
                        break

            # Apply pref payment (TypeID 1019 = Preferred Return distribution)
            is_pref_payment = (type_id == 1019.0)
            if not is_pref_payment:
                is_pref_payment = "preferred return" in tname or "pref return" in tname
            if is_pref_payment and "distri" in major:
                payment = abs(amt)
                if pref_compounded > 0:
                    pay = min(payment, pref_compounded)
                    pref_compounded -= pay
                    payment -= pay
                if pref_cy > 0 and payment > 0:
                    pay = min(payment, pref_cy)
                    pref_cy -= pay

            # Excess CF also reduces accrued pref
            is_excess_cf = "excess cash" in tname or type_id == 1020.0
            if is_excess_cf and "distri" in major and (pref_compounded + pref_cy) > 0:
                payment = abs(amt)
                if pref_compounded > 0:
                    pay = min(payment, pref_compounded)
                    pref_compounded -= pay
                    payment -= pay
                if pref_cy > 0 and payment > 0:
                    pay = min(payment, pref_cy)
                    pref_cy -= pay

            # Update capital balance
            if "contrib" in major:
                capital += abs(amt)
            elif "distri" in major:
                if "return of capital" in tname or "realized gain" in tname:
                    capital = max(0.0, capital - abs(amt))

            prev_date = evt_date

        # Accrue from last transaction to report_date (Act/Act)
        if prev_date is not None and capital > 0 and report_date > prev_date:
            cur = prev_date
            while cur < report_date:
                year_end = date(cur.year, 12, 31)
                next_stop = min(report_date, year_end)
                days = (next_stop - cur).days
                if days > 0:
                    diy = _days_in_year(cur.year)
                    base = max(0.0, capital + pref_compounded)
                    pref_cy += base * pref_rate * (days / diy)
                if next_stop == year_end and next_stop < report_date:
                    pref_compounded += pref_cy
                    pref_cy = 0.0
                    cur = date(cur.year + 1, 1, 1)
                else:
                    break

        total_accrued += pref_compounded + pref_cy

    return total_accrued


def build_roe_summary_row(
    vcode: str,
    deal_name: str,
    acct: pd.DataFrame,
    inv_map: pd.DataFrame,
    report_date: date,
    wf_steps: Optional[pd.DataFrame] = None,
    seed_states: Optional[dict] = None,
    isbs_raw: Optional[pd.DataFrame] = None,
) -> Optional[dict]:
    """Build one ROE summary row from accounting data through report_date.

    Returns dict with: Deal Name, Total Funded, Return of Capital,
    Current Balance, Wtd Avg Balance, CF Received, Accrued Pref, ITD ROE,
    U/W ITD ROE (from Projected IS accounts 7071/7073).
    """
    from loaders import build_investmentid_to_vcode
    from metrics import calculate_roe, calculate_roe_detailed

    vcode_str = str(vcode).strip()

    inv_to_vcode = build_investmentid_to_vcode(inv_map)
    deal_iids = [iid for iid, vc in inv_to_vcode.items() if str(vc) == vcode_str]
    if not deal_iids:
        return None

    acct_norm = acct.copy()
    normalize_columns(acct_norm)
    acct_norm["InvestmentID"] = acct_norm["InvestmentID"].astype(str).str.strip()
    acct_norm["EffectiveDate"] = pd.to_datetime(acct_norm["EffectiveDate"], errors="coerce")

    deal_acct = acct_norm[
        (acct_norm["InvestmentID"].isin(deal_iids))
        & (acct_norm["EffectiveDate"].dt.date <= report_date)
    ].copy()

    if deal_acct.empty:
        return None

    deal_acct["MajorType"] = deal_acct["MajorType"].fillna("").astype(str).str.strip()
    deal_acct["Amt"] = pd.to_numeric(deal_acct["Amt"], errors="coerce").fillna(0.0)
    if "TypeName" not in deal_acct.columns and "Typename" in deal_acct.columns:
        deal_acct["TypeName"] = deal_acct["Typename"]
    elif "TypeName" not in deal_acct.columns:
        deal_acct["TypeName"] = ""
    deal_acct["TypeName"] = deal_acct["TypeName"].fillna("").astype(str).str.strip()
    deal_acct["InvestorID"] = deal_acct["InvestorID"].astype(str).str.strip()

    funded = 0.0
    roc = 0.0
    capital_events = []
    cf_distributions = []

    for _, row in deal_acct.iterrows():
        if row["InvestorID"].upper().startswith("OP"):
            continue
        if row.get("is_commitment", False):
            continue
        major = row["MajorType"].lower()
        tname = row["TypeName"].lower()
        amt = float(row["Amt"])
        evt_date = row["EffectiveDate"].date() if pd.notna(row["EffectiveDate"]) else None
        if evt_date is None:
            continue

        if "contrib" in major:
            funded += abs(amt)
            capital_events.append((evt_date, -abs(amt)))
        elif "distri" in major:
            if "return of capital" in tname or "realized gain" in tname:
                capital_events.append((evt_date, amt))
                roc += amt
            elif "acquisition fee" not in tname:
                # CF distribution — preserve sign so negative corrections
                # reduce ROE instead of inflating it.
                if amt >= 0:
                    capital_events.append((evt_date, amt))
                cf_distributions.append((evt_date, amt))
            # else: acquisition fee — excluded from both capital_events
            # and cf_distributions (no effect on ROE)

    if not capital_events:
        return None

    inception = min(d for d, _ in capital_events)
    detail = calculate_roe_detailed(capital_events, cf_distributions, inception, report_date)

    current_balance = funded - roc

    # Compute accrued pref directly from accounting + waterfall pref rates
    pref_rates = {}
    if wf_steps is not None and not wf_steps.empty:
        wf_deal = wf_steps[wf_steps["vcode"] == vcode_str] if "vcode" in wf_steps.columns else wf_steps
        pref_rows = wf_deal[wf_deal["vState"] == "Pref"] if "vState" in wf_deal.columns else pd.DataFrame()
        rate_col = "nPercent_dec" if "nPercent_dec" in pref_rows.columns else "nPercent"
        for _, pr in pref_rows.iterrows():
            pc = str(pr.get("PropCode", "")).strip()
            r = float(pr[rate_col]) if pd.notna(pr.get(rate_col)) else 0.0
            if pc and r > 0 and pc not in pref_rates:
                pref_rates[pc] = r

    accrued = _compute_accrued_pref(deal_acct, report_date, pref_rates)

    # ---- U/W ROE to Date (ISBS Projected IS only — no actual accounting) ----
    # 7073: positive = contribution, negative = return of capital
    # 7071: underwritten distributions (ROE numerator)
    uw_roe = 0.0
    uw_detail_rows = []
    uw_cf_total = 0.0
    uw_roc_total = 0.0

    if isbs_raw is not None and not isbs_raw.empty:
        from one_pager import _get_uw_pe_periodic, _get_uw_7073_signed, UW_PE_DIST_ACCT, UW_PE_ROC_ACCT

        # Use earliest contribution from 7073 as inception, or fall back to actual
        uw_capital_events_raw = _get_uw_pe_periodic(
            isbs_raw, vcode, date(2000, 1, 1), report_date, UW_PE_ROC_ACCT
        )
        uw_dists = _get_uw_pe_periodic(
            isbs_raw, vcode, date(2000, 1, 1), report_date, UW_PE_DIST_ACCT
        )

        if uw_capital_events_raw or uw_dists:
            # Build capital events from 7073:
            # _get_uw_pe_periodic returns abs(periodic) — we need sign convention:
            # positive original = contribution, negative original = return of capital
            # Re-read raw 7073 to get signs
            uw_capital_events = _get_uw_7073_signed(
                isbs_raw, vcode, date(2000, 1, 1), report_date
            )

            if uw_capital_events or uw_dists:
                # Determine inception from earliest event
                all_dates = [d for d, _ in uw_capital_events] + [d for d, _ in uw_dists]
                if not all_dates:
                    all_dates = [d for d, _ in capital_events] if capital_events else [report_date]
                inception_dt = min(all_dates)

                # Filter to report date range
                uw_capital_events = [(d, a) for d, a in uw_capital_events if d <= report_date]
                uw_dists = [(d, a) for d, a in uw_dists if d >= inception_dt and d <= report_date]

                uw_cf_total = sum(a for _, a in uw_dists)
                uw_roc_total = sum(a for _, a in uw_capital_events if a > 0)

                uw_roe = calculate_roe(uw_capital_events, uw_dists, inception_dt, report_date)

                # Build detail rows
                uw_events = []
                for d, amt in uw_capital_events:
                    if amt < 0:
                        uw_events.append((d, "U/W Contribution (7073)", abs(amt), abs(amt)))
                    else:
                        uw_events.append((d, "U/W Return of Capital (7073)", amt, -amt))
                for d, amt in uw_dists:
                    uw_events.append((d, "U/W Distribution (7071)", amt, 0.0))

                uw_events.sort(key=lambda x: x[0])

                uw_balance = 0.0
                uw_prev = inception_dt
                for evt_date, event_type, amount, balance_change in uw_events:
                    days = (evt_date - uw_prev).days
                    weighted = uw_balance * days
                    new_bal = max(0.0, uw_balance + balance_change)
                    uw_detail_rows.append({
                        "Date": evt_date,
                        "Event": event_type,
                        "Amount": amount,
                        "Days": days,
                        "Capital Balance": uw_balance,
                        "Weighted Capital": weighted,
                        "New Balance": new_bal,
                    })
                    uw_balance = new_bal
                    uw_prev = evt_date

                # Final period
                uw_final_days = (report_date - uw_prev).days
                if uw_final_days > 0:
                    uw_detail_rows.append({
                        "Date": report_date,
                        "Event": "(Report Date)",
                        "Amount": 0.0,
                        "Days": uw_final_days,
                        "Capital Balance": uw_balance,
                        "Weighted Capital": uw_balance * uw_final_days,
                        "New Balance": uw_balance,
                    })

    # Build event-by-event detail for Excel audit trail.
    # Use the original accounting rows (with known MajorType/TypeName) so event
    # classification is authoritative — avoids same-date mis-tagging.
    acct_events = []
    for _, row in deal_acct.iterrows():
        if row["InvestorID"].upper().startswith("OP"):
            continue
        if row.get("is_commitment", False):
            continue
        evt_date = row["EffectiveDate"].date() if pd.notna(row["EffectiveDate"]) else None
        if evt_date is None:
            continue
        amt = float(row["Amt"])
        major = row["MajorType"].lower()
        tname = row["TypeName"].lower()

        if "contrib" in major:
            acct_events.append((evt_date, "Contribution", abs(amt), abs(amt)))
        elif "distri" in major:
            if "return of capital" in tname or "realized gain" in tname:
                acct_events.append((evt_date, "Capital Return", abs(amt), -abs(amt)))
            elif "acquisition fee" in tname:
                acct_events.append((evt_date, "Acquisition Fee", abs(amt), 0.0))
            else:
                acct_events.append((evt_date, "CF Distribution", amt, 0.0))

    acct_events.sort(key=lambda x: x[0])

    detail_rows = []
    running_balance = 0.0
    prev_dt = inception = min(d for d, _ in capital_events)
    total_weighted = 0.0

    for evt_date, event_type, amount, balance_change in acct_events:
        days = (evt_date - prev_dt).days
        weighted = running_balance * days
        total_weighted += weighted

        new_balance = max(0.0, running_balance + balance_change)
        detail_rows.append({
            "Date": evt_date,
            "Event": event_type,
            "Amount": amount,
            "Days": days,
            "Capital Balance": running_balance,
            "Weighted Capital": weighted,
            "New Balance": new_balance,
        })
        running_balance = new_balance
        prev_dt = evt_date

    # Final period to report_date
    final_days = (report_date - prev_dt).days
    final_weighted = running_balance * final_days
    total_weighted += final_weighted
    if final_days > 0:
        detail_rows.append({
            "Date": report_date,
            "Event": "(Report Date)",
            "Amount": 0.0,
            "Days": final_days,
            "Capital Balance": running_balance,
            "Weighted Capital": final_weighted,
            "New Balance": running_balance,
        })

    return {
        "Deal Name": deal_name,
        "Total Funded": funded,
        "Return of Capital": roc,
        "Current Balance": current_balance,
        "Wtd Avg Balance": detail["weighted_avg_capital"],
        "CF Received": detail["total_cf_distributions"],
        "Accrued Pref": accrued,
        "ITD ROE": detail["roe"],
        "U/W ITD ROE": uw_roe,
        "_detail_rows": detail_rows,
        "_uw_detail_rows": uw_detail_rows,
        "_uw_cf_total": uw_cf_total,
        "_uw_roc_total": uw_roc_total,
        "_years": detail["years"],
        "_total_days": (report_date - inception).days,
        "_inception": inception,
    }


def generate_roe_summary_excel(df: pd.DataFrame, all_rows: list = None) -> bytes:
    """Generate formatted Excel for ROE Summary report.

    Sheet 1: ROE Summary table.
    Per-deal sheets: event-by-event weighted capital calculation with formulas.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ROE Summary"

    display_cols = [c for c in df.columns if not c.startswith("_")]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    bold_font = Font(bold=True)
    label_font = Font(bold=True, italic=True)

    for col_idx, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    currency_cols = {"Total Funded", "Return of Capital", "Current Balance",
                     "Wtd Avg Balance", "CF Received", "Accrued Pref"}
    pct_cols = {"ITD ROE", "U/W ITD ROE"}

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(display_cols, 1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name in currency_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = "$#,##0"
            elif col_name in pct_cols:
                cell.value = float(val) if pd.notna(val) else None
                cell.number_format = "0.00%"
            else:
                cell.value = val

    for col_idx, col_name in enumerate(display_cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(col_name) + 2, 16)

    # ---- Per-deal detail sheets ----
    if all_rows:
        for row_data in all_rows:
            detail = row_data.get("_detail_rows")
            if not detail:
                continue
            deal_name = row_data["Deal Name"]
            # Sheet name max 31 chars, no invalid chars
            sheet_name = deal_name[:28].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "").replace(":", "")
            ds = wb.create_sheet(title=sheet_name)

            # Header section — deal-level metrics
            ds.cell(row=1, column=1, value="Deal:").font = label_font
            ds.cell(row=1, column=2, value=deal_name)
            ds.cell(row=2, column=1, value="Inception:").font = label_font
            ds.cell(row=2, column=2, value=row_data.get("_inception"))
            ds.cell(row=2, column=2).number_format = "M/D/YYYY"
            ds.cell(row=2, column=3, value="Total Days:").font = label_font
            ds.cell(row=2, column=4, value=row_data.get("_total_days", 0))
            ds.cell(row=2, column=5, value="Years:").font = label_font
            ds.cell(row=2, column=6, value=row_data.get("_years", 0.0))
            ds.cell(row=2, column=6).number_format = "0.00"

            # Summary metrics
            ds.cell(row=4, column=1, value="Total Funded:").font = label_font
            ds.cell(row=4, column=2, value=row_data["Total Funded"])
            ds.cell(row=4, column=2).number_format = "$#,##0"
            ds.cell(row=4, column=3, value="Return of Capital:").font = label_font
            ds.cell(row=4, column=4, value=row_data["Return of Capital"])
            ds.cell(row=4, column=4).number_format = "$#,##0"
            ds.cell(row=4, column=5, value="Current Balance:").font = label_font
            ds.cell(row=4, column=6, value=row_data["Current Balance"])
            ds.cell(row=4, column=6).number_format = "$#,##0"

            ds.cell(row=5, column=1, value="CF Received:").font = label_font
            ds.cell(row=5, column=2, value=row_data["CF Received"])
            ds.cell(row=5, column=2).number_format = "$#,##0"
            ds.cell(row=5, column=3, value="Wtd Avg Balance:").font = label_font
            ds.cell(row=5, column=4, value=row_data["Wtd Avg Balance"])
            ds.cell(row=5, column=4).number_format = "$#,##0"
            ds.cell(row=5, column=5, value="Accrued Pref:").font = label_font
            ds.cell(row=5, column=6, value=row_data["Accrued Pref"])
            ds.cell(row=5, column=6).number_format = "$#,##0"

            # ROE formula explanation
            ds.cell(row=7, column=1, value="ROE Formula:").font = label_font
            roe_val = row_data["ITD ROE"]
            ds.cell(row=7, column=2,
                    value=f"(CF Received / Wtd Avg Balance) / Years = "
                          f"({row_data['CF Received']:,.0f} / {row_data['Wtd Avg Balance']:,.0f}) / {row_data.get('_years', 0):.2f}")
            ds.cell(row=7, column=6, value=roe_val)
            ds.cell(row=7, column=6).number_format = "0.00%"
            ds.cell(row=7, column=6).font = bold_font

            # Event detail table
            detail_start = 9
            detail_cols = ["Date", "Event", "Amount", "Days",
                           "Capital Balance", "Weighted Capital", "New Balance"]
            for ci, col_name in enumerate(detail_cols, 1):
                cell = ds.cell(row=detail_start, column=ci, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            currency_detail = {"Amount", "Capital Balance", "Weighted Capital", "New Balance"}
            for ri, evt in enumerate(detail, 1):
                r = detail_start + ri
                for ci, col_name in enumerate(detail_cols, 1):
                    cell = ds.cell(row=r, column=ci)
                    val = evt[col_name]
                    if col_name == "Date":
                        cell.value = val
                        cell.number_format = "M/D/YYYY"
                    elif col_name in currency_detail:
                        cell.value = float(val)
                        cell.number_format = "$#,##0"
                    else:
                        cell.value = val

            # Totals row
            total_row = detail_start + len(detail) + 1
            ds.cell(row=total_row, column=1, value="TOTAL").font = bold_font
            top_border = Border(top=Side(style="thin"))
            # Sum of Weighted Capital
            sum_col = detail_cols.index("Weighted Capital") + 1
            days_col = detail_cols.index("Days") + 1
            for ci in [days_col, sum_col]:
                cell = ds.cell(row=total_row, column=ci)
                first_data = detail_start + 1
                last_data = detail_start + len(detail)
                col_letter = cell.column_letter
                cell.value = f"=SUM({col_letter}{first_data}:{col_letter}{last_data})"
                cell.font = bold_font
                cell.border = top_border
                if ci == sum_col:
                    cell.number_format = "$#,##0"

            # Wtd Avg Capital formula
            avg_row = total_row + 1
            ds.cell(row=avg_row, column=1, value="Wtd Avg Capital =").font = label_font
            sum_letter = ds.cell(row=total_row, column=sum_col).column_letter
            days_letter = ds.cell(row=total_row, column=days_col).column_letter
            ds.cell(row=avg_row, column=2,
                    value=f"={sum_letter}{total_row}/{days_letter}{total_row}")
            ds.cell(row=avg_row, column=2).number_format = "$#,##0"

            # Column widths
            widths = [12, 16, 14, 8, 16, 18, 16]
            for ci, w in enumerate(widths, 1):
                ds.column_dimensions[ds.cell(row=1, column=ci).column_letter].width = w

        # ---- U/W ROE detail sheets ----
        for row_data in all_rows:
            uw_detail = row_data.get("_uw_detail_rows")
            if not uw_detail:
                continue
            deal_name = row_data["Deal Name"]
            safe_name = deal_name[:25].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "").replace(":", "")
            ds = wb.create_sheet(title=f"UW {safe_name}")

            uw_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

            ds.cell(row=1, column=1, value="Deal:").font = label_font
            ds.cell(row=1, column=2, value=deal_name)
            ds.cell(row=1, column=4, value="U/W ROE to Date").font = Font(bold=True, color="548235", size=12)
            ds.cell(row=2, column=1, value="Inception:").font = label_font
            ds.cell(row=2, column=2, value=row_data.get("_inception"))
            ds.cell(row=2, column=2).number_format = "M/D/YYYY"
            ds.cell(row=2, column=3, value="Total Days:").font = label_font
            ds.cell(row=2, column=4, value=row_data.get("_total_days", 0))
            ds.cell(row=2, column=5, value="Years:").font = label_font
            ds.cell(row=2, column=6, value=row_data.get("_years", 0.0))
            ds.cell(row=2, column=6).number_format = "0.00"

            ds.cell(row=4, column=1, value="Total Funded (Actual):").font = label_font
            ds.cell(row=4, column=2, value=row_data["Total Funded"])
            ds.cell(row=4, column=2).number_format = "$#,##0"
            ds.cell(row=4, column=3, value="U/W Distributions (7071):").font = label_font
            ds.cell(row=4, column=4, value=row_data.get("_uw_cf_total", 0.0))
            ds.cell(row=4, column=4).number_format = "$#,##0"
            ds.cell(row=4, column=5, value="U/W ROC (7073):").font = label_font
            ds.cell(row=4, column=6, value=row_data.get("_uw_roc_total", 0.0))
            ds.cell(row=4, column=6).number_format = "$#,##0"

            ds.cell(row=5, column=1, value="Capital Structure:").font = label_font
            ds.cell(row=5, column=2, value="Actual contributions + actual capital returns + U/W ROC (7073)")

            # U/W ROE formula
            uw_roe_val = row_data.get("U/W ITD ROE", 0.0)
            years = row_data.get("_years", 0.0)
            ds.cell(row=7, column=1, value="U/W ROE Formula:").font = label_font
            ds.cell(row=7, column=2,
                    value="(U/W Distributions / Wtd Avg Capital) / Years")
            ds.cell(row=7, column=6, value=uw_roe_val)
            ds.cell(row=7, column=6).number_format = "0.00%"
            ds.cell(row=7, column=6).font = bold_font

            # Event detail table
            detail_start = 9
            detail_cols = ["Date", "Event", "Amount", "Days",
                           "Capital Balance", "Weighted Capital", "New Balance"]
            for ci, col_name in enumerate(detail_cols, 1):
                cell = ds.cell(row=detail_start, column=ci, value=col_name)
                cell.font = header_font
                cell.fill = uw_fill
                cell.alignment = Alignment(horizontal="center")

            currency_detail = {"Amount", "Capital Balance", "Weighted Capital", "New Balance"}
            for ri, evt in enumerate(uw_detail, 1):
                r = detail_start + ri
                for ci, col_name in enumerate(detail_cols, 1):
                    cell = ds.cell(row=r, column=ci)
                    val = evt[col_name]
                    if col_name == "Date":
                        cell.value = val
                        cell.number_format = "M/D/YYYY"
                    elif col_name in currency_detail:
                        cell.value = float(val)
                        cell.number_format = "$#,##0"
                    else:
                        cell.value = val

            # Totals row
            total_row = detail_start + len(uw_detail) + 1
            ds.cell(row=total_row, column=1, value="TOTAL").font = bold_font
            top_border = Border(top=Side(style="thin"))
            sum_col = detail_cols.index("Weighted Capital") + 1
            days_col = detail_cols.index("Days") + 1
            for ci in [days_col, sum_col]:
                cell = ds.cell(row=total_row, column=ci)
                first_data = detail_start + 1
                last_data = detail_start + len(uw_detail)
                col_letter = cell.column_letter
                cell.value = f"=SUM({col_letter}{first_data}:{col_letter}{last_data})"
                cell.font = bold_font
                cell.border = top_border
                if ci == sum_col:
                    cell.number_format = "$#,##0"

            avg_row = total_row + 1
            ds.cell(row=avg_row, column=1, value="Wtd Avg Capital =").font = label_font
            sum_letter = ds.cell(row=total_row, column=sum_col).column_letter
            days_letter = ds.cell(row=total_row, column=days_col).column_letter
            ds.cell(row=avg_row, column=2,
                    value=f"={sum_letter}{total_row}/{days_letter}{total_row}")
            ds.cell(row=avg_row, column=2).number_format = "$#,##0"

            widths = [12, 30, 14, 8, 16, 18, 16]
            for ci, w in enumerate(widths, 1):
                ds.column_dimensions[ds.cell(row=1, column=ci).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Pref Balance Detail Report
# ---------------------------------------------------------------------------

def _days_in_year(yr: int) -> int:
    """Return 366 for leap years, 365 otherwise (Act/Act convention)."""
    return 366 if calendar.isleap(yr) else 365


def get_deal_pe_investors(
    vcode: str,
    acct: pd.DataFrame,
    inv_map: pd.DataFrame,
) -> list[dict]:
    """Return list of PE investors for a deal (excludes OP partners)."""
    from loaders import build_investmentid_to_vcode

    vcode_str = str(vcode).strip()
    inv_to_vcode = build_investmentid_to_vcode(inv_map)
    deal_iids = [iid for iid, vc in inv_to_vcode.items()
                 if str(vc).strip().upper() == vcode_str.upper()]
    if not deal_iids:
        return []

    acct_norm = acct.copy()
    normalize_columns(acct_norm)
    acct_norm["InvestorID"] = acct_norm["InvestorID"].astype(str).str.strip()
    acct_norm["InvestmentID"] = acct_norm["InvestmentID"].astype(str).str.strip()

    deal_acct = acct_norm[acct_norm["InvestmentID"].isin(deal_iids)]
    # Collapse casing variants of one investor.  The accounting feed carries the
    # same InvestorID under two casings on some deals (e.g. PPI11 and ppi11),
    # and build_pref_balance_detail() matches InvestorID case-insensitively — so
    # each variant returns the *whole* combined ledger.  Left un-deduped, every
    # caller that iterates this list counts that ledger once per variant, which
    # doubled the One Pager's accrued pref balance.  Sorted order puts the
    # upper-case variant first, so that is the label kept.
    canonical: dict[str, str] = {}
    for iid in sorted(deal_acct["InvestorID"].unique()):
        canonical.setdefault(iid.upper(), iid)
    return [{"investor_id": iid} for iid in canonical.values()]


def _quarter_ends_between(start: date, end: date) -> list[date]:
    """Return all quarter-end dates strictly between start and end."""
    qends = []
    yr = start.year
    while yr <= end.year + 1:
        for m in (3, 6, 9, 12):
            d = date(yr, m, calendar.monthrange(yr, m)[1])
            if start < d < end:
                qends.append(d)
        yr += 1
    return qends


def build_pref_balance_detail(
    vcode: str,
    investor_id: str,
    report_date: date,
    acct: pd.DataFrame,
    inv_map: pd.DataFrame,
    wf_steps: Optional[pd.DataFrame] = None,
) -> dict:
    """Build pref balance detail matching the PE_Pref_Balances Excel layout.

    Returns dict with 'header' (summary info) and 'rows' (transaction detail).
    Uses Act/Act day count convention (366 for leap years).
    """
    from loaders import build_investmentid_to_vcode

    vcode_str = str(vcode).strip()
    inv_to_vcode = build_investmentid_to_vcode(inv_map)
    deal_iids = [iid for iid, vc in inv_to_vcode.items()
                 if str(vc).strip().upper() == vcode_str.upper()]

    # Get pref rate — priority: deal_terms pe_coupon > waterfall
    pref_rate = 0.0

    # 1. deal_terms pe_coupon (authoritative contractual rate)
    try:
        from database import _sa_engine
        if _sa_engine is not None:
            dt_df = pd.read_sql(
                "SELECT pe_coupon FROM deal_terms WHERE UPPER(vcode) = :vc",
                _sa_engine, params={"vc": vcode_str.upper()},
            )
            if not dt_df.empty and pd.notna(dt_df.iloc[0]["pe_coupon"]):
                pref_rate = float(dt_df.iloc[0]["pe_coupon"])
    except Exception:
        pass

    # 2. Waterfall Pref step matching this investor
    if pref_rate <= 0 and wf_steps is not None and not wf_steps.empty:
        wf_deal = wf_steps[wf_steps["vcode"] == vcode_str] if "vcode" in wf_steps.columns else wf_steps
        pref_rows = wf_deal[wf_deal["vState"] == "Pref"] if "vState" in wf_deal.columns else pd.DataFrame()
        rate_col = "nPercent_dec" if "nPercent_dec" in pref_rows.columns else "nPercent"
        for _, pr in pref_rows.iterrows():
            pc = str(pr.get("PropCode", "")).strip()
            r = float(pr[rate_col]) if pd.notna(pr.get(rate_col)) else 0.0
            if pc.upper() == investor_id.upper() and r > 0:
                pref_rate = r
                break
        # 3. Any Pref rate for this deal
        if pref_rate <= 0:
            for _, pr in pref_rows.iterrows():
                r = float(pr[rate_col]) if pd.notna(pr.get(rate_col)) else 0.0
                if r > 0:
                    pref_rate = r
                    break

    # Normalize accounting data
    acct_norm = acct.copy()
    normalize_columns(acct_norm)
    acct_norm["InvestorID"] = acct_norm["InvestorID"].astype(str).str.strip()
    acct_norm["InvestmentID"] = acct_norm["InvestmentID"].astype(str).str.strip()
    acct_norm["EffectiveDate"] = pd.to_datetime(acct_norm["EffectiveDate"], errors="coerce")
    acct_norm["Amt"] = pd.to_numeric(acct_norm["Amt"], errors="coerce").fillna(0.0)
    if "TypeName" not in acct_norm.columns and "Typename" in acct_norm.columns:
        acct_norm["TypeName"] = acct_norm["Typename"]
    elif "TypeName" not in acct_norm.columns:
        acct_norm["TypeName"] = ""
    acct_norm["TypeName"] = acct_norm["TypeName"].fillna("").astype(str).str.strip()
    acct_norm["MajorType"] = acct_norm["MajorType"].fillna("").astype(str).str.strip()

    deal_acct = acct_norm[
        (acct_norm["InvestmentID"].isin(deal_iids))
        & (acct_norm["InvestorID"].str.upper() == investor_id.upper())
        & (acct_norm["EffectiveDate"].dt.date <= report_date)
    ].sort_values("EffectiveDate").copy()

    if deal_acct.empty:
        return {"header": {}, "rows": []}

    # Detect TypeID column
    has_typeid = "TypeID" in deal_acct.columns

    # Build event list: actual transactions + generated quarter-end markers
    events = []  # (date, amt, typename, major_type, is_generated, type_id)
    for _, r in deal_acct.iterrows():
        if r.get("is_commitment", False):
            continue
        evt_date = r["EffectiveDate"].date()
        tname_raw = r["TypeName"]
        tid = None
        if has_typeid:
            try:
                tid = float(r["TypeID"])
            except (ValueError, TypeError):
                pass
        events.append((
            evt_date, float(r["Amt"]), tname_raw, r["MajorType"],
            False, tid,
        ))

    # Insert generated quarter-end rows between events and at report_date
    if events:
        first_date = events[0][0]
        all_dates = set(e[0] for e in events)
        qends = _quarter_ends_between(first_date, report_date)
        for qe in qends:
            if qe not in all_dates:
                events.append((qe, 0, "Generated", "", True, None))
        # Add report_date if not already present and not a quarter-end already added
        added_dates = set(e[0] for e in events)
        if report_date not in added_dates:
            events.append((report_date, 0, "Generated", "", True, None))

    def _evt_sort_key(e):
        """Sort: date, then contributions first, pref return, excess CF, generated last."""
        dt, _, tn, mt, is_gen, _ = e
        tn_lower = tn.lower()
        if is_gen:
            order = 90
        elif "contrib" in mt.lower():
            order = 10
        elif "preferred return" in tn_lower or "pref return" in tn_lower:
            order = 20
        elif "return of capital" in tn_lower or "realized gain" in tn_lower:
            order = 25
        elif "excess cash" in tn_lower:
            order = 30
        else:
            order = 50
        return (dt, order)

    events.sort(key=_evt_sort_key)

    # Walk through events row-by-row matching Excel PE_Pref_Balances formulas.
    # All "balance" columns use negative sign convention (owed = negative).
    # Excel columns: F=InvBal, G=CompPref, H=Inv+Comp, I=Days, J=CurrDue,
    #   K=AccrPref, L=TotalDue, M=PrefPaid, N=Remaining
    # Key formulas:
    #   H(row) = F(prior) + G(prior)          — accrual base is PRIOR row's values
    #   J(row) = H(row) * rate / diy * days
    #   K(row) = N(prior)                     — carry forward prior Remaining
    #   L(row) = J + K                        — total owed this period
    #   M(row) = payment if pref/excess CF    — positive (reduces debt)
    #   N(row) = MIN(0, L + M)                — remaining after payment
    #   G(row) = at 12/31: N(row)             — all unpaid pref compounds
    #            else: MIN(0, G(prior) + M + J) if (M+J)>0, else G(prior)

    inv_id = deal_iids[0] if deal_iids else ""

    # State variables matching Excel row values (stored as positive magnitudes)
    inv_bal = 0.0       # F: Investment Balance (positive = capital outstanding)
    comp_pref = 0.0     # G: Compounded Pref (positive = compounded pref owed)
    remaining = 0.0     # N: Remaining accrual (positive = pref owed)
    prev_date = None
    result_rows = []

    for evt_date, amt, typename, major_type, is_generated, type_id in events:
        major = major_type.lower()
        tname = typename.lower()

        # --- Update capital balance FIRST (affects InvBal for this row) ---
        if "contrib" in major:
            inv_bal += abs(amt)
        elif "distri" in major:
            if "return of capital" in tname or "realized gain" in tname:
                inv_bal = max(0.0, inv_bal - abs(amt))

        # --- H: Inv+Comp = PRIOR row's InvBal + PRIOR row's CompPref ---
        # On the first row there is no prior, so Inv+Comp = 0 (matches Excel)
        if not result_rows:
            inv_plus_comp = 0.0
        else:
            prior = result_rows[-1]
            inv_plus_comp = abs(prior["Investment_Balance"]) + abs(prior["Compounded Pref"])

        # --- I: Days since last event ---
        days_since = (evt_date - prev_date).days if prev_date else 0

        # --- J: Current Due = Inv+Comp * rate / diy * days ---
        # Excel uses simple formula: total days × rate / days_in_year(event year)
        # No splitting across year boundaries — matches Excel exactly.
        curr_due = 0.0
        if prev_date is not None and inv_plus_comp > 0 and days_since > 0:
            diy = _days_in_year(evt_date.year)
            curr_due = inv_plus_comp * pref_rate * (days_since / diy)

        # --- K: Accrued Pref = prior row's Remaining ---
        accr_pref = remaining  # carry forward from prior row

        # --- L: Total Due = Current Due + Accrued Pref ---
        total_due = curr_due + accr_pref

        # --- M: Pref Paid (positive = payment that reduces owed balance) ---
        pref_paid = 0.0
        is_pref_payment = (type_id == 1019.0) or "preferred return" in tname or "pref return" in tname
        is_excess_cf = (type_id == 1020.0) or "excess cash" in tname
        if "distri" in major and (is_pref_payment or is_excess_cf):
            pref_paid = abs(amt)

        # --- N: Remaining = MIN(0, TotalDue + PrefPaid) ---
        # TotalDue is positive (owed), PrefPaid is positive (paid).
        # In Excel sign convention both are negative/positive respectively,
        # but we track magnitudes: remaining_owed = max(0, total_due - pref_paid)
        remaining = max(0.0, total_due - pref_paid)

        # --- G: Compounded Pref ---
        is_year_end = evt_date.month == 12 and evt_date.day == 31
        if is_year_end:
            # At 12/31, all remaining unpaid pref compounds
            comp_pref = remaining
        else:
            # If (PrefPaid + CurrDue) net reduces the owed balance (payment > accrual),
            # the excess reduces compounded pref. Otherwise comp_pref unchanged.
            # Excel: MIN(0, G_prior + M + J) when (M+J)>0, else G_prior
            # In our magnitude convention: payment exceeds accrual → reduce comp_pref
            net_payment = pref_paid - curr_due  # positive if payment > accrual
            if net_payment > 0:
                comp_pref = max(0.0, comp_pref - net_payment)

        # --- Sign convention for display: negative = owed/invested ---
        display_amt = -abs(amt) if "contrib" in major else (abs(amt) if amt != 0 else 0)

        row = {
            "InvestmentID": inv_id,
            "InvestorID": investor_id,
            "EffectiveDate": evt_date.isoformat(),
            "Amt": display_amt,
            "Typename": typename,
            "Investment_Balance": -inv_bal if inv_bal > 0 else 0,
            "Compounded Pref": -comp_pref if comp_pref > 0 else 0,
            "Inv + Comp": -inv_plus_comp if inv_plus_comp > 0 else 0,
            "DaysSinceLast": days_since,
            "Current Due": -curr_due if curr_due > 0 else 0,
            "Accrued Pref": -accr_pref if accr_pref > 0 else 0,
            "Total Due": -total_due if total_due > 0 else 0,
            "Pref Paid": pref_paid,
            "Remaining Accrual": -remaining if remaining > 0 else 0,
        }
        result_rows.append(row)
        prev_date = evt_date

    # Build header
    annual_pref_est = inv_bal * pref_rate if inv_bal > 0 and pref_rate > 0 else 0
    total_accrued = remaining

    header = {
        "vcode": vcode_str,
        "investor_id": investor_id,
        "investment_id": inv_id,
        "pref_rate": pref_rate,
        "report_date": report_date.isoformat(),
        "investment_balance": inv_bal,
        "accrued_pref": total_accrued,
        "total": inv_bal + total_accrued,
        "annual_pref_est": annual_pref_est,
    }

    return {"header": header, "rows": result_rows}


def generate_pref_balance_excel(header: dict, rows: list[dict]) -> bytes:
    """Generate formatted Excel for Pref Balance Detail report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pref Balance Detail"

    bold = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    curr_fmt = '$#,##0'
    pct_fmt = '0.00%'

    # Header section
    ws.cell(row=1, column=1, value="Property #:").font = bold
    ws.cell(row=1, column=2, value=header.get("vcode", ""))
    ws.cell(row=1, column=4, value="As of:").font = bold
    ws.cell(row=1, column=5, value=header.get("report_date", ""))

    ws.cell(row=2, column=1, value="Investment ID:").font = bold
    ws.cell(row=2, column=2, value=header.get("investment_id", ""))
    ws.cell(row=2, column=4, value="Investment Balance:").font = bold
    c = ws.cell(row=2, column=5, value=header.get("investment_balance", 0))
    c.number_format = curr_fmt
    ws.cell(row=2, column=7, value="Annual Pref Est").font = bold
    c = ws.cell(row=2, column=8, value=header.get("annual_pref_est", 0))
    c.number_format = curr_fmt

    ws.cell(row=3, column=1, value="Investor:").font = bold
    ws.cell(row=3, column=2, value=header.get("investor_id", ""))
    ws.cell(row=3, column=4, value="Accrued:").font = bold
    c = ws.cell(row=3, column=5, value=header.get("accrued_pref", 0))
    c.number_format = curr_fmt

    ws.cell(row=4, column=1, value="Pref Rate:").font = bold
    c = ws.cell(row=4, column=2, value=header.get("pref_rate", 0))
    c.number_format = pct_fmt
    ws.cell(row=4, column=4, value="Total:").font = bold
    c = ws.cell(row=4, column=5, value=header.get("total", 0))
    c.number_format = curr_fmt

    # Column headers at row 6
    columns = [
        "InvestmentID", "InvestorID", "EffectiveDate", "Amt", "Typename",
        "Investment_Balance", "Compounded Pref", "Inv + Comp",
        "DaysSinceLast", "Current Due", "Accrued Pref", "Total Due",
        "Pref Paid", "Remaining Accrual",
    ]
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=6, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    currency_cols = {"Amt", "Investment_Balance", "Compounded Pref", "Inv + Comp",
                     "Current Due", "Accrued Pref", "Total Due", "Pref Paid",
                     "Remaining Accrual"}

    for ri, row in enumerate(rows, 7):
        for ci, col in enumerate(columns, 1):
            val = row.get(col)
            cell = ws.cell(row=ri, column=ci)
            if col in currency_cols:
                cell.value = float(val) if val is not None else 0.0
                cell.number_format = curr_fmt
            elif col == "DaysSinceLast":
                cell.value = int(val) if val is not None else 0
            else:
                cell.value = val

    # Auto-width
    for ci, col in enumerate(columns, 1):
        ws.column_dimensions[ws.cell(row=6, column=ci).column_letter].width = max(len(col) + 4, 16)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
