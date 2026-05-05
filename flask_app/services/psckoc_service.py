"""PSCKOC service — upstream entity analysis for PSCKOC holding entity.

Extracts pure logic from psckoc_ui.py (no Streamlit dependency).
Traces deal cash flows through PPI entities to PSCKOC members (PSC1, KCREIT, PCBLE).
"""

import pandas as pd
from io import BytesIO
from datetime import date
from typing import Optional

from ownership_tree import load_relationships
from loaders import load_waterfalls
from waterfall import run_recursive_upstream_waterfalls
from metrics import xirr, calculate_roe

# ── Constants ────────────────────────────────────────────────────────────
PSCKOC_ENTITY = "PSCKOC"
PSCKOC_MEMBERS = ["PSC1", "KCREIT", "PCBLE"]

# Module-level cache for PSCKOC results
_psckoc_cache: dict = {}


def find_psckoc_deals(inv: pd.DataFrame, wf: pd.DataFrame,
                      relationships_raw: pd.DataFrame) -> list[dict]:
    """Find deals whose upstream waterfall chain passes through PSCKOC.

    Returns list of deal info dicts with vcode, name, asset type, PPI linkage.
    """
    if relationships_raw is None or relationships_raw.empty:
        return []
    if wf is None or wf.empty:
        return []
    if inv is None or inv.empty:
        return []

    rel = relationships_raw.copy()
    rel["InvestorID"] = rel["InvestorID"].astype(str).str.strip()
    rel["InvestmentID"] = rel["InvestmentID"].astype(str).str.strip()

    # Step 1: entities where PSCKOC is an investor
    psckoc_investments = set(
        rel[rel["InvestorID"] == PSCKOC_ENTITY]["InvestmentID"].unique()
    )
    if not psckoc_investments:
        return []

    # Step 2: find deal vcodes whose waterfall steps reference a PSCKOC investment entity
    wf_norm = wf.copy()
    wf_norm["PropCode"] = wf_norm["PropCode"].fillna("").astype(str).str.strip()
    wf_norm["vcode"] = wf_norm["vcode"].fillna("").astype(str).str.strip()

    deal_vcodes_set = set(inv["vcode"].astype(str).str.strip())
    wf_vcodes_set = set(wf_norm["vcode"].unique())

    deal_vcodes = set()
    ppi_for_deal = {}  # deal_vcode -> (ppi_entity, ownership_pct)

    # Build PPI ownership percentages
    psckoc_inv = rel[rel["InvestorID"] == PSCKOC_ENTITY]
    ppi_pcts = {}
    for _, r in psckoc_inv.iterrows():
        ppi = r["InvestmentID"]
        pct = float(r.get("OwnershipPct", 0))
        if pct > 1:
            pct = pct / 100.0
        ppi_pcts[ppi] = pct

    for ppi in psckoc_investments:
        refs = wf_norm[wf_norm["PropCode"] == ppi]["vcode"].unique()
        for vc in refs:
            if vc in deal_vcodes_set and vc in wf_vcodes_set:
                deal_vcodes.add(vc)
                ppi_for_deal[vc] = (ppi, ppi_pcts.get(ppi, 0))

    # Remove PSCKOC itself, members, and PPI intermediaries
    deal_vcodes -= {PSCKOC_ENTITY} | set(PSCKOC_MEMBERS) | psckoc_investments

    # Build deal info list
    result = []
    for vc in sorted(deal_vcodes):
        row_info = {"vcode": vc, "name": vc}
        match = inv[inv["vcode"].astype(str).str.strip() == vc]
        if not match.empty:
            r = match.iloc[0]
            row_info["name"] = str(r.get("Investment_Name", vc))
            row_info["asset_type"] = str(r.get("Asset_Type", ""))
            row_info["sale_status"] = str(r.get("Sale_Status", "Active") or "Active")
        if vc in ppi_for_deal:
            ppi, pct = ppi_for_deal[vc]
            row_info["ppi_entity"] = ppi
            row_info["psckoc_pct"] = pct
        result.append(row_info)

    return result


def run_psckoc_computation(deal_vcodes: list[str], data: dict,
                           start_year: int, horizon_years: int,
                           pro_yr_base: int) -> dict:
    """Run deal computations and upstream waterfalls for PSCKOC.

    Uses compute_service.get_cached_deal_result() per deal, then
    run_recursive_upstream_waterfalls() for CF and Cap.

    Returns structured results dict.
    """
    from flask_app.services.compute_service import get_cached_deal_result

    inv = data["inv"]
    wf = data["wf"]
    wf_steps = load_waterfalls(wf)

    all_cf_allocs = []
    all_cap_allocs = []
    deal_results = {}
    errors = []

    actuals_through = data.get("actuals_through")

    for vcode in deal_vcodes:
        try:
            result = get_cached_deal_result(
                vcode=vcode,
                start_year=start_year,
                horizon_years=horizon_years,
                pro_yr_base=pro_yr_base,
                data=data,
                actuals_through=actuals_through,
            )
        except Exception as e:
            errors.append(f"Deal {vcode}: {e}")
            continue

        if "error" in result:
            errors.append(f"Deal {vcode}: {result['error']}")
            continue

        deal_results[vcode] = result

        cf_alloc = result.get("cf_alloc")
        cap_alloc = result.get("cap_alloc")

        if cf_alloc is not None and not cf_alloc.empty:
            cf = cf_alloc.copy()
            cf["vcode"] = vcode
            all_cf_allocs.append(cf)

        if cap_alloc is not None and not cap_alloc.empty:
            ca = cap_alloc.copy()
            ca["vcode"] = vcode
            all_cap_allocs.append(ca)

    if not all_cf_allocs and not all_cap_allocs:
        return {"error": "No deal allocations produced", "errors": errors}

    # Build relationships
    relationships_raw = data["relationships_raw"]
    if relationships_raw is not None and not relationships_raw.empty:
        relationships = load_relationships(relationships_raw)
    else:
        relationships = pd.DataFrame()

    # Combined allocations
    combined_cf = pd.concat(all_cf_allocs, ignore_index=True) if all_cf_allocs else pd.DataFrame()
    combined_cap = pd.concat(all_cap_allocs, ignore_index=True) if all_cap_allocs else pd.DataFrame()

    # Pre-compute AMFee exclusion capital for upstream waterfalls
    from waterfall import build_amfee_exclusions
    _acct = data.get("acct")
    _excl = build_amfee_exclusions(_acct, relationships) if _acct is not None else {}

    # Run CF upstream waterfalls
    cf_upstream_alloc = pd.DataFrame()
    cf_entity_states = {}
    cf_beneficiary_totals = {}
    if not combined_cf.empty:
        cf_upstream_alloc, cf_entity_states, cf_beneficiary_totals = \
            run_recursive_upstream_waterfalls(
                deal_allocations=combined_cf,
                wf_steps=wf_steps,
                relationships=relationships,
                wf_type="CF_WF",
                amfee_exclusions=_excl,
            )

    # Run Cap upstream waterfalls
    cap_upstream_alloc = pd.DataFrame()
    cap_entity_states = {}
    cap_beneficiary_totals = {}
    if not combined_cap.empty:
        cap_upstream_alloc, cap_entity_states, cap_beneficiary_totals = \
            run_recursive_upstream_waterfalls(
                deal_allocations=combined_cap,
                wf_steps=wf_steps,
                relationships=relationships,
                wf_type="Cap_WF",
                amfee_exclusions=_excl,
            )

    # Build structured results
    results = _build_psckoc_results(
        cf_upstream_alloc=cf_upstream_alloc,
        cap_upstream_alloc=cap_upstream_alloc,
        cf_entity_states=cf_entity_states,
        cap_entity_states=cap_entity_states,
        cf_beneficiary_totals=cf_beneficiary_totals,
        cap_beneficiary_totals=cap_beneficiary_totals,
        deal_results=deal_results,
        deal_vcodes=deal_vcodes,
        inv=inv,
    )
    results["errors"] = errors
    results["deals_computed"] = len(deal_results)

    # Cache results
    _psckoc_cache["results"] = results
    return results


def get_cached_results() -> Optional[dict]:
    """Return cached PSCKOC results, or None."""
    return _psckoc_cache.get("results")


def clear_cache():
    """Clear PSCKOC results cache."""
    _psckoc_cache.clear()


def _build_psckoc_results(cf_upstream_alloc, cap_upstream_alloc,
                           cf_entity_states, cap_entity_states,
                           cf_beneficiary_totals, cap_beneficiary_totals,
                           deal_results, deal_vcodes, inv):
    """Build structured results from upstream waterfall outputs."""

    # ── Income schedule: cash arriving at PSCKOC ──────────────────────
    income_rows = []
    for wf_type, alloc_df in [("CF", cf_upstream_alloc), ("Cap", cap_upstream_alloc)]:
        if isinstance(alloc_df, pd.DataFrame) and alloc_df.empty:
            continue
        psckoc_income = alloc_df[alloc_df["PropCode"].astype(str) == PSCKOC_ENTITY].copy()
        for _, row in psckoc_income.iterrows():
            path_str = str(row.get("Path", ""))
            income_rows.append({
                "Date": str(row.get("event_date", "")),
                "Source Entity": str(row.get("Entity", "")),
                "Source Deal": path_str.split("->")[0] if path_str else "",
                "Type": wf_type,
                "vState": str(row.get("vState", "")),
                "vtranstype": str(row.get("vtranstype", "")),
                "Amount": float(row.get("Allocated", 0)),
            })

    # ── Member allocations: cash leaving PSCKOC to members ────────────
    member_alloc_rows = []
    for wf_type, alloc_df in [("CF", cf_upstream_alloc), ("Cap", cap_upstream_alloc)]:
        if isinstance(alloc_df, pd.DataFrame) and alloc_df.empty:
            continue
        psckoc_dists = alloc_df[alloc_df["Entity"].astype(str) == PSCKOC_ENTITY].copy()
        for _, row in psckoc_dists.iterrows():
            member_alloc_rows.append({
                "Date": str(row.get("event_date", "")),
                "Member": str(row.get("PropCode", "")),
                "Type": wf_type,
                "iOrder": int(row.get("iOrder", 0)),
                "vState": str(row.get("vState", "")),
                "vtranstype": str(row.get("vtranstype", "")),
                "FXRate": float(row.get("FXRate", 0)),
                "Amount": float(row.get("Allocated", 0)),
            })

    # ── Partner returns from entity states ────────────────────────────
    partner_returns = []
    for member in PSCKOC_MEMBERS:
        cf_st = cf_entity_states.get(member)
        cap_st = cap_entity_states.get(member)
        state = cf_st or cap_st
        if not state:
            continue

        # Combined cashflows from both waterfall types
        combined_cfs = []
        seen = set()
        for st_obj in [cf_st, cap_st]:
            if st_obj and st_obj.cashflows:
                for cf in st_obj.cashflows:
                    key = (cf[0], cf[1])
                    if key not in seen:
                        combined_cfs.append(cf)
                        seen.add(key)
        combined_cfs.sort(key=lambda x: x[0])

        # Calculate metrics
        irr_val = xirr(combined_cfs) if combined_cfs else None
        contributions = sum(amt for _, amt in combined_cfs if amt < 0)
        distributions = sum(amt for _, amt in combined_cfs if amt > 0)

        # CF distributions for ROE
        cf_dists = []
        if cf_st and cf_st.cf_distributions:
            cf_dists = cf_st.cf_distributions

        roe_val = calculate_roe(
            combined_cfs,
            cf_dists,
            combined_cfs[0][0] if combined_cfs else date.today(),
            combined_cfs[-1][0] if combined_cfs else date.today(),
        ) if combined_cfs else 0.0

        moic_val = distributions / abs(contributions) if contributions < 0 else 0.0

        # CF vs Cap distribution totals from member allocations
        member_alloc_df = pd.DataFrame(member_alloc_rows) if member_alloc_rows else pd.DataFrame()
        cf_dist_total = 0.0
        cap_dist_total = 0.0
        if not member_alloc_df.empty:
            cf_member = member_alloc_df[
                (member_alloc_df["Member"] == member) & (member_alloc_df["Type"] == "CF")
            ]
            cf_dist_total = cf_member["Amount"].sum()
            cap_member = member_alloc_df[
                (member_alloc_df["Member"] == member) & (member_alloc_df["Type"] == "Cap")
            ]
            cap_dist_total = cap_member["Amount"].sum()

        # Serialize cashflows (date -> string for JSON)
        cashflows_serial = [
            {"date": str(d), "amount": float(a)}
            for d, a in combined_cfs
        ]

        partner_returns.append({
            "partner": member,
            "contributions": contributions,
            "cf_distributions": cf_dist_total,
            "cap_distributions": cap_dist_total,
            "total_distributions": distributions,
            "irr": irr_val,
            "roe": roe_val,
            "moic": moic_val,
            "combined_cashflows": cashflows_serial,
        })

    # ── AM Fee schedule ───────────────────────────────────────────────
    am_fee_rows = []
    for wf_type, alloc_df in [("CF", cf_upstream_alloc), ("Cap", cap_upstream_alloc)]:
        if isinstance(alloc_df, pd.DataFrame) and alloc_df.empty:
            continue
        am_fees = alloc_df[
            (alloc_df["Entity"].astype(str) == PSCKOC_ENTITY) &
            (alloc_df["vState"].astype(str) == "AMFee")
        ].copy()
        for _, row in am_fees.iterrows():
            am_fee_rows.append({
                "Date": str(row.get("event_date", "")),
                "Type": wf_type,
                "Recipient": str(row.get("PropCode", "")),
                "Amount": float(row.get("Allocated", 0)),
            })

    # ── Deal summary ──────────────────────────────────────────────────
    all_combined = []
    for pr in partner_returns:
        for cf in pr["combined_cashflows"]:
            all_combined.append((cf["date"], cf["amount"]))
    all_combined.sort(key=lambda x: x[0])

    total_contributions = sum(amt for _, amt in all_combined if amt < 0)
    total_distributions = sum(amt for _, amt in all_combined if amt > 0)

    # Need actual date objects for xirr
    from datetime import datetime
    xirr_cfs = []
    for d_str, amt in all_combined:
        try:
            d = datetime.strptime(d_str, "%Y-%m-%d").date() if isinstance(d_str, str) else d_str
            xirr_cfs.append((d, amt))
        except (ValueError, TypeError):
            pass
    deal_irr = xirr(xirr_cfs) if xirr_cfs else None
    deal_moic = total_distributions / abs(total_contributions) if total_contributions < 0 else 0.0

    deal_summary = {
        "deal_irr": deal_irr,
        "deal_moic": deal_moic,
        "total_contributions": total_contributions,
        "total_distributions": total_distributions,
    }

    return {
        "income_schedule": income_rows,
        "member_allocations": member_alloc_rows,
        "partner_results": partner_returns,
        "deal_summary": deal_summary,
        "am_fee_schedule": am_fee_rows,
        "deal_vcodes": deal_vcodes,
    }


def generate_psckoc_excel(results: dict) -> bytes:
    """Generate 4-sheet PSCKOC Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        return b""

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    currency_fmt = '#,##0'
    pct_fmt = '0.00%'
    mult_fmt = '0.00"x"'
    bold_font = Font(bold=True)
    top_border = Border(top=Side(style="thin"))

    # ── Sheet 1: Partner Returns ──────────────────────────────────────
    ws = wb.active
    ws.title = "Partner Returns"

    headers = ["Partner", "Contributions", "CF Distributions", "Cap Distributions",
               "Total Distributions", "IRR", "ROE", "MOIC"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    partner_results = results.get("partner_results", [])
    deal_summary = results.get("deal_summary", {})

    row_idx = 2
    for pr in partner_results:
        ws.cell(row=row_idx, column=1, value=pr["partner"])
        ws.cell(row=row_idx, column=2, value=pr["contributions"]).number_format = currency_fmt
        ws.cell(row=row_idx, column=3, value=pr["cf_distributions"]).number_format = currency_fmt
        ws.cell(row=row_idx, column=4, value=pr["cap_distributions"]).number_format = currency_fmt
        ws.cell(row=row_idx, column=5, value=pr["total_distributions"]).number_format = currency_fmt
        if pr["irr"] is not None:
            ws.cell(row=row_idx, column=6, value=pr["irr"]).number_format = pct_fmt
        if pr["roe"] is not None:
            ws.cell(row=row_idx, column=7, value=pr["roe"]).number_format = pct_fmt
        ws.cell(row=row_idx, column=8, value=pr["moic"]).number_format = mult_fmt
        row_idx += 1

    # Total row
    for col in range(1, 9):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = bold_font
        cell.border = top_border
    ws.cell(row=row_idx, column=1, value="PSCKOC Total")
    ws.cell(row=row_idx, column=2, value=deal_summary.get("total_contributions", 0)).number_format = currency_fmt
    ws.cell(row=row_idx, column=5, value=deal_summary.get("total_distributions", 0)).number_format = currency_fmt
    if deal_summary.get("deal_irr") is not None:
        ws.cell(row=row_idx, column=6, value=deal_summary["deal_irr"]).number_format = pct_fmt
    ws.cell(row=row_idx, column=8, value=deal_summary.get("deal_moic", 0)).number_format = mult_fmt

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for r in range(2, row_idx + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 25)

    # ── Sheet 2: Income Schedule ──────────────────────────────────────
    income_rows = results.get("income_schedule", [])
    if income_rows:
        ws2 = wb.create_sheet("Income Schedule")
        inc_headers = ["Date", "Source Entity", "Source Deal", "Type", "vState", "vtranstype", "Amount"]
        for col, h in enumerate(inc_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for r_idx, row in enumerate(income_rows, 2):
            for c_idx, h in enumerate(inc_headers, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=row.get(h, ""))
                if h == "Amount":
                    cell.number_format = currency_fmt

    # ── Sheet 3: AM Fee Schedule ──────────────────────────────────────
    am_fee_rows = results.get("am_fee_schedule", [])
    if am_fee_rows:
        ws3 = wb.create_sheet("AM Fee Schedule")
        am_headers = ["Date", "Type", "Recipient", "Amount"]
        for col, h in enumerate(am_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for r_idx, row in enumerate(am_fee_rows, 2):
            for c_idx, h in enumerate(am_headers, 1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=row.get(h, ""))
                if h == "Amount":
                    cell.number_format = currency_fmt

    # ── Sheet 4: XIRR Cash Flows ─────────────────────────────────────
    ws4 = wb.create_sheet("XIRR Cash Flows")
    cf_headers = ["Date", "Partner", "Amount"]
    for col, h in enumerate(cf_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    r_idx = 2
    for pr in partner_results:
        for cf in pr.get("combined_cashflows", []):
            ws4.cell(row=r_idx, column=1, value=cf["date"])
            ws4.cell(row=r_idx, column=2, value=pr["partner"])
            ws4.cell(row=r_idx, column=3, value=cf["amount"]).number_format = currency_fmt
            r_idx += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
