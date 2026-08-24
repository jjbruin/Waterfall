"""
cashflow_parser.py
Generic parser for partner Excel/CSV cash flow models.

Reads a spreadsheet and extracts monthly or annual cash flow rows
(revenue, expenses, capex, NOI) into a list of dicts compatible
with the prospect_cashflows table schema.

No DB or Flask dependencies. Takes file bytes, returns structured dicts.
"""

import io
import logging
import re
from datetime import date
from typing import Dict, Any, List, Optional, Tuple

import pandas as pd
import numpy as np

from utils import month_end

log = logging.getLogger(__name__)


# ============================================================
# COLUMN DETECTION PATTERNS
# ============================================================
# Partners label columns inconsistently. These patterns match
# common variants for each cash flow category.

_DATE_PATTERNS = [
    r'^date$', r'^period', r'^month', r'^year', r'^time',
]

_REVENUE_PATTERNS = [
    r'revenue', r'income', r'gross.?rev', r'egi', r'effective.?gross',
    r'rental.?income', r'total.?revenue', r'total.?income',
    r'potential.?rent', r'base.?rent', r'gross.?income',
    r'gross.?potential', r'collected.?rent',
]

_EXPENSE_PATTERNS = [
    r'expense', r'opex', r'operating.?exp', r'total.?exp',
    r'total.?opex', r'operating.?cost',
]

_NOI_PATTERNS = [
    r'\bnoi\b', r'net.?operating', r'net.?income',
]

_CAPEX_PATTERNS = [
    r'capex', r'capital.?exp', r'capital.?reserve', r'cap.?ex',
    r'tenant.?improve', r'leasing.?comm', r'\bti\b.*\blc\b',
    r'capital.?cost',
]


def _match_column(col_name: str, patterns: list) -> bool:
    """Check if a column name matches any of the given regex patterns."""
    col_lower = str(col_name).lower().strip()
    return any(re.search(p, col_lower) for p in patterns)


def _find_column(columns: list, patterns: list) -> Optional[str]:
    """Find the first column matching the patterns."""
    for col in columns:
        if _match_column(col, patterns):
            return col
    return None


def _detect_date_column(df: pd.DataFrame) -> Optional[str]:
    """Find the date/period column, trying name patterns then dtype."""
    # Try name patterns first
    for col in df.columns:
        if _match_column(col, _DATE_PATTERNS):
            return col

    # Try the first column if it looks like dates
    first_col = df.columns[0]
    sample = df[first_col].dropna().head(5)
    if len(sample) > 0:
        try:
            parsed = pd.to_datetime(sample, format='mixed', dayfirst=False)
            if parsed.notna().sum() >= len(sample) * 0.6:
                return first_col
        except Exception:
            pass

    return None


# ============================================================
# LINE-ITEM ROW MATCHING (for horizontal/transposed layouts)
# ============================================================

_ROW_REVENUE_PATTERNS = [
    r'effective.?gross.?revenue', r'total.?tenant.?revenue',
    r'potential.?gross.?revenue', r'total.?revenue',
    r'gross.?revenue', r'total.?income', r'gross.?income',
    r'effective.?gross.?income',
]

_ROW_EXPENSE_PATTERNS = [
    r'total.?operating.?expense', r'operating.?expense',
]

_ROW_NOI_PATTERNS = [
    r'net.?operating.?income', r'\bnoi\b',
    r'cash.?flow.?before.?debt',
]

_ROW_CAPEX_PATTERNS = [
    r'total.?leasing.*capital', r'total.?capital.?exp',
    r'capital.?expenditure',
]

_ROW_VACANCY_PATTERNS = [
    r'total.?vacancy', r'vacancy.*credit.*loss',
]

# Labels to skip — these look like expenses but are revenue sub-items
_ROW_SKIP_PATTERNS = [
    r'recover', r'reimburs',
]


def _match_row_label(label: str, patterns: list) -> bool:
    """Check if a row label matches patterns."""
    label_lower = str(label).lower().strip()
    return any(re.search(p, label_lower) for p in patterns)


def _should_skip_row(label: str) -> bool:
    """Check if a row should be skipped (false positives)."""
    return _match_row_label(label, _ROW_SKIP_PATTERNS)


def _detect_horizontal_dates(wb_bytes: bytes) -> Optional[dict]:
    """Detect horizontal layout by scanning rows for date-like values.

    Returns {sheet_name, date_row_idx (0-based), label_col_idx, dates: list,
             raw_df: full DataFrame read with no header} or None.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True, read_only=True)
    except Exception:
        return None

    ws = wb[wb.sheetnames[0]]
    sheet_name = ws.title

    # Scan rows 0-15 for a row where most cells parse as dates
    all_rows = []
    for row in ws.iter_rows(max_row=20, values_only=True):
        all_rows.append(list(row))
    wb.close()

    for row_idx, row_vals in enumerate(all_rows):
        if len(row_vals) < 5:
            continue
        # Skip the first cell (usually a label like "For the Months")
        data_cells = row_vals[1:]
        non_empty = [v for v in data_cells if v is not None and str(v).strip()]
        if len(non_empty) < 4:
            continue

        # Try parsing as dates
        parsed_count = 0
        parsed_dates = []
        for v in data_cells:
            if v is None:
                parsed_dates.append(None)
                continue
            try:
                dt = pd.to_datetime(v, format='mixed', dayfirst=False)
                if pd.notna(dt) and 2000 <= dt.year <= 2060:
                    parsed_count += 1
                    parsed_dates.append(dt)
                else:
                    parsed_dates.append(None)
            except Exception:
                # Try "Mon-YYYY" format (e.g. "Oct-2026")
                try:
                    dt = pd.to_datetime(str(v), format='%b-%Y')
                    parsed_count += 1
                    parsed_dates.append(dt)
                except Exception:
                    parsed_dates.append(None)

        # Need at least 60% of non-empty cells to be dates
        if len(non_empty) > 0 and parsed_count / len(non_empty) >= 0.6:
            valid = [d for d in parsed_dates if d is not None]
            log.info("Horizontal dates detected: row=%d, %d dates parsed (%d total cells), "
                     "first=%s, last=%s",
                     row_idx, len(valid), len(parsed_dates),
                     valid[0] if valid else None, valid[-1] if valid else None)
            return {
                'sheet_name': sheet_name,
                'date_row_idx': row_idx,
                'label_col_idx': 0,
                'dates': parsed_dates,
                'header_rows': all_rows[:row_idx],
            }

    return None


def _parse_horizontal_cashflow(
    file_bytes: bytes,
    filename: str,
    horiz_info: dict,
) -> Dict[str, Any]:
    """Parse a horizontal/transposed cash flow (dates across columns, items down rows).

    Scans row labels against revenue/expense/NOI/capex patterns, extracts the
    matching rows, and transposes into the standard vertical output format.
    """
    import openpyxl
    # Use read_only=False to get consistent row widths (read_only can trim trailing cells)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[horiz_info['sheet_name']]

    # Read all data rows (after the date header row)
    date_row_idx = horiz_info['date_row_idx']
    dates = horiz_info['dates']
    num_data_cols = len(dates)

    # Collect all rows with their labels
    labeled_rows = []
    for row_num, row in enumerate(ws.iter_rows(values_only=True)):
        if row_num <= date_row_idx:
            continue
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        if not label:
            continue
        # Extract numeric values from data columns — pad row for safety
        row = list(row) + [None] * max(0, num_data_cols + 1 - len(row))
        values = []
        for col_idx in range(1, num_data_cols + 1):
            v = row[col_idx]
            if v is None:
                values.append(0.0)
            elif isinstance(v, (int, float)):
                values.append(float(v))
            else:
                try:
                    cleaned = str(v).replace(',', '').replace('$', '')
                    cleaned = cleaned.replace('(', '-').replace(')', '')
                    values.append(float(cleaned))
                except (ValueError, TypeError):
                    values.append(0.0)
        labeled_rows.append((label, values))

    wb.close()

    # Match rows to categories — prefer summary rows (Total/Net lines)
    rev_row = None
    exp_row = None
    noi_row = None
    capex_row = None
    vacancy_row = None

    def _has_data(values: list) -> bool:
        """Check if a row has any non-zero numeric values."""
        return any(v != 0.0 for v in values[:20])

    for label, values in labeled_rows:
        if _should_skip_row(label):
            continue
        if not _has_data(values):
            continue  # Skip section headers with no data
        if _match_row_label(label, _ROW_NOI_PATTERNS) and noi_row is None:
            noi_row = (label, values)
        elif _match_row_label(label, _ROW_REVENUE_PATTERNS) and rev_row is None:
            rev_row = (label, values)
        elif _match_row_label(label, _ROW_EXPENSE_PATTERNS) and exp_row is None:
            exp_row = (label, values)
        elif _match_row_label(label, _ROW_CAPEX_PATTERNS) and capex_row is None:
            capex_row = (label, values)
        elif _match_row_label(label, _ROW_VACANCY_PATTERNS) and vacancy_row is None:
            vacancy_row = (label, values)

    columns_detected = {
        'revenue': rev_row[0] if rev_row else None,
        'expenses': exp_row[0] if exp_row else None,
        'noi': noi_row[0] if noi_row else None,
        'capex': capex_row[0] if capex_row else None,
        'vacancy': vacancy_row[0] if vacancy_row else None,
    }

    if not rev_row and not noi_row:
        # Return available row labels so the UI can offer mapping
        available = [label for label, _ in labeled_rows]
        return {
            'error': 'Could not detect revenue or NOI rows in horizontal layout',
            'columns_detected': columns_detected,
            'available_columns': available,
            'layout': 'horizontal',
        }

    # Build monthly cashflows
    cashflows = []
    for i, dt in enumerate(dates):
        if dt is None:
            continue
        period_date = month_end(date(dt.year, dt.month, 1))

        rev = abs(rev_row[1][i]) if rev_row and i < len(rev_row[1]) else 0.0
        exp = abs(exp_row[1][i]) if exp_row and i < len(exp_row[1]) else 0.0
        noi_val = noi_row[1][i] if noi_row and i < len(noi_row[1]) else 0.0
        capex_val = abs(capex_row[1][i]) if capex_row and i < len(capex_row[1]) else 0.0

        # Derive missing
        if rev and exp and not noi_row:
            noi_val = rev - exp
        elif noi_val and not rev_row and not exp_row:
            rev = abs(noi_val) / 0.60
            exp = rev - abs(noi_val)

        cashflows.append({
            'period_date': str(period_date),
            'revenue': round(rev, 2),
            'expenses': round(exp, 2),
            'capex': round(capex_val, 2),
            'noi': round(noi_val, 2),
        })

    if not cashflows:
        return {'error': 'No valid date periods found in horizontal layout'}

    return {
        'periods': len(cashflows),
        'frequency': 'monthly',
        'cashflows': cashflows,
        'columns_detected': columns_detected,
        'layout': 'horizontal',
        'metadata': {
            'filename': filename,
            'rows_parsed': len(cashflows),
            'monthly_rows': len(cashflows),
        },
    }


def _is_annual(dates: pd.Series) -> bool:
    """Detect if dates represent annual periods (gaps > 300 days between consecutive)."""
    sorted_dates = dates.dropna().sort_values()
    if len(sorted_dates) < 2:
        return True
    diffs = sorted_dates.diff().dropna()
    median_gap = diffs.dt.days.median()
    return median_gap > 300


def _annual_to_monthly(rows: List[Dict], year: int) -> List[Dict]:
    """Spread an annual row across 12 monthly rows."""
    monthly = []
    rev = float(rows[0].get('revenue') or 0)
    exp = float(rows[0].get('expenses') or 0)
    capex = float(rows[0].get('capex') or 0)
    noi = float(rows[0].get('noi') or 0)

    for m in range(1, 13):
        monthly.append({
            'period_date': str(month_end(date(year, m, 1))),
            'revenue': round(rev / 12, 2),
            'expenses': round(exp / 12, 2),
            'capex': round(capex / 12, 2),
            'noi': round(noi / 12, 2),
        })
    return monthly


def _suggest_coa(label: str, coa_rows: Optional[List[Dict]] = None) -> Tuple[Optional[int], Optional[str]]:
    """Suggest a COA account for a line item label.

    Uses ARGUS_COA_MAP keyword matching first, then falls back to fuzzy matching
    against COA table descriptions if provided.
    """
    from argus_parser import map_to_coa
    account, category = map_to_coa(label)
    if account is not None:
        return account, category

    if not coa_rows:
        return None, None

    # Fuzzy match against COA vdescription and vMisc
    label_lower = label.strip().lower()
    label_words = set(re.split(r'\W+', label_lower)) - {'', 'total', 'net', 'gross'}
    best_score = 0
    best_account = None
    best_category = None

    for row in coa_rows:
        vcode = row.get('vcode')
        desc = str(row.get('vdescription') or '').lower()
        misc = str(row.get('vmisc') or row.get('vMisc') or '').lower()
        acct_type = str(row.get('vaccounttype') or row.get('vAccountType') or '')

        desc_words = set(re.split(r'\W+', desc)) - {'', 'total', 'net', 'gross'}
        misc_words = set(re.split(r'\W+', misc)) - {'', 'total', 'net', 'gross'}

        overlap = len(label_words & (desc_words | misc_words))
        if overlap > best_score and overlap >= 2:
            best_score = overlap
            try:
                best_account = int(vcode)
            except (ValueError, TypeError):
                continue
            if acct_type.lower().startswith('revenue'):
                best_category = 'revenue'
            elif acct_type.lower().startswith('expense'):
                best_category = 'expense'
            else:
                best_category = 'other'

    return best_account, best_category


# Summary-line patterns for identifying total/subtotal rows
_SUMMARY_PATTERNS = [
    r'^total\b', r'^net\b', r'^subtotal', r'^effective\s+gross',
    r'^gross\s+potential', r'^gross\s+revenue', r'^gross\s+income',
    r'net\s+operating\s+income', r'\bnoi\b',
]


def _is_summary_row(label: str) -> bool:
    """Check if a row label looks like a summary/total row."""
    label_lower = label.strip().lower()
    return any(re.search(p, label_lower) for p in _SUMMARY_PATTERNS)


def parse_cashflow_line_items(
    file_bytes: bytes,
    filename: str,
    coa_rows: Optional[List[Dict]] = None,
) -> Dict[str, Any]:
    """Parse an Excel/CSV cash flow model and extract ALL individual line items.

    Returns each line item with its original label, suggested COA account mapping,
    and monthly values. Used for the two-step import flow where analysts review
    and edit COA assignments before importing.

    Returns:
        {
            'line_items': [
                {
                    'label': str,           # Original row label from Excel
                    'suggested_coa': int,   # Suggested vAccount number (or null)
                    'suggested_category': str, # 'revenue'|'expense'|'capex'|'other'|null
                    'is_summary': bool,     # True if this looks like a total/subtotal row
                    'values': [{'period_date': str, 'amount': float}, ...]
                },
                ...
            ],
            'periods': [str, ...],   # List of period dates
            'frequency': 'monthly' | 'annual',
            'layout': 'horizontal' | 'vertical',
            'metadata': {...},
        }
    """
    fname_lower = filename.lower()

    # Try horizontal layout first (most common for partner models)
    if not fname_lower.endswith('.csv'):
        horiz = _detect_horizontal_dates(file_bytes)
        if horiz is not None:
            return _parse_horizontal_line_items(file_bytes, filename, horiz, coa_rows)

    # Fall back to vertical layout — try to extract individual columns
    return _parse_vertical_line_items(file_bytes, filename, coa_rows)


def _parse_horizontal_line_items(
    file_bytes: bytes,
    filename: str,
    horiz_info: dict,
    coa_rows: Optional[List[Dict]] = None,
) -> Dict[str, Any]:
    """Extract all line items from a horizontal/transposed Excel layout."""
    import openpyxl
    # Use read_only=False to get consistent row widths (read_only can trim trailing cells)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[horiz_info['sheet_name']]

    date_row_idx = horiz_info['date_row_idx']
    dates = horiz_info['dates']
    num_data_cols = len(dates)

    # Build period date list
    periods = []
    for dt in dates:
        if dt is not None:
            periods.append(str(month_end(date(dt.year, dt.month, 1))))
        else:
            periods.append(None)

    # Collect all rows with their labels and values
    line_items = []
    for row_num, row in enumerate(ws.iter_rows(values_only=True)):
        if row_num <= date_row_idx:
            continue
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        if not label:
            continue

        # Extract numeric values — pad row to ensure we cover all date columns
        row = list(row) + [None] * max(0, num_data_cols + 1 - len(row))
        values = []
        has_data = False
        for col_idx in range(1, num_data_cols + 1):
            v = row[col_idx]
            if v is None:
                values.append(0.0)
            elif isinstance(v, (int, float)):
                values.append(float(v))
                if float(v) != 0.0:
                    has_data = True
            else:
                try:
                    cleaned = str(v).replace(',', '').replace('$', '')
                    cleaned = cleaned.replace('(', '-').replace(')', '')
                    val = float(cleaned)
                    values.append(val)
                    if val != 0.0:
                        has_data = True
                except (ValueError, TypeError):
                    values.append(0.0)

        if not has_data:
            continue  # Skip section headers with no data

        # Suggest COA mapping
        suggested_coa, suggested_category = _suggest_coa(label, coa_rows)
        is_summary = _is_summary_row(label)

        # Build per-period values
        period_values = []
        for i, period_date in enumerate(periods):
            if period_date is None:
                continue
            amount = values[i] if i < len(values) else 0.0
            period_values.append({
                'period_date': period_date,
                'amount': round(amount, 2),
            })

        line_items.append({
            'label': label,
            'suggested_coa': suggested_coa,
            'suggested_category': suggested_category,
            'is_summary': is_summary,
            'values': period_values,
        })

    wb.close()

    valid_periods = [p for p in periods if p is not None]
    log.info("Parsed %d line items with %d valid periods: %s .. %s",
             len(line_items), len(valid_periods),
             valid_periods[0] if valid_periods else None,
             valid_periods[-1] if valid_periods else None)
    if line_items:
        first = line_items[0]
        log.info("  First item '%s': %d values, sum=%.2f",
                 first['label'], len(first['values']),
                 sum(v['amount'] for v in first['values']))

    is_annual = len(valid_periods) >= 2 and (
        pd.to_datetime(valid_periods[-1]) - pd.to_datetime(valid_periods[0])
    ).days / max(len(valid_periods) - 1, 1) > 300

    return {
        'line_items': line_items,
        'periods': valid_periods,
        'frequency': 'annual' if is_annual else 'monthly',
        'layout': 'horizontal',
        'metadata': {
            'filename': filename,
            'line_items_found': len(line_items),
            'periods_found': len(valid_periods),
        },
    }


def _parse_vertical_line_items(
    file_bytes: bytes,
    filename: str,
    coa_rows: Optional[List[Dict]] = None,
) -> Dict[str, Any]:
    """Extract line items from a vertical layout (columns = categories).

    In vertical layouts, each numeric column is treated as a line item.
    """
    fname_lower = filename.lower()
    try:
        if fname_lower.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_bytes))
        else:
            xls = pd.ExcelFile(io.BytesIO(file_bytes))
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=xls.sheet_names[0])
    except Exception as e:
        return {'error': f'Failed to read file: {e}'}

    if df.empty:
        return {'error': 'File is empty'}

    df = df.dropna(how='all').dropna(axis=1, how='all')

    date_col = _detect_date_column(df)
    if not date_col:
        return {'error': 'Could not detect a date/period column'}

    try:
        dates = pd.to_datetime(df[date_col], format='mixed', dayfirst=False)
    except Exception:
        try:
            years = pd.to_numeric(df[date_col], errors='coerce')
            dates = years.apply(
                lambda y: pd.Timestamp(year=int(y), month=12, day=31)
                if pd.notna(y) and 1990 <= y <= 2060 else pd.NaT
            )
        except Exception:
            return {'error': f'Could not parse dates from column "{date_col}"'}

    valid_mask = dates.notna()
    df = df[valid_mask].copy()
    dates = dates[valid_mask]

    if df.empty:
        return {'error': 'No valid date rows found'}

    # Each numeric column becomes a line item
    line_items = []
    periods = [str(month_end(date(dt.year, dt.month, 1))) for dt in dates]

    for col in df.columns:
        if col == date_col:
            continue
        # Try parsing as numeric
        s = df[col].astype(str).str.replace(',', '').str.replace('$', '')
        s = s.str.replace('(', '-', regex=False).str.replace(')', '', regex=False)
        numeric = pd.to_numeric(s, errors='coerce')
        if numeric.notna().sum() < len(numeric) * 0.5:
            continue  # Not a numeric column

        has_data = (numeric.fillna(0).abs() > 0).any()
        if not has_data:
            continue

        suggested_coa, suggested_category = _suggest_coa(str(col), coa_rows)
        is_summary = _is_summary_row(str(col))

        period_values = []
        for i, period_date in enumerate(periods):
            amount = float(numeric.iloc[i]) if pd.notna(numeric.iloc[i]) else 0.0
            period_values.append({
                'period_date': period_date,
                'amount': round(amount, 2),
            })

        line_items.append({
            'label': str(col),
            'suggested_coa': suggested_coa,
            'suggested_category': suggested_category,
            'is_summary': is_summary,
            'values': period_values,
        })

    is_annual = _is_annual(dates)

    return {
        'line_items': line_items,
        'periods': periods,
        'frequency': 'annual' if is_annual else 'monthly',
        'layout': 'vertical',
        'metadata': {
            'filename': filename,
            'line_items_found': len(line_items),
            'periods_found': len(periods),
        },
    }


def parse_cashflow_excel(
    file_bytes: bytes,
    filename: str,
) -> Dict[str, Any]:
    """Parse a partner's Excel/CSV cash flow model.

    Auto-detects column mapping and date format. Handles both monthly
    and annual data (annual is spread to monthly).

    Returns:
        {
            'periods': int,
            'frequency': 'monthly' | 'annual',
            'cashflows': [{'period_date', 'revenue', 'expenses', 'capex', 'noi'}, ...],
            'columns_detected': {'revenue': col_name, ...},
            'metadata': {'filename': ..., 'sheet': ..., 'rows_parsed': ...},
        }
    """
    # Read file
    fname_lower = filename.lower()
    try:
        if fname_lower.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_bytes))
        else:
            # Try first sheet, skip common header rows
            xls = pd.ExcelFile(io.BytesIO(file_bytes))
            sheet_name = xls.sheet_names[0]

            # Try reading with different header offsets
            df = None
            for skip in range(6):
                candidate = pd.read_excel(
                    io.BytesIO(file_bytes), sheet_name=sheet_name,
                    header=skip, dtype=str,
                )
                # Check if we found numeric data columns
                num_cols = 0
                for col in candidate.columns:
                    sample = candidate[col].dropna().head(5)
                    try:
                        pd.to_numeric(sample.str.replace(',', '').str.replace('$', '').str.replace('(', '-').str.replace(')', ''))
                        num_cols += 1
                    except Exception:
                        pass
                if num_cols >= 2:
                    df = candidate
                    break

            if df is None:
                df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    except Exception as e:
        return {'error': f'Failed to read file: {e}'}

    if df.empty:
        return {'error': 'File is empty'}

    # Drop fully empty rows and columns
    df = df.dropna(how='all').dropna(axis=1, how='all')

    # Detect columns
    date_col = _detect_date_column(df)
    rev_col = _find_column(df.columns, _REVENUE_PATTERNS)
    exp_col = _find_column(df.columns, _EXPENSE_PATTERNS)
    noi_col = _find_column(df.columns, _NOI_PATTERNS)
    capex_col = _find_column(df.columns, _CAPEX_PATTERNS)

    columns_detected = {
        'date': date_col,
        'revenue': rev_col,
        'expenses': exp_col,
        'noi': noi_col,
        'capex': capex_col,
    }

    if not date_col:
        # Try horizontal/transposed layout (dates across columns, items down rows)
        if not fname_lower.endswith('.csv'):
            horiz = _detect_horizontal_dates(file_bytes)
            if horiz is not None:
                return _parse_horizontal_cashflow(file_bytes, filename, horiz)

        return {
            'error': 'Could not detect a date/period column',
            'columns_detected': columns_detected,
            'available_columns': list(df.columns),
        }

    # Need at least revenue+expenses OR noi
    if not rev_col and not noi_col:
        return {
            'error': 'Could not detect revenue or NOI columns',
            'columns_detected': columns_detected,
            'available_columns': list(df.columns),
        }

    # Parse numeric values
    def to_num(series):
        if series is None:
            return pd.Series([0.0] * len(df))
        s = series.astype(str).str.replace(',', '').str.replace('$', '')
        s = s.str.replace('(', '-', regex=False).str.replace(')', '', regex=False)
        return pd.to_numeric(s, errors='coerce').fillna(0)

    # Parse dates
    try:
        dates = pd.to_datetime(df[date_col], format='mixed', dayfirst=False)
    except Exception:
        try:
            # Try interpreting as years
            years = pd.to_numeric(df[date_col], errors='coerce')
            dates = years.apply(lambda y: pd.Timestamp(year=int(y), month=12, day=31) if pd.notna(y) and 1990 <= y <= 2060 else pd.NaT)
        except Exception:
            return {'error': f'Could not parse dates from column "{date_col}"'}

    valid_mask = dates.notna()
    df = df[valid_mask].copy()
    dates = dates[valid_mask]

    if df.empty:
        return {'error': 'No valid date rows found'}

    rev_data = to_num(df[rev_col]) if rev_col else None
    exp_data = to_num(df[exp_col]) if exp_col else None
    noi_data = to_num(df[noi_col]) if noi_col else None
    capex_data = to_num(df[capex_col]) if capex_col else None

    # Derive missing columns
    if rev_data is not None and exp_data is not None and noi_data is None:
        noi_data = rev_data - exp_data.abs()
    elif noi_data is not None and rev_data is None and exp_data is None:
        # Only NOI provided — estimate revenue as NOI/0.60, expenses as remainder
        rev_data = noi_data.abs() / 0.60
        exp_data = rev_data - noi_data.abs()

    # Normalize signs: revenue positive, expenses positive (stored as positive in prospect_cashflows)
    if rev_data is not None:
        rev_data = rev_data.abs()
    if exp_data is not None:
        exp_data = exp_data.abs()
    if capex_data is not None:
        capex_data = capex_data.abs()
    if noi_data is not None:
        # Keep NOI sign as-is (should be positive for profitable properties)
        pass

    # Detect frequency
    annual = _is_annual(dates)
    frequency = 'annual' if annual else 'monthly'

    # Build rows
    raw_rows = []
    for i in range(len(df)):
        dt = dates.iloc[i]
        raw_rows.append({
            'date': dt,
            'year': dt.year,
            'revenue': float(rev_data.iloc[i]) if rev_data is not None else 0,
            'expenses': float(exp_data.iloc[i]) if exp_data is not None else 0,
            'capex': float(capex_data.iloc[i]) if capex_data is not None else 0,
            'noi': float(noi_data.iloc[i]) if noi_data is not None else 0,
        })

    # Convert to monthly if annual
    cashflows = []
    if annual:
        for row in raw_rows:
            monthly = _annual_to_monthly([row], row['year'])
            cashflows.extend(monthly)
    else:
        for row in raw_rows:
            dt = row['date']
            period_date = month_end(date(dt.year, dt.month, 1))
            cashflows.append({
                'period_date': str(period_date),
                'revenue': round(row['revenue'], 2),
                'expenses': round(row['expenses'], 2),
                'capex': round(row['capex'], 2),
                'noi': round(row['noi'], 2),
            })

    return {
        'periods': len(raw_rows),
        'frequency': frequency,
        'cashflows': cashflows,
        'columns_detected': columns_detected,
        'metadata': {
            'filename': filename,
            'rows_parsed': len(raw_rows),
            'monthly_rows': len(cashflows),
        },
    }
