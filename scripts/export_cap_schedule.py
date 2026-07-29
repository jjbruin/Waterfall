"""Export capitalization schedule for all active deals to Excel.

Usage (from project root):
    python -m scripts.export_cap_schedule
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, numbers, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask_app import create_app
from flask_app.db import get_engine
from flask_app.services.data_service import load_all, get_inv_display
from compute import prepare_cap_lookups, get_deal_capitalization
from consolidation import get_property_vcodes_for_deal


def main():
    app = create_app()
    with app.app_context():
        _run()


def _run():
    engine = get_engine()
    data = load_all(str(engine.url))

    inv = data["inv"]
    acct = data["acct"]
    wf = data["wf"]
    mri_val = data.get("mri_val")
    mri_loans = data.get("mri_loans_raw")
    isbs_raw = data.get("isbs_raw")
    rels = data.get("rels")

    inv_disp = get_inv_display(inv)
    lookups = prepare_cap_lookups(acct, inv, mri_val, mri_loans)
    prop_map = lookups["prop_map"]

    rows = []
    total = len(inv_disp)
    for i, (_, row) in enumerate(inv_disp.iterrows()):
        vcode = str(row["vcode"])
        name = row.get("Investment_Name", vcode)
        investment_id = str(row.get("InvestmentID", ""))
        print(f"  [{i+1}/{total}] {name}")
        try:
            prop_vcodes = prop_map.get(vcode, []) or None
            cap = get_deal_capitalization(
                acct, inv, wf, mri_val, mri_loans,
                deal_vcode=vcode,
                property_vcodes=prop_vcodes,
                lookups=lookups,
                isbs_raw=isbs_raw,
            )
            rows.append({
                "vCode": vcode,
                "Investment Name": name,
                "InvestmentID": investment_id,
                "Debt": cap.get("debt", 0),
                "Pref Equity": cap.get("pref_equity", 0),
                "Ptr Equity": cap.get("partner_equity", 0),
                "Total Cap": cap.get("total_cap", 0),
            })
        except Exception as e:
            print(f"    ERROR: {e}")
            rows.append({
                "vCode": vcode,
                "Investment Name": name,
                "InvestmentID": investment_id,
                "Debt": 0, "Pref Equity": 0, "Ptr Equity": 0, "Total Cap": 0,
            })

    df = pd.DataFrame(rows).sort_values("Investment Name", key=lambda s: s.str.lower())

    # Build Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Capitalization Schedule"

    headers = ["vCode", "Investment Name", "InvestmentID", "Debt", "Pref Equity", "Ptr Equity", "Total Cap"]
    header_font = Font(bold=True)
    thin_border = Border(bottom=Side(style="thin"))
    currency_fmt = '#,##0'
    col_widths = {1: 14, 2: 40, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18}

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.border = thin_border
        if col_idx >= 4:
            cell.alignment = Alignment(horizontal="right")

    for row_idx, (_, data_row) in enumerate(df.iterrows(), 2):
        for col_idx, h in enumerate(headers, 1):
            val = data_row[h]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx >= 4:
                cell.number_format = currency_fmt
                cell.alignment = Alignment(horizontal="right")

    # Total row
    total_row = len(df) + 2
    ws.cell(row=total_row, column=2, value="Total").font = Font(bold=True)
    top_border = Border(top=Side(style="thin"))
    for col_idx in range(4, 8):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=total_row, column=col_idx)
        cell.value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        cell.number_format = currency_fmt
        cell.font = Font(bold=True)
        cell.border = top_border
        cell.alignment = Alignment(horizontal="right")

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    out_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            "capitalization_schedule.xlsx")
    wb.save(out_path)
    print(f"\nExported {len(df)} deals to {out_path}")


if __name__ == "__main__":
    main()
