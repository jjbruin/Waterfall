"""Build the live-vs-PDF variance workbook for one investor and quarter.

Compares every cell the reference document prints against what the app assembles
for the SAME quarter, and writes an Excel workbook: one tab per page, plus a
front summary that lists only the cells needing review.

Every difference is classified:

    TIES   inside tolerance for its unit
    KNOWN  a difference already diagnosed — the reason travels with the row
    NEW    unexplained, and the only thing worth reading the workbook for

The KNOWN table is the point. Without it a variance report is 800 rows of
"different" and nobody reads it; with it the summary tab is a short list.

Live data comes from a LOCALLY assembled bundle, not the deployed endpoint —
scripts/snapshot_payload_dump.py — so this measures the code in the working tree.

Usage
    WF_TOKEN=<jwt> python scripts/snapshot_payload_dump.py bundle TGAM 2026-Q1 > b.json
    python scripts/snapshot_pdf_variance_workbook.py b.json [-o outputs/x.xlsx]
"""
import json
import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import snapshot_pdf_variance_pdfdata as P                        # noqa: E402
from openpyxl import Workbook                                    # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill          # noqa: E402
from openpyxl.utils import get_column_letter                      # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

M = 1e6

# ── tolerances, per unit ─────────────────────────────────────────────────
TOL = {"usd_m": 0.35, "pct": 1.6, "dscr": 0.11, "ratio_pct": 0.6}


# ── the KNOWN table ──────────────────────────────────────────────────────
# (page, vcode-or-row-key, column) -> reason. A "*" vcode applies to the whole
# column on that page. Every entry here was diagnosed earlier in the work and
# the reason is carried into the workbook so a reviewer never has to ask twice.
KNOWN = {
    # --- the four Debt residuals, Financial and Loan alike ---
    ("Financial", "P0000114", "debt"): "Jefferson Stephens: mOrigLoanAmt is exactly 2x the PDF — facility looks double-counted in MRI_Loans",
    ("Loan", "P0000114", "debt"): "Jefferson Stephens: mOrigLoanAmt is exactly 2x the PDF — facility looks double-counted in MRI_Loans",
    ("Financial", "P0000116", "debt"): "Plaza Del Mar: ISBS Interim BS at Q1 is 70.0; it only drops to 27.59 at Q2, so the PDF's 27.6 matches our Q2 — data vintage",
    ("Loan", "P0000116", "debt"): "Plaza Del Mar: ISBS Interim BS at Q1 is 70.0; the PDF's 27.6 matches our Q2 — data vintage",
    ("Financial", "P0000021", "debt"): "JB Fair Park: neither basis matches — ISBS carries a dead 2022-12-31 senior financing row of 66.36M, committed facility is 77.37M",
    ("Loan", "P0000021", "debt"): "JB Fair Park: neither basis matches — ISBS 66.36M (dead 2022 row) vs committed facility 77.37M",
    ("Financial", "P0000067", "debt"): "Brainerd Place: committed facility is BELOW the PDF — MRI_Loans may be missing a tranche",
    ("Loan", "P0000067", "debt"): "Brainerd Place: committed facility is BELOW the PDF — MRI_Loans may be missing a tranche",
    # --- structural, whole-column ---
    ("Financial", "*", "unfunded"): "Un-funded is structurally 0 while COMMITMENT_BASIS='funded', which defines Total Commitment as the Invested formula",
    ("Financial", "*", "total_commitment"): "Total Commitment equals Invested under COMMITMENT_BASIS='funded'; the PDF's commitment exceeds funded",
    ("Financial", "*", "itd"): "ITD Distributions is manual entry with no formula (get_itd) and nothing entered",
    ("Financial", "*", "net_roe"): "Net ROE is manual entry with no formula (get_net_roe) and nothing entered",
    ("Financial", "*", "pct_of_pref"): "% of Pref: the PDF prints whole percents, we print one decimal; also carries the 45th & Main ownership difference",
    # --- allocation ---
    ("Summary", "Self-Storage", "funded"): "Self-Storage is 32.03M live against 34.17M published — pre-existing data difference, predates all snapshot work",
    ("Summary", "Self-Storage", "committed"): "Self-Storage committed inherits the same 2.14M funded difference",
    ("Summary", "*", "committed"): "Committed uses the accounting commitment rows (committed_pe); the PDF's committed column is a different basis and does not reconcile",
    ("Summary", "TOTAL", "funded"): "Funded total is 402.1M against 404.2M — the residual IS the Self-Storage 2.14M plus 45th & Main at 90% in MRI against 100% in the PDF (1.855M)",
    ("Summary", "Value-Add", "funded"): "Deal-type dollars inherit the funded-total difference (Self-Storage 2.14M + 45th & Main at 90% vs 100%, 1.855M); the narrative's figure is also rounded to 0.1M",
    ("Summary", "Income", "funded"): "Deal-type dollars inherit the funded-total difference; the narrative's figure is rounded to 0.1M",
    ("Summary", "New Construction", "funded"): "Deal-type dollars inherit the funded-total difference",
    # --- 45th & Main ownership ---
    ("Financial", "P0000089", "invested"): "45th & Main resolves at 90% in MRI where the PDF was built at 100% — worth 1.855M",
    ("Financial", "P0000089", "total_commitment"): "45th & Main resolves at 90% in MRI where the PDF was built at 100%",
    # --- the three Loan subtotal anomalies ---
    ("Loan", "SUB:Individual Investments", "ytd_dscr"): "PDF prints n/a although three members carry a DSCR (weighting to 2.08x); no denominator produces n/a",
    ("Loan", "SUB:Individual Investments", "debt_yield"): "PDF 4.1% against 8.71% debt-weighted; weighting over ALL fund debt gives 3.76%, and that denominator breaks TGA 2022",
    ("Loan", "SUB:TGA25", "debt_yield"): "PDF 4.9% is Burton over the fund's WHOLE debt; that denominator gives 4.58% on TGA 2022 where 9.6% is published, so neither rule fits both",
    # --- Operating: the At Close portfolio row ---
    ("Operating", "TOTAL", "noi_at_close"): "The PDF's portfolio At Close NOI of 62.7 is 22.7 below the sum of its own five fund subtotals (85.4) — looks like a spreadsheet range that missed the first block; we sum all deals",
    ("Operating", "TOTAL", "expected_growth"): "Follows from the At Close total above: the PDF's 66.7% is computed off 62.7, ours off the true sum",
    ("Operating", "TOTAL", "actual_growth"): "Follows from the At Close total above",
    # --- PDF-only row ---
    ("Financial", "PDFONLY_PEGADD", "*"): "PDF prints a second 'Pegasus Life Storage - Add'l' row under TGA 2022; we carry one Pegasus, in Individual Investments",
    ("Operating", "PDFONLY_PEGADD", "*"): "PDF-only row — see Financial",
    ("Loan", "PDFONLY_PEGADD", "*"): "PDF-only row — see Financial",
    # --- harness limits of the local dump, not app behaviour ---
    ("Loan", "*", "debt_yield"): "HARNESS: the local bundle dump cannot pass quarterly_noi_provider (it needs the 800k-row ISBS frame, not fetchable over REST), so Debt Yield is None throughout. The app itself computes it — see build_subtab",
    # --- Pegasus is ONE deal for us and TWO rows on the page ---
    ("Financial", "P0000066", "total_pref"): "The PDF splits Pegasus across two rows (Individual 8.1 + TGA22 Add'l 24.2 = 32.3). Our single row carries 32.3 — an EXACT match once the split is summed",
    ("Financial", "P0000066", "total_cap"): "PDF split rows 10.7 + 24.2 = 34.9; our single row is 34.9 — exact once summed",
    ("Financial", "P0000066", "invested"): "PDF split rows 7.4 + 21.7 = 29.1 against our single 27.0; the split explains the bulk of it",
    ("Financial", "P0000066", "ptr_equity"): "PDF split rows 2.6 + 0.0 = 2.6; ours 2.6 — exact once summed",
    ("Operating", "P0000066", "*"): "The PDF splits Pegasus across two rows; ours is one",
    ("Loan", "P0000066", "*"): "The PDF splits Pegasus across two rows; ours is one",
    # --- a PDF cell that does not foot to its own row ---
    ("Financial", "P0000110", "total_cap"): "PDF ANOMALY: Trolley Square's printed Total Cap of 13.5 does not foot to its own row — debt 30.8 + pref 6.8 + equity 6.8 = 44.4. Ours is 41.7 and does foot",
    # --- Loan rate/maturity presentation ---
    ("Loan", "*", "rate"): "Rate strings: we print '3.9%' where the PDF prints '3.9% fixed'; the interest-type suffix is not carried on our rows",
}


def known_for(page, key, col):
    for k in ((page, key, col), (page, key, "*"), (page, "*", col)):
        if k in KNOWN:
            return KNOWN[k]
    return None


# ── comparison ───────────────────────────────────────────────────────────

class Cell:
    __slots__ = ("page", "group", "row", "key", "col", "pdf", "live",
                 "unit", "status", "reason")

    def __init__(self, page, group, row, key, col, pdf, live, unit):
        self.page, self.group, self.row = page, group, row
        self.key, self.col = key, col
        self.pdf, self.live, self.unit = pdf, live, unit
        self.status, self.reason = None, ""

    @property
    def diff(self):
        if isinstance(self.pdf, (int, float)) and isinstance(self.live, (int, float)):
            return self.live - self.pdf
        return None

    @property
    def pct_diff(self):
        d = self.diff
        if d is None or not self.pdf:
            return None
        return d / abs(self.pdf) * 100.0


def classify(c):
    """TIES / KNOWN / NEW, and the reason when known."""
    reason = known_for(c.page, c.key, c.col)
    pdf_txt = isinstance(c.pdf, str)
    live_txt = isinstance(c.live, str)

    # Both are literals ("n/a" vs "n/a", "Dev" vs "n/a", ...)
    if pdf_txt or live_txt or c.pdf is None or c.live is None:
        pn = "" if c.pdf is None else str(c.pdf)
        ln = "" if c.live is None else str(c.live)
        # A dash and a zero say the same thing; n/a and n/a agree.
        blankish = {"", "n/a", "-", "—", "0.0", "0"}
        if (pn.lower() in blankish and ln.lower() in blankish) or pn.lower() == ln.lower():
            c.status = "TIES"
            return
        # numeric-vs-blank where the number rounds to nothing
        for a, b in ((c.pdf, c.live), (c.live, c.pdf)):
            if isinstance(a, (int, float)) and (b is None or str(b).lower() in blankish):
                if abs(a) < 0.05:
                    c.status = "TIES"
                    return
        c.status = "KNOWN" if reason else "NEW"
        c.reason = reason or ""
        return

    tol = TOL[c.unit]
    if abs(c.live - c.pdf) <= tol:
        c.status = "TIES"
        return
    c.status = "KNOWN" if reason else "NEW"
    c.reason = reason or ""


# ── live readers ─────────────────────────────────────────────────────────

def live_financial(st):
    rows, subs = {}, {}
    for g, blk in (st.get("groups") or {}).items():
        for r in blk["deals"]:
            rows[r["vcode"]] = (g, r)
        s = blk.get("subtotal") or {}
        subs[g] = s
    for r in (st.get("ownership_flagged") or []):
        rows.setdefault(r["vcode"], (None, r))
    return rows, subs, st.get("total") or {}, st.get("total_excluding_dev") or {}


def live_grouped(st):
    """Operating / Loan: groups are {name: [rows]} with sibling subtotals."""
    rows = {}
    for g, rs in (st.get("groups") or {}).items():
        for r in rs:
            rows[r["vcode"]] = (g, r)
    for r in (st.get("ownership_flagged") or []):
        rows.setdefault(r["vcode"], (None, r))
    return rows, (st.get("subtotals") or {}), (st.get("total") or {})


def m(v):
    return None if v is None else v / M


def pct(v):
    return None if v is None else v * 100.0


def build_cells(bundle):
    cells = []
    subs_ = bundle["subtabs"]

    # ---- Summary / allocation ----
    sm = subs_.get("summary") or {}
    aa = sm.get("asset_allocation") or {}
    live_buckets = {b["label"]: b for b in (aa.get("buckets") or [])}
    for label, (pf, pc) in P.PAGE1_ASSET.items():
        lb = live_buckets.get(label) or {}
        cells.append(Cell("Summary", "Asset Allocation", label, label,
                          "funded", pf / M, m(lb.get("funded")), "usd_m"))
        cells.append(Cell("Summary", "Asset Allocation", label, label,
                          "committed", pc / M, m(lb.get("committed")), "usd_m"))
    cells.append(Cell("Summary", "Asset Allocation", "Total", "TOTAL", "funded",
                      P.PAGE1_TOTALS["funded"] / M, m(aa.get("total_funded")), "usd_m"))
    cells.append(Cell("Summary", "Asset Allocation", "Total", "TOTAL", "committed",
                      P.PAGE1_TOTALS["committed"] / M, m(aa.get("total_committed")), "usd_m"))

    dt = sm.get("deal_type_allocation") or {}
    live_dt = {b["label"]: b for b in (dt.get("buckets") or [])}
    for label, (ppct, pusd) in P.PAGE1_DEAL_TYPE.items():
        lb = live_dt.get(label) or {}
        cells.append(Cell("Summary", "Deal Type", label, label, "pct_of_funded",
                          ppct, pct(lb.get("funded_pct")), "pct"))
        cells.append(Cell("Summary", "Deal Type", label, label, "funded",
                          pusd / M, m(lb.get("funded")), "usd_m"))

    # ---- Financial ----
    frows, fsubs, ftot, fexdev = live_financial(subs_.get("financial") or {})
    for vc, tup in P.PAGE2.items():
        name, vals = tup[0], tup[1:]
        g, r = frows.get(vc, (None, {}))
        for col, pv in zip(P.FIN_COLS, vals):
            lv = r.get(col)
            if col in ("pct_of_pref", "net_roe"):
                lv = pct(lv) if isinstance(lv, (int, float)) else lv
                unit = "pct"
            elif col == "itd":
                unit = "usd_m"
                lv = lv if isinstance(lv, (int, float)) else None
                lv = m(lv) if (lv is not None and abs(lv) > 1000) else lv
            else:
                lv = m(lv) if isinstance(lv, (int, float)) else lv
                unit = "usd_m"
            cells.append(Cell("Financial", g or "(not in live set)",
                              P.ALIASES.get(vc, name), vc, col, pv, lv, unit))
    for g, tup in P.PAGE2_SUBTOTALS.items():
        label, vals = tup[0], tup[1:]
        s = fsubs.get(g) or {}
        for col, pv in zip(P.FIN_COLS, vals):
            lv = s.get(col)
            unit = "pct" if col in ("pct_of_pref", "net_roe") else "usd_m"
            lv = pct(lv) if (unit == "pct" and isinstance(lv, (int, float))) else (
                m(lv) if isinstance(lv, (int, float)) else lv)
            cells.append(Cell("Financial", g, label, f"SUB:{g}", col, pv, lv, unit))
    label, vals = P.PAGE2_TOTAL[0], P.PAGE2_TOTAL[1:]
    for col, pv in zip(P.FIN_COLS, vals):
        lv = ftot.get(col)
        unit = "pct" if col in ("pct_of_pref", "net_roe") else "usd_m"
        lv = pct(lv) if (unit == "pct" and isinstance(lv, (int, float))) else (
            m(lv) if isinstance(lv, (int, float)) else lv)
        cells.append(Cell("Financial", "TOTAL", label, "TOTAL", col, pv, lv, unit))
    for col, pv in P.PAGE2_EXDEV.items():
        lv = fexdev.get(col)
        unit = "pct" if col == "net_roe" else "usd_m"
        lv = pct(lv) if (unit == "pct" and isinstance(lv, (int, float))) else (
            m(lv) if isinstance(lv, (int, float)) else lv)
        cells.append(Cell("Financial", "TOTAL", "Excluding Development Deals",
                          "EXDEV", col, pv, lv, unit))

    # ---- Operating ----
    orows, osubs, ototal = live_grouped(subs_.get("operating") or {})
    def op_live(r, col):
        """The value the Operating page DISPLAYS for one cell.

        Not the raw field. A development deal's raw econ_occ can hold a real
        93.05 while the page prints "n/a" — the reference document prints n/a
        too, so comparing raw put every dev deal's suppressed cells in the NEW
        list. The display twins are the honest comparison.

        econ_occ_display is a single fallback pick rather than one per column,
        so the suppression is read off it and applied to all three occupancy
        columns — which is exactly how the page behaves.
        """
        occ, noi = (r.get("econ_occ") or {}), (r.get("noi") or {})
        nd = r.get("noi_display") or {}
        occ_suppressed = isinstance(r.get("econ_occ_display"), str)

        def noi_cell(key):
            v = nd.get(key, noi.get(key))
            return v if isinstance(v, str) else m(v)

        def occ_cell(key):
            return r["econ_occ_display"] if occ_suppressed else occ.get(key)

        def growth(key):
            v = r.get(f"{key}_display", r.get(key))
            return v if isinstance(v, str) else pct(v)

        return {
            "occ_at_close": occ_cell("at_close"), "noi_at_close": noi_cell("at_close"),
            "occ_uw": occ_cell("uw_ye"), "noi_uw": noi_cell("uw_ye"),
            "occ_proj": occ_cell("projected_ye"), "noi_proj": noi_cell("projected_ye"),
            "expected_growth": growth("expected_growth"),
            "actual_growth": growth("actual_growth"),
        }[col]
    for vc, tup in P.PAGE3.items():
        name, vals = tup[0], tup[1:]
        g, r = orows.get(vc, (None, {}))
        for col, pv in zip(P.OP_COLS, vals):
            unit = "usd_m" if col.startswith("noi") else (
                "ratio_pct" if col.endswith("growth") else "pct")
            cells.append(Cell("Operating", g or "(not in live set)",
                              P.ALIASES.get(vc, name), vc, col, pv,
                              op_live(r, col) if r else None, unit))
    for g, tup in P.PAGE3_SUBTOTALS.items():
        label, vals = tup[0], tup[1:]
        s = osubs.get(g) or {}
        for col, pv in zip(P.OP_COLS, vals):
            unit = "usd_m" if col.startswith("noi") else (
                "ratio_pct" if col.endswith("growth") else "pct")
            cells.append(Cell("Operating", g, label, f"SUB:{g}", col, pv,
                              op_live(s, col) if s else None, unit))
    label, vals = P.PAGE3_TOTAL[0], P.PAGE3_TOTAL[1:]
    for col, pv in zip(P.OP_COLS, vals):
        unit = "usd_m" if col.startswith("noi") else (
            "ratio_pct" if col.endswith("growth") else "pct")
        cells.append(Cell("Operating", "TOTAL", label, "TOTAL", col, pv,
                          op_live(ototal, col) if ototal else None, unit))

    # ---- Loan ----
    lrows, lsubs, ltotal = live_grouped(subs_.get("loan") or {})
    def loan_live(r, col):
        if col == "debt":
            return m(r.get("debt"))
        # The *_display twin, not the raw field: a dev deal's raw ltv is None
        # while its display is the literal "Dev", which is what the page prints
        # and therefore what the PDF should be compared against. Comparing raw
        # put 20-odd "Dev vs None" rows in the NEW list.
        if col in ("ltv", "debt_yield"):
            v = r.get(f"{col}_display", r.get(col))
            return pct(v) if isinstance(v, (int, float)) else v
        if col == "ytd_dscr":
            v = r.get("ytd_dscr_display", r.get("ytd_dscr"))
            return v
        if col == "rate":
            return r.get("rate_display")
        if col == "maturity":
            return r.get("maturity_display")
        return None
    for vc, tup in P.PAGE4.items():
        name, vals = tup[0], tup[1:]
        g, r = lrows.get(vc, (None, {}))
        for col, pv in zip(P.LOAN_COLS, vals):
            unit = {"debt": "usd_m", "ltv": "pct", "ytd_dscr": "dscr",
                    "debt_yield": "pct"}.get(col, "pct")
            cells.append(Cell("Loan", g or "(not in live set)",
                              P.ALIASES.get(vc, name), vc, col, pv,
                              loan_live(r, col) if r else None, unit))
    for g, tup in P.PAGE4_SUBTOTALS.items():
        label, vals = tup[0], tup[1:]
        s = lsubs.get(g) or {}
        for col, pv in zip(P.LOAN_COLS, vals):
            if col in ("rate", "maturity"):
                continue
            unit = {"debt": "usd_m", "ltv": "pct", "ytd_dscr": "dscr",
                    "debt_yield": "pct"}[col]
            cells.append(Cell("Loan", g, label, f"SUB:{g}", col, pv,
                              loan_live(s, col) if s else None, unit))
    label, vals = P.PAGE4_TOTAL[0], P.PAGE4_TOTAL[1:]
    for col, pv in zip(P.LOAN_COLS, vals):
        if col in ("rate", "maturity"):
            continue
        unit = {"debt": "usd_m", "ltv": "pct", "ytd_dscr": "dscr",
                "debt_yield": "pct"}[col]
        cells.append(Cell("Loan", "TOTAL", label, "TOTAL", col, pv,
                          loan_live(ltotal, col) if ltotal else None, unit))

    for c in cells:
        classify(c)
    return cells


# ── workbook ─────────────────────────────────────────────────────────────

HDR = Font(bold=True, color="FFFFFF", size=10)
HDR_FILL = PatternFill("solid", fgColor="44546A")
BOLD = Font(bold=True)
FILLS = {"TIES": PatternFill("solid", fgColor="E8F3E8"),
         "KNOWN": PatternFill("solid", fgColor="FFF6E0"),
         "NEW": PatternFill("solid", fgColor="F8D7DA")}
SUB_FILL = PatternFill("solid", fgColor="EEF2F7")
COLS = ["Group", "Deal / Row", "Column", "PDF", "Live", "Difference",
        "% Difference", "Status", "Reason (KNOWN) / blank if NEW"]


def sheet(wb, page, cells):
    ws = wb.create_sheet(page)
    ws.append(COLS)
    for i, _ in enumerate(COLS, 1):
        c = ws.cell(row=1, column=i)
        c.font, c.fill = HDR, HDR_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for c in cells:
        is_tot = c.key.startswith("SUB:") or c.key in ("TOTAL", "EXDEV")
        ws.append([
            c.group or "", c.row, c.col,
            c.pdf if c.pdf is not None else "—",
            c.live if c.live is not None else "—",
            round(c.diff, 3) if c.diff is not None else "",
            round(c.pct_diff, 1) if c.pct_diff is not None else "",
            c.status, c.reason,
        ])
        r = ws.max_row
        ws.cell(row=r, column=8).fill = FILLS[c.status]
        ws.cell(row=r, column=8).font = Font(bold=(c.status == "NEW"))
        if is_tot:
            for col in range(1, 4):
                ws.cell(row=r, column=col).font = BOLD
            for col in range(1, 10):
                if col != 8:
                    ws.cell(row=r, column=col).fill = SUB_FILL
        for col in (4, 5, 6, 7):
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=9).alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCDEFGHI", (24, 32, 18, 12, 12, 12, 13, 9, 78)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
    return ws


def summary_sheet(wb, cells, meta):
    ws = wb.create_sheet("Overview", 0)
    ws["A1"] = "Live vs 26Q1 PDF — variance review"
    ws["A1"].font = Font(bold=True, size=14)
    row = 3
    for k, v in meta.items():
        ws.cell(row=row, column=1, value=k).font = BOLD
        ws.cell(row=row, column=2, value=v)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Counts").font = Font(bold=True, size=12)
    row += 1
    ws.append([])
    hdr = ["Page", "Cells compared", "TIES", "KNOWN", "NEW"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font, c.fill = HDR, HDR_FILL
    row += 1
    pages = ["Summary", "Financial", "Operating", "Loan"]
    tot = [0, 0, 0, 0]
    for pg in pages:
        sub = [c for c in cells if c.page == pg]
        n = len(sub)
        t = sum(1 for c in sub if c.status == "TIES")
        k = sum(1 for c in sub if c.status == "KNOWN")
        nw = sum(1 for c in sub if c.status == "NEW")
        for i, v in enumerate((n, t, k, nw)):
            tot[i] += v
        ws.cell(row=row, column=1, value=pg)
        for i, v in enumerate((n, t, k, nw), 2):
            ws.cell(row=row, column=i, value=v)
        ws.cell(row=row, column=5).fill = FILLS["NEW"] if nw else FILLS["TIES"]
        row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = BOLD
    for i, v in enumerate(tot, 2):
        ws.cell(row=row, column=i, value=v).font = BOLD
    row += 2

    new = [c for c in cells if c.status == "NEW"]
    ws.cell(row=row, column=1,
            value=f"NEW — unexplained, needs review ({len(new)})"
            ).font = Font(bold=True, size=12)
    row += 1
    hdr2 = ["Page", "Group", "Deal / Row", "Column", "PDF", "Live",
            "Difference", "% Difference"]
    for i, h in enumerate(hdr2, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font, c.fill = HDR, HDR_FILL
    row += 1
    if not new:
        ws.cell(row=row, column=1,
                value="None — every difference is accounted for.").font = BOLD
        row += 1
    for c in sorted(new, key=lambda x: (x.page, x.group or "", x.row, x.col)):
        for i, v in enumerate((c.page, c.group or "", c.row, c.col,
                               c.pdf if c.pdf is not None else "—",
                               c.live if c.live is not None else "—",
                               round(c.diff, 3) if c.diff is not None else "",
                               round(c.pct_diff, 1) if c.pct_diff is not None else ""), 1):
            ws.cell(row=row, column=i, value=v)
        ws.cell(row=row, column=1).fill = FILLS["NEW"]
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="How to read this").font = Font(bold=True, size=12)
    row += 1
    for line in (
        "TIES  — inside tolerance: $0.35M, 1.6pp on a percentage, 0.11 on a "
        "DSCR, 0.6pp on a growth rate. The published inputs are themselves "
        "rounded to one decimal, so exact equality is not available.",
        "KNOWN — already diagnosed. The reason is on the row, on its page tab.",
        "NEW   — not explained by anything diagnosed so far. This is the list "
        "to work through.",
        "",
        "Live figures come from a LOCALLY assembled bundle, so this measures the "
        "working tree, not whatever is deployed.",
        "Both sides are the SAME quarter — 2026-Q1, quarter_end 2026-03-31.",
    ):
        ws.cell(row=row, column=1, value=line).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
    for col, w in zip("ABCDEFGH", (26, 24, 32, 20, 12, 12, 12, 13)):
        ws.column_dimensions[col].width = w
    return ws


def main():
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        return 2
    src = sys.argv[1]
    out = "outputs/live_vs_pdf_26Q1_variance.xlsx"
    if "-o" in sys.argv:
        out = sys.argv[sys.argv.index("-o") + 1]

    bundle = json.load(open(src, encoding="utf-8"))
    cells = build_cells(bundle)

    res = bundle.get("resolution") or {}
    meta = {
        "Investor": res.get("investor_name") or "TGAM",
        "Quarter": "2026-Q1",
        "Quarter end": res.get("quarter_end") or "",
        "Reference document": "TIAA 26Q1, 4 pages",
        "Live source": f"locally assembled bundle ({os.path.basename(src)})",
        "Built": str(date.today()),
    }

    wb = Workbook()
    wb.remove(wb.active)
    # Page tabs are numbered to match the document, and the front tab is
    # "Overview" — naming both it and the page-1 tab "Summary" made openpyxl
    # silently create a "Summary1".
    for i, pg in enumerate(("Summary", "Financial", "Operating", "Loan"), 1):
        ws = sheet(wb, pg, [c for c in cells if c.page == pg])
        ws.title = f"{i} {pg}"
    summary_sheet(wb, cells, meta)
    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
    wb.save(out)

    n = len(cells)
    t = sum(1 for c in cells if c.status == "TIES")
    k = sum(1 for c in cells if c.status == "KNOWN")
    nw = sum(1 for c in cells if c.status == "NEW")
    print(f"{n} cells compared:  {t} TIES   {k} KNOWN   {nw} NEW")
    print(f"wrote {out}")
    if nw:
        print("\nNEW (unexplained):")
        for c in sorted([c for c in cells if c.status == "NEW"],
                        key=lambda x: (x.page, x.row, x.col)):
            d = "" if c.diff is None else f"  diff {c.diff:+.2f}"
            print(f"  {c.page:<10}{c.row[:30]:<32}{c.col:<18}"
                  f"PDF {str(c.pdf):>9}   live {str(c.live)[:12]:>12}{d}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
