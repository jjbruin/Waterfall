"""
psckoc_ui.py
PSCKOC Entity Analysis tab — upstream waterfall computation and member returns.

PSCKOC is a holding entity in the upstream waterfall chain:
    deals → PIG5/PIG6 → PSCKOC → members (PSC1, KCREIT, PCBLE)

This tab orchestrates:
1. Compute underlying deal results via get_cached_deal_result()
2. Run upstream waterfalls to trace cash into PSCKOC and through to members
3. Display income schedule, waterfall allocations, partner returns,
   AM fee schedule, and XIRR cash flows
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date
from typing import Dict, List, Optional

from compute import get_cached_deal_result
from waterfall import (
    run_recursive_upstream_waterfalls,
    calculate_entity_through_flow,
)
from loaders import load_waterfalls
from metrics import xirr, calculate_roe


# ── Constants ────────────────────────────────────────────────────────────
PSCKOC_ENTITY = "PSCKOC"
PSCKOC_MEMBERS = ["PSC1", "KCREIT", "PCBLE"]


# ── Public entry point ───────────────────────────────────────────────────

def render_psckoc_tab(inv, wf, acct, fc, coa, relationships_raw,
                      mri_loans_raw, mri_supp, mri_val,
                      capital_calls_raw, isbs_raw,
                      start_year, horizon_years, pro_yr_base):
    """Render the PSCKOC Entity Analysis tab."""
    _psckoc_fragment(inv, wf, acct, fc, coa, relationships_raw,
                     mri_loans_raw, mri_supp, mri_val,
                     capital_calls_raw, isbs_raw,
                     start_year, horizon_years, pro_yr_base)


@st.fragment
def _psckoc_fragment(inv, wf, acct, fc, coa, relationships_raw,
                     mri_loans_raw, mri_supp, mri_val,
                     capital_calls_raw, isbs_raw,
                     start_year, horizon_years, pro_yr_base):
    """Fragment-isolated PSCKOC analysis body."""

    st.subheader("PSCKOC Entity Analysis")
    st.caption("Upstream waterfall analysis for PSCKOC holding entity — "
               "traces deal cash flows through to PSC1, KCREIT, and PCBLE.")

    # ── 1. Identify underlying deals ──────────────────────────────────
    deal_vcodes = _find_psckoc_deals(inv, wf, relationships_raw)

    if not deal_vcodes:
        st.warning("No deals linked to PSCKOC found. Ensure PSCKOC waterfall "
                   "steps or relationship records exist.")
        return

    # Show deal portfolio table
    deal_info = _build_deal_portfolio_table(deal_vcodes, inv, wf, relationships_raw)
    st.markdown("**Deal Portfolio**")
    st.dataframe(deal_info, use_container_width=True, hide_index=True)

    # ── 2. Compute button ─────────────────────────────────────────────
    if st.button("Compute PSCKOC Returns", key="_psckoc_compute", type="primary"):
        _run_psckoc_computation(
            deal_vcodes, inv, wf, acct, fc, coa, relationships_raw,
            mri_loans_raw, mri_supp, mri_val,
            capital_calls_raw, isbs_raw,
            start_year, horizon_years, pro_yr_base,
        )

    # ── 3. Display results if available ───────────────────────────────
    results = st.session_state.get("_psckoc_results")
    if not results:
        st.info("Click **Compute PSCKOC Returns** to run the analysis.")
        return

    _display_results(results)


# ── Deal discovery ────────────────────────────────────────────────────────

def _find_psckoc_deals(inv, wf, relationships_raw):
    """Find deals whose upstream waterfall chain passes through PSCKOC.

    Chain: deal waterfall → PPI entity (PropCode) → PSCKOC (investor via relationships)

    Strategy:
    1. From relationships, find all entities where PSCKOC is an investor (PPI entities)
    2. From waterfalls, find deals whose PropCodes include any of those PPI entities
    3. Filter to actual deals with waterfall definitions
    """
    if relationships_raw is None or relationships_raw.empty:
        return []
    if wf is None or wf.empty:
        return []
    if inv is None or inv.empty:
        return []

    rel = relationships_raw.copy()
    rel["InvestorID"] = rel["InvestorID"].astype(str).str.strip()
    rel["InvestmentID"] = rel["InvestmentID"].astype(str).str.strip()

    # Step 1: entities where PSCKOC is an investor
    psckoc_investments = set(
        rel[rel["InvestorID"] == PSCKOC_ENTITY]["InvestmentID"].unique()
    )

    if not psckoc_investments:
        return []

    # Step 2: find deal vcodes whose waterfall steps reference any PSCKOC investment entity
    wf_norm = wf.copy()
    wf_norm["PropCode"] = wf_norm["PropCode"].fillna("").astype(str).str.strip()
    wf_norm["vcode"] = wf_norm["vcode"].fillna("").astype(str).str.strip()

    deal_vcodes_set = set(inv["vcode"].astype(str).str.strip())
    wf_vcodes_set = set(wf_norm["vcode"].unique())

    deal_vcodes = set()
    for ppi in psckoc_investments:
        # Find vcodes whose waterfall references this PPI entity
        refs = wf_norm[wf_norm["PropCode"] == ppi]["vcode"].unique()
        for vc in refs:
            # Only add if it's a real deal in the deals table AND has a waterfall
            if vc in deal_vcodes_set and vc in wf_vcodes_set:
                deal_vcodes.add(vc)

    # Remove PSCKOC itself, its members, and PPI intermediaries from deal list
    deal_vcodes -= {PSCKOC_ENTITY} | set(PSCKOC_MEMBERS) | psckoc_investments

    return sorted(deal_vcodes)


def _build_deal_portfolio_table(deal_vcodes, inv, wf=None, relationships_raw=None):
    """Build a summary table of underlying deals with PPI linkage."""

    # Build PPI → deal mapping and PSCKOC ownership %
    ppi_for_deal = {}  # deal_vcode -> (ppi_entity, ownership_pct)
    if wf is not None and relationships_raw is not None:
        rel = relationships_raw.copy()
        rel["InvestorID"] = rel["InvestorID"].astype(str).str.strip()
        rel["InvestmentID"] = rel["InvestmentID"].astype(str).str.strip()

        psckoc_inv = rel[rel["InvestorID"] == PSCKOC_ENTITY]
        ppi_pcts = {}
        for _, r in psckoc_inv.iterrows():
            ppi = r["InvestmentID"]
            pct = float(r.get("OwnershipPct", 0))
            if pct > 1:
                pct = pct / 100.0
            ppi_pcts[ppi] = pct

        wf_norm = wf.copy()
        wf_norm["PropCode"] = wf_norm["PropCode"].fillna("").astype(str).str.strip()
        wf_norm["vcode"] = wf_norm["vcode"].fillna("").astype(str).str.strip()

        for ppi, pct in ppi_pcts.items():
            refs = wf_norm[wf_norm["PropCode"] == ppi]["vcode"].unique()
            for vc in refs:
                if vc in deal_vcodes:
                    ppi_for_deal[vc] = (ppi, pct)

    rows = []
    for vc in deal_vcodes:
        row = {"vcode": vc, "Investment Name": vc}
        if inv is not None and not inv.empty:
            match = inv[inv["vcode"].astype(str).str.strip() == vc]
            if not match.empty:
                r = match.iloc[0]
                row["Investment Name"] = str(r.get("Investment_Name", vc))
                row["Asset Type"] = str(r.get("Asset_Type", ""))
                row["Sale Status"] = str(r.get("Sale_Status", "Active") or "Active")
        if vc in ppi_for_deal:
            ppi, pct = ppi_for_deal[vc]
            row["PPI Entity"] = ppi
            row["PSCKOC %"] = f"{pct:.1%}"
        rows.append(row)
    return pd.DataFrame(rows)


# ── Computation ───────────────────────────────────────────────────────────

def _run_psckoc_computation(deal_vcodes, inv, wf, acct, fc, coa,
                            relationships_raw, mri_loans_raw, mri_supp, mri_val,
                            capital_calls_raw, isbs_raw,
                            start_year, horizon_years, pro_yr_base):
    """Run deal computations and upstream waterfalls for PSCKOC."""

    wf_steps = load_waterfalls(wf)
    all_cf_allocs = []
    all_cap_allocs = []
    deal_results = {}
    errors = []

    progress = st.progress(0, text="Computing deals...")

    for i, vcode in enumerate(deal_vcodes):
        progress.progress((i + 1) / len(deal_vcodes),
                          text=f"Computing {vcode} ({i+1}/{len(deal_vcodes)})...")

        # Look up deal metadata
        match = inv[inv["vcode"].astype(str) == vcode]
        if match.empty:
            errors.append(f"Deal {vcode}: not found in deals table")
            continue

        row = match.iloc[0]
        inv_id = str(row.get("InvestmentID", ""))
        sale_date_raw = row.get("Sale_Date", None)

        try:
            result = get_cached_deal_result(
                vcode=vcode,
                start_year=start_year,
                horizon_years=horizon_years,
                pro_yr_base=pro_yr_base,
                deal_investment_id=inv_id,
                sale_date_raw=sale_date_raw,
                inv=inv, wf=wf, acct=acct, fc=fc, coa=coa,
                mri_loans_raw=mri_loans_raw,
                mri_supp=mri_supp,
                mri_val=mri_val,
                relationships_raw=relationships_raw,
                capital_calls_raw=capital_calls_raw,
                isbs_raw=isbs_raw,
            )
        except Exception as e:
            errors.append(f"Deal {vcode}: {e}")
            continue

        if "error" in result:
            errors.append(f"Deal {vcode}: {result['error']}")
            continue

        deal_results[vcode] = result

        # Collect allocations for upstream processing
        cf_alloc = result.get("cf_alloc")
        cap_alloc = result.get("cap_alloc")

        if cf_alloc is not None and not cf_alloc.empty:
            cf = cf_alloc.copy()
            cf["vcode"] = vcode
            all_cf_allocs.append(cf)

        if cap_alloc is not None and not cap_alloc.empty:
            ca = cap_alloc.copy()
            ca["vcode"] = vcode
            all_cap_allocs.append(ca)

    progress.empty()

    if errors:
        for e in errors:
            st.warning(e)

    if not all_cf_allocs and not all_cap_allocs:
        st.error("No deal allocations produced. Cannot run upstream waterfalls.")
        return

    # ── Run upstream waterfalls ───────────────────────────────────────
    with st.spinner("Running upstream waterfalls..."):
        # Combine all deal allocations
        combined_cf = pd.concat(all_cf_allocs, ignore_index=True) if all_cf_allocs else pd.DataFrame()
        combined_cap = pd.concat(all_cap_allocs, ignore_index=True) if all_cap_allocs else pd.DataFrame()

        # Build relationships for upstream processing
        if relationships_raw is not None and not relationships_raw.empty:
            from ownership_tree import load_relationships
            relationships = load_relationships(relationships_raw)
        else:
            relationships = pd.DataFrame()

        # Run CF upstream waterfalls
        cf_upstream_alloc = pd.DataFrame()
        cf_entity_states = {}
        cf_beneficiary_totals = {}
        if not combined_cf.empty:
            cf_upstream_alloc, cf_entity_states, cf_beneficiary_totals = \
                run_recursive_upstream_waterfalls(
                    deal_allocations=combined_cf,
                    wf_steps=wf_steps,
                    relationships=relationships,
                    wf_type="CF_WF",
                )

        # Run Cap upstream waterfalls
        cap_upstream_alloc = pd.DataFrame()
        cap_entity_states = {}
        cap_beneficiary_totals = {}
        if not combined_cap.empty:
            cap_upstream_alloc, cap_entity_states, cap_beneficiary_totals = \
                run_recursive_upstream_waterfalls(
                    deal_allocations=combined_cap,
                    wf_steps=wf_steps,
                    relationships=relationships,
                    wf_type="Cap_WF",
                )

    # ── Build results dict ────────────────────────────────────────────
    results = _build_psckoc_results(
        cf_upstream_alloc=cf_upstream_alloc,
        cap_upstream_alloc=cap_upstream_alloc,
        cf_entity_states=cf_entity_states,
        cap_entity_states=cap_entity_states,
        cf_beneficiary_totals=cf_beneficiary_totals,
        cap_beneficiary_totals=cap_beneficiary_totals,
        deal_results=deal_results,
        deal_vcodes=deal_vcodes,
        inv=inv,
    )

    st.session_state["_psckoc_results"] = results
    st.success(f"PSCKOC analysis complete — {len(deal_results)} deals processed.")


def _build_psckoc_results(cf_upstream_alloc, cap_upstream_alloc,
                          cf_entity_states, cap_entity_states,
                          cf_beneficiary_totals, cap_beneficiary_totals,
                          deal_results, deal_vcodes, inv):
    """Build structured results from upstream waterfall outputs."""

    # ── Income schedule: cash arriving at PSCKOC ──────────────────────
    income_rows = []
    for wf_type, alloc_df in [("CF", cf_upstream_alloc), ("Cap", cap_upstream_alloc)]:
        if alloc_df.empty:
            continue
        # PSCKOC receives cash where it appears as PropCode in upstream allocations
        psckoc_income = alloc_df[alloc_df["PropCode"].astype(str) == PSCKOC_ENTITY].copy()
        for _, row in psckoc_income.iterrows():
            income_rows.append({
                "Date": row.get("event_date"),
                "Source Entity": str(row.get("Entity", "")),
                "Source Deal": str(row.get("Path", "")).split("->")[0] if row.get("Path") else "",
                "Type": wf_type,
                "vState": str(row.get("vState", "")),
                "vtranstype": str(row.get("vtranstype", "")),
                "Amount": float(row.get("Allocated", 0)),
            })
    income_df = pd.DataFrame(income_rows)

    # ── Member allocations: cash leaving PSCKOC to members ────────────
    member_alloc_rows = []
    for wf_type, alloc_df in [("CF", cf_upstream_alloc), ("Cap", cap_upstream_alloc)]:
        if alloc_df.empty:
            continue
        # Steps where Entity == PSCKOC (i.e., PSCKOC's waterfall distributing to members)
        psckoc_dists = alloc_df[alloc_df["Entity"].astype(str) == PSCKOC_ENTITY].copy()
        for _, row in psckoc_dists.iterrows():
            member_alloc_rows.append({
                "Date": row.get("event_date"),
                "Member": str(row.get("PropCode", "")),
                "Type": wf_type,
                "iOrder": int(row.get("iOrder", 0)),
                "vState": str(row.get("vState", "")),
                "vtranstype": str(row.get("vtranstype", "")),
                "FXRate": float(row.get("FXRate", 0)),
                "Amount": float(row.get("Allocated", 0)),
            })
    member_alloc_df = pd.DataFrame(member_alloc_rows)

    # ── Partner returns from entity states ────────────────────────────
    partner_returns = []
    for member in PSCKOC_MEMBERS:
        cf_st = cf_entity_states.get(member)
        cap_st = cap_entity_states.get(member)
        state = cf_st or cap_st
        if not state:
            continue

        # Combined cashflows from both waterfall types
        combined_cfs = []
        seen = set()
        for st_obj in [cf_st, cap_st]:
            if st_obj and st_obj.cashflows:
                for cf in st_obj.cashflows:
                    key = (cf[0], cf[1])
                    if key not in seen:
                        combined_cfs.append(cf)
                        seen.add(key)

        # Sort by date
        combined_cfs.sort(key=lambda x: x[0])

        # Calculate metrics
        irr_val = xirr(combined_cfs) if combined_cfs else None
        contributions = sum(amt for _, amt in combined_cfs if amt < 0)
        distributions = sum(amt for _, amt in combined_cfs if amt > 0)

        # CF distributions for ROE
        cf_dists = []
        if cf_st and cf_st.cf_distributions:
            cf_dists = cf_st.cf_distributions

        roe_val = calculate_roe(
            combined_cfs,
            cf_dists,
            combined_cfs[0][0] if combined_cfs else date.today(),
            combined_cfs[-1][0] if combined_cfs else date.today(),
        ) if combined_cfs else 0.0

        moic_val = (distributions + abs(contributions)) / abs(contributions) if contributions < 0 else 0.0
        # MOIC = (distributions + unrealized) / |contributions|
        # For upstream entities, unrealized is 0 (terminal beneficiaries)
        moic_val = distributions / abs(contributions) if contributions < 0 else 0.0

        # CF vs Cap distribution totals
        cf_dist_total = 0.0
        if not member_alloc_df.empty:
            cf_member = member_alloc_df[
                (member_alloc_df["Member"] == member) & (member_alloc_df["Type"] == "CF")
            ]
            cf_dist_total = cf_member["Amount"].sum()

        cap_dist_total = 0.0
        if not member_alloc_df.empty:
            cap_member = member_alloc_df[
                (member_alloc_df["Member"] == member) & (member_alloc_df["Type"] == "Cap")
            ]
            cap_dist_total = cap_member["Amount"].sum()

        partner_returns.append({
            "partner": member,
            "contributions": contributions,
            "cf_distributions": cf_dist_total,
            "cap_distributions": cap_dist_total,
            "total_distributions": distributions,
            "irr": irr_val,
            "roe": roe_val,
            "moic": moic_val,
            "combined_cashflows": combined_cfs,
            "cf_only_distributions": cf_dists,
        })

    # ── AM Fee schedule ───────────────────────────────────────────────
    am_fee_rows = []
    for wf_type, alloc_df in [("CF", cf_upstream_alloc), ("Cap", cap_upstream_alloc)]:
        if alloc_df.empty:
            continue
        am_fees = alloc_df[
            (alloc_df["Entity"].astype(str) == PSCKOC_ENTITY) &
            (alloc_df["vState"].astype(str) == "AMFee")
        ].copy()
        for _, row in am_fees.iterrows():
            am_fee_rows.append({
                "Date": row.get("event_date"),
                "Type": wf_type,
                "Recipient": str(row.get("PropCode", "")),
                "Amount": float(row.get("Allocated", 0)),
            })
    am_fee_df = pd.DataFrame(am_fee_rows)

    # ── Deal summary ──────────────────────────────────────────────────
    all_combined = []
    for pr in partner_returns:
        all_combined.extend(pr["combined_cashflows"])
    all_combined.sort(key=lambda x: x[0])

    total_contributions = sum(amt for _, amt in all_combined if amt < 0)
    total_distributions = sum(amt for _, amt in all_combined if amt > 0)
    deal_irr = xirr(all_combined) if all_combined else None
    deal_moic = total_distributions / abs(total_contributions) if total_contributions < 0 else 0.0

    deal_summary = {
        "deal_irr": deal_irr,
        "deal_moic": deal_moic,
        "total_contributions": total_contributions,
        "total_distributions": total_distributions,
    }

    return {
        "income_schedule": income_df,
        "member_allocations": member_alloc_df,
        "partner_results": partner_returns,
        "deal_summary": deal_summary,
        "am_fee_schedule": am_fee_df,
        "deal_vcodes": deal_vcodes,
        "deal_results": {vc: True for vc in deal_results},  # Just flags, don't store full results
    }


# ── Display ───────────────────────────────────────────────────────────────

def _display_results(results):
    """Display all PSCKOC analysis sections."""

    partner_results = results.get("partner_results", [])
    deal_summary = results.get("deal_summary", {})
    income_df = results.get("income_schedule", pd.DataFrame())
    member_alloc_df = results.get("member_allocations", pd.DataFrame())
    am_fee_df = results.get("am_fee_schedule", pd.DataFrame())

    # ── Partner Returns (KPI cards) ───────────────────────────────────
    st.markdown("---")
    st.markdown("### Partner Returns")

    if partner_results:
        cols = st.columns(len(partner_results) + 1)

        for i, pr in enumerate(partner_results):
            with cols[i]:
                st.markdown(f"**{pr['partner']}**")
                irr_str = f"{pr['irr']:.2%}" if pr['irr'] is not None else "N/A"
                st.metric("IRR", irr_str)
                st.metric("Distributions", f"${pr['total_distributions']:,.0f}")
                moic_str = f"{pr['moic']:.2f}x" if pr['moic'] else "N/A"
                st.metric("MOIC", moic_str)

        with cols[-1]:
            st.markdown("**PSCKOC Total**")
            irr_str = f"{deal_summary.get('deal_irr', 0):.2%}" if deal_summary.get('deal_irr') is not None else "N/A"
            st.metric("IRR", irr_str)
            st.metric("Distributions", f"${deal_summary.get('total_distributions', 0):,.0f}")
            moic_str = f"{deal_summary.get('deal_moic', 0):.2f}x"
            st.metric("MOIC", moic_str)

    # ── Partner Returns Table ─────────────────────────────────────────
    if partner_results:
        table_rows = []
        for pr in partner_results:
            table_rows.append({
                "Partner": pr["partner"],
                "Contributions": pr["contributions"],
                "CF Distributions": pr["cf_distributions"],
                "Cap Distributions": pr["cap_distributions"],
                "Total Distributions": pr["total_distributions"],
                "IRR": pr["irr"],
                "ROE": pr["roe"],
                "MOIC": pr["moic"],
            })
        # Add deal total row
        table_rows.append({
            "Partner": "PSCKOC Total",
            "Contributions": deal_summary.get("total_contributions", 0),
            "CF Distributions": sum(pr["cf_distributions"] for pr in partner_results),
            "Cap Distributions": sum(pr["cap_distributions"] for pr in partner_results),
            "Total Distributions": deal_summary.get("total_distributions", 0),
            "IRR": deal_summary.get("deal_irr"),
            "ROE": None,
            "MOIC": deal_summary.get("deal_moic", 0),
        })

        returns_df = pd.DataFrame(table_rows)

        styled = returns_df.style.format({
            "Contributions": "${:,.0f}",
            "CF Distributions": "${:,.0f}",
            "Cap Distributions": "${:,.0f}",
            "Total Distributions": "${:,.0f}",
            "IRR": lambda v: f"{v:.2%}" if pd.notna(v) and v is not None else "N/A",
            "ROE": lambda v: f"{v:.2%}" if pd.notna(v) and v is not None else "N/A",
            "MOIC": lambda v: f"{v:.2f}x" if pd.notna(v) and v is not None else "N/A",
        }).apply(
            lambda row: ["font-weight: bold; border-top: 2px solid #333"] * len(row)
            if row["Partner"] == "PSCKOC Total" else [""] * len(row),
            axis=1,
        )
        st.dataframe(styled, use_container_width=True, hide_index=True)

    # ── Income Schedule ───────────────────────────────────────────────
    with st.expander("Income Schedule — Cash arriving at PSCKOC", expanded=False):
        if income_df.empty:
            st.info("No income allocated to PSCKOC.")
        else:
            # Summary by source deal and type
            summary = income_df.groupby(["Source Deal", "Type"])["Amount"].sum().reset_index()
            summary["Amount"] = summary["Amount"].map(lambda x: f"${x:,.2f}")
            st.dataframe(summary, use_container_width=True, hide_index=True)

            # Detail table
            with st.expander("Full Detail"):
                display = income_df.copy()
                display["Amount"] = display["Amount"].map(lambda x: f"${x:,.2f}")
                st.dataframe(display, use_container_width=True, hide_index=True)

    # ── Waterfall Allocations ─────────────────────────────────────────
    with st.expander("Waterfall Allocations — PSCKOC to Members", expanded=False):
        if member_alloc_df.empty:
            st.info("No member allocations found.")
        else:
            for wf_type in ["CF", "Cap"]:
                type_df = member_alloc_df[member_alloc_df["Type"] == wf_type]
                if type_df.empty:
                    continue
                st.markdown(f"**{wf_type} Waterfall**")
                summary = type_df.groupby(["iOrder", "Member", "vState", "vtranstype"])["Amount"].sum().reset_index()
                summary = summary.sort_values("iOrder")
                summary["Amount"] = summary["Amount"].map(lambda x: f"${x:,.2f}")
                st.dataframe(summary, use_container_width=True, hide_index=True)

    # ── AM Fee Schedule ───────────────────────────────────────────────
    with st.expander("AM Fee Schedule", expanded=False):
        if am_fee_df.empty:
            st.info("No AM fee entries found. (AMFee steps will appear after "
                    "waterfall is configured with AMFee vState.)")
        else:
            am_fee_df_display = am_fee_df.copy()
            am_fee_df_display["Amount"] = am_fee_df_display["Amount"].map(lambda x: f"${x:,.2f}")
            st.dataframe(am_fee_df_display, use_container_width=True, hide_index=True)
            st.metric("Total AM Fees", f"${am_fee_df['Amount'].sum():,.2f}")

    # ── XIRR Cash Flows ──────────────────────────────────────────────
    with st.expander("XIRR Cash Flows — Combined per Member", expanded=False):
        if not partner_results:
            st.info("No cashflow data available.")
        else:
            for pr in partner_results:
                st.markdown(f"**{pr['partner']}**")
                cfs = pr.get("combined_cashflows", [])
                if cfs:
                    cf_df = pd.DataFrame(cfs, columns=["Date", "Amount"])
                    cf_df = cf_df.sort_values("Date")
                    cf_df["Amount"] = cf_df["Amount"].map(lambda x: f"${x:,.2f}")
                    st.dataframe(cf_df, use_container_width=True, hide_index=True)
                else:
                    st.caption("No cashflows recorded.")

    # ── Excel Export ──────────────────────────────────────────────────
    xlsx_data = _generate_psckoc_excel(results)
    if xlsx_data:
        st.download_button(
            "Download PSCKOC Analysis (Excel)",
            xlsx_data,
            file_name="psckoc_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="_psckoc_download",
        )


# ── Excel export ──────────────────────────────────────────────────────────

def _generate_psckoc_excel(results):
    """Generate Excel workbook with PSCKOC analysis results."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, numbers, Border, Side
    except ImportError:
        return None

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    currency_fmt = '#,##0'
    pct_fmt = '0.00%'
    mult_fmt = '0.00"x"'
    bold_font = Font(bold=True)
    top_border = Border(top=Side(style="thin"))

    # ── Sheet 1: Partner Returns ──────────────────────────────────────
    ws = wb.active
    ws.title = "Partner Returns"

    headers = ["Partner", "Contributions", "CF Distributions", "Cap Distributions",
               "Total Distributions", "IRR", "ROE", "MOIC"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    partner_results = results.get("partner_results", [])
    deal_summary = results.get("deal_summary", {})

    row_idx = 2
    for pr in partner_results:
        ws.cell(row=row_idx, column=1, value=pr["partner"])
        ws.cell(row=row_idx, column=2, value=pr["contributions"]).number_format = currency_fmt
        ws.cell(row=row_idx, column=3, value=pr["cf_distributions"]).number_format = currency_fmt
        ws.cell(row=row_idx, column=4, value=pr["cap_distributions"]).number_format = currency_fmt
        ws.cell(row=row_idx, column=5, value=pr["total_distributions"]).number_format = currency_fmt
        if pr["irr"] is not None:
            ws.cell(row=row_idx, column=6, value=pr["irr"]).number_format = pct_fmt
        ws.cell(row=row_idx, column=7, value=pr["roe"]).number_format = pct_fmt
        ws.cell(row=row_idx, column=8, value=pr["moic"]).number_format = mult_fmt
        row_idx += 1

    # Total row
    for col in range(1, 9):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = bold_font
        cell.border = top_border
    ws.cell(row=row_idx, column=1, value="PSCKOC Total")
    ws.cell(row=row_idx, column=2, value=deal_summary.get("total_contributions", 0)).number_format = currency_fmt
    ws.cell(row=row_idx, column=5, value=deal_summary.get("total_distributions", 0)).number_format = currency_fmt
    if deal_summary.get("deal_irr") is not None:
        ws.cell(row=row_idx, column=6, value=deal_summary["deal_irr"]).number_format = pct_fmt
    ws.cell(row=row_idx, column=8, value=deal_summary.get("deal_moic", 0)).number_format = mult_fmt

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for r in range(2, row_idx + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 25)

    # ── Sheet 2: Income Schedule ──────────────────────────────────────
    income_df = results.get("income_schedule", pd.DataFrame())
    if not income_df.empty:
        ws2 = wb.create_sheet("Income Schedule")
        inc_headers = list(income_df.columns)
        for col, h in enumerate(inc_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for r_idx, (_, row) in enumerate(income_df.iterrows(), 2):
            for c_idx, h in enumerate(inc_headers, 1):
                val = row[h]
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                if h == "Amount":
                    cell.number_format = currency_fmt

    # ── Sheet 3: AM Fee Schedule ──────────────────────────────────────
    am_fee_df = results.get("am_fee_schedule", pd.DataFrame())
    if not am_fee_df.empty:
        ws3 = wb.create_sheet("AM Fee Schedule")
        am_headers = list(am_fee_df.columns)
        for col, h in enumerate(am_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for r_idx, (_, row) in enumerate(am_fee_df.iterrows(), 2):
            for c_idx, h in enumerate(am_headers, 1):
                val = row[h]
                cell = ws3.cell(row=r_idx, column=c_idx, value=val)
                if h == "Amount":
                    cell.number_format = currency_fmt

    # ── Sheet 4: XIRR Cash Flows ─────────────────────────────────────
    ws4 = wb.create_sheet("XIRR Cash Flows")
    cf_headers = ["Date", "Partner", "Amount"]
    for col, h in enumerate(cf_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    r_idx = 2
    for pr in partner_results:
        for cf_date, cf_amt in pr.get("combined_cashflows", []):
            ws4.cell(row=r_idx, column=1, value=cf_date)
            ws4.cell(row=r_idx, column=2, value=pr["partner"])
            ws4.cell(row=r_idx, column=3, value=cf_amt).number_format = currency_fmt
            r_idx += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
