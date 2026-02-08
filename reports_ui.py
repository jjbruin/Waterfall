"""
reports_ui.py
Reports tab UI — report type selector, population selector, generation, preview, and Excel export.

Initial report: Projected Returns Summary (Contributions, CF Distributions, Capital Distributions, IRR, ROE, MOIC).
"""

import streamlit as st
import pandas as pd
from io import BytesIO

from compute import compute_deal_analysis


def _build_partner_returns(deal_result: dict, deal_name: str) -> list[dict]:
    """Extract partner-level and deal-level return metrics from a computed deal result.

    Thin wrapper around pre-computed partner_results / deal_summary produced by
    build_partner_results() in compute.py.  Maps to the report row format.

    Returns a list of dicts — partner rows followed by a deal-level total row.
    The deal-level row has Partner='Deal Total' and _is_deal_total=True (internal flag).
    """
    partner_results = deal_result.get('partner_results', [])
    deal_summary = deal_result.get('deal_summary', {})

    if not partner_results:
        return []

    rows = []
    for pr in partner_results:
        rows.append({
            'Deal Name': deal_name,
            'Partner': pr['partner'],
            'Contributions': pr['contributions'],
            'CF Distributions': pr['cf_distributions'],
            'Capital Distributions': pr['cap_distributions'],
            'IRR': pr['irr'],
            'ROE': pr['roe'],
            'MOIC': pr['moic'],
            '_is_deal_total': False,
        })

    rows.append({
        'Deal Name': deal_name,
        'Partner': 'Deal Total',
        'Contributions': deal_summary.get('total_contributions', 0.0),
        'CF Distributions': deal_summary.get('total_cf_distributions', 0.0),
        'Capital Distributions': deal_summary.get('total_cap_distributions', 0.0),
        'IRR': deal_summary.get('deal_irr'),
        'ROE': deal_summary.get('deal_roe', 0.0),
        'MOIC': deal_summary.get('deal_moic', 0.0),
        '_is_deal_total': True,
    })

    return rows


def _generate_excel(df: pd.DataFrame) -> bytes:
    """Create a formatted Excel workbook from the projected returns DataFrame.

    Deal-total rows get bold text and a solid top border.

    Returns:
        Bytes suitable for st.download_button.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Projected Returns"

    # Exclude the internal flag column from output
    display_cols = [c for c in df.columns if c != '_is_deal_total']

    # --- Header row ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, col_name in enumerate(display_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # --- Styles for deal-total rows ---
    bold_font = Font(bold=True)
    top_border = Border(top=Side(style='medium'))

    # --- Data rows ---
    currency_cols = {'Contributions', 'CF Distributions', 'Capital Distributions'}
    pct_cols = {'IRR', 'ROE'}
    moic_cols = {'MOIC'}

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        is_total = bool(row.get('_is_deal_total', False))

        for col_idx, col_name in enumerate(display_cols, start=1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name in currency_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = '$#,##0'
            elif col_name in pct_cols:
                cell.value = float(val) if pd.notna(val) else None
                cell.number_format = '0.00%'
            elif col_name in moic_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = '0.00"x"'
            else:
                cell.value = val

            if is_total:
                cell.font = bold_font
                cell.border = top_border

    # --- Auto-width ---
    for col_idx, col_name in enumerate(display_cols, start=1):
        max_len = len(str(col_name))
        for row_idx in range(2, len(df) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_deal_lookup(inv, wf):
    """Build deal label/vcode/eligible lookup tables once per render.

    Returns (inv_disp, eligible_deals, eligible_labels, wf_norm) where:
    - inv_disp: deals DataFrame with DealLabel column, child properties excluded
    - eligible_deals: subset of inv_disp that have waterfall definitions
    - eligible_labels: sorted list of eligible deal labels
    - wf_norm: waterfalls DataFrame with normalised 'vcode' column
    """
    inv_disp = inv.copy()
    inv_disp["Investment_Name"] = inv_disp["Investment_Name"].fillna("").astype(str)
    inv_disp["vcode"] = inv_disp["vcode"].astype(str)

    name_counts = inv_disp["Investment_Name"].value_counts()
    inv_disp["DealLabel"] = inv_disp.apply(
        lambda r: (
            f"{r['Investment_Name']} ({r['vcode']})"
            if name_counts.get(r['Investment_Name'], 0) > 1
            else r['Investment_Name']
        ),
        axis=1,
    )

    # --- Fast child-property filter (vectorised, no per-row function call) ---
    # A deal is a child property if its Portfolio_Name matches another deal's
    # Investment_Name AND it is not that deal itself.
    if "Portfolio_Name" in inv_disp.columns:
        inv_disp["Portfolio_Name"] = inv_disp["Portfolio_Name"].fillna("").astype(str).str.strip()
        parent_names = set(inv_disp["Investment_Name"].str.strip())
        is_child = (
            inv_disp["Portfolio_Name"].isin(parent_names)
            & (inv_disp["Portfolio_Name"] != inv_disp["Investment_Name"].str.strip())
            & (inv_disp["Portfolio_Name"] != "")
        )
        inv_disp = inv_disp[~is_child].copy()

    # --- Waterfalls normalisation ---
    wf_norm = wf.copy()
    wf_norm.columns = [str(c).strip() for c in wf_norm.columns]
    if "vCode" in wf_norm.columns and "vcode" not in wf_norm.columns:
        wf_norm = wf_norm.rename(columns={"vCode": "vcode"})
    wf_norm["vcode"] = wf_norm["vcode"].astype(str)

    wf_vcodes = set(wf_norm["vcode"].unique())
    eligible_deals = inv_disp[inv_disp["vcode"].isin(wf_vcodes)]
    eligible_labels = sorted(
        eligible_deals["DealLabel"].dropna().unique().tolist(),
        key=lambda x: x.lower(),
    )

    return inv_disp, eligible_deals, eligible_labels, wf_norm


def render_reports(
    inv, wf, acct, fc, coa,
    mri_loans_raw, mri_supp, mri_val,
    relationships_raw, capital_calls_raw, isbs_raw,
    start_year: int, horizon_years: int, pro_yr_base: int,
):
    """Render the Reports tab contents."""

    # --- Report type selector ---
    report_type = st.selectbox(
        "Report Type",
        ["Projected Returns Summary"],
        key="report_type_selector",
    )

    # --- Pre-compute deal/label lookup (fast, no per-row calls) ---
    inv_disp, eligible_deals, eligible_labels, wf_norm = _build_deal_lookup(inv, wf)

    # --- Population selector ---
    population_options = [
        "Current Deal",
        "Select Deals",
        "By Partner",
        "By Upstream Investor",
        "All Deals",
    ]
    population = st.selectbox(
        "Population",
        population_options,
        key="report_population_selector",
    )

    selected_report_labels = []

    if population == "Current Deal":
        current_label = st.session_state.get("deal_selector", "")
        if current_label:
            st.caption(f"Using currently selected deal: **{current_label}**")
            selected_report_labels = [current_label]
        else:
            st.warning("No deal selected in the Deal Analysis tab.")
            return

    elif population == "Select Deals":
        selected_report_labels = st.multiselect(
            "Select deals to include",
            eligible_labels,
            key="report_deal_multiselect",
        )

    elif population == "By Partner":
        # Build partner → eligible deal labels from waterfalls PropCode
        partner_to_labels = {}
        eligible_vcodes = set(eligible_deals["vcode"].astype(str))
        for _, r in wf_norm.iterrows():
            vc = str(r["vcode"])
            pc = str(r.get("PropCode", "")).strip()
            if vc in eligible_vcodes and pc:
                partner_to_labels.setdefault(pc, set()).add(vc)

        # Resolve vcodes → labels
        vcode_to_label = dict(zip(
            eligible_deals["vcode"].astype(str),
            eligible_deals["DealLabel"],
        ))
        for pc in partner_to_labels:
            partner_to_labels[pc] = sorted(
                [vcode_to_label[vc] for vc in partner_to_labels[pc] if vc in vcode_to_label],
                key=lambda x: x.lower(),
            )

        all_partners = sorted(partner_to_labels.keys(), key=lambda x: x.lower())
        selected_partner = st.selectbox(
            "Select partner",
            all_partners,
            key="report_partner_selector",
        )
        if selected_partner:
            matched = partner_to_labels.get(selected_partner, [])
            st.caption(f"Partner **{selected_partner}** appears in {len(matched)} deal(s).")
            selected_report_labels = matched

    elif population == "By Upstream Investor":
        if relationships_raw is None or relationships_raw.empty:
            st.warning("No relationship data available. Load MRI_IA_Relationship to use this filter.")
            return

        from ownership_tree import load_relationships, build_ownership_tree, get_ultimate_investors
        from loaders import build_investmentid_to_vcode

        relationships = load_relationships(relationships_raw)
        nodes = build_ownership_tree(relationships)
        inv_to_vcode = build_investmentid_to_vcode(inv)

        eligible_vcodes = set(eligible_deals["vcode"].astype(str))
        vcode_to_label = dict(zip(
            eligible_deals["vcode"].astype(str),
            eligible_deals["DealLabel"],
        ))

        # Walk each eligible deal's ownership tree to find upstream investors
        investor_to_vcodes: dict[str, set[str]] = {}
        for inv_id, vc in inv_to_vcode.items():
            if str(vc) not in eligible_vcodes:
                continue
            if inv_id not in nodes:
                continue
            ultimate = get_ultimate_investors(inv_id, nodes, normalize=True)
            for investor_id, _ in ultimate:
                investor_to_vcodes.setdefault(investor_id, set()).add(str(vc))

        # Also include direct investors from relationships
        for _, rel_row in relationships.iterrows():
            inv_id = str(rel_row.get("InvestmentID", "")).strip()
            investor_id = str(rel_row.get("InvestorID", "")).strip()
            vc = inv_to_vcode.get(inv_id)
            if vc and str(vc) in eligible_vcodes and investor_id:
                investor_to_vcodes.setdefault(investor_id, set()).add(str(vc))

        # Build labels and sort
        investor_to_labels = {}
        for investor_id, vcs in investor_to_vcodes.items():
            labels = sorted(
                [vcode_to_label[vc] for vc in vcs if vc in vcode_to_label],
                key=lambda x: x.lower(),
            )
            if labels:
                investor_to_labels[investor_id] = labels

        # Add investor name if available from nodes
        investor_options = []
        for iid in sorted(investor_to_labels.keys(), key=lambda x: x.lower()):
            node = nodes.get(iid)
            name = node.name if node and hasattr(node, 'name') and node.name else ""
            display = f"{iid} — {name}" if name and name != iid else iid
            investor_options.append((iid, display))

        selected_investor_display = st.selectbox(
            "Select upstream investor",
            [d for _, d in investor_options],
            key="report_upstream_investor_selector",
        )
        if selected_investor_display:
            # Resolve display back to id
            selected_investor_id = next(
                (iid for iid, d in investor_options if d == selected_investor_display),
                None,
            )
            if selected_investor_id:
                matched = investor_to_labels.get(selected_investor_id, [])
                st.caption(
                    f"Investor **{selected_investor_display}** has exposure to {len(matched)} deal(s)."
                )
                selected_report_labels = matched

    else:  # All Deals
        selected_report_labels = eligible_labels
        st.caption(f"{len(eligible_labels)} deals with waterfall definitions will be included.")

    # --- Generate button ---
    if st.button("Generate Report", type="primary", key="generate_report_btn"):
        if not selected_report_labels:
            st.warning("No deals selected.")
            return

        all_rows = []
        errors = []

        progress = st.progress(0, text="Generating report...")

        for i, label in enumerate(selected_report_labels):
            progress.progress(
                (i + 1) / len(selected_report_labels),
                text=f"Processing {label}...",
            )

            # Resolve label → vcode + investment_id + sale_date
            match = inv_disp[inv_disp["DealLabel"] == label]
            if match.empty:
                errors.append(f"Could not find deal: {label}")
                continue

            row = match.iloc[0]
            vcode = str(row["vcode"])
            inv_id = str(row.get("InvestmentID", ""))
            sale_date_raw = row.get("Sale_Date", None)

            # Use cached result when available
            _deal_key = f"{vcode}|{start_year}|{horizon_years}|{pro_yr_base}"
            if st.session_state.get('_deal_cache_key') == _deal_key:
                result = st.session_state['_deal_cache']
            else:
                try:
                    result = compute_deal_analysis(
                        deal_vcode=vcode,
                        deal_investment_id=inv_id,
                        sale_date_raw=sale_date_raw,
                        inv=inv, wf=wf, acct=acct, fc=fc, coa=coa,
                        mri_loans_raw=mri_loans_raw,
                        mri_supp=mri_supp,
                        mri_val=mri_val,
                        relationships_raw=relationships_raw,
                        capital_calls_raw=capital_calls_raw,
                        isbs_raw=isbs_raw,
                        start_year=start_year,
                        horizon_years=horizon_years,
                        pro_yr_base=pro_yr_base,
                    )
                except Exception as e:
                    errors.append(f"{label}: {e}")
                    continue

            if 'error' in result:
                errors.append(f"{label}: {result['error']}")
                continue

            partner_rows = _build_partner_returns(result, label)
            all_rows.extend(partner_rows)

        progress.empty()

        if all_rows:
            report_df = pd.DataFrame(all_rows)
            st.session_state['_report_df'] = report_df
            st.session_state['_report_errors'] = errors
        else:
            st.warning("No partner data found for selected deals.")
            if errors:
                with st.expander("Errors"):
                    for err in errors:
                        st.text(err)
            return

    # --- Display results (persists across reruns) ---
    report_df = st.session_state.get('_report_df')
    errors = st.session_state.get('_report_errors', [])

    if report_df is not None and not report_df.empty:
        if errors:
            with st.expander(f"{len(errors)} deal(s) skipped"):
                for err in errors:
                    st.text(err)

        st.subheader("Projected Returns Summary")

        # Build display DataFrame (drop internal flag)
        display_df = report_df.drop(columns=['_is_deal_total'], errors='ignore')

        # Style: bold + top border for deal-total rows
        total_set = set(report_df.index[report_df['_is_deal_total'].fillna(False).astype(bool)])

        def _highlight_totals(row):
            if row.name in total_set:
                return ['font-weight: bold; border-top: 2px solid #333'] * len(row)
            return [''] * len(row)

        styled = display_df.style.format({
            'Contributions': '${:,.0f}',
            'CF Distributions': '${:,.0f}',
            'Capital Distributions': '${:,.0f}',
            'IRR': lambda v: f'{v:.2%}' if pd.notna(v) else 'N/A',
            'ROE': '{:.2%}',
            'MOIC': '{:.2f}x',
        }).apply(_highlight_totals, axis=1)

        st.dataframe(
            styled,
            use_container_width=True,
            hide_index=True,
        )

        # --- Download Excel ---
        excel_bytes = _generate_excel(report_df)
        st.download_button(
            label="Download Excel",
            data=excel_bytes,
            file_name="projected_returns.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
            key="download_projected_returns",
        )
