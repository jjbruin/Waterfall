"""
sold_portfolio_ui.py
Sold Portfolio tab UI — historical returns for sold deals computed from accounting_feed.

Unlike active deals that run compute_deal_analysis() (forecast waterfalls), sold deals
compute returns directly from accounting history: contributions, distributions, and
capital events.
"""

import streamlit as st
import pandas as pd
from io import BytesIO

from metrics import xirr, calculate_roe
from loaders import build_investmentid_to_vcode, normalize_accounting_feed


def render_sold_portfolio(inv, acct):
    """Render the Sold Portfolio tab contents.

    Delegates to a @st.fragment so widgets only rerun this tab.
    """
    _sold_portfolio_fragment(inv, acct)


@st.fragment
def _sold_portfolio_fragment(inv, acct):
    """Fragment-isolated Sold Portfolio body."""

    st.subheader("Sold Portfolio — Historical Returns")

    # --- Filter to sold deals ---
    if "Sale_Status" not in inv.columns:
        st.info("No sold deals found.")
        return

    inv_sold = inv[inv["Sale_Status"].fillna("").str.upper() == "SOLD"].copy()

    # Exclude child properties (same vectorised filter as _build_deal_lookup)
    if "Portfolio_Name" in inv_sold.columns:
        inv_sold["Investment_Name"] = inv_sold["Investment_Name"].fillna("").astype(str).str.strip()
        inv_sold["Portfolio_Name"] = inv_sold["Portfolio_Name"].fillna("").astype(str).str.strip()
        parent_names = set(inv.loc[inv["Sale_Status"].fillna("").str.upper() == "SOLD", "Investment_Name"]
                          .fillna("").astype(str).str.strip())
        is_child = (
            inv_sold["Portfolio_Name"].isin(parent_names)
            & (inv_sold["Portfolio_Name"] != inv_sold["Investment_Name"])
            & (inv_sold["Portfolio_Name"] != "")
        )
        inv_sold = inv_sold[~is_child].copy()

    if inv_sold.empty:
        st.info("No sold deals found.")
        return

    # --- Compute returns ---
    result_df = _compute_all_sold_returns(inv_sold, acct, inv)

    if result_df.empty:
        st.info("No accounting data found for sold deals.")
        return

    # --- Display styled table ---
    display_df = result_df.drop(columns=["_is_deal_total"], errors="ignore")

    total_set = set(result_df.index[result_df["_is_deal_total"].fillna(False).astype(bool)])

    def _highlight_totals(row):
        if row.name in total_set:
            return ["font-weight: bold; border-top: 2px solid #333"] * len(row)
        return [""] * len(row)

    styled = display_df.style.format({
        "Total Contributions": "${:,.0f}",
        "Total Distributions": "${:,.0f}",
        "IRR": lambda v: f"{v:.2%}" if pd.notna(v) else "N/A",
        "ROE": lambda v: f"{v:.2%}" if pd.notna(v) else "N/A",
        "MOIC": "{:.2f}x",
    }).apply(_highlight_totals, axis=1)

    st.dataframe(styled, use_container_width=True, hide_index=True)

    # --- Excel download ---
    excel_bytes = _generate_sold_excel(result_df)
    st.download_button(
        label="Download Excel",
        data=excel_bytes,
        file_name="sold_portfolio_returns.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        key="download_sold_portfolio",
    )

    # --- Deal detail drill-down ---
    st.markdown("---")
    deal_names = result_df.loc[~result_df["_is_deal_total"], "Investment Name"].tolist()
    selected_deal = st.selectbox(
        "Select deal to view activity detail",
        deal_names,
        key="sold_deal_detail_selector",
    )

    if selected_deal:
        detail_df = _build_deal_detail(selected_deal, inv_sold, acct, inv)
        if detail_df is not None and not detail_df.empty:
            # Pull the summary row for this deal
            deal_summary = result_df[
                (~result_df["_is_deal_total"])
                & (result_df["Investment Name"] == selected_deal)
            ]

            with st.expander(f"Financial Activity — {selected_deal}", expanded=True):
                st.dataframe(
                    detail_df.style.format({
                        "Amount": "${:,.2f}",
                        "Cashflow (XIRR)": "${:,.2f}",
                        "Capital Balance": "${:,.2f}",
                    }),
                    use_container_width=True,
                    hide_index=True,
                )

                # Show computed metrics below detail for reference
                if not deal_summary.empty:
                    row = deal_summary.iloc[0]
                    irr_str = f"{row['IRR']:.2%}" if pd.notna(row["IRR"]) else "N/A"
                    roe_str = f"{row['ROE']:.2%}" if pd.notna(row["ROE"]) else "N/A"
                    c1, c2, c3 = st.columns(3)
                    c1.metric("IRR", irr_str)
                    c2.metric("ROE", roe_str)
                    c3.metric("MOIC", f"{row['MOIC']:.2f}x")

                detail_xlsx = _generate_detail_excel(detail_df, selected_deal, deal_summary)
                st.download_button(
                    label="Download Activity Detail",
                    data=detail_xlsx,
                    file_name=f"sold_activity_{selected_deal.replace(' ', '_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
                    key="download_sold_detail",
                )
        else:
            st.info("No pref equity accounting activity found for this deal.")


def _compute_all_sold_returns(inv_sold, acct, inv):
    """Compute pref-equity returns for all sold deals from accounting.

    Only includes preferred equity members (InvestorID not starting with "OP").
    Groups InvestorIDs case-insensitively to handle mixed-case entity IDs.

    Returns a DataFrame with columns:
        Investment Name, Acquisition Date, Sale Date,
        Total Contributions, Total Distributions, IRR, ROE, MOIC, _is_deal_total
    """
    # Normalize accounting if needed (raw acct lacks is_contribution/is_distribution/is_capital)
    if "is_contribution" not in acct.columns:
        acct = normalize_accounting_feed(acct)

    inv_to_vcode = build_investmentid_to_vcode(inv)

    # Reverse map: vcode → list of InvestmentIDs
    vcode_to_iids = {}
    for iid, vc in inv_to_vcode.items():
        vcode_to_iids.setdefault(vc, []).append(iid)

    all_rows = []

    # Collect all pref cashflows across all deals for portfolio total
    portfolio_cashflows = []
    portfolio_capital_events = []
    portfolio_cf_distributions = []

    for _, deal_row in inv_sold.iterrows():
        deal_name = str(deal_row.get("Investment_Name", "")).strip()
        deal_vcode = str(deal_row.get("vcode", "")).strip()

        # Acquisition and sale dates from deals table
        acq_raw = deal_row.get("Acquisition_Date", None)
        sale_raw = deal_row.get("Sale_Date", None)
        acq_date = pd.to_datetime(acq_raw, errors="coerce")
        sale_date = pd.to_datetime(sale_raw, errors="coerce")
        acq_str = acq_date.strftime("%m/%d/%Y") if pd.notna(acq_date) else ""
        sale_str = sale_date.strftime("%m/%d/%Y") if pd.notna(sale_date) else ""

        # Find all InvestmentIDs for this deal's vcode (handles sub-portfolio children)
        deal_iids = set(vcode_to_iids.get(deal_vcode, []))
        # Also include the InvestmentID from this row directly
        direct_iid = str(deal_row.get("InvestmentID", "")).strip()
        if direct_iid:
            deal_iids.add(direct_iid)

        if not deal_iids or acct is None or acct.empty:
            continue

        # Filter accounting to this deal's InvestmentIDs
        deal_acct = acct[acct["InvestmentID"].isin(deal_iids)].copy()
        if deal_acct.empty:
            continue

        # Filter to preferred equity only (exclude OP partners)
        deal_acct = deal_acct[~deal_acct["InvestorID"].str.upper().str.startswith("OP")].copy()
        if deal_acct.empty:
            continue

        # Case-insensitive grouping: normalise to upper for grouping key
        deal_acct["_investor_key"] = deal_acct["InvestorID"].str.upper()

        # Aggregate all pref equity cashflows for the deal
        pref_cashflows = []
        pref_capital_events = []
        pref_cf_distributions = []

        for _, grp in deal_acct.groupby("_investor_key"):
            _, _, _, _, _, cashflows, capital_events, cf_dists = (
                _compute_partner_metrics(grp)
            )
            pref_cashflows.extend(cashflows)
            pref_capital_events.extend(capital_events)
            pref_cf_distributions.extend(cf_dists)

        if not pref_cashflows:
            continue

        contribs = sum(abs(a) for _, a in pref_cashflows if a < 0)
        distribs = sum(a for _, a in pref_cashflows if a > 0)
        irr_val = xirr(pref_cashflows) if len(pref_cashflows) >= 2 else None

        start = min(d for d, _ in pref_cashflows)
        end = max(d for d, _ in pref_cashflows)
        roe_val = calculate_roe(pref_capital_events, pref_cf_distributions, start, end)
        moic_val = distribs / contribs if contribs > 0 else 0.0

        all_rows.append({
            "Investment Name": deal_name,
            "Acquisition Date": acq_str,
            "Sale Date": sale_str,
            "Total Contributions": contribs,
            "Total Distributions": distribs,
            "IRR": irr_val,
            "ROE": roe_val,
            "MOIC": moic_val,
            "_is_deal_total": False,
        })

        # Accumulate for portfolio total
        portfolio_cashflows.extend(pref_cashflows)
        portfolio_capital_events.extend(pref_capital_events)
        portfolio_cf_distributions.extend(pref_cf_distributions)

    if not all_rows:
        return pd.DataFrame()

    # Portfolio total row — IRR/ROE/MOIC from combined cashflow pool
    port_contribs = sum(abs(a) for _, a in portfolio_cashflows if a < 0)
    port_distribs = sum(a for _, a in portfolio_cashflows if a > 0)
    port_irr = xirr(portfolio_cashflows) if len(portfolio_cashflows) >= 2 else None

    port_start = min(d for d, _ in portfolio_cashflows)
    port_end = max(d for d, _ in portfolio_cashflows)
    port_roe = calculate_roe(portfolio_capital_events, portfolio_cf_distributions, port_start, port_end)
    port_moic = port_distribs / port_contribs if port_contribs > 0 else 0.0

    all_rows.append({
        "Investment Name": "Portfolio Total",
        "Acquisition Date": "",
        "Sale Date": "",
        "Total Contributions": port_contribs,
        "Total Distributions": port_distribs,
        "IRR": port_irr,
        "ROE": port_roe,
        "MOIC": port_moic,
        "_is_deal_total": True,
    })

    return pd.DataFrame(all_rows)


def _compute_partner_metrics(partner_acct):
    """Compute metrics for a single partner from accounting data.

    Returns:
        (contributions, distributions, irr, roe, moic,
         cashflows, capital_events, cf_distributions)
    """
    cashflows = []
    capital_events = []
    cf_distributions = []

    for _, r in partner_acct.iterrows():
        d = r["EffectiveDate"]
        amt = float(r["Amt"])

        if r["is_contribution"]:
            cf = -abs(amt)  # negative = investment
            cashflows.append((d, cf))
            capital_events.append((d, cf))
        elif r["is_distribution"]:
            cf = abs(amt)   # positive = return
            cashflows.append((d, cf))
            if r["is_capital"]:
                capital_events.append((d, cf))  # capital return
            else:
                cf_distributions.append((d, cf))  # CF only (for ROE)

    contributions = sum(abs(a) for _, a in cashflows if a < 0)
    distributions = sum(a for _, a in cashflows if a > 0)

    irr_val = xirr(cashflows) if len(cashflows) >= 2 else None

    if cashflows:
        start = min(d for d, _ in cashflows)
        end = max(d for d, _ in cashflows)
        roe_val = calculate_roe(capital_events, cf_distributions, start, end)
    else:
        roe_val = 0.0

    moic_val = distributions / contributions if contributions > 0 else 0.0

    return contributions, distributions, irr_val, roe_val, moic_val, cashflows, capital_events, cf_distributions


def _build_deal_detail(deal_name, inv_sold, acct, inv):
    """Build a cashflow detail table for a single sold deal.

    Returns a DataFrame sorted by date with columns:
        Date, InvestorID, MajorType, Typename, Capital, Amount,
        Cashflow (XIRR), Capital Balance
    """
    if "is_contribution" not in acct.columns:
        acct = normalize_accounting_feed(acct)

    inv_to_vcode = build_investmentid_to_vcode(inv)
    vcode_to_iids = {}
    for iid, vc in inv_to_vcode.items():
        vcode_to_iids.setdefault(vc, []).append(iid)

    match = inv_sold[inv_sold["Investment_Name"].fillna("").astype(str).str.strip() == deal_name]
    if match.empty:
        return None

    deal_row = match.iloc[0]
    deal_vcode = str(deal_row.get("vcode", "")).strip()

    deal_iids = set(vcode_to_iids.get(deal_vcode, []))
    direct_iid = str(deal_row.get("InvestmentID", "")).strip()
    if direct_iid:
        deal_iids.add(direct_iid)

    if not deal_iids:
        return None

    deal_acct = acct[acct["InvestmentID"].isin(deal_iids)].copy()
    if deal_acct.empty:
        return None

    # Pref equity only
    deal_acct = deal_acct[~deal_acct["InvestorID"].str.upper().str.startswith("OP")].copy()
    if deal_acct.empty:
        return None

    # Build detail rows
    rows = []
    for _, r in deal_acct.iterrows():
        amt = float(r["Amt"])
        if r["is_contribution"]:
            cashflow = -abs(amt)
        elif r["is_distribution"]:
            cashflow = abs(amt)
        else:
            cashflow = 0.0

        rows.append({
            "Date": r["EffectiveDate"],
            "InvestorID": r["InvestorID"],
            "MajorType": r["MajorType"],
            "Typename": r.get("Typename", ""),
            "Capital": r["Capital"],
            "Amount": amt,
            "Cashflow (XIRR)": cashflow,
        })

    if not rows:
        return None

    df = pd.DataFrame(rows).sort_values("Date").reset_index(drop=True)

    # Running capital balance: contributions increase, capital returns decrease
    balance = 0.0
    balances = []
    for _, r in df.iterrows():
        cf = r["Cashflow (XIRR)"]
        if cf < 0:
            # contribution → increases capital outstanding
            balance += abs(cf)
        elif r["Capital"] == "Y" and cf > 0:
            # capital return → decreases capital outstanding
            balance = max(0.0, balance - cf)
        balances.append(balance)
    df["Capital Balance"] = balances

    return df


def _generate_detail_excel(detail_df, deal_name, deal_summary):
    """Create an Excel workbook with the deal activity detail and summary metrics.

    Returns:
        Bytes suitable for st.download_button.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Detail"

    cols = list(detail_df.columns)

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    currency_cols = {"Amount", "Cashflow (XIRR)", "Capital Balance"}
    date_cols = {"Date"}

    for ri, (_, row) in enumerate(detail_df.iterrows(), start=2):
        for ci, col in enumerate(cols, 1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci)
            if col in date_cols:
                cell.value = val
                cell.number_format = "MM/DD/YYYY"
            elif col in currency_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = '$#,##0.00'
            else:
                cell.value = val

    # Summary section below the data
    summary_row = len(detail_df) + 3
    bold = Font(bold=True)
    ws.cell(row=summary_row, column=1, value="Computed Returns").font = bold

    if deal_summary is not None and not deal_summary.empty:
        sr = deal_summary.iloc[0]
        labels = [
            ("Total Contributions", sr.get("Total Contributions", 0), '$#,##0'),
            ("Total Distributions", sr.get("Total Distributions", 0), '$#,##0'),
            ("IRR", sr.get("IRR"), '0.00%'),
            ("ROE", sr.get("ROE"), '0.00%'),
            ("MOIC", sr.get("MOIC", 0), '0.00"x"'),
        ]
        for i, (label, val, fmt) in enumerate(labels):
            r = summary_row + 1 + i
            ws.cell(row=r, column=1, value=label).font = bold
            cell = ws.cell(row=r, column=2)
            cell.value = float(val) if pd.notna(val) else None
            cell.number_format = fmt

    # Auto-width
    for ci, col in enumerate(cols, 1):
        max_len = len(str(col))
        for ri in range(2, len(detail_df) + 2):
            cv = ws.cell(row=ri, column=ci).value
            if cv is not None:
                max_len = max(max_len, len(str(cv)))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 4, 30)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_sold_excel(df):
    """Create a formatted Excel workbook for sold portfolio returns.

    Returns:
        Bytes suitable for st.download_button.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Sold Portfolio Returns"

    display_cols = [c for c in df.columns if c != "_is_deal_total"]

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, col_name in enumerate(display_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    bold_font = Font(bold=True)
    top_border = Border(top=Side(style="medium"))

    currency_cols = {"Total Contributions", "Total Distributions"}
    pct_cols = {"IRR", "ROE"}
    moic_cols = {"MOIC"}

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        is_total = bool(row.get("_is_deal_total", False))

        for col_idx, col_name in enumerate(display_cols, start=1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name in currency_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = "$#,##0"
            elif col_name in pct_cols:
                cell.value = float(val) if pd.notna(val) else None
                cell.number_format = "0.00%"
            elif col_name in moic_cols:
                cell.value = float(val) if pd.notna(val) else 0.0
                cell.number_format = '0.00"x"'
            else:
                cell.value = val

            if is_total:
                cell.font = bold_font
                cell.border = top_border

    # Auto-width
    for col_idx, col_name in enumerate(display_cols, start=1):
        max_len = len(str(col_name))
        for row_idx in range(2, len(df) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
